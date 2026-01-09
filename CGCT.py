"""
COORDINATE CONVENTION DOCUMENTATION
====================================

This script uses 1-indexed (1-based), inclusive coordinates throughout:

COORDINATE SYSTEM STANDARDS:
  - GFF3 (Generic Feature Format v3): 1-indexed, inclusive [start, end]
    https://github.com/The-Sequence-Ontology/Specifications/blob/master/gff3.md
  - GenBank (GBK): 0-indexed, half-open [start, end)
  - BED: 0-indexed, half-open [start, end)
  - This program: 1-indexed, inclusive (start, end)

STORAGE & PROCESSING:
  - GFF/GBK features: Converted to 1-indexed inclusive on import
    * GenBank: 0-indexed half-open → converted (start+1, end) ✓
    * GFF3: Already 1-indexed; verified and offset applied if gffutils returns 0-indexed
    * Validation: All intergenic regions checked against GFF features for overlaps
  - Custom labels: Stored as 1-indexed (start, end) tuples
  - Subsequence positions: Input as 1-indexed from UI
  - BED files: 0-indexed half-open [start, end) - converted for overlap detection
  - Mauve LCBs: 0-indexed → converted to 1-indexed via _apply_mauve_coordinate_conversion()

DISPLAY (User-Facing):
  - All tooltips show 1-indexed coordinates
  - All labels show 1-indexed coordinates
  - All exported files show 1-indexed coordinates in headers

INTERNAL ARRAYS (Python):
  - Sequence identity scores: 0-indexed (align with Python string indexing)
  - When displaying: convert to 1-indexed

EXPORT:
  - FASTA subsequence headers: Shows 1-indexed range
  - Excel reports: All coordinates shown as 1-indexed inclusive
  - GFF output: All coordinates are 1-indexed inclusive (standard)

KEY CONVERSION POINTS:
  - User input → 1-indexed (validate >= 1)
  - Python slicing ← 0-indexed (e.g., seq[0:10] for positions 1-10)
  - GraphicFeature ← 1-indexed (dna_features_viewer expects 1-indexed)
  - Display/tooltip ← 1-indexed (as stored)
  - Mauve coords ← Apply +1 to start (0-indexed → 1-indexed)
"""
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
from matplotlib import patches as mpatches
from matplotlib.patches import PathPatch
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.collections import LineCollection
from matplotlib.backend_bases import MouseEvent
from Bio import SeqIO
from Bio.Seq import Seq
from dna_features_viewer import CircularGraphicRecord, GraphicFeature
import subprocess, threading, traceback, warnings, queue, json, os, sys, datetime, time, random, io, platform, glob
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.path as mpath
import matplotlib.transforms as mtransforms
import matplotlib.colors as mcolors
import matplotlib.cm as cm
import matplotlib.collections as mcoll
import numpy as np
import pandas as pd
import gffutils, Bio, shutil, tempfile, shlex, openpyxl
from collections import defaultdict
from openpyxl import load_workbook
import tkinter.font as tkfont

matplotlib.use("TkAgg")
warnings.filterwarnings("ignore", category=Bio.BiopythonDeprecationWarning)
warnings.filterwarnings("ignore", message=".*main thread is not in main loop.*")

# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# Subprocess Helper for Windows GUI (No Console Popups)
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
def run_quiet(args, **kwargs):
    """
    Runs a subprocess without flashing a console window on Windows packaged GUI apps.
    Works fine on macOS/Linux too. Doesn't hide WSL popup.
    """
    if platform.system() == "Windows":
        # CREATE_NO_WINDOW prevents the console popup
        kwargs.setdefault("creationflags", subprocess.CREATE_NO_WINDOW)
        # Hide the window startup info
        si = subprocess.STARTUPINFO()
        si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        si.wShowWindow = subprocess.SW_HIDE
        kwargs.setdefault("startupinfo", si)
    return subprocess.run(args, **kwargs)

GFFUTILS_COORD_OFFSET_GLOBAL = 0

# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# Coordinate Validation Functions
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────

def _fix_intergenic_boundaries(intergenic_data, gff_features):
    """
    Fix boundary overlaps: if an intergenic region ends at position X and a feature
    starts at position X, reduce the intergenic end to X-1 (exclude the boundary).

    Both coordinates assumed to be 1-indexed inclusive.
    """
    feature_starts = {int(f.start) for f in gff_features}
    fixed_count = 0
    for ig_fragment in intergenic_data:
        ig_end = ig_fragment.get("query_end", 0)
        if ig_end in feature_starts and ig_end > 1:
            ig_fragment["query_end"] = ig_end - 1
            fixed_count += 1

    return fixed_count


def _validate_intergenic_no_gff_overlap(intergenic_data, gff_features, log_queue=None):
    """
    Validate that intergenic regions don't overlap with GFF features.

    COORDINATE SYSTEMS:
    - intergenic_data: query_start, query_end are 1-indexed inclusive (from string search in FASTA)
    - gff_features: .start, .end should be 1-indexed inclusive after load_gene_features() conversion
      * If GFFUTILS_COORD_OFFSET_GLOBAL == 0: GFF3 format (features were already 1-indexed)
      * If GFFUTILS_COORD_OFFSET_GLOBAL == 1: old GFF format (features were 0-indexed, now 1-indexed)

    Both coordinate systems are compared as 1-indexed inclusive.
    - Ranges [a,b] and [c,d] overlap if: a <= d AND b >= c
    - Touching boundary: only 1bp overlap (feature_start == intergenic_end, etc.)

    Args:
        intergenic_data: List of dicts with 'query_start', 'query_end' (1-indexed inclusive)
        gff_features: List of feature objects with .start and .end (1-indexed inclusive after loading)
        log_queue: Optional queue for status messages

    Returns:
        (is_valid, problems_list) where is_valid=False only if TRUE overlaps (>1bp) exist
    """
    global GFFUTILS_COORD_OFFSET_GLOBAL

    def status(msg, color="black"):
        if log_queue: log_queue.put(("status", (msg, color)))

    problems = []
    gff_intervals = [(int(f.start), int(f.end)) for f in gff_features]

    print("\n[VALIDATION] Coordinate system info:")
    print("  - Intergenic regions: 1-indexed inclusive (from FASTA string search)")
    if GFFUTILS_COORD_OFFSET_GLOBAL == 0:
        print("  - GFF features: 1-indexed inclusive (GFF3 format, no conversion needed)")
    else:
        print("  - GFF features: 1-indexed inclusive (old GFF format, converted from 0-indexed)")
    print(f"  - Checking {len(intergenic_data)} intergenic vs {len(gff_intervals)} features")
    print("  - Boundary touches (1bp) are flagged as WARNING, not CRITICAL\n")

    overlap_count, touching_count = 0, 0

    for idx, ig_region in enumerate(intergenic_data):
        ig_start, ig_end = int(ig_region.get("query_start", 0)), int(ig_region.get("query_end", 0))

        if ig_start <= 0 or ig_end <= 0: continue

        for feat_idx, (feat_start, feat_end) in enumerate(gff_intervals):
            if ig_start <= feat_end and ig_end >= feat_start:
                if ig_end == feat_start or ig_start == feat_end:
                    touching_count += 1
                    problems.append({
                        "intergenic_idx": idx,
                        "intergenic_range": (ig_start, ig_end),
                        "feature_idx": feat_idx,
                        "feature_range": (feat_start, feat_end),
                        "overlap_start": max(ig_start, feat_start),
                        "overlap_end": min(ig_end, feat_end),
                        "overlap_length": 1,
                        "is_touching": True
                    })
                else:
                    overlap_count += 1
                    problems.append({
                        "intergenic_idx": idx,
                        "intergenic_range": (ig_start, ig_end),
                        "feature_idx": feat_idx,
                        "feature_range": (feat_start, feat_end),
                        "overlap_start": max(ig_start, feat_start),
                        "overlap_end": min(ig_end, feat_end),
                        "overlap_length": min(ig_end, feat_end) - max(ig_start, feat_start) + 1,
                        "is_touching": False
                    })

    if overlap_count > 0:
        msg = f"[VALIDATION FAILED] Found {overlap_count} TRUE OVERLAPS + {touching_count} BOUNDARY TOUCHES"
        print(msg)
        status(msg, "red")

        true_overlaps = [p for p in problems if not p.get("is_touching", False)]
        if true_overlaps:
            print("\n  TRUE OVERLAPS (intergenic includes coding positions):")
            for p in true_overlaps:
                detail = (
                    f"    Intergenic {p['intergenic_range']} overlaps feature {p['feature_range']} "
                    f"at positions {p['overlap_start']}-{p['overlap_end']} ({p['overlap_length']}bp)"
                )
                print(detail)
                status(detail, "red")
        touching = [p for p in problems if p.get("is_touching", False)]
        if touching:
            print(f"\n  BOUNDARY TOUCHES ({len(touching)} intergenic regions end exactly where features begin):")
            for p in touching:
                detail = (
                    f"    Intergenic {p['intergenic_range']} ends where feature {p['feature_range']} begins "
                    f"at position {p['overlap_start']}"
                )
                print(detail)
                status(detail, "orange")
            print("\n  ⚠ Boundary touches indicate intergenic spans should exclude their end position")

    elif touching_count > 0:
        msg = f"[VALIDATION WARNING] Found {touching_count} BOUNDARY TOUCHES (no true overlaps)"
        print(msg)
        status(msg, "orange")
        touching = [p for p in problems if p.get("is_touching", False)]
        for p in touching:
            detail = (
                f"    Intergenic {p['intergenic_range']} touches feature {p['feature_range']} "
                f"at position {p['overlap_start']}"
            )
            print(detail)
            status(detail, "orange")
        print("\n  ⚠ These touches indicate intergenic spans should exclude their end position")
    else:
        msg = f"[VALIDATION OK] All {len(intergenic_data)} intergenic regions are properly non-overlapping"
        print(msg)
        status(msg, "green")

    print()
    return overlap_count == 0, problems

# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# Configuration File Handler
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────

def _load_config():
    """
    Load configuration from config.txt in the script's directory.
    Returns a dictionary with config values, using defaults/auto-detection for empty values.
    """
    config = {
        "WSL_DISTRO": "Ubuntu",
        "WSL_USER": "",
        "WSL_SIBELIAZ_ENV": "",
        "USE_DOCKER_SIBELIAZ": "false",
        "DOCKER_SIBELIAZ_IMAGE": "aaronmauve/sibeliaz-conda:latest"
    }

    script_dir = Path(__file__).parent
    config_file = script_dir / "config.txt"

    if config_file.exists():
        try:
            with open(config_file, 'r') as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith('#'): continue

                    if '=' in line:
                        key, value = line.split('=', 1)
                        key = key.strip()
                        value = value.strip()

                        if key in config and value: config[key] = value
            print(f"[CONFIG] Loaded configuration from {config_file}")
        except Exception as e: print(f"[CONFIG] Error reading config file {config_file}: {e}")
    else: print(f"[CONFIG] Config file not found at {config_file}, using defaults")
    return config


def _create_bounded_entry_validator(var, min_val, max_val, param_name, is_float=False, is_odd=False):
    """Create a focus-out validator function for a bounded entry."""

    def validator(event=None):
        try:
            if is_float: val = float(var.get())
            else: val = int(var.get())

            if is_odd and val % 2 == 0:
                messagebox.showwarning(
                    "Invalid Value",
                    f"{param_name} must be an odd number.\nEntered: {val}",
                    parent=None
                )
                return

            if val < min_val or val > max_val:
                messagebox.showwarning(
                    "Out of Range",
                    f"{param_name}: {val}\nAcceptable range: {min_val} - {max_val}",
                    parent=None
                )
        except (ValueError, tk.TclError): pass
    return validator


def _create_positive_validator(var, param_name, is_float=False):
    """Create a focus-out validator for positive values."""

    def validator(event=None):
        try:
            val = float(var.get()) if is_float else int(var.get())
            if val <= 0:
                messagebox.showwarning(
                    "Invalid Value",
                    f"{param_name} must be positive.\nEntered: {val}",
                    parent=None)
        except (ValueError, tk.TclError): pass
    return validator


def _create_nonnegative_validator(var, param_name, is_float=False):
    """Create a focus-out validator for non-negative values."""
    def validator(event=None):
        try:
            val = float(var.get()) if is_float else int(var.get())
            if val < 0:
                messagebox.showwarning(
                    "Invalid Value",
                    f"{param_name} must be non-negative.\nEntered: {val}",
                    parent=None)
        except (ValueError, tk.TclError): pass
    return validator

class SafeNavigationToolbar2Tk(NavigationToolbar2Tk):
    """Custom toolbar that prevents AttributeError when _zoom_info is None."""
    def drag_zoom(self, event):
        """Override drag_zoom to safely handle None _zoom_info while preserving visual feedback."""
        if self._zoom_info is None: return
        try: super().drag_zoom(event)
        except AttributeError as e:
            if "_zoom_info" in str(e): return
            else: raise

# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# Globals & Configuration
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────

RESOLUTION_OPTIONS = {
    "Full HD (1920x1080)": (1920, 1080),
    "2K QHD (2560x1440)": (2560, 1440),
    "4K UHD (3840x2160)": (3840, 2160),
    "6K (6144x3240)": (6144, 3240),
    "8K UHD (7680x4320)": (7680, 4320),
    "12K (11520x6480)": (11520, 6480),
    "16K (15360x8640)": (15360, 8640),
    "20K (19200x10800)": (19200, 10800),
    "Custom Resolution": None
}

def _get_cpu_count():
    try: return str(os.cpu_count() or 1)
    except Exception: return "1"

BLAST_DEFAULTS = {
    "megablast": {
        "word_size": "28", "reward": "1", "penalty": "-2", "evalue": "10",
        "max_target_seqs": "500", "num_threads": _get_cpu_count(), "gapopen": "", "gapextend": "",
        "dust": "yes", "soft_masking": "true", "strand": "both", "perc_identity": "", "culling_limit": ""
    },
    "blastn": {
        "word_size": "11", "reward": "2", "penalty": "-3", "evalue": "10",
        "max_target_seqs": "500", "num_threads": _get_cpu_count(), "gapopen": "5", "gapextend": "2",
        "dust": "yes", "soft_masking": "true", "strand": "both", "perc_identity": "", "culling_limit": ""
    },
    "dc-megablast": {
        "word_size": "11", "reward": "2", "penalty": "-3", "evalue": "10",
        "max_target_seqs": "500", "num_threads": _get_cpu_count(), "gapopen": "5", "gapextend": "2",
        "dust": "yes", "soft_masking": "true", "strand": "both", "perc_identity": "", "culling_limit": ""
    },
    "blastn-short": {
        "word_size": "7", "reward": "1", "penalty": "-3", "evalue": "1000",
        "max_target_seqs": "500", "num_threads": _get_cpu_count(), "gapopen": "5", "gapextend": "2",
        "dust": "no", "soft_masking": "true", "strand": "both", "perc_identity": "", "culling_limit": ""
    }
}

# ──────────────────────────────────────────────────────────────────────────────
# WSL launcher (Windows-only) for SibeliaZ
# ──────────────────────────────────────────────────────────────────────────────

_cfg = _load_config()
WSL_DISTRO = _cfg["WSL_DISTRO"]

def _detect_wsl_user():
    """
    Attempt to automatically detect the WSL user.
    """
    # Build command based on whether a specific distro is defined
    cmd = ["wsl"]
    if WSL_DISTRO: cmd.extend(["-d", WSL_DISTRO])
    cmd.append("whoami")

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=5)
        if result.returncode == 0:
            user = result.stdout.strip().split("\\")[-1]
            if user:
                print(f"[WSL] Detected user: {user}")
                return user
    except Exception as e: print(f"[WSL] Could not auto-detect user: {e}")

    # Fallback guessing logic
    common_defaults = ["root", "user", "ubuntu"]
    for guess_user in common_defaults:
        try:
            # Build command again for the guess
            guess_cmd = ["wsl"]
            if WSL_DISTRO: guess_cmd.extend(["-d", WSL_DISTRO])
            guess_cmd.extend(["-u", guess_user, "whoami"])

            result = subprocess.run(guess_cmd, capture_output=True, text=True, timeout=5)
            if result.returncode == 0:
                print(f"[WSL] Using best guess username: {guess_user}")
                return guess_user
        except Exception: continue

    print("[WSL] Could not determine user, using 'root' as fallback")
    return "root"

if platform.system() == "Windows":
    # Simply load what is in the config. If empty, we will let WSL decide.
    WSL_USER = _cfg["WSL_USER"]
    WSL_SIBELIAZ_ENV = _cfg["WSL_SIBELIAZ_ENV"]
    USE_DOCKER_SIBELIAZ = _cfg["USE_DOCKER_SIBELIAZ"].lower() in ["true", "1", "yes"]
    DOCKER_SIBELIAZ_IMAGE = _cfg["DOCKER_SIBELIAZ_IMAGE"] or "sibeliaz-conda:latest"
else:
    WSL_USER, WSL_SIBELIAZ_ENV = None, None
    USE_DOCKER_SIBELIAZ = False
    DOCKER_SIBELIAZ_IMAGE = "sibeliaz-conda:latest"

# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# External Tool Wrappers (WGA & BLAST)
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────

def _win_to_wsl_path(p: str) -> str:
    """Convert Windows path to WSL mount path (/mnt/<drive>/...)."""
    p = os.path.abspath(p)
    drive, rest = p[0].lower(), p[2:].replace("\\", "/")
    return f"/mnt/{drive}{rest}"


def _run_in_wsl_bash(cmd: str, timeout: int | None = None):
    """
    Run a command in WSL bash with debug logging.
    """
    full = ["wsl"]

    # Only append distro/user if explicitly set
    if WSL_DISTRO: full.extend(["-d", WSL_DISTRO])
    if WSL_USER: full.extend(["-u", WSL_USER])

    full.extend(["bash", "-l", "-c", cmd])

    print(f"[DEBUG] WSL Executing: {' '.join(full)}")

    try: proc = subprocess.run(full, capture_output=True, text=True, timeout=timeout)
    except FileNotFoundError as e:
        print(f"[FILE_NOT_FOUND] WSL subprocess.run: {e}")
        raise
    except OSError as e:
        print(f"[OSERROR] WSL subprocess.run: {e}")
        raise
    except subprocess.TimeoutExpired as e:
        print(f"[TIMEOUT] WSL command timed out after {e.timeout} seconds")
        raise

    return proc.returncode, proc.stdout, proc.stderr


def _run_sibeliaz_wsl(query_fa_win: str, ref_fa_win: str, out_dir_win: str, extra_args: list[str] | None = None,
                      timeout: int | None = None):
    """
    Call SibeliaZ inside WSL.
    Explicitly attempts to source Conda and activate 'base' (or specified env)
    because non-interactive shells often skip .bashrc PATH updates.
    """
    try:
        safe_query_path_win = os.path.join(out_dir_win, "query.fasta")
        safe_ref_path_win = os.path.join(out_dir_win, "ref.fasta")

        try: shutil.copy(query_fa_win, safe_query_path_win)
        except Exception as e: raise RuntimeError(f"Failed copy query: {e}")

        try: shutil.copy(ref_fa_win, safe_ref_path_win)
        except Exception as e: raise RuntimeError(f"Failed copy ref: {e}")

    except Exception as e: raise RuntimeError(f"Failed to copy FASTA files to temp directory: {e}")

    q_wsl = _win_to_wsl_path(safe_query_path_win)
    r_wsl = _win_to_wsl_path(safe_ref_path_win)
    o_wsl = _win_to_wsl_path(out_dir_win)

    sibeliaz_command_parts = ["sibeliaz", "-o", o_wsl]
    sibeliaz_command_parts.extend([q_wsl, r_wsl])
    if extra_args: sibeliaz_command_parts.extend(extra_args)
    sibeliaz_cmd_str = " ".join(shlex.quote(arg) for arg in sibeliaz_command_parts)

    conda_locations = [
        "~/miniconda3/etc/profile.d/conda.sh",
        "~/anaconda3/etc/profile.d/conda.sh",
        "/opt/conda/etc/profile.d/conda.sh"
    ]

    source_chain = " || ".join([f"source {loc} 2>/dev/null" for loc in conda_locations]) + " || true"

    if WSL_SIBELIAZ_ENV:
        bash = (
            f"{source_chain} && "
            f"conda activate {shlex.quote(WSL_SIBELIAZ_ENV)} && "
            f"{sibeliaz_cmd_str}"
        )
    else:
        bash = (
            f"{source_chain}; "
            f"conda activate base 2>/dev/null || true; "
            f"{sibeliaz_cmd_str}"
        )

    rc, so, se = _run_in_wsl_bash(bash, timeout=timeout)

    if rc != 0:
        if rc == 127:
            raise RuntimeError(
                f"SibeliaZ was not found in WSL (rc=127).\n"
                f"The script attempted to source Conda and activate 'base' but failed.\n"
                f"Please ensure 'sibeliaz' is installed and accessible.\n"
                f"STDERR:\n{se}\nSTDOUT:\n{so}")
        raise RuntimeError(f"SibeliaZ (WSL) failed (rc={rc}).\nSTDERR:\n{se}\nSTDOUT:\n{so}")
    return so, se


def _run_sibeliaz_docker(query_fasta, ref_fasta, output_dir, extra_args=None, timeout=None):
    """
    Runs SibeliaZ inside Docker (Mac-friendly).
    Pulls public image, supports Apple Silicon (M1/M2), and respects active Docker context.
    """
    print("\n" + "="*80)
    print("ENTERING _run_sibeliaz_docker")
    print("="*80)
    print(f"[DEBUG SIBELIAZ] Input parameters:")
    print(f"  query_fasta: {query_fasta}")
    print(f"  ref_fasta: {ref_fasta}")
    print(f"  output_dir: {output_dir}")
    print(f"  extra_args: {extra_args}")
    print(f"  timeout: {timeout}")

    # 1. PATH Setup
    print("\n[DEBUG SIBELIAZ] Setting up PATH environment...")
    original_path = os.environ.get("PATH", "")
    print(f"  Original PATH: {original_path[:150]}...")

    os.environ["PATH"] = (
            "/usr/local/bin:/opt/homebrew/bin:/Applications/Docker.app/Contents/MacOS:"
            + os.environ.get("PATH", "")
    )
    print(f"  Updated PATH: {os.environ['PATH'][:150]}...")

    # 2. Detect Architecture
    arch = platform.machine().lower()
    print(f"\n[DEBUG SIBELIAZ] Detected system architecture: {arch}")
    print(f"[DEBUG SIBELIAZ] Platform system: {platform.system()}")
    print(f"[DEBUG SIBELIAZ] Platform release: {platform.release()}")

    # 3. Find Docker Executable
    print(f"\n[DEBUG SIBELIAZ] Searching for Docker executable...")
    try:
        docker_path = shutil.which("docker")
        print(f"[DEBUG SIBELIAZ] shutil.which('docker') returned: {docker_path}")
    except Exception as e:
        print(f"[ERROR SIBELIAZ] shutil.which(docker): {type(e).__name__}: {e}")
        traceback.print_exc()
        raise

    if not docker_path:
        print("[ERROR SIBELIAZ] docker executable not found in system PATH")
        raise FileNotFoundError("Docker executable not found — ensure Docker Desktop/Colima/OrbStack is running.")
    print(f"[DEBUG SIBELIAZ] Using Docker binary: {docker_path}")

    # 4. Prepare Paths
    print(f"\n[DEBUG SIBELIAZ] Preparing file paths...")
    mount_base = os.path.dirname(query_fasta)
    query_name = os.path.basename(query_fasta)
    ref_name = os.path.basename(ref_fasta)
    print(f"  mount_base: {mount_base}")
    print(f"  query_name: {query_name}")
    print(f"  ref_name: {ref_name}")
    print(f"  Files exist check:")
    print(f"    query_fasta exists: {os.path.exists(query_fasta)}")
    print(f"    ref_fasta exists: {os.path.exists(ref_fasta)}")
    print(f"    output_dir exists: {os.path.exists(output_dir)}")

    # 5. Construct Command inside container
    print(f"\n[DEBUG SIBELIAZ] Constructing SibeliaZ command...")
    sibeliaz_command_parts = ["sibeliaz", "-o", "/workspace/output"]
    sibeliaz_command_parts.extend([f"/workspace/input/{query_name}", f"/workspace/input/{ref_name}"])
    if extra_args:
        sibeliaz_command_parts.extend(extra_args)
        print(f"  Added extra args: {extra_args}")
    sibeliaz_cmd_str = " ".join(shlex.quote(arg) for arg in sibeliaz_command_parts)
    print(f"  SibeliaZ command string: {sibeliaz_cmd_str}")

    # Docker image and command setup - using config variable or default
    docker_image = DOCKER_SIBELIAZ_IMAGE
    print(f"\n[DEBUG SIBELIAZ] Using Docker image: {docker_image}")

    # Set Docker context for macOS
    os.environ["DOCKER_CONTEXT"] = "desktop-linux"
    print(f"[DEBUG SIBELIAZ] Set DOCKER_CONTEXT to: desktop-linux")

    # Build the Docker Run Command
    print(f"[DEBUG SIBELIAZ] Building Docker run command...")
    cmd = [
        docker_path, "run", "--rm", "--pull", "never",

        # Mounts
        "-v", f"{mount_base}:/workspace/input",
        "-v", f"{output_dir}:/workspace/output",

        # Image - using local prebuilt image
        docker_image,

        "bash", "-c",
        (
            "set -e;"
            f"{sibeliaz_cmd_str};"
            "echo '=== SibeliaZ finished successfully inside Docker ==='"
        )
    ]
    print(f"[DEBUG SIBELIAZ] Docker mounts:")
    print(f"  Input mount: {mount_base} -> /workspace/input")
    print(f"  Output mount: {output_dir} -> /workspace/output")
    print("\n" + "=" * 40)
    print(f"DEBUG: Full Docker Command:\n{' '.join(cmd)}")
    print("=" * 40 + "\n")

    try:
        print(f"\n[DEBUG SIBELIAZ] Attempting to pull and run image: {docker_image}")
        print(f"[DEBUG SIBELIAZ] Subprocess timeout set to: {timeout or 3600} seconds")
        try:
            print(f"[DEBUG SIBELIAZ] Executing subprocess.run...")
            print(f"[DEBUG SIBELIAZ] Full command: {' '.join(cmd)}")
            result = run_quiet(cmd, capture_output=True, text=True, timeout=timeout or 3600)
            print(f"[DEBUG SIBELIAZ] subprocess.run completed")
            print("\n=== DEBUG: RAW STDOUT ===")
            print(result.stdout)
            print("=========================\n")

            print("\n=== DEBUG: RAW STDERR ===")
            print(result.stderr)
            print("=========================\n")
        except FileNotFoundError as e:
            print(f"[FILE_NOT_FOUND] subprocess.run(docker): {e}")
            raise
        except OSError as e:
            print(f"[OSERROR] subprocess.run(docker): {e}")
            raise

        print("\n=== Docker process completed ===")
        print(f"Exit code: {result.returncode}")

        if result.stdout.strip():
            print("=== STDOUT (trimmed) ===")
            print("\n".join(result.stdout.splitlines()[-20:]))
        else:
            print("STDOUT: [empty]")

        if result.stderr.strip():
            print("=== STDERR (trimmed) ===")
            print("\n".join(result.stderr.splitlines()[-20:]))

        if result.returncode != 0:
            print(f"[ERROR SIBELIAZ] Docker exited with non-zero code: {result.returncode}")
            raise RuntimeError(f"Docker exited with code {result.returncode}")

        try:
            output_files = [f for f in os.listdir(output_dir) if os.path.isfile(os.path.join(output_dir, f))]
        except Exception as e:
            print(f"[ERROR] listing output directory: {e}")
            output_files = []

        if output_files:
            print("Generated output files:")
            for f in output_files: print(" -", f)
        else:
            print("WARNING: No output files generated by Docker run.")

    except subprocess.TimeoutExpired:
        raise TimeoutError("SibeliaZ Docker execution timed out.")
    except Exception as e:
        raise RuntimeError(f"Docker execution crashed: {e}")


def _run_progressive_mauve(query_fasta: Path, ref_fasta: Path, output_xmfa_path: Path, mauve_options=None):
    """
    Runs progressiveMauve using a local binary located in a Tools folder.
    """
    if mauve_options is None: mauve_options = {}
    output_xmfa_path.parent.mkdir(exist_ok=True)

    # 1. Detect System and Construct Path
    system = platform.system().lower()
    script_dir = Path(__file__).parent.absolute()

    exe_name = "progressiveMauve"
    if system == "windows":
        exe_name = "progressiveMauve.exe"

    # Look for the executable in the Tools folder
    tools_dir = script_dir / "Tools"
    mauve_exe_path = tools_dir / exe_name

    print(f"\n[DEBUG MAUVE] Detected System: {system}")
    print(f"[DEBUG MAUVE] Tools folder: {tools_dir}")
    print(f"[DEBUG MAUVE] Selected Mauve binary path: {mauve_exe_path}")
    print(f"[DEBUG MAUVE] Binary exists: {mauve_exe_path.exists()}")
    print(f"[DEBUG MAUVE] Tools folder exists: {tools_dir.exists()}")

    if not mauve_exe_path.exists():
        raise FileNotFoundError(
            f"progressiveMauve executable not found.\n"
            f"Looked for: {mauve_exe_path}\n"
            f"Please ensure the 'Tools' folder exists in the same directory as this program, containing '{exe_name}'."
        )

    # 2. Ensure Execution Permissions (Linux/Mac)
    if system != "windows":
        print(f"\n[DEBUG MAUVE] Checking execution permissions...")
        try:
            current_perms = os.stat(mauve_exe_path).st_mode
            print(f"[DEBUG MAUVE] Current permissions: {oct(current_perms)}")
            print(f"[DEBUG MAUVE] Has execute permission: {bool(current_perms & os.X_OK)}")
            if not (current_perms & os.X_OK):
                print(f"[DEBUG MAUVE] Setting execute permission on {mauve_exe_path}")
                os.chmod(mauve_exe_path, 0o755)
                new_perms = os.stat(mauve_exe_path).st_mode
                print(f"[DEBUG MAUVE] New permissions: {oct(new_perms)}")
        except Exception as e:
            print(f"[WARNING MAUVE] Could not set execute permissions on {mauve_exe_path}: {e}")
            traceback.print_exc()

    # 4. Construct Command (rest of your function remains the same)
    base_cmd_parts = [str(mauve_exe_path), f"--output={str(output_xmfa_path)}"]

    if mauve_options.get("disable_backbone"): base_cmd_parts.append("--disable-backbone")
    if mauve_options.get("mums"): base_cmd_parts.append("--mums")
    if mauve_options.get("collinear"): base_cmd_parts.append("--collinear")
    if mauve_options.get("skip_refinement"): base_cmd_parts.append("--skip-refinement")
    if mauve_options.get("skip_gapped_alignment"): base_cmd_parts.append("--skip-gapped-alignment")
    if mauve_options.get("seed_family"): base_cmd_parts.append("--seed-family")
    if mauve_options.get("no_weight_scaling"): base_cmd_parts.append("--no-weight-scaling")
    if mauve_options.get("mem_clean"): base_cmd_parts.append("--mem-clean")

    # Integer/Float options
    for key, val in mauve_options.items():
        if not val or key in ["disable_backbone", "mums", "collinear", "skip_refinement",
                              "skip_gapped_alignment", "seed_family", "no_weight_scaling", "mem_clean"]:
            continue

        # Specific flag handling
        if key == "seed_weight": base_cmd_parts.append(f"--seed-weight={int(val)}")
        elif key == "max_gapped_aligner_length": base_cmd_parts.append(f"--max-gapped-aligner-length={int(val)}")
        elif key == "max_breakpoint_distance_scale": base_cmd_parts.append(f"--max-breakpoint-distance-scale={val}")
        elif key == "conservation_distance_scale": base_cmd_parts.append(f"--conservation-distance-scale={val}")
        elif key == "bp_dist_estimate_min_score": base_cmd_parts.append(f"--bp-dist-estimate-min-score={int(val)}")
        elif key == "gap_open": base_cmd_parts.append(f"--gap-open={int(val)}")
        elif key == "gap_extend": base_cmd_parts.append(f"--gap-extend={int(val)}")
        elif key == "weight": base_cmd_parts.append(f"--weight={int(val)}")
        elif key == "min_scaled_penalty": base_cmd_parts.append(f"--min-scaled-penalty={int(val)}")
        elif key == "hmm_p_go_homologous": base_cmd_parts.append(f"--hmm-p-go-homologous={val}")
        elif key == "hmm_p_go_unrelated": base_cmd_parts.append(f"--hmm-p-go-unrelated={val}")

    base_cmd_parts.append(str(query_fasta))
    base_cmd_parts.append(str(ref_fasta))

    print(f"\n[DEBUG MAUVE] Final command parts:")
    for i, part in enumerate(base_cmd_parts):
        print(f"  [{i}]: {part}")
    print(f"\n[DEBUG MAUVE] Executing progressiveMauve...")
    print(f"[DEBUG MAUVE] Full command: {' '.join(base_cmd_parts)}")

    # 5. Execute
    try:
        print(f"[DEBUG MAUVE] Starting subprocess.run...")
        print(f"[DEBUG MAUVE] Working directory: {Path.cwd()}")
        print(f"[DEBUG MAUVE] Using shell=False")

        # On Windows, try to ensure paths are properly quoted
        proc = run_quiet(base_cmd_parts, check=True, capture_output=True, text=True, timeout=1000)

        print(f"[DEBUG MAUVE] subprocess.run completed with return code: {proc.returncode}")
        if proc.stdout:
            print(f"[DEBUG MAUVE] STDOUT ({len(proc.stdout)} chars total):\n{proc.stdout[:500]}")
            if len(proc.stdout) > 500:
                print(f"[DEBUG MAUVE] ... (truncated, showing first 500 of {len(proc.stdout)} chars)")
        else:
            print(f"[DEBUG MAUVE] STDOUT: (empty)")
        if proc.stderr:
            print(f"[DEBUG MAUVE] STDERR ({len(proc.stderr)} chars total):\n{proc.stderr[:500]}")
            if len(proc.stderr) > 500:
                print(f"[DEBUG MAUVE] ... (truncated, showing first 500 of {len(proc.stderr)} chars)")
        else:
            print(f"[DEBUG MAUVE] STDERR: (empty)")

    except subprocess.CalledProcessError as e:
        print(f"[ERROR MAUVE] progressiveMauve failed (rc={e.returncode})")
        print(f"[ERROR MAUVE] Command that failed: {' '.join(base_cmd_parts)}")
        print(f"[ERROR MAUVE] STDERR:\n{e.stderr}")
        print(f"[ERROR MAUVE] STDOUT:\n{e.stdout}")
        traceback.print_exc()
        raise RuntimeError(
            f"progressiveMauve failed (rc={e.returncode}).\n"
            f"STDERR:\n{e.stderr}\nSTDOUT:\n{e.stdout}"
        ) from e
    except subprocess.TimeoutExpired as e:
        print(f"[ERROR MAUVE] progressiveMauve timed out after {e.timeout} seconds")
        print(f"[ERROR MAUVE] Command: {' '.join(base_cmd_parts)}")
        traceback.print_exc()
        raise TimeoutError(
            f"progressiveMauve timed out after {e.timeout} seconds. The genomes may be too large or dissimilar."
        ) from e
    except Exception as e:
        print(f"[ERROR MAUVE] Unexpected error during Mauve execution: {type(e).__name__}: {e}")
        print(f"[ERROR MAUVE] Command: {' '.join(base_cmd_parts)}")
        traceback.print_exc()
        raise

    # 6. Verify Output
    print(f"\n[DEBUG MAUVE] Verifying output file...")
    print(f"[DEBUG MAUVE] Expected output path: {output_xmfa_path}")
    print(f"[DEBUG MAUVE] Output exists: {output_xmfa_path.exists()}")
    if not output_xmfa_path.exists():
        print(f"[DEBUG MAUVE] Expected output file not found, checking alternatives...")
        output_basename = output_xmfa_path.with_suffix('')
        print(f"[DEBUG MAUVE] Checking for file without extension: {output_basename}")
        print(f"[DEBUG MAUVE] Basename exists: {output_basename.exists()}")
        if output_basename.exists():
            print(f"[WARNING MAUVE] Mauve created '{output_basename.name}' "
                  f"instead of the expected .xmfa file. Using this file instead.")
            return output_basename
        else:
            raise FileNotFoundError(f"Mauve ran but did not produce the expected alignment file: {output_xmfa_path}")

    return output_xmfa_path

def _is_path_safe(path: Path) -> bool:
    """
    Checks if a path is 'safe' for command-line tools (no spaces or parentheses).
    """
    path_str = str(path)
    if " " in path_str or "(" in path_str or ")" in path_str: return False
    return True

def _find_blast_executable(exe_name):
    """Find BLAST executable in PATH or common Windows installation locations."""
    found_exe = shutil.which(exe_name)
    if found_exe: return found_exe

    # Windows-specific search paths
    if platform.system() == "Windows":
        ncbi_base = Path("C:\\Program Files\\NCBI")
        if ncbi_base.exists():
            blast_dirs = sorted(glob.glob(str(ncbi_base / "blast-*")), reverse=True)
            for blast_dir in blast_dirs:
                exe_path = Path(blast_dir) / "bin" / f"{exe_name}.exe"
                if exe_path.exists(): return str(exe_path)

        # Other common installation locations
        common_paths = [
            Path(os.environ.get("PROGRAMFILES", "C:\\Program Files")) / "ncbi-blast" / "bin" / f"{exe_name}.exe",
            Path(os.environ.get("PROGRAMFILES", "C:\\Program Files")) / "ncbi" / "blast" / "bin" / f"{exe_name}.exe",
            Path("C:\\ncbi-blast\\bin") / f"{exe_name}.exe",
            Path(os.environ.get("PROGRAMFILES(X86)",
                                "C:\\Program Files (x86)")) / "ncbi-blast" / "bin" / f"{exe_name}.exe",
        ]
        for path in common_paths:
            if path.exists(): return str(path)

    return None


def _validate_bed_vs_features(bed_path, features, seqlen, fasta_id, log_queue):
    """
    Validates BED file contents using vectorized NumPy operations.
    Replaces slow iterrows/nested loops for instant validation.
    """
    if not bed_path or not Path(bed_path).is_file(): return

    issues = []

    try:
        bed_df = pd.read_csv(bed_path, sep="\t", header=None, comment='#')
        if bed_df.shape[1] < 3: return

        total_regions = len(bed_df)
        if total_regions == 0: return

        chroms = bed_df[0].astype(str).values
        starts_0 = bed_df[1].values
        ends_0 = bed_df[2].values

        search_id = fasta_id.split(':')[0]

        id_match_mask = [
            (search_id in c) or (c in search_id)
            for c in chroms
        ]
        id_mismatches = total_regions - np.sum(id_match_mask)

        out_of_bounds_mask = (starts_0 < 0) | (ends_0 > seqlen)
        out_of_bounds = np.sum(out_of_bounds_mask)

        if features:
            b_starts = starts_0 + 1
            b_ends = ends_0
            f_coords = np.array([[int(f.start), int(f.end)] for f in features])
            f_starts = f_coords[:, 0]
            f_ends = f_coords[:, 1]

            cond1 = b_starts[:, np.newaxis] <= f_ends[np.newaxis, :]
            cond2 = f_starts[np.newaxis, :] <= b_ends[:, np.newaxis]

            overlap_matrix = cond1 & cond2
            has_overlap = np.any(overlap_matrix, axis=1)
            no_feature_overlap = np.sum(~has_overlap)
        else:
            no_feature_overlap = total_regions

        if out_of_bounds > 0:
            issues.append(
                f"• Out of Bounds:\n   {out_of_bounds} regions exceed the query sequence length ({seqlen:,} bp).")

        if id_mismatches > 0:
            if id_mismatches == total_regions:
                issues.append(
                    f"• ID Mismatch:\n   None of the {total_regions} BED regions match FASTA ID '{search_id}'.")
            else:
                issues.append(f"• ID Mismatch:\n   {id_mismatches} regions have IDs that don't match '{search_id}'.")

        if no_feature_overlap > 0:
            issues.append(
                f"• Non-Coding Highlights:\n   {no_feature_overlap} "
                f"regions do not overlap any gene in the GFF.\n   "
                f"(This may be intentional for intergenic highlighting).")

    except Exception as e:
        print(f"[WARNING] Could not validate BED file: {e}")
        return

    if issues:
        msg_header = "Highlight Region (BED) Validation Issues\n\n"
        msg_body = "\n\n".join(issues)
        msg_footer = "\n\nDo you want to continue processing with these regions?"

        full_msg = msg_header + msg_body + msg_footer

        if log_queue:
            user_response_event = threading.Event()
            user_response_data = {}

            log_queue.put(("ask_yes_no", (full_msg, user_response_event, user_response_data)))
            user_response_event.wait()

            if not user_response_data.get("result", False):
                print("[INFO] Processing cancelled by user due to BED discrepancies.")
                raise RuntimeError("Processing cancelled by user (BED discrepancies).")
        else:
            print("\n" + "!" * 60)
            print(f"[WARNING] {msg_header}")
            for issue in issues:
                print(issue.replace("• ", "  - "))
            print("!" * 60 + "\n")

def _validate_alignment_files(xmfa_path, maf_path, query_path, ref_path,
                              query_id, ref_id, query_len, ref_len,
                              force_mauve, force_sibeliaz, log_queue):
    xmfa_issues = []
    maf_issues = []

    # XMFA Validation
    if xmfa_path and Path(xmfa_path).is_file() and not force_mauve:
        try:
            q_name = Path(query_path).name
            r_name = Path(ref_path).name

            with open(xmfa_path, 'r') as f:
                for line in f:
                    if line.startswith('#Sequence1File'):
                        xmfa_q_name = Path(line.strip().split('\t')[-1]).name
                        if xmfa_q_name != q_name:
                            xmfa_issues.append(
                                f"• XMFA Filename Mismatch:\n   XMFA expects: '{xmfa_q_name}'\n   Current file: '{q_name}'")
                    elif line.startswith('#Sequence2File'):
                        xmfa_r_name = Path(line.strip().split('\t')[-1]).name
                        if xmfa_r_name != r_name:
                            xmfa_issues.append(
                                f"• XMFA Filename Mismatch:\n   XMFA expects: '{xmfa_r_name}'\n   Current file: '{r_name}'")
                    elif line.startswith('>'):
                        parts = line.strip().split()
                        if len(parts) > 1:
                            coords_part = parts[1]
                            if ':' in coords_part:
                                seq_idx_str, range_str = coords_part.split(':')
                                if '-' in range_str:
                                    start_str, end_str = range_str.split('-')
                                    try:
                                        end_pos = int(end_str)
                                        seq_idx = int(seq_idx_str)

                                        if seq_idx == 1:
                                            if end_pos > query_len:
                                                xmfa_issues.append(
                                                    f"• XMFA Coordinates Out of Bounds (Query):\n   Found pos {end_pos:,} > Query length {query_len:,}")
                                                break
                                        elif seq_idx == 2:
                                            if end_pos > ref_len:
                                                xmfa_issues.append(
                                                    f"• XMFA Coordinates Out of Bounds (Ref):\n   Found pos {end_pos:,} > Ref length {ref_len:,}")
                                                break
                                    except ValueError:
                                        pass
        except Exception as e:
            print(f"[WARNING] Could not validate XMFA: {e}")

    # MAF Validation
    if maf_path and Path(maf_path).is_file() and not force_sibeliaz:
        try:
            with open(maf_path, 'r') as f:
                for line in f:
                    if line.startswith('s '):
                        parts = line.split()
                        if len(parts) >= 6:
                            src = parts[1]
                            start = int(parts[2])
                            size = int(parts[3])
                            srcSize = int(parts[5])

                            is_query = (src == query_id or src in query_id or query_id in src)
                            is_ref = (src == ref_id or src in ref_id or ref_id in src)

                            if is_query:
                                if srcSize != query_len:
                                    maf_issues.append(
                                        f"• MAF Length Mismatch (Query):\n   MAF says {srcSize:,}, FASTA is {query_len:,}")
                                    break
                            elif is_ref:
                                if srcSize != ref_len:
                                    maf_issues.append(
                                        f"• MAF Length Mismatch (Ref):\n   MAF says {srcSize:,}, FASTA is {ref_len:,}")
                                    break

                            if start + size > srcSize:
                                maf_issues.append(
                                    f"• MAF Bounds Error:\n   Segment end {start + size:,} > Total length {srcSize:,} for {src}")
                                break
        except Exception as e:
            print(f"[WARNING] Could not validate MAF: {e}")

    # Construct Message
    all_issues = xmfa_issues + maf_issues

    if all_issues:
        msg_header = "Cached Alignment File Discrepancies Detected!\n\n"
        msg_body = "\n\n".join(all_issues)

        # Build dynamic suggestion based on which files failed
        suggestions = []
        if xmfa_issues: suggestions.append("'Force New Alignment' (progressiveMauve)")
        if maf_issues: suggestions.append("'Force New SibeliaZ Alignment'")

        suggestion_str = " and/or \n".join(suggestions)

        msg_footer = (f"\n\nIt is highly recommended to enable:\n{suggestion_str}\n"
                      f"to regenerate valid data.\n\n"
                      "Do you want to continue with these potentially invalid files?")

        full_msg = msg_header + msg_body + msg_footer

        if log_queue:
            import threading
            user_response_event = threading.Event()
            user_response_data = {}

            log_queue.put(("ask_yes_no", (full_msg, user_response_event, user_response_data)))
            user_response_event.wait()

            if not user_response_data.get("result", False):
                print("[INFO] Processing cancelled by user due to alignment file discrepancies.")
                raise RuntimeError("Processing cancelled by user (Alignment discrepancies).")
        else:
            print("\n" + "!" * 60)
            print(f"[WARNING] {msg_header}")
            for issue in all_issues:
                print(issue.replace("• ", "  - "))
            print("!" * 60 + "\n")

def _validate_blast_tools():
    """Validate that makeblastdb and blastn are available before processing."""
    missing_tools = []

    for tool in ["makeblastdb", "blastn"]:
        found = _find_blast_executable(tool)
        if not found:
            missing_tools.append(tool)
        else:
            print(f"[BLAST] Found {tool} at: {found}")

    if missing_tools:
        error_msg = f"""BLAST tools not found: {', '.join(missing_tools)}

Please install NCBI BLAST+:
  1. Download from: https://ftp.ncbi.nlm.nih.gov/blast/executables/blast+/LATEST/
  2. Install to: C:\\Program Files\\ncbi-blast\\ (or another common location)
  3. Ensure the 'bin' folder is in your system PATH, or
  4. Install via conda: conda install -c bioconda blast

After installation, restart this application."""
        raise RuntimeError(error_msg)


def _run_blastn(query_fasta: Path, ref_fasta: Path, out_tab: Path, task, word_size, reward, penalty, evalue,
                max_target_seqs=None, num_threads=None, gapopen=None, gapextend=None, dust=None,
                soft_masking=None, strand=None, perc_identity=None, culling_limit=None):
    # Validate BLAST tools are available before processing
    _validate_blast_tools()

    # Find the actual executables (for proper paths on Windows)
    makeblastdb_exe = _find_blast_executable("makeblastdb")
    blastn_exe = _find_blast_executable("blastn")

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_dir_path = Path(temp_dir)
        copy_output_needed = False

        # 1. Determine query path (use original if safe, else copy to temp)
        if _is_path_safe(query_fasta):
            query_path_for_blast = query_fasta
        else:
            safe_query_path = temp_dir_path / "query.fasta"
            try:
                shutil.copy(query_fasta, safe_query_path)
            except FileNotFoundError as e:
                print(f"[FILE_NOT_FOUND] BLAST query copy: {e} | src={query_fasta} | dest={safe_query_path}")
                raise
            except Exception as e:
                print(f"[ERROR] BLAST query copy: {type(e).__name__}: {e}")
                raise
            query_path_for_blast = safe_query_path

        # 2. Determine reference path (use original if safe, else copy to temp)
        if _is_path_safe(ref_fasta):
            ref_path_for_makeblastdb = ref_fasta
        else:
            safe_ref_path = temp_dir_path / "ref.fasta"
            try:
                shutil.copy(ref_fasta, safe_ref_path)
            except FileNotFoundError as e:
                print(f"[FILE_NOT_FOUND] BLAST ref copy: {e} | src={ref_fasta} | dest={safe_ref_path}")
                raise
            except Exception as e:
                print(f"[ERROR] BLAST ref copy: {type(e).__name__}: {e}")
                raise
            ref_path_for_makeblastdb = safe_ref_path

        # 3. Determine output path (write directly if safe, else write to temp)
        if _is_path_safe(out_tab):
            out_path_for_blast = out_tab
        else:
            out_path_for_blast = temp_dir_path / "blast_out.tab"
            copy_output_needed = True

        # Database is always built in the temp dir for simplicity
        temp_db = temp_dir_path / "ref_db"

        # Run makeblastdb using the determined safe path
        try:
            run_quiet(
                [makeblastdb_exe, "-in", str(ref_path_for_makeblastdb), "-dbtype", "nucl", "-out", str(temp_db)],
                check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
            )
        except FileNotFoundError as e:
            print(f"[FILE_NOT_FOUND] makeblastdb: {e} | executable=makeblastdb | ref_path={ref_path_for_makeblastdb}")
            raise
        except subprocess.CalledProcessError as e:
            print(f"[SUBPROCESS_ERROR] makeblastdb failed (rc={e.returncode})")
            raise
        except OSError as e:
            print(f"[OSERROR] makeblastdb: {e}")
            raise

        # We specify all 14 columns we want, including qseq and sseq
        outfmt_string = "6 qseqid sseqid pident length mismatch gapopen qstart qend sstart send evalue bitscore qseq sseq"

        blast_cmd = [
            blastn_exe, "-query", str(query_path_for_blast), "-db", str(temp_db),
            "-outfmt", outfmt_string, "-out", str(out_path_for_blast)
        ]

        if task: blast_cmd.extend(("-task", task))
        if word_size: blast_cmd.extend(("-word_size", word_size))
        if reward: blast_cmd.extend(("-reward", reward))
        if penalty: blast_cmd.extend(("-penalty", penalty))
        if evalue: blast_cmd.extend(("-evalue", evalue))
        if max_target_seqs: blast_cmd.extend(("-max_target_seqs", max_target_seqs))
        if num_threads: blast_cmd.extend(("-num_threads", num_threads))
        if gapopen not in (None, "", "0"): blast_cmd.extend(("-gapopen", gapopen))
        if gapextend not in (None, "", "0"): blast_cmd.extend(("-gapextend", gapextend))
        if dust: blast_cmd.extend(("-dust", dust))
        if soft_masking: blast_cmd.extend(("-soft_masking", soft_masking))
        if strand: blast_cmd.extend(("-strand", strand))
        if perc_identity not in (None, "", "0"): blast_cmd.extend(("-perc_identity", perc_identity))
        if culling_limit not in (None, "", "0"): blast_cmd.extend(("-culling_limit", culling_limit))

        try:
            run_quiet(blast_cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except FileNotFoundError as e:
            print(f"[FILE_NOT_FOUND] blastn: {e} | executable=blastn | query={query_path_for_blast}")
            raise
        except subprocess.CalledProcessError as e:
            print(f"[SUBPROCESS_ERROR] blastn failed (rc={e.returncode})")
            raise
        except OSError as e:
            print(f"[OSERROR] blastn: {e}")
            raise

        # 4. Copy temp output to final destination *only if needed*
        if copy_output_needed:
            try:
                shutil.copy(out_path_for_blast, out_tab)
            except FileNotFoundError as e:
                print(f"[FILE_NOT_FOUND] BLAST output copy: {e} | src={out_path_for_blast} | dest={out_tab}")
                raise
            except Exception as e:
                print(f"[ERROR] BLAST output copy: {type(e).__name__}: {e}")
                raise


# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# Core Data Helpers & Utilities
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────

def _get_kmers(sequence, k):
    """Helper to generate k-mers from a sequence."""
    kmers, seq_len = set(), len(sequence)
    if seq_len < k: return kmers
    for i in range(seq_len - k + 1):
        kmers.add(sequence[i:i + k])
    return kmers


def _get_max_exact_match(lcb_data_list):
    """
    Analyzes WGA (Mauve/SibeliaZ) LCBs to find the longest
    continuous exact match (no gaps or mismatches).
    """
    global_max_len = 0
    if not lcb_data_list: return 0

    for lcb_list in lcb_data_list:
        for lcb in lcb_list:
            q_seq = lcb.get('aligned_query_sequence', '')
            r_seq = lcb.get('aligned_reference_sequence', '')
            if not q_seq or not r_seq: continue

            current_match_len = 0
            for q_char, r_char in zip(q_seq, r_seq):
                if q_char == r_char and q_char != '-': current_match_len += 1
                else:
                    if current_match_len > global_max_len: global_max_len = current_match_len
                    current_match_len = 0

            # Check one last time at the end of the LCB
            if current_match_len > global_max_len: global_max_len = current_match_len
    return global_max_len


def _calculate_fragment_identities(aligned_seq1, aligned_seq2):
    """Calculates identity% (excluding gaps) and (including gaps as mismatches)."""
    if not aligned_seq1 or not aligned_seq2:
        return 0.0, 0.0

    # Handle length mismatches - this should NOT happen with properly aligned sequences
    len1, len2 = len(aligned_seq1), len(aligned_seq2)
    if len1 != len2:
        print(f"[ERROR] Aligned sequence length mismatch detected: {len1} vs {len2}")
        print(f"  This indicates a bug in the alignment parsing or fragment extraction.")
        print(f"  Applying emergency padding to prevent 0% identity, but results may be incorrect.")
        print(f"  Query preview: ...{aligned_seq1[-50:] if len1 > 50 else aligned_seq1}")
        print(f"  Ref preview:   ...{aligned_seq2[-50:] if len2 > 50 else aligned_seq2}")

        # Emergency padding (band-aid fix)
        if len1 < len2:
            aligned_seq1 = aligned_seq1 + '-' * (len2 - len1)
        else:
            aligned_seq2 = aligned_seq2 + '-' * (len1 - len2)

    arr1 = np.array(list(aligned_seq1))
    arr2 = np.array(list(aligned_seq2))
    is_match = (arr1 == arr2)
    is_not_gap_either = (arr1 != '-') & (arr2 != '-')
    total_aligned_length = len(arr1)
    matches = (is_match & is_not_gap_either).sum()
    valid_comparisons = is_not_gap_either.sum()
    identity_excl_gaps = (matches / valid_comparisons * 100) if valid_comparisons > 0 else 0.0
    identity_incl_gaps = (matches / total_aligned_length * 100) if total_aligned_length > 0 else 0.0

    # Detect anti-aligned sequences (gaps in one align with bases in the other)
    # This suggests a serious alignment or parsing error
    if identity_excl_gaps < 5.0 and valid_comparisons > 0:
        # Check if this is an "anti-alignment" where gaps oppose bases
        is_gap_seq1 = (arr1 == '-')
        is_gap_seq2 = (arr2 == '-')
        is_base_seq1 = ~is_gap_seq1
        is_base_seq2 = ~is_gap_seq2

        # Count positions where seq1 has base and seq2 has gap, and vice versa
        anti_aligned = ((is_base_seq1 & is_gap_seq2) | (is_gap_seq1 & is_base_seq2)).sum()
        anti_aligned_pct = (anti_aligned / total_aligned_length * 100)

        if anti_aligned_pct > 50.0:
            print(f"[WARNING] Anti-alignment detected: {anti_aligned_pct:.1f}% of positions have gaps opposite bases")
            print(f"  Identity: {identity_excl_gaps:.2f}% (excl gaps), {identity_incl_gaps:.2f}% (incl gaps)")
            print(f"  This may indicate incorrect alignment or strand orientation.")
            print(f"  Preview (first 100 chars):")
            print(f"    Seq1: {aligned_seq1[:100]}")
            print(f"    Seq2: {aligned_seq2[:100]}")

    return identity_excl_gaps, identity_incl_gaps


def _is_coding(pos, coding_intervals):
    """Checks if a position falls within any coding interval."""
    for start, end in coding_intervals:
        if start <= pos <= end: return True
    return False


def _get_attr_value(attrs, key):
    if not attrs or key not in attrs: return None
    v = attrs[key]
    return v[0] if isinstance(v, list) and v else str(v)


def _nice_label(s, max_len=60):
    if not s: return None
    return s if len(s) <= max_len else (s[:max_len].rsplit(" ", 1)[0] or s[:max_len]) + "…"


def _generate_diff_strings(q_seq, r_seq, identity_incl_gaps):
    """
    Generates difference visualization strings based on aligned sequences.
    """
    diff_string, query_mismatch_hash, query_diff_only, ref_diff_only = "", "", "", ""
    seq_len = len(q_seq)

    # Optimization for 100% identity
    if abs(identity_incl_gaps - 100.0) < 1e-6:  # Use small tolerance for float comparison
        diff_string, query_mismatch_hash = q_seq, q_seq
        query_diff_only, ref_diff_only = "#" * seq_len, "#" * seq_len
        return diff_string, query_mismatch_hash, query_diff_only, ref_diff_only

    in_diff_block = False
    diff_q_buffer, diff_r_buffer = "", ""

    for i in range(seq_len):
        q_char, r_char = q_seq[i], r_seq[i]

        if q_char == r_char:  # Match (or both gaps, although trim should remove terminal ones)
            if in_diff_block:
                # Close the previous diff block
                diff_string += f"({diff_q_buffer}/{diff_r_buffer})"
                in_diff_block = False
                diff_q_buffer, diff_r_buffer = "", ""
            diff_string += q_char
            query_mismatch_hash += q_char
            query_diff_only += "#"
            ref_diff_only += "#"
        else:
            if not in_diff_block:
                # Start a new diff block
                in_diff_block = True
            diff_q_buffer += q_char
            diff_r_buffer += r_char
            query_mismatch_hash += "#"
            query_diff_only += q_char
            ref_diff_only += r_char

    # Close any open diff block at the end
    if in_diff_block: diff_string += f"({diff_q_buffer}/{diff_r_buffer})"

    return diff_string, query_mismatch_hash, query_diff_only, ref_diff_only


def get_unique_filepath(filepath: Path) -> Path:
    if not filepath.exists(): return filepath
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return filepath.parent / f"{filepath.stem}_{timestamp}{filepath.suffix}"


def _calculate_genome_stats(blast_df):
    """
    Calculates Length-Weighted ANI and dDDH-like approximation.
    IMPROVED: Uses weighted average instead of simple arithmetic mean.
    """
    if blast_df.empty: return 0.0, 0.0

    # 1. Calculate Length-Weighted ANI (Standard approach)
    # Formula: Sum(Identity * Length) / Sum(Length)
    # This prevents short, low-quality hits from dragging down the score.
    total_length = blast_df['length'].sum()

    if total_length > 0:
        # Calculate weighted sum of identities
        weighted_identity_sum = (blast_df['pident'] * blast_df['length']).sum()
        ani_value = weighted_identity_sum / total_length

        # dDDH approximation (Total identical bases / Total aligned length)
        # Note: In this simple formula, Weighted ANI and this dDDH approx are mathematically
        # identical, but we keep the structure for compatibility.
        identical_bases = (blast_df['pident'] / 100.0 * blast_df['length']).sum()
        dDDH_value = (identical_bases / total_length) * 100.0
    else:
        ani_value = 0.0
        dDDH_value = 0.0

    return ani_value, dDDH_value


# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# Annotation & Feature Processing
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────

def _label_for_feature(feature, allow_code_fallback: bool):
    attrs = getattr(feature, "attributes", {}) or {}
    for key in ["product", "gene", "Name"]:
        if v := _get_attr_value(attrs, key): return _nice_label(v, 60)
    if allow_code_fallback:
        for key in ["ID", "locus_tag"]:
            if v := _get_attr_value(attrs, key): return _nice_label(v, 60)
    return None


def _generate_contextual_tooltip(feature_obj, start, end, feature_identity):
    """Generate context-aware tooltip based on whether feature is baseline or offset."""

    is_baseline = getattr(feature_obj, 'is_baseline', True)
    attrs = getattr(feature_obj, 'attributes', {}) or {}

    if is_baseline:
        # Baseline: Compact single-line format
        coords_str = f"{start}-{end}"
        label = _label_for_feature(feature_obj, False)
        if not label:
            label = _get_attr_value(attrs, "gene") or _get_attr_value(attrs, "product") or f"feature:{start}-{end}"

        if feature_identity and feature_identity > 0:
            return f"{label} [{coords_str}] | Identity: {feature_identity:.1f}%"
        else: return f"{label} [{coords_str}]"

    else:
        # Offset: Multi-line with hierarchy
        lines = []

        # Show parent context if available
        parent_info = getattr(feature_obj, 'parent_info', None)
        if parent_info:
            parent_coords = f"{parent_info['start']}-{parent_info['end']}"
            lines.append(f"Parent: {parent_info['label']} ({parent_info['type']}) [{parent_coords}]")
            lines.append("─" * 30)

        # Show self
        label = _label_for_feature(feature_obj, False)
        if not label:
            label = _get_attr_value(attrs, "product") or _get_attr_value(attrs, "gene") or str(
                getattr(feature_obj, 'id', 'feature'))

        feat_type = getattr(feature_obj, 'featuretype', 'feature').upper()
        coords_str = f"{start}-{end}"
        lines.append(f"{feat_type}: {label} [{coords_str}]")

        # Show direct exons if any
        direct_exons = getattr(feature_obj, 'direct_exons', [])
        if direct_exons:
            lines.append("Exons:")
            for exon in direct_exons:
                exon_id = exon.get('id', '')
                exon_coords = f"{exon['start']}-{exon['end']}"
                # Only show ID if not auto-generated looking
                if exon_id and not exon_id.startswith('exon-'): lines.append(f"  • {exon_id} ({exon_coords})")
                else: lines.append(f"  • ({exon_coords})")

        # Show siblings if any
        siblings = getattr(feature_obj, 'siblings', [])
        if siblings:
            lines.append("Siblings:")
            for sib in siblings:
                sib_type = sib.get('type', 'feature').upper()
                sib_coords = f"{sib['start']}-{sib['end']}"
                lines.append(f"  • {sib_type} [{sib_coords}]")

        return "\n".join(lines)


def _create_feature_from_gffutils(feature_for_coords, feature_for_label, highlight_intervals, allow_code_fallback: bool,
                                  blast_df=None,
                                  homology_threshold=0.0, manual_find=None, manual_replace=None):
    is_highlighted = False
    if highlight_intervals:
        # Use coordinates from the feature_for_coords
        fs, fe = int(feature_for_coords.start), int(feature_for_coords.end)
        if any(fs <= int(be) and int(bs) < fe for bs, be in highlight_intervals): is_highlighted = True

    best_identity = 0.0
    is_homologous = False

    if blast_df is not None and not blast_df.empty:
        # Use coordinates from the feature_for_coords
        feature_start, feature_end = int(feature_for_coords.start), int(feature_for_coords.end)
        q_starts = blast_df[['qstart', 'qend']].min(axis=1)
        q_ends = blast_df[['qstart', 'qend']].max(axis=1)
        overlapping_hits_df = blast_df[(q_starts < feature_end) & (q_ends > feature_start)]

        if not overlapping_hits_df.empty:
            best_hit = overlapping_hits_df.loc[overlapping_hits_df['pident'].idxmax()]
            best_identity = best_hit.pident
            if best_identity >= homology_threshold: is_homologous = True
            best_sstart, best_send, best_sseqid = best_hit.sstart, best_hit.send, best_hit.sseqid
            best_qseqid = best_hit.qseqid

    if is_highlighted:
        color = "#e53935"
        # Get label from the feature_for_label
        label = _label_for_feature(feature_for_label, allow_code_fallback)
        if not label:
            attrs = getattr(feature_for_label, "attributes", {}) or {}
            label = _nice_label(
                _get_attr_value(attrs, "gene") or _get_attr_value(attrs, "product") or
                _get_attr_value(attrs, "Name") or _get_attr_value(attrs, "ID") or
                _get_attr_value(attrs, "locus_tag") or f"feature:{feature_for_coords.start}-{feature_for_coords.end}"
            )
    else:
        color = "#43a047" if is_homologous else "#cccccc"
        label = _label_for_feature(feature_for_label, allow_code_fallback)

    # Get attributes from the feature_for_label
    attrs = getattr(feature_for_label, "attributes", {}) or {}
    gene_name = _get_attr_value(attrs, "gene")
    if manual_find and manual_replace and gene_name and gene_name.lower() == manual_find.lower():
        label = manual_replace

    graphic_feature = GraphicFeature(
        start=int(feature_for_coords.start), end=int(feature_for_coords.end),
        strand=1 if getattr(feature_for_coords, "strand", "+") == "+" else -1,
        color=color, label=label
    )
    graphic_feature.attributes = attrs
    graphic_feature.best_identity = best_identity
    if 'best_sstart' in locals():
        graphic_feature.best_sstart = best_sstart
        graphic_feature.best_send = best_send
        graphic_feature.best_sseqid = best_sseqid
        graphic_feature.best_qseqid = best_qseqid
    return graphic_feature


def load_gene_features(annot_file: Path, fasta_id: str, highlight_intervals, allow_code_fallback: bool, blast_df=None,
                       homology_threshold=0.0, manual_find=None, manual_replace=None, seq: str = "", log_queue=None,
                       seqlen=None):
    features = []

    # GENBANK handling
    if annot_file.suffix.lower() in {".gb", ".gbk"}:
        # Attempt Strict Match
        for record in SeqIO.parse(annot_file, "genbank"):
            if record.id == fasta_id:
                for feat in record.features:
                    if feat.type in {"gene", "CDS"} and feat.location is not None:
                        mock = type("MockFeature", (), {
                            "start": int(feat.location.start) + 1, "end": int(feat.location.end),
                            "strand": "+" if getattr(feat, "strand", 1) in (1, "+") else "-",
                            "attributes": getattr(feat, "qualifiers", {}) or {}
                        })()
                        features.append(
                            _create_feature_from_gffutils(mock, mock, highlight_intervals, allow_code_fallback,
                                                          blast_df=None, homology_threshold=homology_threshold,
                                                          manual_find=manual_find, manual_replace=manual_replace))
                if blast_df is not None and not blast_df.empty:
                    try:
                        _batch_assign_homology(features, blast_df, homology_threshold)
                    except Exception as e:
                        print(f"[WARNING] Batch homology assignment failed: {e}")

                return features

        # Fallback for GenBank mismatch
        print(f"[INFO] Strict GenBank ID match failed. Loading first record from {annot_file.name}...")
        try:
            record = next(SeqIO.parse(annot_file, "genbank"))
            for feat in record.features:
                if feat.type in {"gene", "CDS"} and feat.location is not None:
                    mock = type("MockFeature", (), {
                        "start": int(feat.location.start) + 1, "end": int(feat.location.end),
                        "strand": "+" if getattr(feat, "strand", 1) in (1, "+") else "-",
                        "attributes": getattr(feat, "qualifiers", {}) or {}
                    })()
                    features.append(
                        _create_feature_from_gffutils(mock, mock, highlight_intervals, allow_code_fallback,
                                                      blast_df=None, homology_threshold=homology_threshold,
                                                      manual_find=manual_find, manual_replace=manual_replace))
            if blast_df is not None and not blast_df.empty:
                try:
                    _batch_assign_homology(features, blast_df, homology_threshold)
                except Exception as e:
                    print(f"[WARNING] Batch homology assignment failed: {e}")

            return features
        except Exception as e:
            raise RuntimeError(f"GenBank record not found or parse error: {e}")

    # GFF3 handling
    unique_id = f"{int(time.time())}_{random.randint(1000, 9999)}"
    db_path = annot_file.with_name(f"{annot_file.stem}_{unique_id}.gff.db")
    db = None
    try:
        if not annot_file or str(annot_file).strip() == "":
            print("!!! ERROR: No annotation file path provided")
            return []

        # Create DB
        db = gffutils.create_db(str(annot_file), dbfn=str(db_path), force=True, keep_order=True, merge_strategy="merge")
        db_seqids = list(db.seqids())

        # Validation logic
        issues = []
        target_gff_id = fasta_id

        # Check 1: Sequence ID Mismatch
        if fasta_id not in db_seqids:
            if len(db_seqids) > 0:
                target_gff_id = db_seqids[0]
                issues.append(f"• Sequence ID mismatch:\n   FASTA: '{fasta_id}'\n   GFF:   '{target_gff_id}'")
            else:
                raise RuntimeError("No SeqIDs found in GFF database.")

        # Check 2: Coordinates Out of Range
        if seqlen is not None and seqlen > 0:
            try:
                cursor = db.conn.execute("SELECT MAX(end) FROM features")
                result = cursor.fetchone()
                max_gff_end = result[0] if result else 0

                if max_gff_end > seqlen:
                    issues.append(
                        f"• Coordinates out of range:\n   Max GFF coordinate ({max_gff_end:,}) exceeds FASTA length ({seqlen:,}).")
            except Exception as e:
                print(f"[WARNING] Could not validate GFF coordinates: {e}")

        # Show popup if issues exist
        if issues:
            msg_header = "Annotation / Sequence Discrepancies Detected!\n\n"
            msg_body = "\n\n".join(issues)
            msg_footer = ("\n\nThe tool can attempt to process using the first available ID and ignoring bounds, "
                          "but results may be invalid or incomplete.\n\nDo you want to continue?")

            full_msg = msg_header + msg_body + msg_footer

            if log_queue:
                import threading
                user_response_event = threading.Event()
                user_response_data = {}

                # Send request to main thread
                log_queue.put(("ask_yes_no", (full_msg, user_response_event, user_response_data)))

                # Block until user responds
                user_response_event.wait()

                if not user_response_data.get("result", False):
                    # User clicked No
                    print("[INFO] Processing cancelled by user due to GFF discrepancies.")
                    raise RuntimeError("Processing cancelled by user (GFF discrepancies).")
            else:
                # Console fallback
                print("\n" + "!" * 60)
                print(f"[WARNING] {msg_header}")
                for issue in issues:
                    print(issue.replace("• ", "  - "))
                print("!" * 60 + "\n")

        # Detect Coordinate System
        GFFUTILS_COORD_OFFSET = 0

        # Sample first feature to detect if GFF is 0-indexed
        try:
            sample_feature = next(db.features_of_type("gene", order_by="start"), None)
            if not sample_feature:
                sample_feature = next(db.all_features(), None)

            if sample_feature and sample_feature.seqid == target_gff_id:
                sample_start = int(sample_feature.start)
                # If feature starts at 0, it's likely 0-indexed (should be 1-indexed in GFF3)
                if sample_start == 0:
                    GFFUTILS_COORD_OFFSET = 1
                    print("[DETECTION] GFF appears to be 0-indexed; applying +1 offset")
                else:
                    print(f"[DETECTION] GFF appears to be 1-indexed; no offset needed (sample start: {sample_start})")
        except Exception as e:
            print(f"[WARNING] Could not auto-detect GFF coordinate system: {e}")

        global GFFUTILS_COORD_OFFSET_GLOBAL
        GFFUTILS_COORD_OFFSET_GLOBAL = GFFUTILS_COORD_OFFSET

        # Feature loading loop
        primary_types = ["gene", "ORF", "rRNA", "tRNA", "ncRNA"]
        processed_feature_ids = set()

        for ptype in primary_types:
            for parent_feature in db.features_of_type(ptype, order_by="start"):

                if parent_feature.seqid != target_gff_id:
                    continue

                if parent_feature.id in processed_feature_ids:
                    continue

                parent_start = int(parent_feature.start) + GFFUTILS_COORD_OFFSET
                parent_end = int(parent_feature.end)

                try:
                    all_children = list(db.children(parent_feature, order_by="start"))
                except:
                    all_children = []

                child_priority_types = ["CDS", "tRNA", "rRNA", "ORF", "mRNA"]
                full_span_child = None

                for child_type in child_priority_types:
                    try:
                        candidates = list(db.children(parent_feature, featuretype=child_type, order_by="start"))
                        for candidate in candidates:
                            c_start = int(candidate.start) + GFFUTILS_COORD_OFFSET
                            c_end = int(candidate.end)
                            if c_start == parent_start and c_end == parent_end:
                                full_span_child = candidate
                                break
                    except:
                        pass
                    if full_span_child: break

                if full_span_child:
                    feature_for_coords = type("OffsetGFFFeature", (), {
                        "start": int(full_span_child.start) + GFFUTILS_COORD_OFFSET,
                        "end": int(full_span_child.end),
                        "strand": getattr(full_span_child, "strand", "+"),
                        "attributes": getattr(full_span_child, "attributes", {})
                    })()
                    feature_for_label = full_span_child
                    processed_feature_ids.add(full_span_child.id)
                else:
                    feature_for_coords = type("OffsetGFFFeature", (), {
                        "start": parent_start,
                        "end": parent_end,
                        "strand": getattr(parent_feature, "strand", "+"),
                        "attributes": getattr(parent_feature, "attributes", {})
                    })()
                    feature_for_label = parent_feature

                baseline_feature = _create_feature_from_gffutils(
                    feature_for_coords, feature_for_label, highlight_intervals,
                    allow_code_fallback, blast_df=None, homology_threshold=homology_threshold,
                    manual_find=manual_find, manual_replace=manual_replace
                )
                baseline_feature.is_baseline = True
                baseline_feature.parent_info = None
                baseline_feature.offset_children = []
                baseline_feature.direct_exons = []

                features.append(baseline_feature)
                processed_feature_ids.add(parent_feature.id)

                for child in all_children:
                    if child.id in processed_feature_ids: continue
                    if child.featuretype == "exon":
                        baseline_feature.direct_exons.append({
                            "id": child.id, "start": int(child.start), "end": int(child.end),
                            "featuretype": child.featuretype
                        })
                        processed_feature_ids.add(child.id)
                        continue

                    child_start = int(child.start) + GFFUTILS_COORD_OFFSET
                    child_end = int(child.end)

                    child_wrapper = type("OffsetGFFFeature", (), {
                        "start": child_start, "end": child_end,
                        "strand": getattr(child, "strand", "+"), "attributes": getattr(child, "attributes", {})
                    })()

                    offset_feature = _create_feature_from_gffutils(
                        child_wrapper, child, highlight_intervals, allow_code_fallback,
                        blast_df, homology_threshold, manual_find, manual_replace
                    )
                    offset_feature.label = None
                    offset_feature.is_baseline = False
                    offset_feature.parent_info = {"type": ptype, "label": str(feature_for_label.id),
                                                  "start": parent_start, "end": parent_end}
                    offset_feature.direct_exons = []
                    offset_feature.siblings = []

                    features.append(offset_feature)
                    baseline_feature.offset_children.append(
                        {"type": child.featuretype, "start": child_start, "end": child_end})
                    processed_feature_ids.add(child.id)

        for orphan_type in ["CDS", "tRNA", "rRNA", "ORF"]:
            try:
                for orphan_feature in db.features_of_type(orphan_type, order_by="start"):
                    if orphan_feature.seqid != target_gff_id:
                        continue

                    if orphan_feature.id in processed_feature_ids:
                        continue

                    orphan_wrapper = type("OffsetGFFFeature", (), {
                        "start": int(orphan_feature.start) + GFFUTILS_COORD_OFFSET,
                        "end": int(orphan_feature.end),
                        "strand": getattr(orphan_feature, "strand", "+"),
                        "attributes": getattr(orphan_feature, "attributes", {})
                    })()

                    orphan_obj = _create_feature_from_gffutils(
                        orphan_wrapper, orphan_feature, highlight_intervals,
                        allow_code_fallback, blast_df, homology_threshold, manual_find, manual_replace
                    )
                    orphan_obj.is_baseline = True
                    orphan_obj.parent_info = None
                    orphan_obj.offset_children = []
                    orphan_obj.direct_exons = []

                    features.append(orphan_obj)
                    processed_feature_ids.add(orphan_feature.id)
            except:
                pass

    finally:
        if db and hasattr(db, "conn"): db.conn.close()
        if db_path.exists():
            try:
                db_path.unlink()
            except Exception as e:
                print(f"Warning: Could not delete temporary database {db_path}: {e}")

    if blast_df is not None and not blast_df.empty:
        try:
            _batch_assign_homology(features, blast_df, homology_threshold)
        except Exception as e:
            print(f"[WARNING] Batch homology assignment failed: {e}")

    return features


# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# Alignment Data Parsers
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────

def _run_mauve_and_parse(query_fasta_path, ref_fasta_path, full_seqlen, log_queue=None,
                         xmfa_file=None, force_mauve=False,
                         mauve_disable_backbone=False, mauve_mums=False, mauve_collinear=False,
                         mauve_skip_refinement=False, mauve_skip_gapped_alignment=False, mauve_seed_family=False,
                         mauve_no_weight_scaling=False, mauve_mem_clean=False, mauve_seed_weight="",
                         mauve_max_gapped_aligner_length="", mauve_max_breakpoint_distance_scale="",
                         mauve_conservation_distance_scale="", mauve_bp_dist_estimate_min_score="",
                         mauve_gap_open="", mauve_gap_extend="", mauve_weight="", mauve_min_scaled_penalty="",
                         mauve_hmm_p_go_homologous="", mauve_hmm_p_go_unrelated=""):
    """
    Runs progressiveMauve (or uses cache/input), parses XMFA, builds identity array,
    and extracts LCB data with dual identity calculations.

    IMPORTANT:
      - We assume a pairwise Mauve alignment (query vs reference).
      - For each LCB, we treat the first sequence as query, second as reference.
      - We use Biopython's rec.annotations['start'], ['end'], ['strand'] for coords/strand.
      - We do NOT reverse-complement any sequences here; we keep Mauve's orientation.
    """

    def status(msg, color="black"):
        if log_queue:
            log_queue.put(("status", (msg, color)))

    # Decide XMFA path
    if not force_mauve and xmfa_file and Path(xmfa_file).is_file():
        xmfa_file_path = Path(xmfa_file)
        status(f"Using user-provided Mauve alignment: {xmfa_file_path.name}", "blue")
    else:
        query_stem = query_fasta_path.stem
        ref_stem = ref_fasta_path.stem
        output_dir = query_fasta_path.parent / "mauve_alignments"
        output_dir.mkdir(parents=True, exist_ok=True)
        xmfa_file_path = output_dir / f"{query_stem}_vs_{ref_stem}.xmfa"

        run_mauve_now = False
        if force_mauve:
            status("Forcing new Mauve alignment...", "orange")
            run_mauve_now = True
        elif not xmfa_file_path.exists():
            status("No cache found. Running progressiveMauve WGA...", "black")
            run_mauve_now = True
        else:
            status(f"Using cached Mauve alignment: {xmfa_file_path.name}", "blue")

        if run_mauve_now:
            try:
                # Use parameters directly - they're already passed to this function
                mauve_options = {
                    "disable_backbone": mauve_disable_backbone,
                    "mums": mauve_mums,
                    "collinear": mauve_collinear,
                    "skip_refinement": mauve_skip_refinement,
                    "skip_gapped_alignment": mauve_skip_gapped_alignment,
                    "seed_family": mauve_seed_family,
                    "no_weight_scaling": mauve_no_weight_scaling,
                    "mem_clean": mauve_mem_clean,
                    "seed_weight": mauve_seed_weight,
                    "max_gapped_aligner_length": mauve_max_gapped_aligner_length,
                    "max_breakpoint_distance_scale": mauve_max_breakpoint_distance_scale,
                    "conservation_distance_scale": mauve_conservation_distance_scale,
                    "bp_dist_estimate_min_score": mauve_bp_dist_estimate_min_score,
                    "gap_open": mauve_gap_open,
                    "gap_extend": mauve_gap_extend,
                    "weight": mauve_weight,
                    "min_scaled_penalty": mauve_min_scaled_penalty,
                    "hmm_p_go_homologous": mauve_hmm_p_go_homologous,
                    "hmm_p_go_unrelated": mauve_hmm_p_go_unrelated,
                }
                print(f"DEBUG: About to run progressiveMauve with file: {xmfa_file_path}")
                _run_progressive_mauve(query_fasta_path, ref_fasta_path, xmfa_file_path, mauve_options)
                print(f"DEBUG: progressiveMauve completed, checking if file exists: {xmfa_file_path.exists()}")
                if xmfa_file_path.exists():
                    file_size = xmfa_file_path.stat().st_size
                    print(f"DEBUG: Mauve output file created, size: {file_size} bytes")
                status("progressiveMauve analysis complete.", "blue")
            except Exception as e:
                print(f"[ERROR] progressiveMauve execution failed with exception:")
                print(f"  Exception type: {type(e).__name__}")
                print(f"  Exception message: {e}")
                import traceback
                traceback.print_exc()
                status(f"progressiveMauve failed: {e}", "red")
                if xmfa_file_path.exists():
                    try:
                        xmfa_file_path.unlink()
                        print("[DEBUG] Removed partial/failed cache file.")
                    except Exception as e2:
                        print(f"[DEBUG] Warning, could not remove partial cache file: {e2}")
                return None, [], 0.0

    print(f"\nDEBUG: About to parse Mauve alignment from: {xmfa_file_path}")
    print(f"DEBUG: File exists: {xmfa_file_path.exists()}")
    if not xmfa_file_path.exists():
        print("DEBUG: ERROR - File does not exist! Returning empty data.")
        status(f"ERROR: Mauve XMFA file not found: {xmfa_file_path}", "red")
        return None, [], 0.0

    status("Parsing XMFA and mapping LCBs (Mauve)...")

    g_query_identity = np.zeros(full_seqlen)
    lcb_data = []
    total_lcb_matches = 0
    total_lcb_valid_comparisons = 0

    try:
        alignments = Bio.AlignIO.parse(xmfa_file_path, "mauve")
        lcb_count = 0

        for lcb in alignments:
            lcb_count += 1
            if lcb_count <= 3 or lcb_count % 100 == 0:
                print(f"DEBUG: Processing Mauve LCB #{lcb_count}, sequences: {len(lcb)}")

            # We expect pairwise: [query, reference]
            if len(lcb) < 2:
                print(f"  SKIPPED LCB {lcb_count}: less than 2 sequences in this block.")
                continue

            q_rec = lcb[0]
            r_rec = lcb[1]

            # Debug headers/annotations once
            if lcb_count == 1:
                print("DEBUG: Example Mauve record IDs and annotations:")
                print("  Query rec.id:", repr(q_rec.id))
                print("  Query annotations:", q_rec.annotations)
                print("  Ref   rec.id:", repr(r_rec.id))
                print("  Ref   annotations:", r_rec.annotations)

            # Coordinates and strands from annotations
            try:
                q_start_1idx = int(q_rec.annotations.get("start"))
                q_end_1idx = int(q_rec.annotations.get("end"))
                q_strand_ann = int(q_rec.annotations.get("strand", 1))

                r_start_1idx = int(r_rec.annotations.get("start"))
                r_end_1idx = int(r_rec.annotations.get("end"))
                r_strand_ann = int(r_rec.annotations.get("strand", 1))
            except Exception as e:
                print(f"  SKIPPED LCB {lcb_count}: Could not read start/end/strand from annotations. Error: {e}")
                continue

            q_strand = '+' if q_strand_ann >= 0 else '-'
            r_strand = '+' if r_strand_ann >= 0 else '-'

            # Normalized genomic ranges (min..max, 1-based)
            q_start_norm = min(q_start_1idx, q_end_1idx)
            q_end_norm = max(q_start_1idx, q_end_1idx)
            r_start_norm = min(r_start_1idx, r_end_1idx)
            r_end_norm = max(r_start_1idx, r_end_1idx)

            aln_q = str(q_rec.seq).upper().strip()
            aln_r = str(r_rec.seq).upper().strip()
            aln_len_q = len(aln_q)
            aln_len_r = len(aln_r)

            if aln_len_r != aln_len_q:
                print(f"  SKIPPED LCB {lcb_count}: query/ref alignment lengths differ ({aln_len_q} vs {aln_len_r}).")
                # Debug: show first/last 20 chars to identify issue
                if abs(aln_len_q - aln_len_r) <= 5:
                    print(f"    Query ends with: {repr(aln_q[-20:])}")
                    print(f"    Ref ends with: {repr(aln_r[-20:])}")
                continue

            aln_len = aln_len_q

            # Per-LCB identity
            identity_pct_ignore_gaps, identity_pct_gaps_as_mismatch = _calculate_fragment_identities(aln_q, aln_r)

            lcb_data.append(
                {
                    "query_start": q_start_norm,
                    "query_end": q_end_norm,
                    "query_strand": q_strand,
                    "ref_start": r_start_norm,
                    "ref_stop": r_end_norm,
                    "ref_strand": r_strand,
                    "identity_excl_gaps": identity_pct_ignore_gaps,
                    "identity_incl_gaps": identity_pct_gaps_as_mismatch,
                    "aligned_query_sequence": aln_q,
                    "aligned_reference_sequence": aln_r,
                }
            )

            # Global WGA identity accumulators
            arr1 = np.array(list(aln_q))
            arr2 = np.array(list(aln_r))
            is_match = (arr1 == arr2)
            is_not_gap_either = (arr1 != '-') & (arr2 != '-')
            matches = (is_match & is_not_gap_either).sum()
            valid_comparisons = is_not_gap_either.sum()

            total_lcb_matches += matches
            total_lcb_valid_comparisons += valid_comparisons

            M = (is_match & is_not_gap_either).astype(int) * 100  # 100 for matches, 0 otherwise

            q_bytes = np.frombuffer(aln_q.encode('ascii'), dtype='S1')
            non_gap_mask = (q_bytes != b'-')

            # Scores for non-gap positions
            scores_to_apply = M[non_gap_mask].astype(int)

            # Calculate genomic indices all at once
            num_bases = len(scores_to_apply)
            if num_bases > 0:
                if q_strand == '+':
                    start_idx = q_start_norm - 1
                    indices = start_idx + np.arange(num_bases)
                else:
                    start_idx = q_end_norm - 1
                    indices = start_idx - np.arange(num_bases)

                # Boundary check (just in case)
                valid_mask = (indices >= 0) & (indices < full_seqlen)
                valid_indices = indices[valid_mask]
                valid_scores = scores_to_apply[valid_mask]

                # Instant array update using maximum logic
                np.maximum.at(g_query_identity, valid_indices, valid_scores)

    except Exception as e:
        print(f"ERROR during Mauve XMFA parsing: {e}")
        import traceback
        traceback.print_exc()
        return None, [], 0.0

    global_wga_id_pct = (
        total_lcb_matches / total_lcb_valid_comparisons * 100
        if total_lcb_valid_comparisons > 0
        else 0.0
    )
    status(f"Finished parsing Mauve XMFA: {lcb_count} LCBs processed.", "blue")
    print(f"DEBUG: Mauve LCBs parsed and kept: {len(lcb_data)}")

    return g_query_identity, lcb_data, global_wga_id_pct


def _run_sibeliaz_and_parse(query_fasta_path, ref_fasta_path, full_seqlen,
                            query_seq_full, ref_seq_main, ref_seqlen, log_queue=None,
                            maf_file=None, force_sibeliaz=False):
    """
    Runs SibeliaZ (or uses cache/input), parses MAF, builds identity array,
    and extracts LCB data with dual identity calculations.
    """
    # 1. Define status helper first
    def status(msg, color="black"):
        if log_queue: log_queue.put(("status", (msg, color)))

    query_stem = query_fasta_path.stem
    ref_stem = ref_fasta_path.stem
    alignments_dir = query_fasta_path.parent / "sibeliaz_alignments"
    alignments_dir.mkdir(exist_ok=True)
    target_maf_file_path = alignments_dir / f"{query_stem}_vs_{ref_stem}.maf"
    run_sibeliaz_now = False
    maf_file_to_parse = None

    # 2. Check Cache
    if not force_sibeliaz and maf_file and Path(maf_file).is_file():
        status(f"Using user-provided SibeliaZ alignment: {Path(maf_file).name}", "blue")
        maf_file_to_parse = Path(maf_file)
    elif not force_sibeliaz and target_maf_file_path.is_file():
        status(f"Using cached SibeliaZ alignment: {target_maf_file_path.name}", "blue")
        maf_file_to_parse = target_maf_file_path
    else:
        if force_sibeliaz:
            status("Forcing new SibeliaZ alignment...", "orange")
        else:
            status("No cache found. Running SibeliaZ WGA...", "black")
        run_sibeliaz_now = True

    # 3. Execution Logic
    if run_sibeliaz_now:
        status("Running SibeliaZ WGA...")

        with tempfile.TemporaryDirectory(prefix="sibeliaz_run_") as temp_dir:
            temp_dir_path = Path(temp_dir)
            safe_query_path = temp_dir_path / "query.fasta"
            safe_ref_path = temp_dir_path / "ref.fasta"
            safe_output_dir = temp_dir_path / "output"
            safe_output_dir.mkdir()

            try:
                shutil.copy(query_fasta_path, safe_query_path)
                shutil.copy(ref_fasta_path, safe_ref_path)
            except Exception as e:
                print(f"[ERROR] Copy failed: {e}")
                status(f"SibeliaZ setup failed: {e}", "red")
                return None, [], 0.0

            try:
                if sys.platform.startswith("win"):
                    if USE_DOCKER_SIBELIAZ:
                        _run_sibeliaz_docker(str(safe_query_path), str(safe_ref_path), str(safe_output_dir), [])
                    else:
                        _run_sibeliaz_wsl(str(safe_query_path), str(safe_ref_path), str(safe_output_dir), [])
                elif sys.platform == "darwin":
                    _run_sibeliaz_docker(str(safe_query_path), str(safe_ref_path), str(safe_output_dir), [])
                else:
                    sibeliaz_exe = shutil.which("sibeliaz")
                    if not sibeliaz_exe:
                        raise FileNotFoundError("SibeliaZ executable not found in PATH.")
                    run_quiet([sibeliaz_exe, "-o", str(safe_output_dir), str(safe_query_path), str(safe_ref_path)],
                                   check=True, capture_output=True, text=True)
            except Exception as e:
                print(f"\n!!! SIBELIAZ EXECUTION ERROR !!!\n{e}\n")
                status(f"SibeliaZ execution failed: {e}", "red")
                return None, [], 0.0

            source_maf = safe_output_dir / "alignment.maf"
            if not source_maf.exists():
                print(f"[ERROR] SibeliaZ ran but {source_maf} was not created.")
                status("SibeliaZ output missing.", "red")
                return None, [], 0.0

            try:
                shutil.copy(source_maf, target_maf_file_path)
                print(f"DEBUG: Saved SibeliaZ alignment to cache: {target_maf_file_path}")
            except Exception as e:
                print(f"[ERROR] Cache copy failed: {e}")
                status(f"FATAL: Could not copy MAF to cache: {e}", "red")
                return None, [], 0.0

        maf_file_to_parse = target_maf_file_path

    # 4. Parsing Logic
    print(f"\nDEBUG: Parsing SibeliaZ MAF file: {maf_file_to_parse}")
    status("Parsing MAF and mapping LCBs (SibeliaZ)...")

    g_query_identity = np.zeros(full_seqlen)
    lcb_data = []
    total_lcb_matches = 0
    total_lcb_valid_comparisons = 0

    def _fasta_ids(path, cap=2000):
        ids = set()
        with open(path, "r") as fh:
            for line in fh:
                if line.startswith(">"): ids.add(line[1:].strip().split()[0]);
                if len(ids) >= cap: break
        return ids

    query_idset = _fasta_ids(query_fasta_path)
    ref_idset = _fasta_ids(ref_fasta_path)

    try:
        alignments = Bio.AlignIO.parse(maf_file_to_parse, "maf")
        lcb_count = 0
        for lcb in alignments:
            lcb_count += 1
            q_idx = [i for i, rec in enumerate(lcb) if
                     rec.id in query_idset or any(rec.id.startswith(q) for q in query_idset)]
            r_idx = [i for i, rec in enumerate(lcb) if
                     rec.id in ref_idset or any(rec.id.startswith(r) for r in ref_idset)]
            if not q_idx or not r_idx: continue

            def aln_len(i): return len(str(lcb[i].seq))

            ref_idx, query_idx = max(((ri, qi) for ri in r_idx for qi in q_idx),
                                     key=lambda pr: aln_len(pr[0]) + aln_len(pr[1]))

            query_record = lcb[query_idx]
            ref_record = lcb[ref_idx]

            aln_q_raw = str(query_record.seq).upper().strip()
            aln_r_raw = str(ref_record.seq).upper().strip()

            q_strand = query_record.annotations.get("strand", "+")
            ref_strand = ref_record.annotations.get("strand", "+")

            aln_q = aln_q_raw
            aln_r = aln_r_raw
            aln_len_q = len(aln_q)
            aln_len_r = len(aln_r)

            if aln_len_q != aln_len_r:
                print(f"  SKIPPED LCB {lcb_count}: query/ref alignment lengths differ ({aln_len_q} vs {aln_len_r}).")
                # Debug: show first/last 20 chars to identify issue
                if abs(aln_len_q - aln_len_r) <= 5:
                    print(f"    Query ends with: {repr(aln_q[-20:])}")
                    print(f"    Ref ends with: {repr(aln_r[-20:])}")
                continue

            aln_len = aln_len_q

            # Coordinates
            q_start_0idx_maf = query_record.annotations["start"]
            q_size_maf = query_record.annotations["size"]
            q_srcSize = query_record.annotations.get("srcSize", full_seqlen)

            if str(q_strand) in ("-", "-1"):
                q_start_0idx_plus = q_srcSize - q_start_0idx_maf - q_size_maf
                q_start_1idx = q_start_0idx_plus + 1
                q_end_1idx = q_start_1idx + q_size_maf - 1
            else:
                q_start_0idx_plus = q_start_0idx_maf
                q_start_1idx = q_start_0idx_maf + 1
                q_end_1idx = q_start_1idx + q_size_maf - 1

            r_start_0idx_maf = ref_record.annotations["start"]
            r_size_maf = ref_record.annotations["size"]
            r_srcSize = ref_record.annotations.get("srcSize", ref_seqlen)

            if str(ref_strand) in ("-", "-1") and r_srcSize > 0:
                r_start_0idx_plus = r_srcSize - r_start_0idx_maf - r_size_maf
                r_start_1idx = r_start_0idx_plus + 1
                r_end_1idx = r_start_1idx + r_size_maf - 1
            else:
                r_start_1idx = r_start_0idx_maf + 1
                r_end_1idx = r_start_1idx + r_size_maf - 1

            # Verification Logic
            try:
                ungapped_q = aln_q.replace("-", "")
                ungapped_r = aln_r.replace("-", "")

                search_seq_q = str(Seq(ungapped_q).reverse_complement()) if str(q_strand) in ("-", "-1") else ungapped_q
                search_seq_r = str(Seq(ungapped_r).reverse_complement()) if str(ref_strand) in ("-", "-1") else ungapped_r

                if query_seq_full and search_seq_q:
                    found_q_idx = query_seq_full.find(search_seq_q)
                    if found_q_idx != -1:
                        found_q_start_1idx = found_q_idx + 1
                        diff_q = abs(found_q_start_1idx - q_start_1idx)
                        if diff_q > 50:
                            print(f"DEBUG [LCB {lcb_count}]: Query string match ({found_q_start_1idx}) is >50bp from calculated coordinate ({q_start_1idx}). Diff: {diff_q}bp")

                if ref_seq_main and search_seq_r:
                    found_r_idx = ref_seq_main.find(search_seq_r)
                    if found_r_idx != -1:
                        found_r_start_1idx = found_r_idx + 1
                        diff_r = abs(found_r_start_1idx - r_start_1idx)
                        if diff_r > 50:
                            print(f"DEBUG [LCB {lcb_count}]: Ref string match ({found_r_start_1idx}) is >50bp from calculated coordinate ({r_start_1idx}). Diff: {diff_r}bp")
            except Exception as e_verify:
                print(f"DEBUG [LCB {lcb_count}]: Error during string verification: {e_verify}")

            identity_pct_ignore_gaps, identity_pct_gaps_as_mismatch = _calculate_fragment_identities(aln_q, aln_r)

            lcb_data.append({
                "query_start": q_start_1idx, "query_end": q_end_1idx, "query_strand": q_strand,
                "ref_start": r_start_1idx, "ref_stop": r_end_1idx, "ref_strand": ref_strand,
                "identity_excl_gaps": identity_pct_ignore_gaps,
                "identity_incl_gaps": identity_pct_gaps_as_mismatch,
                "aligned_query_sequence": aln_q, "aligned_reference_sequence": aln_r,
            })

            arr1 = np.array(list(aln_q))
            arr2 = np.array(list(aln_r))
            is_match = (arr1 == arr2)
            is_not_gap_either = (arr1 != '-') & (arr2 != '-')
            matches = (is_match & is_not_gap_either).sum()
            valid_comparisons = is_not_gap_either.sum()
            total_lcb_matches += matches
            total_lcb_valid_comparisons += valid_comparisons
            M = (is_match & is_not_gap_either).astype(int)
            i_lcb_scores = M * 100
            q_bytes = np.frombuffer(aln_q.encode('ascii'), dtype='S1')
            non_gap_mask = (q_bytes != b'-')
            scores_to_apply = i_lcb_scores[non_gap_mask]
            num_bases = len(scores_to_apply)

            if num_bases > 0:
                indices = q_start_0idx_plus + np.arange(num_bases)

                valid_mask = (indices >= 0) & (indices < full_seqlen)
                valid_indices = indices[valid_mask]
                valid_scores = scores_to_apply[valid_mask]

                np.maximum.at(g_query_identity, valid_indices, valid_scores)

    except Exception as e:
        if "No alignments found" in str(e) or "empty file" in str(e):
            status("Warning: SibeliaZ MAF file was empty. No LCBs loaded.", "orange")
            return None, [], 0.0
        print(f"ERROR during SibeliaZ MAF parsing: {e}")
        import traceback
        traceback.print_exc()
        return None, [], 0.0

    global_wga_id_pct = (
            total_lcb_matches / total_lcb_valid_comparisons * 100) if total_lcb_valid_comparisons > 0 else 0.0
    status(f"Finished parsing SibeliaZ MAF: {lcb_count} LCBs processed.", "blue")

    return g_query_identity, lcb_data, global_wga_id_pct


def _merge_intervals(intervals):
    """
    Merge overlapping [start, end) intervals (0-based, end-exclusive).
    Returns a list of non-overlapping intervals sorted by start.
    """
    if not intervals:
        return []
    intervals = sorted(intervals)
    merged = [list(intervals[0])]
    for s, e in intervals[1:]:
        last_s, last_e = merged[-1]
        if s <= last_e:
            # Overlap or touching
            merged[-1][1] = max(last_e, e)
        else:
            merged.append([s, e])
    return [(s, e) for s, e in merged]


def _apply_mauve_coordinate_conversion(frag, apply_proactive_plus_one=True):
    """
    Apply proactive coordinate system conversion from Mauve output to 1-indexed.

    Mauve outputs coordinates in 0-indexed, half-open [start, end) format.
    This function converts them to 1-indexed, inclusive (start, end] coordinates.

    The key insight from extensive testing:
    - Mauve start coordinates need +1 (convert from 0-indexed to 1-indexed)
    - Mauve end coordinates stay the same (half-open end == inclusive end value)

    Args:
        frag: Fragment dict with keys: query_start, query_end, ref_start, ref_stop
        apply_proactive_plus_one: If True, add +1 to start coordinates (default True)

    Returns:
        Updated fragment dict with converted coordinates
    """
    if apply_proactive_plus_one:
        try:
            # Convert from 0-indexed [start, end) to 1-indexed (start, end]
            frag["query_start"] = int(frag.get("query_start", 0)) + 1
            frag["ref_start"] = int(frag.get("ref_start", 0)) + 1
            # End coordinates remain the same (half-open end == inclusive end value)
            frag["query_end"] = int(frag.get("query_end", 0))
            frag["ref_stop"] = int(frag.get("ref_stop", 0))
        except (TypeError, ValueError):
            pass  # If conversion fails, leave coordinates as-is

    return frag


def _attach_original_sequences_to_lcbs(lcb_list, query_seq, ref_seq, skip_coordinate_conversion=False):
    """
    Attach original genomic sequences to each LCB/fragment using 1-indexed coordinates.

    This function:
      1. OPTIONALLY converts coordinates: adds +1 to start (0-indexed → 1-indexed) for Mauve
      2. Extracts sequences using the 1-indexed inclusive coordinates
      3. FALLBACK: Verifies extraction via string matching if length mismatch occurs

    Args:
        lcb_list: List of LCB/fragment dictionaries with coordinate and sequence info
        query_seq: Full query sequence string
        ref_seq: Full reference sequence string
        skip_coordinate_conversion: If True, skip +1 conversion (for pre-converted SibeliaZ data).
                                   If False (default), apply +1 to start coords (for Mauve data).

    Coordinate Convention:
      - Mauve input: 0-indexed, half-open [start, end) format → add +1 to start
      - SibeliaZ input: already 1-indexed, inclusive (start, end] format → no conversion needed
      - Storage: 1-indexed, inclusive (start, end] format
      - Extraction: Add +1 to start coords (for Mauve), keep end coords unchanged

    Why Fallback Verification?
      String matching via `find()` validates that extracted sequences match the ungapped
      alignment, correcting for any coordinate inaccuracies.
    """
    if not lcb_list:
        return
    q_len = len(query_seq)
    r_len = len(ref_seq)

    for frag in lcb_list:
        try:
            # Step 1: PROACTIVE - Apply coordinate conversion from 0-indexed to 1-indexed (Mauve only)
            if not skip_coordinate_conversion:
                frag = _apply_mauve_coordinate_conversion(frag, apply_proactive_plus_one=True)

            q_start = int(frag.get("query_start", 0))
            q_end = int(frag.get("query_end", 0))
            r_start = int(frag.get("ref_start", 0))
            r_end = int(frag.get("ref_stop", 0))
            q_strand = frag.get("query_strand", "+")
            r_strand = frag.get("ref_strand", "+")

        except (TypeError, ValueError):
            continue

        # Get aligned sequences for fallback verification
        aln_q = frag.get("aligned_query_sequence", "")
        aln_r = frag.get("aligned_reference_sequence", "")
        ungapped_q = aln_q.replace("-", "") if aln_q else ""
        ungapped_r = aln_r.replace("-", "") if aln_r else ""

        # Query sequence extraction
        extracted_q = ""
        if 0 < q_start <= q_len and 0 < q_end <= q_len:
            # Normalize to get the leftmost and rightmost positions (1-indexed)
            q_left = min(q_start, q_end)
            q_right = max(q_start, q_end)

            # Extract using 1-indexed inclusive coordinates
            # Convert to 0-indexed for Python slicing: [q_left-1, q_right)
            q_s = q_left - 1
            q_e = q_right
            extracted_q = query_seq[q_s:q_e]

            # FALLBACK VERIFICATION: Only verify if we have ungapped alignment sequence
            if ungapped_q and len(extracted_q) != len(ungapped_q):
                print(
                    f"DEBUG: Query length mismatch after proactive +1 conversion. "
                    f"Expected {len(ungapped_q)}, got {len(extracted_q)}. "
                    f"Coords: {q_start}-{q_end}, ungapped sequence length mismatch.")

                # Try string matching as fallback
                if str(q_strand) in ["+", "1"]:
                    search_seq = ungapped_q
                else:
                    search_seq = str(Seq(ungapped_q).reverse_complement())

                # Search in window around expected position
                window_start = max(0, q_left - 100)
                window_end = min(q_len, q_right + 100)
                q_segment = query_seq[window_start:window_end]
                local_idx = q_segment.find(search_seq)

                if local_idx != -1:
                    # Found via string match! Update coordinates
                    actual_start = window_start + local_idx + 1  # Convert to 1-based
                    actual_end = actual_start + len(search_seq) - 1
                    extracted_q = query_seq[actual_start - 1:actual_end]
                    print(f"DEBUG: Query string match found! Updated from {q_start}-{q_end} "
                          f"to {actual_start}-{actual_end}")
                    frag["query_start"] = actual_start
                    frag["query_end"] = actual_end
                else:
                    print("DEBUG: Query string match failed. Using best-effort extraction.")

        frag["original_query_sequence"] = extracted_q

        extracted_r = ""
        if 0 < r_start <= r_len and 0 < r_end <= r_len:
            r_left = min(r_start, r_end)
            r_right = max(r_start, r_end)
            r_s = r_left - 1
            r_e = r_right
            extracted_r = ref_seq[r_s:r_e]
            if ungapped_r and len(extracted_r) != len(ungapped_r):
                print(
                    f"DEBUG: Reference length mismatch after proactive +1 conversion. "
                    f"Expected {len(ungapped_r)}, got {len(extracted_r)}. "
                    f"Coords: {r_start}-{r_end}, ungapped sequence length mismatch.")
                if str(r_strand) in ["+", "1"]: search_seq = ungapped_r
                else: search_seq = str(Seq(ungapped_r).reverse_complement())

                window_start = max(0, r_left - 100)
                window_end = min(r_len, r_right + 100)
                r_segment = ref_seq[window_start:window_end]
                local_idx = r_segment.find(search_seq)

                if local_idx != -1:
                    # Found via string match! Update coordinates
                    actual_start = window_start + local_idx + 1  # Convert to 1-based
                    actual_end = actual_start + len(search_seq) - 1
                    extracted_r = ref_seq[actual_start - 1:actual_end]
                    print(f"DEBUG: Reference string match found! Updated from {r_start}-{r_end} "
                          f"to {actual_start}-{actual_end}")
                    frag["ref_start"] = actual_start
                    frag["ref_stop"] = actual_end
                else:
                    print("DEBUG: Reference string match failed. Using best-effort extraction.")

        frag["original_reference_sequence"] = extracted_r


def _generate_coding_fragments(lcb_data, coding_intervals_0based, query_seq, ref_seq, features,
                               allow_code_fallback=False, log_queue=None):
    """
    From full LCBs, generate coding (CDS) alignment fragments with gene names.

    For each LCB:
      1. Build alignment-index -> query genomic coordinate map.
      2. Find all coding spans within the LCB.
      3. For each coding span, select alignment columns whose query genomic
         pos lies inside that span.
      4. Build aligned fragments and compute identities.
      5. Find overlapping gene and add gene name to fragment.
    """

    def status(msg, color="black"):
        if log_queue:
            log_queue.put(("status", (msg, color)))

    if not lcb_data:
        status("No LCB data provided for coding filtering.", "orange")
        return []

    status("Filtering alignment blocks for coding regions...", "black")
    coding_lcb_data = []

    # Merge coding intervals once
    merged_coding = _merge_intervals(coding_intervals_0based)

    # Get sequence length
    seq_length = 0
    if lcb_data:
        for lcb in lcb_data:
            q_end = max(int(lcb["query_start"]), int(lcb["query_end"]))
            seq_length = max(seq_length, q_end)

    print("\n=== DEBUG: _generate_coding_fragments ===")
    print(f"Total LCBs: {len(lcb_data)}")
    print(f"Total merged coding intervals: {len(merged_coding)}")

    feature_list = []
    for feature in features:
        start_0based = int(feature.start)
        end_0based = int(feature.end)
        feature_list.append((start_0based, end_0based, feature))

    for lcb_idx, lcb in enumerate(lcb_data):
        q_start_1idx = int(lcb["query_start"])
        q_end_1idx = int(lcb["query_end"])
        q_strand = lcb["query_strand"]
        r_start_1idx = int(lcb["ref_start"])
        r_end_1idx = int(lcb["ref_stop"])
        r_strand = lcb["ref_strand"]

        aln_q = lcb["aligned_query_sequence"]
        aln_r = lcb["aligned_reference_sequence"]
        aln_len = len(aln_q)
        if aln_len == 0:
            continue

        # Build alignment-index -> query genomic position map
        q_idx2pos = {}
        q_left = min(q_start_1idx, q_end_1idx)
        q_right = max(q_start_1idx, q_end_1idx)

        if str(q_strand) in ["+", "1"]:
            current_q_pos = q_left
            for i in range(aln_len):
                if aln_q[i] != "-":
                    q_idx2pos[i] = current_q_pos
                    current_q_pos += 1
        else:
            current_q_pos = q_right
            for i in range(aln_len):
                if aln_q[i] != "-":
                    q_idx2pos[i] = current_q_pos
                    current_q_pos -= 1

        # Build reference position map
        r_idx2pos = {}
        r_left = min(r_start_1idx, r_end_1idx)
        r_right = max(r_start_1idx, r_end_1idx)

        if str(r_strand) in ["+", "1"]:
            current_r_pos = r_left
            for i in range(aln_len):
                if aln_r[i] != "-":
                    r_idx2pos[i] = current_r_pos
                    current_r_pos += 1
        else:
            current_r_pos = r_right
            for i in range(aln_len):
                if aln_r[i] != "-":
                    r_idx2pos[i] = current_r_pos
                    current_r_pos -= 1

        if not q_idx2pos:
            print(f"[LCB {lcb_idx + 1}] No ungapped query positions in alignment, skipping.")
            continue

        # LCB span on query in 0-based coords
        q_positions_0based = [pos - 1 for pos in q_idx2pos.values()]
        lcb_start_0 = min(q_positions_0based)
        lcb_end_0 = max(q_positions_0based) + 1

        q_lcb_left = min(q_start_1idx, q_end_1idx)
        q_lcb_right = max(q_start_1idx, q_end_1idx) + 1
        r_lcb_left = min(r_start_1idx, r_end_1idx)
        r_lcb_right = max(r_start_1idx, r_end_1idx) + 1

        # Find overlapping coding regions
        overlapping_cds = []
        for cds_start, cds_end in merged_coding:
            if cds_start < lcb_end_0 and cds_end > lcb_start_0:
                overlap_start = max(cds_start, lcb_start_0)
                overlap_end = min(cds_end, lcb_end_0)
                if overlap_start < overlap_end:
                    overlapping_cds.append((overlap_start, overlap_end))

        if not overlapping_cds:
            continue

        overlapping_cds.sort()

        # Process each coding span
        for coding_start_0, coding_end_0 in overlapping_cds:
            indices_in_span = []
            for idx, pos in q_idx2pos.items():
                pos_0 = pos - 1
                if coding_start_0 <= pos_0 < coding_end_0: indices_in_span.append(idx)

            if not indices_in_span: continue

            aln_start_idx = min(indices_in_span)
            aln_end_idx_excl = max(indices_in_span) + 1

            fragment_q_seq = aln_q[aln_start_idx:aln_end_idx_excl]
            fragment_r_seq = aln_r[aln_start_idx:aln_end_idx_excl]
            if not fragment_q_seq: continue

            # Trim leading/trailing gap columns
            start_trim, end_trim = 0, 0
            L = len(fragment_q_seq)
            for k in range(L):
                if fragment_q_seq[k] == "-" and fragment_r_seq[k] == "-": start_trim += 1
                else: break
            for k in range(L - 1, -1, -1):
                if fragment_q_seq[k] == "-" and fragment_r_seq[k] == "-": end_trim += 1
                else: break

            frag_start_idx_global = aln_start_idx + start_trim
            frag_end_idx_global = aln_end_idx_excl - end_trim

            if frag_start_idx_global >= frag_end_idx_global: continue

            trimmed_q_seq = aln_q[frag_start_idx_global:frag_end_idx_global]
            trimmed_r_seq = aln_r[frag_start_idx_global:frag_end_idx_global]

            if not trimmed_q_seq: continue

            ungapped_q, ungapped_r = trimmed_q_seq.replace("-", ""), trimmed_r_seq.replace("-", "")
            q_len_no_gaps, r_len_no_gaps = len(ungapped_q), len(ungapped_r)

            # Filter short fragments
            if q_len_no_gaps < 10 or r_len_no_gaps < 10: continue

            # Map fragment back to FASTA coordinates
            if str(q_strand) in ["-", "-1"]: query_search_seq = str(Seq(ungapped_q).reverse_complement())
            else: query_search_seq = ungapped_q

            q_segment = query_seq[q_lcb_left - 1: q_lcb_right]
            local_q_idx = q_segment.find(query_search_seq)

            if local_q_idx == -1:
                global_q_idx = query_seq.find(query_search_seq)
                if global_q_idx == -1:
                    print(f"[LCB {lcb_idx + 1}] Could not locate query fragment in FASTA.")
                    continue
                q_start_adj = global_q_idx + 1
            else: q_start_adj = q_lcb_left + local_q_idx

            q_end_adj = q_start_adj + len(query_search_seq) - 1

            # Reference mapping
            if str(r_strand) in ["-", "-1"]: ref_search_seq = str(Seq(ungapped_r).reverse_complement())
            else: ref_search_seq = ungapped_r

            r_segment = ref_seq[r_lcb_left - 1: r_lcb_right]
            local_r_idx = r_segment.find(ref_search_seq)

            if local_r_idx == -1:
                global_r_idx = ref_seq.find(ref_search_seq)
                if global_r_idx == -1:
                    print(f"[LCB {lcb_idx + 1}] Could not locate reference fragment in FASTA.")
                    continue
                r_start_adj = global_r_idx + 1
            else: r_start_adj = r_lcb_left + local_r_idx

            r_end_adj = r_start_adj + len(ref_search_seq) - 1

            # Identity calculations
            id_excl, id_incl = _calculate_fragment_identities(trimmed_q_seq, trimmed_r_seq)
            diff_str, q_mismatch_hash, q_diff_only, r_diff_only = _generate_diff_strings(
                trimmed_q_seq, trimmed_r_seq, id_incl)

            # Original sequences
            q_orig_seq, r_orig_seq = "", ""
            if 1 <= q_start_adj <= len(query_seq) and 1 <= q_end_adj <= len(query_seq):
                q_orig_seq = query_seq[q_start_adj - 1: q_end_adj]

            if 1 <= r_start_adj <= len(ref_seq) and 1 <= r_end_adj <= len(ref_seq):
                r_orig_seq = ref_seq[r_start_adj - 1: r_end_adj]

            # Find gene name and ID using the same fallback logic as plot labels
            gene_name = None
            gene_id = None
            for fstart, fend, feature in feature_list:
                # Check if coding fragment overlaps with this feature
                if fstart <= q_start_adj <= fend or fstart <= q_end_adj <= fend or (
                        q_start_adj <= fstart and q_end_adj >= fend):
                    label = _label_for_feature(feature, allow_code_fallback)
                    if label:
                        gene_name = label
                        # Extract gene ID from feature attributes
                        attrs = getattr(feature, "attributes", {}) or {}
                        gene_id = _get_attr_value(attrs, "ID")
                        break

            if not gene_name: gene_name = "unknown"
            if not gene_id: gene_id = "unknown"

            coding_lcb_data.append({
                "gene_name": gene_name,
                "gene_id": gene_id,
                "query_start": q_start_adj,
                "query_end": q_end_adj,
                "query_strand": q_strand,
                "ref_start": r_start_adj,
                "ref_stop": r_end_adj,
                "ref_strand": r_strand,
                "identity_excl_gaps": id_excl,
                "identity_incl_gaps": id_incl,
                "aligned_query_sequence": trimmed_q_seq,
                "aligned_reference_sequence": trimmed_r_seq,
                "difference_string": diff_str,
                "query_mismatches_marked": q_mismatch_hash,
                "query_differences_only": q_diff_only,
                "ref_differences_only": r_diff_only,
                "original_query_sequence": q_orig_seq,
                "original_reference_sequence": r_orig_seq,
                "region_type": "coding",
            })

    status(f"Generated {len(coding_lcb_data)} coding alignment fragments.", "blue")
    return coding_lcb_data


def _generate_intergenic_fragments(lcb_data, coding_intervals_0based, query_seq, ref_seq, log_queue=None):
    """
    From full LCBs, generate intergenic alignment fragments.

    For each LCB:
      1. Build alignment-index -> query genomic coordinate map (respecting query strand).
      2. Compute intergenic spans (in genomic coords on the query).
      3. For each intergenic span, select alignment columns whose query genomic
         pos lies inside that span.
      4. From those columns, build aligned fragments (query+ref with gaps).
      5. For each fragment:
         - Remove gaps from query and reference.
         - Map fragment back to original FASTA coordinates by string search:
             * If strand '+': search ungapped seq in FASTA.
             * If strand '-': search reverse-complement of ungapped seq in FASTA.
         - Use that exact match to set query_start/query_end and ref_start/ref_stop.
         - Extract original (plus-orientation) genomic sequences for Excel.
    """

    def status(msg, color="black"):
        if log_queue: log_queue.put(("status", (msg, color)))

    if not lcb_data:
        status("No LCB data provided for intergenic filtering.", "orange")
        return []

    status("Filtering alignment blocks for intergenic regions...", "black")
    intergenic_lcb_data = []

    # Merge coding intervals once
    merged_coding = _merge_intervals(coding_intervals_0based)

    # Add synthetic coding regions at start and end to cover sequence boundaries
    seq_length = 0
    if lcb_data:
        for lcb in lcb_data:
            q_end = max(int(lcb["query_start"]), int(lcb["query_end"]))
            seq_length = max(seq_length, q_end)

    # Add synthetic regions at boundaries (in 0-based coords)
    if merged_coding:
        first_coding_start = merged_coding[0][0]
        last_coding_end = merged_coding[-1][1]

        if first_coding_start > 0: merged_coding.insert(0, (0, first_coding_start))
        if last_coding_end < seq_length: merged_coding.append((last_coding_end, seq_length))

        merged_coding = sorted(merged_coding)

    for lcb_idx, lcb in enumerate(lcb_data):
        q_start_1idx = int(lcb["query_start"])
        q_end_1idx = int(lcb["query_end"])
        q_strand = lcb["query_strand"]
        r_start_1idx = int(lcb["ref_start"])
        r_end_1idx = int(lcb["ref_stop"])
        r_strand = lcb["ref_strand"]

        aln_q = lcb["aligned_query_sequence"]
        aln_r = lcb["aligned_reference_sequence"]
        aln_len = len(aln_q)
        if aln_len == 0:
            continue

        q_idx2pos = {}
        q_left = min(q_start_1idx, q_end_1idx)
        q_right = max(q_start_1idx, q_end_1idx)

        if str(q_strand) in ["+", "1"]:
            current_q_pos = q_left
            for i in range(aln_len):
                if aln_q[i] != "-":
                    q_idx2pos[i] = current_q_pos
                    current_q_pos += 1
        else:
            current_q_pos = q_right
            for i in range(aln_len):
                if aln_q[i] != "-":
                    q_idx2pos[i] = current_q_pos
                    current_q_pos -= 1

        # Also build reference position map for accurate window calculation
        r_idx2pos = {}
        r_left = min(r_start_1idx, r_end_1idx)
        r_right = max(r_start_1idx, r_end_1idx)

        if str(r_strand) in ["+", "1"]:
            current_r_pos = r_left
            for i in range(aln_len):
                if aln_r[i] != "-":
                    r_idx2pos[i] = current_r_pos
                    current_r_pos += 1
        else:
            current_r_pos = r_right
            for i in range(aln_len):
                if aln_r[i] != "-":
                    r_idx2pos[i] = current_r_pos
                    current_r_pos -= 1

        if not q_idx2pos:
            print(f"[LCB {lcb_idx + 1}] No ungapped query positions in alignment, skipping.")
            continue

        # LCB span on query in 0-based coords
        q_positions_0based = [pos - 1 for pos in q_idx2pos.values()]
        lcb_start_0 = min(q_positions_0based)
        lcb_end_0 = max(q_positions_0based) + 1  # end-exclusive

        # LCB span on reference in 0-based coords
        if r_idx2pos:
            r_positions_0based = [pos - 1 for pos in r_idx2pos.values()]
            r_lcb_start_0 = min(r_positions_0based)
            r_lcb_end_0 = max(r_positions_0based) + 1  # end-exclusive
        else: r_lcb_start_0, r_lcb_end_0 = r_left - 1, r_right

        overlapping_cds = []
        for cds_start, cds_end in merged_coding:
            if cds_start < lcb_end_0 and cds_end > lcb_start_0:
                overlap_start = max(cds_start, lcb_start_0)
                overlap_end = min(cds_end, lcb_end_0)
                if overlap_start < overlap_end: overlapping_cds.append((overlap_start, overlap_end))

        intergenic_spans_0based = []
        current_pos = lcb_start_0
        overlapping_cds.sort()
        for i, (cds_start, cds_end) in enumerate(overlapping_cds):
            if current_pos < cds_start: intergenic_spans_0based.append((current_pos, cds_start))
            current_pos = cds_end
        if current_pos < lcb_end_0: intergenic_spans_0based.append((current_pos, lcb_end_0))

        if not intergenic_spans_0based: continue

        # Use calculated coordinates (from actual alignment positions)
        q_lcb_left = lcb_start_0 + 1  # Convert to 1-based
        q_lcb_right = lcb_end_0
        r_lcb_left = r_lcb_start_0 + 1  # Convert to 1-based
        r_lcb_right = r_lcb_end_0

        # For each intergenic span, extract fragment from alignment
        for ig_start_0, ig_end_0 in intergenic_spans_0based:
            if ig_start_0 >= ig_end_0: continue

            # collect all alignment indices whose query genomic pos falls in this span
            indices_in_span = []
            for idx, pos_1idx in q_idx2pos.items():
                pos_0 = pos_1idx - 1
                if ig_start_0 <= pos_0 < ig_end_0: indices_in_span.append(idx)

            if not indices_in_span: continue

            aln_start_idx, aln_end_idx_excl = min(indices_in_span), max(indices_in_span) + 1

            fragment_q_seq = aln_q[aln_start_idx:aln_end_idx_excl]
            fragment_r_seq = aln_r[aln_start_idx:aln_end_idx_excl]
            if not fragment_q_seq: continue

            # Trim leading/trailing columns where both query and ref are gaps
            start_trim, end_trim = 0, 0
            L = len(fragment_q_seq)
            for k in range(L):
                if fragment_q_seq[k] == "-" and fragment_r_seq[k] == "-": start_trim += 1
                else: break
            for k in range(L - 1, -1, -1):
                if fragment_q_seq[k] == "-" and fragment_r_seq[k] == "-": end_trim += 1
                else: break

            frag_start_idx_global = aln_start_idx + start_trim
            frag_end_idx_global = aln_end_idx_excl - end_trim

            if frag_start_idx_global >= frag_end_idx_global: continue

            trimmed_q_seq = aln_q[frag_start_idx_global:frag_end_idx_global]
            trimmed_r_seq = aln_r[frag_start_idx_global:frag_end_idx_global]

            if not trimmed_q_seq: continue

            ungapped_q, ungapped_r = trimmed_q_seq.replace("-", ""), trimmed_r_seq.replace("-", "")
            q_len_no_gaps, r_len_no_gaps = len(ungapped_q), len(ungapped_r)

            # Filter short fragments if you want (can relax if needed)
            if q_len_no_gaps < 10 or r_len_no_gaps < 10: continue

            # QUERY mapping
            if str(q_strand) in ["-", "-1"]: query_search_seq = str(Seq(ungapped_q).reverse_complement())
            else: query_search_seq = ungapped_q

            # Use calculated LCB window for search
            q_segment = query_seq[q_lcb_left - 1: q_lcb_right]
            local_q_idx = q_segment.find(query_search_seq)

            if local_q_idx == -1:
                global_q_idx = query_seq.find(query_search_seq)
                if global_q_idx == -1:
                    continue
                q_start_adj = global_q_idx + 1
            else: q_start_adj = q_lcb_left + local_q_idx

            q_end_adj = q_start_adj + len(query_search_seq) - 1

            # REFERENCE mapping
            if str(r_strand) in ["-", "-1"]: ref_search_seq = str(Seq(ungapped_r).reverse_complement())
            else: ref_search_seq = ungapped_r

            r_segment = ref_seq[r_lcb_left - 1: r_lcb_right]
            local_r_idx = r_segment.find(ref_search_seq)

            if local_r_idx == -1:
                global_r_idx = ref_seq.find(ref_search_seq)
                if global_r_idx == -1: continue

                print(
                    f"[LCB {lcb_idx + 1}]   WARNING: Reference found globally at {global_r_idx + 1}, "
                    f"but LCB says {r_start_1idx}-{r_end_1idx}. Trusting sequence.")
                r_start_adj = global_r_idx + 1
            else: r_start_adj = r_lcb_left + local_r_idx

            r_end_adj = r_start_adj + len(ref_search_seq) - 1

            # Identity and diff strings (still computed on aligned strings)
            id_excl, id_incl = _calculate_fragment_identities(trimmed_q_seq, trimmed_r_seq)
            diff_str, q_mismatch_hash, q_diff_only, r_diff_only = _generate_diff_strings(
                trimmed_q_seq, trimmed_r_seq, id_incl
            )

            # Original sequences (always in plus orientation, as in FASTA)
            q_orig_seq, r_orig_seq = "", ""
            if 1 <= q_start_adj <= len(query_seq) and 1 <= q_end_adj <= len(query_seq):
                extracted = query_seq[q_start_adj - 1: q_end_adj]
                q_orig_seq = extracted

            if 1 <= r_start_adj <= len(ref_seq) and 1 <= r_end_adj <= len(ref_seq):
                extracted = ref_seq[r_start_adj - 1: r_end_adj]
                r_orig_seq = extracted

            intergenic_lcb_data.append(
                {
                    "query_start": q_start_adj,
                    "query_end": q_end_adj,
                    "query_strand": q_strand,
                    "ref_start": r_start_adj,
                    "ref_stop": r_end_adj,
                    "ref_strand": r_strand,
                    "identity_excl_gaps": id_excl,
                    "identity_incl_gaps": id_incl,
                    "aligned_query_sequence": trimmed_q_seq,
                    "aligned_reference_sequence": trimmed_r_seq,
                    "difference_string": diff_str,
                    "query_mismatches_marked": q_mismatch_hash,
                    "query_differences_only": q_diff_only,
                    "ref_differences_only": r_diff_only,
                    "original_query_sequence": q_orig_seq,
                    "original_reference_sequence": r_orig_seq,
                    "region_type": "non-coding",
                }
            )

    status(f"Generated {len(intergenic_lcb_data)} intergenic alignment fragments.", "blue")
    return intergenic_lcb_data


def _calculate_global_intergenic_wga(intergenic_lcb_data, total_intergenic_bases, log_queue=None):
    """
    Calculates a global WGA identity for intergenic regions:

      - Numerator: matches in intergenic regions covered by LCBs (gaps = mismatches).
      - Denominator: TOTAL intergenic bases in the query genome
                     (including regions not covered by any LCB, which
                      are effectively counted as 0% identity).

    This matches your requirement that unaligned intergenic regions
    are treated as all mismatches.
    """

    def status(msg, color="black"):
        if log_queue: log_queue.put(("status", (msg, color)))

    if total_intergenic_bases <= 0:
        status("No intergenic bases defined for global WGA ID.", "orange")
        return 0.0

    if not intergenic_lcb_data:
        status("No intergenic fragments to calculate global WGA ID.", "orange")
        return 0.0

    status("Calculating global intergenic identity (gaps as mismatch)...", "black")

    total_intergenic_matches = 0

    for frag in intergenic_lcb_data:
        q_seq = frag["aligned_query_sequence"]
        r_seq = frag["aligned_reference_sequence"]
        if not q_seq or not r_seq or len(q_seq) != len(r_seq): continue

        arr1, arr2 = np.array(list(q_seq)), np.array(list(r_seq))
        is_match = (arr1 == arr2)
        is_not_gap_either = (arr1 != '-') & (arr2 != '-')
        matches = (is_match & is_not_gap_either).sum()
        total_intergenic_matches += matches

    wga_id_intergenic_incl_gaps = (total_intergenic_matches / total_intergenic_bases) * 100.0
    status(f"Global intergenic WGA ID (gaps as mismatches): {wga_id_intergenic_incl_gaps:.2f}%", "blue")
    return wga_id_intergenic_incl_gaps


# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# BLAST Data Processing
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────

def _enrich_blast_df(blast_df):
    """
    Calculates identities and diff strings for the BLAST dataframe.
    Optimized: Uses list iteration (zip) instead of slow iterrows().
    """
    if blast_df.empty or "qseq" not in blast_df.columns: return blast_df

    # Extract columns to lists for fast iteration
    q_seqs = blast_df['qseq'].astype(str).str.upper().values
    s_seqs = blast_df['sseq'].astype(str).str.upper().values
    s_starts = blast_df['sstart'].values
    s_ends = blast_df['send'].values

    # Pre-calculate data
    identities_excl = []
    identities_incl = []
    diff_strings = []
    q_mm_hashes = []
    q_diff_only = []
    r_diff_only = []

    # Fast iteration
    for q, r in zip(q_seqs, s_seqs):
        # 1. Calculate identities
        id_excl, id_incl = _calculate_fragment_identities(q, r)
        identities_excl.append(id_excl)
        identities_incl.append(id_incl)

        # 2. Generate diff strings
        ds, qmh, qdo, rdo = _generate_diff_strings(q, r, id_incl)
        diff_strings.append(ds)
        q_mm_hashes.append(qmh)
        q_diff_only.append(qdo)
        r_diff_only.append(rdo)

    # 3. Vectorized strand calculation
    # query_strand is always "+" for BLASTn output
    # ref_strand is "+" if sstart <= send, else "-"
    ref_strands = np.where(s_starts <= s_ends, "+", "-")

    # Assign new columns directly (much faster than row-by-row)
    blast_df['identity_excl_gaps'] = identities_excl
    blast_df['identity_incl_gaps'] = identities_incl
    blast_df['difference_string'] = diff_strings
    blast_df['query_mismatches_marked'] = q_mm_hashes
    blast_df['query_differences_only'] = q_diff_only
    blast_df['ref_differences_only'] = r_diff_only
    blast_df['query_strand'] = "+"
    blast_df['ref_strand'] = ref_strands

    return blast_df


# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# Legacy & K-mer Analysis
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────

def generate_hierarchical_legacy_report(query_seq, ref_records_dict, gene_features, max_k, min_k, log_queue=None):
    """
    Implements "greedy hierarchical" search against ALL reference contigs.
    Finds all matches for k=max_k, "masks" them, then
    finds all matches for k=max_k-1, and so on.
    """

    def status(msg, color="black"):
        if log_queue: log_queue.put(("status", (msg, color)))

    q_len = len(query_seq)
    query_mask = np.zeros(q_len, dtype=bool)  # Tracks "claimed" query bases
    coding_intervals = sorted([(f.start, f.end) for f in gene_features])  # 1-based inclusive

    final_all_blocks, final_intergenic_blocks = [], []

    for k in range(max_k, min_k - 1, -1):
        status(f"Report: Scanning for {k}-mer exact matches...")

        # 1. Build a GLOBAL k-mer map from ALL contigs
        ref_kmer_map = defaultdict(list)
        for contig_id, ref_record in ref_records_dict.items():
            ref_seq = str(ref_record.seq).upper()
            seq_len = len(ref_seq)
            if seq_len < k:
                continue
            for i in range(seq_len - k + 1):
                kmer = ref_seq[i:i + k]
                ref_kmer_map[kmer].append((contig_id, i))

        if not ref_kmer_map: continue

        # 2. Scan the query
        for q_pos_0based in range(q_len - k + 1):

            # 3. Check if any part of this k-mer is already claimed
            if query_mask[q_pos_0based: q_pos_0based + k].any(): continue

            kmer = query_seq[q_pos_0based: q_pos_0based + k]

            # 4. If it's a match, claim it and store all hits
            if kmer in ref_kmer_map:
                # Claim this query region
                query_mask[q_pos_0based: q_pos_0based + k] = True
                q_start_1based = q_pos_0based + 1
                q_end_1based = q_pos_0based + k
                k_start, k_end = q_start_1based, q_end_1based
                is_coding_hit = False
                for gene_start, gene_end in coding_intervals:
                    # Check if the k-mer (k_start to k_end) overlaps with the gene (gene_start to gene_end)
                    if k_start <= gene_end and k_end >= gene_start:  # Overlap
                        is_coding_hit = True
                        break  # Found an overlap, no need to check more

                # Create a block for each ref match
                for contig_id, r_pos_0based in ref_kmer_map[kmer]:
                    r_start_1based = r_pos_0based + 1
                    r_end_1based = r_pos_0based + k

                    block = {
                        "query_start": q_start_1based,
                        "query_end": q_end_1based,
                        "query_strand": "+",
                        "ref_contig": contig_id,
                        "ref_start": r_start_1based,
                        "ref_stop": r_end_1based,
                        "ref_strand": "+",
                        "identity_excl_gaps": 100.0,
                        "identity_incl_gaps": 100.0,
                        "aligned_query_sequence": kmer,
                        "aligned_reference_sequence": kmer,
                        "original_query_sequence": kmer,
                        "original_reference_sequence": kmer
                    }

                    final_all_blocks.append(block)
                    if not is_coding_hit:
                        final_intergenic_blocks.append(block)

    status(f"Generated {len(final_all_blocks):,} hierarchical legacy blocks.", "blue")
    return final_all_blocks, final_intergenic_blocks


def _compare_legacy_to_wga(legacy_intergenic_blocks, mauve_intergenic_lcb_data, sibeliaz_intergenic_lcb_data):
    """
    Compares legacy k-mer blocks to WGA blocks (Mauve/SibeliaZ) to find overlaps.
    This is a test/debug function to print stats to the console.
    """
    print("\n--- Legacy Alignment Comparison ---")
    if not legacy_intergenic_blocks:
        print("No legacy blocks to compare.")
        print("-----------------------------------\n")
        return

    # 1. Create a single list of all WGA intervals
    wga_intervals = []
    for lcb in mauve_intergenic_lcb_data: wga_intervals.append((lcb['query_start'], lcb['query_end']))
    for lcb in sibeliaz_intergenic_lcb_data: wga_intervals.append((lcb['query_start'], lcb['query_end']))

    # 2. Merge all overlapping WGA intervals for efficient lookup
    merged_wga = []
    if wga_intervals:
        wga_intervals.sort(key=lambda x: x[0])
        merged_wga = [list(wga_intervals[0])]
        for current_start, current_end in wga_intervals[1:]:
            last_start, last_end = merged_wga[-1]
            # Merge if the new block starts at or before the last one ends (+1 to merge adjacent blocks)
            if current_start <= last_end + 1: merged_wga[-1][1] = max(last_end, current_end)
            else: merged_wga.append([current_start, current_end])

    # 3. Iterate through legacy blocks and check for any overlap
    overlapping_count, unique_count = 0, 0

    for leg_block in legacy_intergenic_blocks:
        leg_start, leg_end = leg_block['query_start'], leg_block['query_end']
        found_overlap = False

        for wga_start, wga_end in merged_wga:
            # Standard overlap check:
            if leg_start < wga_end and leg_end > wga_start:
                found_overlap = True
                break  # Found an overlap, no need to check other WGA blocks

        if found_overlap: overlapping_count += 1
        else: unique_count += 1


def calculate_legacy_base_comparison_ring_data(query_seq, ref_seq_concatenated, gene_features, window_size, step_size,
                                               log_queue=None):
    """
    Performs sliding window analysis using simple base-by-base comparison.
    Compares query[window] to ref[window] at the same coordinates.
    """

    def status(msg, color="black"):
        if log_queue: log_queue.put(("status", (msg, color)))

    status("Running Legacy (1:1 Base Comparison) analysis for plot...", "black")

    coding_intervals = sorted([(f.start, f.end) for f in gene_features])
    coding_scores, non_coding_scores, all_scores_for_avg = [], [], []
    q_len, r_len = len(query_seq), len(ref_seq_concatenated)

    for w_start in range(0, q_len - window_size + 1, step_size):
        w_end = w_start + window_size
        midpoint = w_start + window_size // 2

        # Check if window is out of bounds for the reference
        if w_end > r_len:
            score = 0.0
        else:
            # Use numpy for fast base-by-base comparison
            query_slice = np.array(list(query_seq[w_start:w_end]))
            ref_slice = np.array(list(ref_seq_concatenated[w_start:w_end]))
            matches = (query_slice == ref_slice).sum()
            score = (matches / window_size) * 100.0
        score_tuple = (midpoint, score)
        all_scores_for_avg.append(score)

        # Check coding status using 1-based midpoint
        if _is_coding(midpoint + 1, coding_intervals): coding_scores.append(score_tuple)
        else: non_coding_scores.append(score_tuple)

    global_avg_identity = np.mean(all_scores_for_avg) if all_scores_for_avg else 0.0

    if log_queue:
        total_windows = len(coding_scores) + len(non_coding_scores)
        log_queue.put(
            ("status", (f"Completed legacy base comparison on {total_windows} windows.", "blue")))

    return coding_scores, non_coding_scores, global_avg_identity


def calculate_global_kmer_ring_data(query_seq, ref_kmers_global_set, gene_features, window_size, step_size,
                                    log_queue=None):
    """
    Performs sliding window analysis over the query genome.
    Score is based on the percentage of a window's k-mers (k=10) that
    exist ANYWHERE in the entire reference genome (passed as a set).
    """
    KMER_SIZE = 10

    # Sliding window over the query sequence
    coding_intervals = sorted([(f.start, f.end) for f in gene_features])
    coding_scores, non_coding_scores = [], []
    seqlen = len(query_seq)

    for w_start in range(0, seqlen - window_size + 1, step_size):
        w_end = w_start + window_size
        midpoint = w_start + window_size // 2

        query_slice = query_seq[w_start:w_end]
        query_window_kmers = _get_kmers(query_slice, KMER_SIZE)

        if not query_window_kmers: score_tuple = (midpoint, 0)
        else:
            # Check query window k-mers against the GLOBAL reference set
            common_kmers = query_window_kmers.intersection(ref_kmers_global_set)
            num_common = len(common_kmers)
            num_query_kmers = len(query_window_kmers)
            normalized_score = (num_common / num_query_kmers) * 100
            score_tuple = (midpoint, normalized_score)

        if _is_coding(midpoint, coding_intervals): coding_scores.append(score_tuple)
        else: non_coding_scores.append(score_tuple)

    all_scores = [s for _, s in coding_scores] + [s for _, s in non_coding_scores]
    global_avg_identity = np.mean(all_scores) if all_scores else 0.0

    if log_queue:
        total_windows = len(coding_scores) + len(non_coding_scores)
        log_queue.put(
            ("status", (f"Completed global k-mer analysis (k={KMER_SIZE}) on {total_windows} windows.", "blue")))
    return coding_scores, non_coding_scores, global_avg_identity


# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# Main Data Pipeline
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────

def calculate_gc_skew_data(sequence, window_size):
    """
    Vectorized calculation of GC content and skew using Numpy.
    ~50x-100x faster than the loop-based method.
    """
    seqlen = len(sequence)
    if seqlen == 0:
        return np.zeros(0), np.zeros(0)

    # 1. Convert string to numpy byte array for fast comparison
    # We use 'S1' (byte string) or uint8 to work with ASCII codes directly
    try:
        # Ensure we have a clean byte string
        if isinstance(sequence, str):
            seq_bytes = sequence.encode('ascii', errors='ignore')
        else:
            seq_bytes = bytes(sequence)
    except Exception:
        # Fallback for Biopython Seq objects or others
        seq_bytes = str(sequence).encode('ascii', errors='ignore')

    seq_arr = np.frombuffer(seq_bytes, dtype='S1')

    # 2. Handle the "Main" body (perfectly divisible by window_size)
    n_windows = seqlen // window_size
    cutoff = n_windows * window_size

    # If genome is smaller than one window, handle it gracefully
    if n_windows == 0:
        g_count = np.sum(seq_arr == b'G')
        c_count = np.sum(seq_arr == b'C')
        gc_sum = g_count + c_count
        if gc_sum == 0:
            return np.array([0.0]), np.array([0.0])
        return np.array([gc_sum / seqlen]), np.array([(g_count - c_count) / gc_sum])

    main_body = seq_arr[:cutoff]

    # Reshape into (n_windows, window_size) - this is the "Vectorization Trick"
    # We create a 2D view of the data without copying memory
    windows = main_body.reshape(n_windows, window_size)

    # Count Gs and Cs (Vectorized summation across rows)
    # This replaces the inner loop entirely
    g_counts = np.sum(windows == b'G', axis=1)
    c_counts = np.sum(windows == b'C', axis=1)

    # Calculate metrics for the main body
    gc_sum = g_counts + c_counts

    # Avoid division by zero warnings
    with np.errstate(divide='ignore', invalid='ignore'):
        # GC Content: (G+C) / Window_Size
        gc_vals = gc_sum / window_size

        # GC Skew: (G-C) / (G+C)
        skew_vals = (g_counts - c_counts) / gc_sum

    skew_vals = np.nan_to_num(skew_vals)

    # 3. Handle the "Tail" (remaining bases at the end)
    if cutoff < seqlen:
        tail = seq_arr[cutoff:]
        g_tail = np.sum(tail == b'G')
        c_tail = np.sum(tail == b'C')
        l_tail = len(tail)

        gc_sum_tail = g_tail + c_tail
        if l_tail > 0:
            gc_tail_val = gc_sum_tail / l_tail
        else:
            gc_tail_val = 0.0

        if gc_sum_tail > 0:
            skew_tail_val = (g_tail - c_tail) / gc_sum_tail
        else:
            skew_tail_val = 0.0

        # Append tail to results
        gc_vals = np.append(gc_vals, gc_tail_val)
        skew_vals = np.append(skew_vals, skew_tail_val)

    # 4. Expand back to full genome resolution for plotting
    # The plot expects an array of size 'seqlen', so we repeat the values
    gc_arr_full = np.repeat(gc_vals, window_size)
    skew_arr_full = np.repeat(skew_vals, window_size)

    # The tail repeat might overshoot or undershoot slightly due to the math,
    # so we clip exactly to seqlen to prevent array shape errors
    return gc_arr_full[:seqlen], skew_arr_full[:seqlen]


def _sample_identity_profile(g_query_identity, full_seqlen, window_size, step_size, gene_features, log_queue=None):
    """
    Optimized sampling using Numpy vectorization (Cumulative Sum approach).
    Replaces the slow loop-based method for instant calculation.
    """

    def status(msg, color="black"):
        if log_queue: log_queue.put(("status", (msg, color)))

    if g_query_identity is None:
        status("Skipping identity sampling: No alignment data.", "orange")
        return [], [], 0.0

    status("Sampling identity profile (Vectorized)...", "black")

    # 1. Validate inputs
    if full_seqlen < window_size:
        return [], [], 0.0

    # 2. Create boolean mask for coding regions
    is_coding_mask = np.zeros(full_seqlen, dtype=bool)
    # gene_features uses 1-based inclusive coordinates
    for feat in gene_features:
        start = max(0, int(feat.start) - 1)
        end = min(full_seqlen, int(feat.end))
        if start < end:
            is_coding_mask[start:end] = True

    # 3. Generate window indices
    # We want windows starting at 0, step, 2*step...
    n_windows = (full_seqlen - window_size) // step_size + 1

    if n_windows <= 0:
        return [], [], 0.0

    starts = np.arange(n_windows) * step_size
    ends = starts + window_size
    # Midpoints for plotting (0-based)
    midpoints = starts + (window_size // 2)

    # 4. Calculate Window Averages using Cumulative Sum (Integral Image)
    # This allows calculating the sum of any window in O(1) time
    # Pad with 0 at the start for easier indexing
    cumsum_vec = np.insert(np.cumsum(g_query_identity), 0, 0)

    # Sum of window [start, end) = cumsum[end] - cumsum[start]
    # Note: cumsum_vec is shifted by 1 index, so cumsum_vec[end] corresponds to sum up to index end-1
    window_sums = cumsum_vec[ends] - cumsum_vec[starts]
    scores = window_sums / window_size

    # 5. Determine coding status for each window (BEFORE gap-filling for accurate statistics)
    # Check the boolean mask at the exact midpoint index
    # We clip to ensure we don't go out of bounds (though logic shouldn't allow it)
    safe_midpoints = np.clip(midpoints, 0, full_seqlen - 1)
    window_is_coding = is_coding_mask[safe_midpoints]

    # 6. Calculate statistics using REAL windowed data only (before gap-filling)
    # Create (midpoint, score) pairs from real data
    real_combined_data = np.column_stack((midpoints, scores))

    # Calculate global average for non-coding regions using only real windows
    noncoding_avg = 0.0
    if np.any(~window_is_coding):
        noncoding_avg = np.mean(real_combined_data[~window_is_coding][:, 1])

    # 7. Fill gaps at start and end by extrapolating the nearest score (for visual continuity only)
    if len(scores) > 0:
        # 1. Fill gap at start
        first_mid = midpoints[0]
        if first_mid > 0:
            # Generate points: first_mid-step, first_mid-2*step... down to 0
            pre_m = np.arange(first_mid - step_size, -1, -step_size)[::-1]

            # Ensure 0 is included
            if len(pre_m) == 0 or pre_m[0] > 0:
                pre_m = np.insert(pre_m, 0, 0)

            pre_s = np.full(len(pre_m), scores[0])
            # Determine coding status for gap-filled start points
            safe_pre_m = np.clip(pre_m, 0, full_seqlen - 1)
            pre_is_coding = is_coding_mask[safe_pre_m]

            midpoints = np.concatenate((pre_m, midpoints))
            scores = np.concatenate((pre_s, scores))
            window_is_coding = np.concatenate((pre_is_coding, window_is_coding))

        # 2. Fill gap at end
        last_mid = midpoints[-1]
        if last_mid < full_seqlen:
            # Generate points: last_mid+step ... up to full_seqlen
            post_m = np.arange(last_mid + step_size, full_seqlen + 1, step_size)

            # Ensure full_seqlen is included
            if len(post_m) == 0 or post_m[-1] < full_seqlen:
                post_m = np.append(post_m, full_seqlen)

            post_s = np.full(len(post_m), scores[-1])
            # Determine coding status for gap-filled end points
            safe_post_m = np.clip(post_m, 0, full_seqlen - 1)
            post_is_coding = is_coding_mask[safe_post_m]

            midpoints = np.concatenate((midpoints, post_m))
            scores = np.concatenate((scores, post_s))
            window_is_coding = np.concatenate((window_is_coding, post_is_coding))

    # 8. Split into coding and non-coding lists (now includes gap-filled data for plotting)
    # Create (midpoint, score) pairs
    combined_data = np.column_stack((midpoints, scores))

    # Filter using the boolean mask
    coding_scores = combined_data[window_is_coding].tolist()
    non_coding_scores = combined_data[~window_is_coding].tolist()

    total_windows = len(coding_scores) + len(non_coding_scores)
    status(f"Identity sampling complete: {total_windows} windows (Fast Mode).", "blue")

    # Convert to tuples for compatibility with existing code structure
    coding_scores = [tuple(x) for x in coding_scores]
    non_coding_scores = [tuple(x) for x in non_coding_scores]

    return coding_scores, non_coding_scores, noncoding_avg


def _rerun_blast_only(data, fasta_path, ref_fasta_path, annot_path, fasta_id, seqlen,
                      blast_task, blast_word_size, blast_reward, blast_penalty, blast_evalue,
                      blast_num_threads, homology_threshold, allow_code_fallback,
                      manual_find, manual_replace, highlight_bed_path, log_queue=None):
    """
    Re-runs BLAST only and updates gene features in-place, preserving all identity ring data.
    OPTIMIZED: Skips GFF re-parsing by updating existing feature objects directly.
    """

    def status(msg, color="black"):
        if log_queue: log_queue.put(("status", (msg, color)))

    status("Re-running BLASTn with new parameters...")
    blast_tab_path = fasta_path.with_suffix(".vs_ref.blast.tab")

    # 1. Run BLAST
    _run_blastn(fasta_path, ref_fasta_path, blast_tab_path,
                task=blast_task, word_size=blast_word_size, reward=blast_reward,
                penalty=blast_penalty, evalue=blast_evalue,
                num_threads=blast_num_threads)

    # 2. Parse BLAST Output
    blast_df = pd.read_csv(blast_tab_path, sep="\t",
                           names=["qseqid", "sseqid", "pident", "length", "mismatch", "gapopen", "qstart", "qend",
                                  "sstart", "send", "evalue", "bitscore", "qseq", "sseq"], comment='#')

    status("Updating gene features with new BLAST results...")

    # 3. Retrieve existing features (Do NOT reload GFF)
    features = data.get("gene_feats", [])

    # 4. Reset Feature State & Identify ROIs
    # We must reset homology/colors first, but preserve ROI (Red) status
    roi_indices = set()

    for i, f in enumerate(features):
        # Detect if this was a Highlighted ROI (Red)
        if getattr(f, 'color', '') == '#e53935':
            roi_indices.add(i)

        # Reset properties to default (Non-homologous Grey)
        f.best_identity = 0.0
        f.color = "#cccccc"
        f.best_sstart = None
        f.best_send = None
        f.best_sseqid = None
        f.best_qseqid = None

        # Re-apply Manual Label Override (in case params changed)
        if manual_find and manual_replace:
            attrs = getattr(f, "attributes", {}) or {}
            gene_name = _get_attr_value(attrs, "gene")
            if gene_name and gene_name.lower() == manual_find.lower():
                f.label = manual_replace

    # 5. Batch Match New BLAST Hits (Vectorized)
    if not blast_df.empty and features:
        # This function updates f.best_identity and sets f.color="#43a047" (Green) if match > threshold
        _batch_assign_homology(features, blast_df, homology_threshold)

    # 6. Restore ROI Colors (Red)
    # _batch_assign_homology sets matches to Green, so we must overwrite that for ROIs
    for i in roi_indices:
        features[i].color = '#e53935'

    # 7. Add is_coding column to blast_df (Vectorized)
    if not blast_df.empty and features:
        gene_starts = np.array([f.start for f in features])
        gene_ends = np.array([f.end for f in features])
        hit_starts_all = blast_df[['qstart', 'qend']].min(axis=1).values
        hit_ends_all = blast_df[['qstart', 'qend']].max(axis=1).values
        is_coding_list = []
        chunk_size = 2000
        for i in range(0, len(hit_starts_all), chunk_size):
            hit_starts_chunk = hit_starts_all[i:i + chunk_size]
            hit_ends_chunk = hit_ends_all[i:i + chunk_size]
            # Check overlap: (Hit_Start < Gene_End) and (Hit_End > Gene_Start)
            overlap_matrix = (hit_starts_chunk[:, np.newaxis] < gene_ends) & (
                    hit_ends_chunk[:, np.newaxis] > gene_starts)
            is_coding_list.extend(overlap_matrix.any(axis=1))
        blast_df['is_coding'] = is_coding_list
    else:
        blast_df['is_coding'] = False

    # 8. Enrich blast_df with additional columns
    blast_df = _enrich_blast_df(blast_df)

    # 9. Update the data dictionary
    data["blast_df"] = blast_df
    data["gene_feats"] = features  # (Updated in-place, but good to be explicit)

    status("BLAST re-run complete. Features updated successfully.", "blue")
    return data


def _process_data(fasta, annot, ref_fasta, window, highlight_bed_path, allow_code_fallback, homology_threshold,
                  manual_find, manual_replace, log_queue=None, nc_window_size=500, nc_step_size=50,
                  blast_task="blastn", blast_word_size="11", blast_reward="2", blast_penalty="-3", blast_evalue="10",
                  blast_num_threads=None,
                  identity_algorithm="Mauve + SibeliaZ Fallback", xmfa_file=None, force_mauve=False,
                  sibeliaz_maf=None, force_sibeliaz=False, run_legacy_report=True,
                  mauve_disable_backbone=False, mauve_mums=False, mauve_collinear=False,
                  mauve_skip_refinement=False, mauve_skip_gapped_alignment=False, mauve_seed_family=False,
                  mauve_no_weight_scaling=False, mauve_mem_clean=False, mauve_seed_weight="",
                  mauve_max_gapped_aligner_length="", mauve_max_breakpoint_distance_scale="",
                  mauve_conservation_distance_scale="", mauve_bp_dist_estimate_min_score="",
                  mauve_gap_open="", mauve_gap_extend="", mauve_weight="", mauve_min_scaled_penalty="",
                  mauve_hmm_p_go_homologous="", mauve_hmm_p_go_unrelated=""):
    def status(msg, color="black"):
        if log_queue: log_queue.put(("status", (msg, color)))

    def _calculate_true_anib(query_seq, ref_fasta_path, blast_num_threads="1", temp_dir_obj=None):
        """
        Calculates True ANIb (Average Nucleotide Identity via BLAST) according to standard protocols.
        Includes safety check for spaces/symbols in reference filename.
        """
        FRAGMENT_SIZE = 1020
        MIN_COVERAGE = 0.7  # 70% of fragment length must align

        # 1. Generate Fragments
        seq_len = len(query_seq)
        num_fragments = (seq_len // FRAGMENT_SIZE) + (1 if seq_len % FRAGMENT_SIZE != 0 else 0)

        if seq_len == 0:
            return 0.0, 0.0, 0, 0

        # Create a temporary directory if one wasn't provided
        local_temp_manager = None
        if temp_dir_obj is None:
            local_temp_manager = tempfile.TemporaryDirectory()
            temp_dir = Path(local_temp_manager.name)
        else:
            temp_dir = Path(temp_dir_obj)

        try:
            fragments_path = temp_dir / "fragments.fasta"

            with open(fragments_path, "w") as f_out:
                for i in range(0, seq_len, FRAGMENT_SIZE):
                    chunk = query_seq[i: i + FRAGMENT_SIZE]
                    # Label fragment with ID and length for easy parsing later
                    f_out.write(f">frag_{i}_{len(chunk)}\n{chunk}\n")

            # 2. Setup BLAST DB (With Safety Check)
            makeblastdb_exe = _find_blast_executable("makeblastdb")
            blastn_exe = _find_blast_executable("blastn")

            if not makeblastdb_exe or not blastn_exe:
                print("[ANIb] BLAST tools not found, skipping ANIb.")
                return 0.0, 0.0, 0, 0

            # Check if the path is safe for command line usage
            if _is_path_safe(ref_fasta_path):
                safe_ref_path_for_db = ref_fasta_path
            else:
                # If not safe (contains spaces/symbols), copy to the temp dir we just created
                # This happens only ONCE per analysis, so overhead is negligible.
                safe_ref_path_for_db = temp_dir / "ref_safe.fasta"
                try:
                    shutil.copy(ref_fasta_path, safe_ref_path_for_db)
                except Exception as e:
                    print(f"[ANIb Error] Could not copy unsafe ref path for DB creation: {e}")
                    return 0.0, 0.0, 0, 0

            ref_db_path = temp_dir / "ref_db"

            # Build temp DB for reference
            cmd_db = [
                makeblastdb_exe,
                "-in", str(safe_ref_path_for_db),  # Use the sanitized path
                "-dbtype", "nucl",
                "-out", str(ref_db_path)
            ]
            run_quiet(cmd_db, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

            # 3. Run BLASTn (Strict ANIb-like settings)
            blast_out_path = temp_dir / "anib.blast.tab"

            cmd_blast = [
                blastn_exe,
                "-query", str(fragments_path),
                "-db", str(ref_db_path),
                "-out", str(blast_out_path),
                "-outfmt", "6 qseqid length pident bitscore qlen",
                "-task", "blastn",
                "-dust", "no",  # Important: Disable masking for accurate ANI
                "-evalue", "1e-15",
                "-num_threads", str(blast_num_threads),
                "-max_target_seqs", "5"  # We only need the top hit
            ]

            run_quiet(cmd_blast, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

            # 4. Parse and Calculate
            if not blast_out_path.exists() or blast_out_path.stat().st_size == 0:
                return 0.0, 0.0, 0, num_fragments

            # Load data
            df = pd.read_csv(blast_out_path, sep="\t", names=["qseqid", "length", "pident", "bitscore", "qlen"])

            if df.empty:
                return 0.0, 0.0, 0, num_fragments

            # Filter: Best Hit Per Fragment
            # Sort by Bitscore descending, then drop duplicates on qseqid to keep top hit
            best_hits = df.sort_values("bitscore", ascending=False).drop_duplicates("qseqid")

            # Filter: Coverage > 70%
            # Coverage = Alignment Length / Fragment Length
            valid_hits = best_hits[(best_hits['length'] / best_hits['qlen']) >= MIN_COVERAGE]

            valid_count = len(valid_hits)

            if valid_count == 0:
                return 0.0, 0.0, 0, num_fragments

            # ANIb Calculation: Mean Identity of Valid Fragments
            ani_score = valid_hits['pident'].mean()

            # Coverage Calculation: (Valid Fragments / Total Fragments)
            coverage_pct = (valid_count / num_fragments) * 100.0

            return ani_score, coverage_pct, valid_count, num_fragments

        except Exception as e:
            print(f"[ANIb Error] {e}")
            return 0.0, 0.0, 0, 0
        finally:
            # Cleanup if we created the temp dir ourselves
            if local_temp_manager:
                local_temp_manager.cleanup()

    status("Reading query and reference FASTA files...")
    fasta_path, annot_path = Path(fasta), Path(annot)
    ref_fasta_path = Path(ref_fasta)
    record = SeqIO.read(fasta_path, "fasta")
    seq, seqlen, fasta_id = str(record.seq).upper(), len(record.seq), record.id
    gc_window_size = int(window)
    ref_records_dict = SeqIO.to_dict(SeqIO.parse(ref_fasta_path, "fasta"))
    if not ref_records_dict: raise ValueError("Reference FASTA file is empty.")
    ref_record = next(iter(ref_records_dict.values()))

    query_seq_full = seq
    ref_seq_main = str(ref_record.seq).upper()
    ref_seqlen = len(ref_seq_main)

    # Validate Alignment Files
    _validate_alignment_files(
        xmfa_path=xmfa_file,
        maf_path=sibeliaz_maf,
        query_path=fasta_path,
        ref_path=ref_fasta_path,
        query_id=fasta_id,
        ref_id=ref_record.id,
        query_len=seqlen,
        ref_len=ref_seqlen,
        force_mauve=force_mauve,
        force_sibeliaz=force_sibeliaz,
        log_queue=log_queue
    )

    highlight_intervals = []
    if highlight_bed_path and Path(highlight_bed_path).is_file():
        try:
            bed_df = pd.read_csv(highlight_bed_path, sep="\t", header=None, usecols=[0, 1, 2])
            search_id = fasta_id.split(':')[0]
            matching = bed_df[bed_df[0].str.contains(search_id, na=False, regex=False)]

            if not matching.empty:
                highlight_intervals = list(matching[[1, 2]].itertuples(index=False, name=None))
                print(f"[INFO] Loaded {len(highlight_intervals)} highlight regions matching '{search_id}'")
            else:
                print(f"[WARNING] No BED regions matched FASTA ID '{search_id}'.")
                print(f"          -> Loading ALL {len(bed_df)} regions from BED file instead.")
                highlight_intervals = list(bed_df[[1, 2]].itertuples(index=False, name=None))

        except Exception as e:
            print(f"[ERROR] Could not read BED file: {e}")
            highlight_intervals = []
    status("Calculating GC content and skew...")
    gc_arr_full, skew_arr_full = calculate_gc_skew_data(seq, gc_window_size)
    status("Running BLASTn...")
    blast_tab_path = fasta_path.with_suffix(".vs_ref.blast.tab")
    _run_blastn(fasta_path, ref_fasta_path, blast_tab_path,
                task=blast_task, word_size=blast_word_size, reward=blast_reward,
                penalty=blast_penalty, evalue=blast_evalue,
                num_threads=blast_num_threads)
    blast_df = pd.read_csv(blast_tab_path, sep="\t",
                           names=["qseqid", "sseqid", "pident", "length", "mismatch", "gapopen", "qstart", "qend",
                                  "sstart", "send", "evalue", "bitscore", "qseq", "sseq"], comment='#')
    status("Loading gene annotations...")

    features = load_gene_features(
        annot_file=annot_path, fasta_id=fasta_id, highlight_intervals=highlight_intervals,
        allow_code_fallback=allow_code_fallback, blast_df=blast_df, homology_threshold=homology_threshold,
        manual_find=manual_find, manual_replace=manual_replace, log_queue=log_queue,
        seqlen=seqlen
    )

    if highlight_bed_path:
        _validate_bed_vs_features(highlight_bed_path, features, seqlen, fasta_id, log_queue)

    coding_intervals_0based = sorted([(int(f.start) - 1, int(f.end)) for f in features])

    merged_coding_intervals_0based = _merge_intervals(coding_intervals_0based)
    total_coding_bases = sum(end - start for start, end in merged_coding_intervals_0based)
    total_intergenic_bases = max(seqlen - total_coding_bases, 0)

    combined_coding_scores, combined_non_coding_scores = None, None
    combined_noncoding_avg_id = 0.0

    if identity_algorithm != "Legacy (Global k-mer %)":
        status("Running Mauve analysis...")
        mauve_g_array, mauve_lcb_data, global_wga_id_mauve = _run_mauve_and_parse(
            fasta_path, ref_fasta_path, seqlen, log_queue, xmfa_file, force_mauve,
            mauve_disable_backbone=mauve_disable_backbone,
            mauve_mums=mauve_mums,
            mauve_collinear=mauve_collinear,
            mauve_skip_refinement=mauve_skip_refinement,
            mauve_skip_gapped_alignment=mauve_skip_gapped_alignment,
            mauve_seed_family=mauve_seed_family,
            mauve_no_weight_scaling=mauve_no_weight_scaling,
            mauve_mem_clean=mauve_mem_clean,
            mauve_seed_weight=mauve_seed_weight,
            mauve_max_gapped_aligner_length=mauve_max_gapped_aligner_length,
            mauve_max_breakpoint_distance_scale=mauve_max_breakpoint_distance_scale,
            mauve_conservation_distance_scale=mauve_conservation_distance_scale,
            mauve_bp_dist_estimate_min_score=mauve_bp_dist_estimate_min_score,
            mauve_gap_open=mauve_gap_open,
            mauve_gap_extend=mauve_gap_extend,
            mauve_weight=mauve_weight,
            mauve_min_scaled_penalty=mauve_min_scaled_penalty,
            mauve_hmm_p_go_homologous=mauve_hmm_p_go_homologous,
            mauve_hmm_p_go_unrelated=mauve_hmm_p_go_unrelated
        )
        if mauve_g_array is None:
            print("[INFO] Mauve returned no data. Check console logs for specific error.")
            mauve_coding_scores = mauve_non_coding_scores = None
            mauve_noncoding_avg_id = 0.0
            mauve_coding_lcb_data = []
            mauve_intergenic_lcb_data = []
        else:
            mauve_coding_scores, mauve_non_coding_scores, mauve_noncoding_avg_id = _sample_identity_profile(
                mauve_g_array, seqlen, nc_window_size, nc_step_size, features, log_queue
            )
            mauve_coding_lcb_data = _generate_coding_fragments(
                mauve_lcb_data,
                coding_intervals_0based,
                query_seq_full,
                ref_seq_main,
                features,
                allow_code_fallback,
                log_queue,
            )
            print(f"Mauve coding fragments generated: {len(mauve_coding_lcb_data)}")
            mauve_intergenic_lcb_data = _generate_intergenic_fragments(
                mauve_lcb_data,
                coding_intervals_0based,
                query_seq_full,
                ref_seq_main,
                log_queue,
            )
            print(f"Mauve intergenic fragments generated: {len(mauve_intergenic_lcb_data)}")

            fixed = _fix_intergenic_boundaries(mauve_intergenic_lcb_data, features)

            is_valid, problems = _validate_intergenic_no_gff_overlap(mauve_intergenic_lcb_data, features, log_queue)
            if not is_valid: print("[CRITICAL] Mauve intergenic validation failed - TRUE overlaps detected!")
            else: print("Mauve intergenic validation passed (no true overlaps)\n")

            _attach_original_sequences_to_lcbs(mauve_lcb_data, query_seq_full, ref_seq_main)

        status("Running SibeliaZ analysis...")
        sibeliaz_g_array, sibeliaz_lcb_data, global_wga_id_sibeliaz = _run_sibeliaz_and_parse(
            query_fasta_path=fasta_path,
            ref_fasta_path=ref_fasta_path,
            full_seqlen=seqlen,
            query_seq_full=query_seq_full,
            ref_seq_main=ref_seq_main,
            ref_seqlen=ref_seqlen,
            log_queue=log_queue,
            maf_file=sibeliaz_maf,
            force_sibeliaz=force_sibeliaz
        )

        if sibeliaz_g_array is None:
            print("[INFO] SibeliaZ returned no data (failed or empty). check console logs above.")
            sibeliaz_coding_scores = sibeliaz_non_coding_scores = None
            sibeliaz_noncoding_avg_id = 0.0
            sibeliaz_coding_lcb_data, sibeliaz_intergenic_lcb_data = [], []
        else:
            sibeliaz_coding_scores, sibeliaz_non_coding_scores, sibeliaz_noncoding_avg_id = _sample_identity_profile(
                sibeliaz_g_array, seqlen, nc_window_size, nc_step_size, features, log_queue
            )
            sibeliaz_coding_lcb_data = _generate_coding_fragments(
                sibeliaz_lcb_data,
                coding_intervals_0based,
                query_seq_full,
                ref_seq_main,
                features,
                allow_code_fallback,
                log_queue,
            )
            print(f"SibeliaZ coding fragments generated: {len(sibeliaz_coding_lcb_data)}")
            sibeliaz_intergenic_lcb_data = _generate_intergenic_fragments(
                sibeliaz_lcb_data,
                coding_intervals_0based,
                query_seq_full,
                ref_seq_main,
                log_queue,
            )
            print(f"SibeliaZ intergenic fragments generated: {len(sibeliaz_intergenic_lcb_data)}")

            fixed = _fix_intergenic_boundaries(sibeliaz_intergenic_lcb_data, features)

            is_valid, problems = _validate_intergenic_no_gff_overlap(sibeliaz_intergenic_lcb_data, features, log_queue)
            if not is_valid: print("[CRITICAL] SibeliaZ intergenic validation failed - TRUE overlaps detected!")
            else: print("SibeliaZ intergenic validation passed (no true overlaps)\n")

            _attach_original_sequences_to_lcbs(sibeliaz_lcb_data, query_seq_full, ref_seq_main,
                                               skip_coordinate_conversion=True)

        if mauve_g_array is not None and sibeliaz_g_array is not None:
            combined_g_array = np.where(mauve_g_array > 0, mauve_g_array, sibeliaz_g_array)
            combined_coding_scores, combined_non_coding_scores, combined_noncoding_avg_id = _sample_identity_profile(
                combined_g_array, seqlen, nc_window_size, nc_step_size, features, log_queue
            )

    else:
        mauve_g_array, mauve_lcb_data, global_wga_id_mauve = None, [], 0.0
        sibeliaz_g_array, sibeliaz_lcb_data, global_wga_id_sibeliaz = None, [], 0.0
        mauve_intergenic_lcb_data, sibeliaz_intergenic_lcb_data = [], []
        mauve_coding_scores = mauve_non_coding_scores = None
        mauve_noncoding_avg_id = 0.0
        sibeliaz_coding_scores = sibeliaz_non_coding_scores = None
        sibeliaz_noncoding_avg_id = 0.0

    status("Classifying BLAST hits by gene overlap...")
    if not blast_df.empty and features:
        gene_starts = np.array([f.start for f in features])
        gene_ends = np.array([f.end for f in features])
        hit_starts_all = blast_df[['qstart', 'qend']].min(axis=1).values
        hit_ends_all = blast_df[['qstart', 'qend']].max(axis=1).values
        is_coding_list = []
        chunk_size = 2000
        for i in range(0, len(hit_starts_all), chunk_size):
            hit_starts_chunk = hit_starts_all[i:i + chunk_size]
            hit_ends_chunk = hit_ends_all[i:i + chunk_size]
            overlap_matrix = (hit_starts_chunk[:, np.newaxis] < gene_ends) & (
                    hit_ends_chunk[:, np.newaxis] > gene_starts)
            is_coding_list.extend(overlap_matrix.any(axis=1))
        blast_df['is_coding'] = is_coding_list
    else: blast_df['is_coding'] = False
    blast_df = _enrich_blast_df(blast_df)

    KMER_SIZE_FOR_PLOT = 10

    status("Concatenating reference contigs...")
    all_ref_seqs = [str(r.seq).upper() for r in ref_records_dict.values()]
    ref_seq_str = "".join(all_ref_seqs)
    status(f"Reference sequence length: {len(ref_seq_str):,} bp", "blue")

    status("Building reference k-mer index for plot...")
    ref_kmers_global_set_plot = _get_kmers(ref_seq_str, KMER_SIZE_FOR_PLOT)
    status(f"Found {len(ref_kmers_global_set_plot):,} unique {KMER_SIZE_FOR_PLOT}-mers in reference.", "blue")

    status("Running Legacy (Global k-mer) analysis for plot...")
    legacy_kmer_coding_scores, legacy_kmer_non_coding_scores, legacy_kmer_global_avg_id = calculate_global_kmer_ring_data(
        seq, ref_kmers_global_set_plot, features, nc_window_size, nc_step_size, log_queue
    )

    status("Running Legacy (1:1 Base Comparison) analysis for plot...", "black")
    legacy_base_coding_scores, legacy_base_non_coding_scores, legacy_base_global_avg_id = calculate_legacy_base_comparison_ring_data(
        seq, ref_seq_str, features, nc_window_size, nc_step_size, log_queue
    )

    MIN_K_FOR_REPORT = 15
    status("Analyzing WGA data for max exact match...")
    max_k_mauve = _get_max_exact_match([mauve_lcb_data])
    max_k_sibeliaz = _get_max_exact_match([sibeliaz_lcb_data])

    max_k = max(max_k_mauve, max_k_sibeliaz, MIN_K_FOR_REPORT)
    max_k = min(max_k, 200)

    if run_legacy_report:
        status(
            f"Max exact match found: {max_k}bp. Running hierarchical scan from k={max_k} down to {MIN_K_FOR_REPORT}bp.",
            "blue")

        legacy_all_blocks, legacy_intergenic_blocks = generate_hierarchical_legacy_report(
            seq, ref_records_dict, features, max_k, MIN_K_FOR_REPORT, log_queue
        )
        print(f"Legacy intergenic blocks generated: {len(legacy_intergenic_blocks)}")

        fixed = _fix_intergenic_boundaries(legacy_intergenic_blocks, features)
        if fixed > 0:
            print(f"[INFO] Fixed {fixed} Legacy intergenic boundaries to exclude coding start positions")

        is_valid, problems = _validate_intergenic_no_gff_overlap(legacy_intergenic_blocks, features, log_queue)
        if not is_valid: print("[CRITICAL] Legacy intergenic validation failed - TRUE overlaps detected!")
        else: print("Legacy intergenic validation passed (no true overlaps)\n")
    else:
        status("Deferring legacy report generation to background thread.", "grey")
        legacy_all_blocks, legacy_intergenic_blocks = [], []

    mauve_wga_id_intergenic = _calculate_global_intergenic_wga(
        mauve_intergenic_lcb_data, total_intergenic_bases, log_queue
    )
    sibeliaz_wga_id_intergenic = _calculate_global_intergenic_wga(
        sibeliaz_intergenic_lcb_data, total_intergenic_bases, log_queue
    )
    combined_wga_id_intergenic = mauve_wga_id_intergenic

    status("Calculating True ANIb...", "black")
    ani_val, ani_cov, _, _ = _calculate_true_anib(seq, ref_fasta_path, blast_num_threads)

    all_data = {
        "record": record, "seqlen": seqlen,
        "ref_record": ref_record,
        "ref_records_dict": ref_records_dict,
        "ref_seq_concatenated": ref_seq_str,
        "ref_seqlen": len(ref_record.seq),
        "gene_feats": features, "blast_df": blast_df,
        "gc_arr": gc_arr_full, "skew_arr": skew_arr_full,
        "gc_window_size": gc_window_size,
        "allow_code_fallback": allow_code_fallback,
        "identity_algorithm_used": identity_algorithm,
        "mauve_g_array": mauve_g_array,
        "sibeliaz_g_array": sibeliaz_g_array,
        "combined_g_array": combined_g_array if 'combined_g_array' in locals() else None,
        "anib_score": ani_val,
        "anib_coverage": ani_cov,

        "mauve_data": {
            "coding_scores": mauve_coding_scores,
            "non_coding_scores": mauve_non_coding_scores,
            "global_wga_id": global_wga_id_mauve,
            "noncoding_avg_id": mauve_noncoding_avg_id,
            "lcb_data": mauve_lcb_data,
            "coding_lcb_data": mauve_coding_lcb_data,
            "intergenic_lcb_data": mauve_intergenic_lcb_data,
            "wga_id_intergenic_incl_gaps": mauve_wga_id_intergenic,
        },
        "sibeliaz_data": {
            "coding_scores": sibeliaz_coding_scores,
            "non_coding_scores": sibeliaz_non_coding_scores,
            "global_wga_id": global_wga_id_sibeliaz,
            "noncoding_avg_id": sibeliaz_noncoding_avg_id,
            "lcb_data": sibeliaz_lcb_data,
            "coding_lcb_data": sibeliaz_coding_lcb_data,
            "intergenic_lcb_data": sibeliaz_intergenic_lcb_data,
            "wga_id_intergenic_incl_gaps": sibeliaz_wga_id_intergenic,
        },
        "combined_data": {
            "coding_scores": combined_coding_scores,
            "non_coding_scores": combined_non_coding_scores,
            "global_wga_id": global_wga_id_mauve,
            "noncoding_avg_id": combined_noncoding_avg_id,
            "lcb_data": mauve_lcb_data,
            "coding_lcb_data": mauve_coding_lcb_data,
            "intergenic_lcb_data": mauve_intergenic_lcb_data,
            "wga_id_intergenic_incl_gaps": combined_wga_id_intergenic,
        },
        "legacy_data": {
            "coding_scores": legacy_kmer_coding_scores,
            "non_coding_scores": legacy_kmer_non_coding_scores,
            "global_wga_id": legacy_kmer_global_avg_id,
            "noncoding_avg_id": legacy_kmer_global_avg_id,
            "lcb_data": legacy_all_blocks,
            "intergenic_lcb_data": legacy_intergenic_blocks,
            "wga_id_intergenic_incl_gaps": 0.0
        },
        "legacy_base_data": {
            "coding_scores": legacy_base_coding_scores,
            "non_coding_scores": legacy_base_non_coding_scores,
            "global_wga_id": legacy_base_global_avg_id,
            "noncoding_avg_id": legacy_base_global_avg_id,
            "lcb_data": [],
            "intergenic_lcb_data": [],
            "wga_id_intergenic_incl_gaps": 0.0
        },
    }
    if identity_algorithm == "Legacy (Global k-mer %)": selected_data_key = "legacy_data"
    elif identity_algorithm == "Legacy (1:1 Base %)": selected_data_key = "legacy_base_data"
    elif identity_algorithm == "Mauve": selected_data_key = "mauve_data"
    elif identity_algorithm == "SibeliaZ": selected_data_key = "sibeliaz_data"
    else: selected_data_key = "combined_data"

    selected_data_valid = (
            all_data.get(selected_data_key) and
            (all_data[selected_data_key]["coding_scores"] or all_data[selected_data_key]["non_coding_scores"])
    )

    if not selected_data_valid:
        if all_data["combined_data"]["coding_scores"] or all_data["combined_data"]["non_coding_scores"]:
            selected_data_key = "combined_data"
        elif all_data["legacy_data"]["coding_scores"] or all_data["legacy_data"]["non_coding_scores"]:
            selected_data_key = "legacy_data"
        else: selected_data_key = "legacy_base_data"

        status(f"Warning: Data for '{identity_algorithm}' not found. Falling back to '{selected_data_key}'.", "orange")
        all_data["identity_algorithm_used"] = selected_data_key

    selected_data = all_data[selected_data_key]
    all_data["coding_scores"] = selected_data["coding_scores"]
    all_data["non_coding_scores"] = selected_data["non_coding_scores"]
    all_data["global_wga_id"] = selected_data["global_wga_id"]
    all_data["noncoding_avg_id"] = selected_data["noncoding_avg_id"]
    all_data["wga_id_intergenic_incl_gaps"] = selected_data.get("wga_id_intergenic_incl_gaps", 0.0)
    all_data["mauve_g_array"] = mauve_g_array
    all_data["sibeliaz_g_array"] = sibeliaz_g_array
    if mauve_g_array is not None and sibeliaz_g_array is not None:
        all_data["combined_g_array"] = np.where(mauve_g_array > 0, mauve_g_array, sibeliaz_g_array)

    return all_data

# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# Legend Management Functions
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────

def _determine_legend_items(data, plot_params, filtered_blast_df, bed_genes, homolog_genes,
                            limit_region=False, region_start=0, region_end=None):
    """
    Determine which legend items should be displayed based on actual data present
    VISIBLE within the current plot region.

    Enforces Order: Genes -> GC Skew (+) -> GC Skew (-) -> GC Content
    """
    legend_items = []

    # 1. Gene Features (ROI, Homologs, Non-Homologs)
    all_genes = data.get('gene_feats', [])
    roi_count, homolog_count, non_homolog_count = 0, 0, 0

    # Helper to check visibility based on plot limits
    def is_visible(feature):
        if not limit_region: return True
        f_start = int(feature.start)
        f_end = int(feature.end)
        r_start = int(region_start)
        r_end = int(region_end) if region_end is not None else float('inf')
        return f_start < r_end and f_end > r_start

    for f in all_genes:
        if not is_visible(f): continue

        c = getattr(f, 'color', '').lower()
        if c == '#e53935': roi_count += 1
        elif c == '#43a047': homolog_count += 1
        elif c == '#cccccc': non_homolog_count += 1

    # Highlighted ROI genes
    if roi_count > 0: legend_items.append(('roi', mpatches.Patch(color='#e53935', label='Highlighted ROI Gene')))

    # Homologous genes
    if homolog_count > 0:
        homology_thresh_min = plot_params.get("homology_thresh", 0)
        homology_thresh_max = plot_params.get("homology_thresh_max", 100)

        def get_val(v):
            if hasattr(v, 'get'): return float(v.get())
            try: return float(v)
            except: return 0.0

        min_val, max_val = get_val(homology_thresh_min), get_val(homology_thresh_max)

        legend_items.append(('homolog', mpatches.Patch(
            color='#43a047',
            label=f'Homologous Gene ({min_val:.0f}-{max_val:.0f}%)'
        )))

    # Non-homologous genes
    if non_homolog_count > 0:
        legend_items.append(('non_homolog', mpatches.Patch(color='#cccccc', label='Non-homologous Gene')))

    # 2. GC Skew (Moved BEFORE GC Content)
    skew_arr = data.get("skew_arr")
    if skew_arr is not None and len(skew_arr) > 0:
        legend_items.append(('gc_skew_pos', plt.Line2D([0], [0], color='#33a02c', lw=1.2, label='GC Skew (+)')))
        legend_items.append(('gc_skew_neg', plt.Line2D([0], [0], color='#fb9a99', lw=1.2, label='GC Skew (-)')))

    # 3. GC Content (Moved to End)
    gc_arr = data.get("gc_arr")
    if gc_arr is not None and len(gc_arr) > 0:
        gc_cmap_raw = plot_params.get("gc_colormap", "Grey")
        gc_cmap_name = gc_cmap_raw.get() if hasattr(gc_cmap_raw, 'get') else str(gc_cmap_raw)

        if gc_cmap_name.strip() == "Grey":
            legend_items.append(('gc_content', mpatches.Patch(
                facecolor='0.7', alpha=0.5, edgecolor='none', label='GC Content (vs. Median)'
            )))

    return legend_items


def _determine_colorbars_to_show(data, plot_params, filtered_blast_df, scores_to_plot,
                                 gc_cmap, gc_norm, nc_cmap, color_vmin, color_vmax, id_label, all_ids=None):
    colorbar_definitions = {}

    if not filtered_blast_df.empty:
        colormap_name = plot_params.get("homology_colormap", "Plasma")
        if hasattr(colormap_name, 'get'): colormap_name = colormap_name.get()

        try:
            is_reversed = " (Reversed)" in str(colormap_name)
            base_name = str(colormap_name).replace(" (Reversed)", "")
            LOWERCASE_CMAPS = {'plasma', 'viridis', 'inferno', 'magma', 'cividis', 'turbo', 'coolwarm', 'gist_earth',
                               'terrain', 'hsv', 'jet'}
            if base_name.lower() in LOWERCASE_CMAPS: base_name = base_name.lower()
            cmap = plt.get_cmap(base_name)
            if is_reversed: cmap = cmap.reversed()
        except:
            cmap = plt.get_cmap("viridis")

        # Prevent matplotlib from expanding zero-range normalizations
        # For percentage data [0-100], adjust range to prevent expansion while respecting bounds
        norm_vmin, norm_vmax = color_vmin, color_vmax
        if norm_vmin >= norm_vmax:  # Zero or inverted range
            # Add epsilon to vmax, but keep vmin >= 0 and vmax <= 100
            if norm_vmax < 100.0:
                norm_vmax = min(100.0, norm_vmin + 0.001)  # Expand upward if possible
            elif norm_vmin > 0.0:
                norm_vmin = max(0.0, norm_vmin - 0.001)  # Expand downward if vmax is already at max
            # If vmin=100 and vmax=100, keep it and let epsilon be tiny
            if norm_vmin >= norm_vmax:
                norm_vmax = norm_vmin + 0.001
        norm = mcolors.Normalize(vmin=norm_vmin, vmax=norm_vmax)
        colorbar_definitions['blast'] = {'label': 'BLAST Hit Identity (%)',
                                         'mappable': cm.ScalarMappable(norm=norm, cmap=cmap)}

    has_identity_data = False
    s_min, s_max = 0, 100

    if plot_params.get("id_ring_precise_mode", False) and all_ids is not None and len(all_ids) > 0:
        has_identity_data = True
        if plot_params.get("nc_auto_fit_color", True):
            s_min = min(all_ids)
            s_max = max(all_ids)
            if s_min > 0 and (s_max - s_min) < 5:
                s_min = max(0, s_min - 5)

    elif all_ids and len(all_ids) > 0:
        # all_ids contains values from actually drawn artists - use this for accurate range
        has_identity_data = True
        if plot_params.get("nc_auto_fit_color", True):
            s_min = min(all_ids)
            s_max = max(all_ids)
    elif scores_to_plot and len(scores_to_plot) > 0:
        # Fallback: if no drawn artists available, use filtered scores
        has_identity_data = True
        vals = [s for _, s in scores_to_plot]
        if vals and plot_params.get("nc_auto_fit_color", True):
            s_min, s_max = min(vals), max(vals)

    if has_identity_data and id_label != "No Identity Data":
        # Ensure we have a valid range
        if s_min >= s_max:
            s_max = s_min + 1.0

        # Prevent matplotlib from expanding zero-range normalizations
        # For percentage data [0-100], adjust range to prevent expansion while respecting bounds
        norm_vmin, norm_vmax = s_min, s_max
        if norm_vmin >= norm_vmax:  # Zero or inverted range
            # Add epsilon to vmax, but keep vmin >= 0 and vmax <= 100
            if norm_vmax < 100.0:
                norm_vmax = min(100.0, norm_vmin + 0.001)  # Expand upward if possible
            elif norm_vmin > 0.0:
                norm_vmin = max(0.0, norm_vmin - 0.001)  # Expand downward if vmax is already at max
            # If vmin=100 and vmax=100, keep it and let epsilon be tiny
            if norm_vmin >= norm_vmax:
                norm_vmax = norm_vmin + 0.001
        norm = mcolors.Normalize(vmin=norm_vmin, vmax=norm_vmax)
        colorbar_definitions['identity'] = {
            'label': f'{id_label} (%)',
            'mappable': cm.ScalarMappable(norm=norm, cmap=nc_cmap)
        }

    if gc_cmap is not None and gc_norm is not None:
        colorbar_definitions['gc'] = {
            'label': 'GC Content (%)',
            'mappable': cm.ScalarMappable(norm=gc_norm, cmap=gc_cmap)
        }

    return colorbar_definitions


def _calculate_safe_vertical_zone(fig, ax, legend_bbox_fig, text_objects):
    renderer, inv_fig = fig.canvas.get_renderer(), fig.transFigure.inverted()

    if legend_bbox_fig: floor_y = legend_bbox_fig.y1 + 0.01
    else: floor_y = 0.02

    plot_bbox = ax.get_window_extent(renderer).transformed(inv_fig)
    ceiling_y = plot_bbox.y0

    if text_objects:
        min_label_y = float('inf')
        for txt in text_objects:
            if txt is None: continue
            try:
                bbox = txt.get_window_extent(renderer).transformed(inv_fig)
                if bbox.y0 < min_label_y: min_label_y = bbox.y0
            except: pass

        if min_label_y != float('inf'): ceiling_y = min(ceiling_y, min_label_y)
    ceiling_y -= 0.01
    return floor_y, ceiling_y


def _create_legend_and_colorbars(fig, ax, data, plot_params, filtered_blast_df, scores_to_plot,
                                 bed_genes, homolog_genes, gc_cmap, gc_norm, nc_cmap,
                                 color_vmin, color_vmax, id_label, all_ids=None,
                                 limit_region=False, region_start=0, region_end=None):
    legend_items_raw = _determine_legend_items(
        data, plot_params, filtered_blast_df, bed_genes, homolog_genes,
        limit_region=limit_region, region_start=region_start, region_end=region_end
    )
    legend_elements = [item[1] for item in legend_items_raw]

    num_legend_items = len(legend_elements)
    legend_bbox_fig, legend = None, None

    if num_legend_items > 0:
        if num_legend_items <= 5: ncol = num_legend_items
        else: ncol = (num_legend_items + 1) // 2

        legend_bottom_y = 0.01
        legend = fig.legend(handles=legend_elements, loc='lower center',
                            bbox_to_anchor=(0.5, legend_bottom_y),
                            bbox_transform=fig.transFigure, ncol=ncol,
                            fontsize=9, borderpad=0.3, labelspacing=0.3)

        fig.canvas.draw()
        try:
            renderer = fig.canvas.get_renderer()
            legend_bbox_fig = legend.get_window_extent(renderer=renderer).transformed(fig.transFigure.inverted())
        except: legend_bbox_fig = None

    colorbar_definitions = _determine_colorbars_to_show(
        data, plot_params, filtered_blast_df, scores_to_plot,
        gc_cmap, gc_norm, nc_cmap, color_vmin, color_vmax, id_label, all_ids
    )

    all_cbars_and_axes = []
    num_cbars = len(colorbar_definitions)

    if num_cbars > 0:
        text_objects = [child for child in ax.get_children() if isinstance(child, matplotlib.text.Text)]
        floor_y, ceiling_y = _calculate_safe_vertical_zone(fig, ax, legend_bbox_fig, text_objects)

        available_height = ceiling_y - floor_y
        ideal_cbar_height = 0.035
        text_padding_estimate = 0.045
        required_height = ideal_cbar_height + text_padding_estimate
        final_cbar_height = ideal_cbar_height

        # Calculate resize factor if space is tight
        if required_height > available_height:
            shrink_factor = max(available_height, 0.001) / required_height

            # Enforce minimum display size (50% of ideal)
            effective_shrink = max(shrink_factor, 0.5)
            final_cbar_height = ideal_cbar_height * effective_shrink

        center_y = floor_y + (available_height / 2)
        cbar_bottom_y = center_y - (final_cbar_height / 2)

        cbar_width = 0.22
        h_gap = 0.06
        total_width = num_cbars * cbar_width + max(0, num_cbars - 1) * h_gap
        start_left = (1.0 - total_width) / 2.0
        current_left = start_left

        cbar_order = ['blast', 'identity', 'gc']

        for key in cbar_order:
            if key in colorbar_definitions:
                cbar_def = colorbar_definitions[key]
                cax = fig.add_axes([current_left, cbar_bottom_y, cbar_width, final_cbar_height])
                cbar = fig.colorbar(cbar_def['mappable'], cax=cax, orientation='horizontal')
                cbar.set_label(cbar_def['label'], fontsize=8)
                cbar.ax.xaxis.set_label_position('top')
                cbar.ax.xaxis.set_ticks_position('bottom')
                cbar.ax.tick_params(labelsize=7, pad=2)

                if key == 'gc':
                    vmin_gc, vmax_gc = cbar_def['mappable'].norm.vmin, cbar_def['mappable'].norm.vmax
                    ticks_gc = np.linspace(vmin_gc, vmax_gc, 4) if vmax_gc > vmin_gc else [vmin_gc]
                    cbar.set_ticks(ticks_gc)
                    cbar.ax.set_xticklabels([f'{t * 100:.0f}' for t in ticks_gc])
                elif key == 'blast':
                    vmin_blast = cbar_def['mappable'].norm.vmin
                    vmax_blast = cbar_def['mappable'].norm.vmax
                    ticks_blast = np.linspace(vmin_blast, vmax_blast, 4) if vmax_blast > vmin_blast else [vmin_blast]
                    cbar.set_ticks(ticks_blast)
                    cbar.ax.set_xticklabels([f'{t:.1f}' for t in ticks_blast])
                elif key == 'identity':
                    vmin_id, vmax_id = cbar_def['mappable'].norm.vmin, cbar_def['mappable'].norm.vmax
                    ticks_id = np.linspace(vmin_id, vmax_id, 5) if vmax_id > vmin_id else [vmin_id]
                    cbar.set_ticks(ticks_id)
                    # Check if integer labels would have duplicates
                    int_labels = [f'{t:.0f}' for t in ticks_id]
                    if len(int_labels) != len(set(int_labels)):
                        # Use 1 decimal place if there are duplicates
                        cbar.ax.set_xticklabels([f'{t:.1f}' for t in ticks_id])
                    else:
                        cbar.ax.set_xticklabels(int_labels)

                all_cbars_and_axes.append((cbar, cax))
                current_left += cbar_width + h_gap

    return legend, all_cbars_and_axes


def _draw_plot(ax, data, plot_params, recalculate_auto_view=False):
    """Render the circular genome comparison plot."""
    ax.clear()
    fig = ax.figure

    # Clear old identity ring artists from the previous plot run
    if hasattr(ax, '_identity_ring_artists'):
        for artist in ax._identity_ring_artists:
            if artist and artist.axes:
                try: artist.remove()
                except ValueError: pass

    gene_feature_patches_local = []

    # Get Full Sequence Data
    record, full_seqlen = data["record"], data["seqlen"]
    ref_record, ref_seqlen = data["ref_record"], data["ref_seqlen"]
    original_gene_feats, original_blast_df = data["gene_feats"], data["blast_df"]
    original_gc_arr, original_skew_arr = data["gc_arr"], data["skew_arr"]

    # Use whatever scores are currently active (from cache or recalculation)
    original_coding_scores = data.get("coding_scores", [])
    original_non_coding_scores = data.get("non_coding_scores", [])

    # Get Region Limiting Params
    limit_region = plot_params.get("limit_to_region", False)
    region_start, region_end = 0, full_seqlen
    try:
        if limit_region:
            start_val = int(plot_params.get("region_start"))
            end_val = int(plot_params.get("region_end"))
            # Convert user 1-based input to 0-based internal
            start_val_0based = start_val - 1
            end_val_0based = end_val

            # Basic validation
            if 0 <= start_val_0based < end_val_0based <= full_seqlen:
                region_start, region_end = start_val_0based, end_val_0based
            else:
                limit_region = False
    except (ValueError, TypeError):
        limit_region = False

    include_coding = plot_params.get("id_ring_include_coding", False)
    show_noncoding = plot_params.get("id_ring_show_noncoding", True)
    autorotate_roi = plot_params.get("autorotate_roi", True)
    manual_rotation_deg = plot_params.get("manual_rot_deg", 0.0)
    zoom_factor = plot_params.get("zoom", 1.0)

    # Default Geometry
    id_ring_radius = plot_params.get("id_ring_radius", 0.6)
    gc_inner = plot_params.get("gc_inner", 0.3)
    skew_inner_radius = plot_params.get("skew_inner_radius", 0.6)
    blast_ring_radius = plot_params.get("blast_ring_radius", 1.0)
    blast_ring_thickness = plot_params.get("blast_ring_thickness", 0.15)
    show_roi_labels = plot_params.get("show_roi_labels", True)
    show_homology_labels = plot_params.get("show_homology_labels", True)
    homology_thresh_min = plot_params.get("homology_thresh", 0.0)
    homology_thresh_max = plot_params.get("homology_thresh_max", 100.0)
    autofit_color_range = plot_params.get("autofit_color_range", True)
    use_smart_layout = plot_params.get("use_smart_layout", True)
    show_blast_separators = plot_params.get("show_blast_separators", False)
    colormap_name = plot_params.get("homology_colormap", "Plasma")
    show_blast_coding = plot_params.get("show_blast_coding", True)
    show_blast_noncoding = plot_params.get("show_blast_noncoding", False)
    id_alpha = 1.0 - (plot_params.get("id_baseline_transparency_pct", 50.0) / 100.0)
    gc_alpha = 1.0 - (plot_params.get("gc_baseline_transparency_pct", 50.0) / 100.0)
    skew_alpha = 1.0 - (plot_params.get("skew_baseline_transparency_pct", 50.0) / 100.0)
    is_initial_load = plot_params.get('is_initial_load', False)
    nc_step_size = plot_params.get("nc_step_size", 50)
    show_start_end_line = plot_params.get("show_start_end_line", False)
    cached_x_center_pan = plot_params.get('cached_x_center_pan', 0.0)
    cached_y_center_pan = plot_params.get('cached_y_center_pan', -1.0)
    original_id_thickness = plot_params.get("id_ring_thickness", 0.2)
    remove_gene_borders = plot_params.get("remove_gene_borders", False)

    # Setup Plot Rotation & Auto-Stretch
    plot_seqlen = full_seqlen
    if limit_region:
        try:
            region_span_bp = region_end - region_start
            arc_fraction = region_span_bp / full_seqlen
            target_angle_deg = 0.0 if arc_fraction < 0.5 else -90.0

            # Create dummy feature to center view on the region midpoint
            region_midpoint_bp = (region_start + region_end) / 2.0
            dummy_feat = type("MockFeature", (), {"start": region_midpoint_bp, "end": region_midpoint_bp})

            plot_top_position = compute_top_position_for_roi(
                [dummy_feat], full_seqlen, base_top=0.0,
                desired_angle_deg=target_angle_deg,
                manual_deg=manual_rotation_deg,
                limit_region=True, region_start=region_start, region_end=region_end
            )
        except Exception:
            plot_top_position = compute_top_position_for_roi(
                [f for f in original_gene_feats if f.color == "#e53935"] if autorotate_roi else [],
                full_seqlen, manual_deg=manual_rotation_deg
            )
    else:
        plot_top_position = compute_top_position_for_roi(
            [f for f in original_gene_feats if f.color == "#e53935"] if autorotate_roi else [],
            full_seqlen, manual_deg=manual_rotation_deg
        )

    # Setup Graphic Record (Invisible base for coordinates)
    graphic = CircularGraphicRecord(sequence_length=plot_seqlen, features=[], feature_level_height=0.1)
    graphic.top_position = plot_top_position
    graphic.plot(ax=ax)

    R = _estimate_dfv_radius(ax, graphic)
    ax.clear()

    if cached_y_center_pan == -1.0: cached_y_center_pan = -R
    id_ring_color_only = plot_params.get("id_ring_color_only", False)

    # Data prep for Identity Ring
    if include_coding and show_noncoding:
        id_label = "Genome ID"
        raw_scores = sorted(original_coding_scores + original_non_coding_scores, key=lambda x: x[0])
    elif include_coding and not show_noncoding:
        id_label = "Coding ID"
        raw_scores = original_coding_scores
    elif not include_coding and show_noncoding:
        id_label = "Non-Coding ID"
        raw_scores = original_non_coding_scores
    else:
        id_label = "No Identity Data"
        raw_scores = []

        # Apply Region Limit to Identity Scores for Legend/Autofit
    if limit_region:
        scores_to_plot = [(p, s) for p, s in raw_scores if region_start <= p < region_end]
    else:
        scores_to_plot = raw_scores

    # Calculate ID thickness geometry first
    if not id_ring_color_only and (plot_params.get("auto_thickness_rings", False) or is_initial_load):
        try:
            # Try to fit ID ring between Skew Inner and Blast Inner
            skew_r_frac = plot_params.get("skew_inner_radius", 0.5)
            blast_r_frac = plot_params.get("blast_ring_radius", 1.0)
            blast_t_frac = plot_params.get("blast_ring_thickness", 0.15)

            blast_inner_edge_r = R * (blast_r_frac - blast_t_frac)

            # Set ID radius to midpoint of available space
            new_id_radius_frac = (skew_r_frac + (blast_r_frac - blast_t_frac)) / 2.0
            id_ring_radius = new_id_radius_frac
            r_baseline = R * id_ring_radius

            # Calculate available space outwards
            inv = ax.transData.inverted()
            one_pixel_gap = abs(inv.transform((0, 1))[1] - inv.transform((0, 0))[1])
            available_space = blast_inner_edge_r - r_baseline
            effective_space = max(0, available_space - one_pixel_gap)
            calculated_id_thickness = max(0.01, effective_space / R)
        except Exception:
            id_ring_radius = plot_params.get("id_ring_radius", 0.6)
            calculated_id_thickness = original_id_thickness
    else:
        id_ring_radius = plot_params.get("id_ring_radius", 0.6)
        calculated_id_thickness = original_id_thickness

    # Auto-thickness Calculation
    if plot_params.get("auto_thickness_rings", False) or is_initial_load:
        gc_thickness, skew_thickness = 1.05, 0.505
        gc_baseline_r = R * gc_inner
        skew_baseline_r = R * skew_inner_radius
        id_ring_baseline_r = R * id_ring_radius

        if original_gc_arr is not None and len(original_gc_arr) > 0 and original_skew_arr is not None and len(
                original_skew_arr) > 0:
            gc_full = original_gc_arr
            skew_full = original_skew_arr
            gc_median = np.median(gc_full)
            space_gc_to_skew = skew_baseline_r - gc_baseline_r
            space_skew_to_id = id_ring_baseline_r - skew_baseline_r
            gap = 0.01 * R

            # Calculate deviations
            gc_deviation = gc_full - gc_median
            max_gc_outward = np.max(gc_deviation)
            skew_clipped = np.clip(skew_full, -1.0, 1.0)
            max_skew_outward = np.max(skew_clipped)

            # Step 1: Initial GC Thickness
            if plot_params.get("gc_ring_color_only", False):
                gc_thickness = plot_params.get("gc_thick", 0.15)
            else:
                # Use 50% of space to skew baseline as a safe start
                if max_gc_outward > 1e-9:
                    gc_thickness = (0.5 * space_gc_to_skew) / (R * max_gc_outward)
                else:
                    gc_thickness = 0.5 * space_gc_to_skew / R

            # Step 2: Skew Thickness Constraint (Avoid GC Baseline)
            negative_skew_mask = skew_clipped < -1e-9
            if np.any(negative_skew_mask):
                if plot_params.get("gc_ring_color_only", False):
                    # Create an array matching the mask size so subsequent indexing works
                    count = np.sum(negative_skew_mask)
                    gc_space_used = np.full(count, R * gc_thickness)
                else:
                    gc_val_at_neg_skew = gc_deviation[negative_skew_mask]
                    gc_space_used = np.maximum(0.0, R * gc_thickness * gc_val_at_neg_skew)

                # Remaining space for Skew
                remaining_space = space_gc_to_skew - gap - gc_space_used
                skew_inward_demand = -skew_clipped[negative_skew_mask]

                valid_demand = skew_inward_demand > 1e-9
                if np.any(valid_demand):
                    allowable_skew_thick = remaining_space[valid_demand] / (R * skew_inward_demand[valid_demand])
                    # Filter positive values
                    valid_thicknesses = allowable_skew_thick[allowable_skew_thick > 0]
                    if len(valid_thicknesses) > 0:
                        skew_thickness_from_collision = np.min(valid_thicknesses)
                    else:
                        skew_thickness_from_collision = 0.01
                else:
                    skew_thickness_from_collision = float('inf')
            else:
                skew_thickness_from_collision = float('inf')

            # STEP 3: Identity Ring Constraint (Outer Limit) ===
            if max_skew_outward > 1e-9:
                skew_thickness_from_identity = (space_skew_to_id - gap) / (R * max_skew_outward)
            else:
                skew_thickness_from_identity = float('inf')

            # Finalize Skew Thickness
            skew_thickness = min(skew_thickness_from_collision, skew_thickness_from_identity)
            skew_thickness = max(0.01, skew_thickness)

            # === Step 4: GC Re-expansion (Fill the Gap)
            if not plot_params.get("gc_ring_color_only", False):
                positive_gc_mask = gc_deviation > 1e-9

                if np.any(positive_gc_mask):
                    # Calculate Skew usage at GC peaks
                    skew_val_at_pos_gc = skew_clipped[positive_gc_mask]
                    # If skew is outward (positive), it uses 0 inward space
                    skew_space_used = R * skew_thickness * np.maximum(0.0, -skew_val_at_pos_gc)

                    available_for_gc = space_gc_to_skew - gap - skew_space_used
                    gc_demand = gc_deviation[positive_gc_mask]

                    valid_points = available_for_gc > 0
                    if np.any(valid_points):
                        max_gc_thick_per_point = available_for_gc[valid_points] / (R * gc_demand[valid_points])
                        new_optimal_gc_thickness = np.min(max_gc_thick_per_point)
                        # Clamp expansion to reasonable limits (e.g. 3x initial)
                        gc_thickness = max(gc_thickness, min(new_optimal_gc_thickness, gc_thickness * 3.0))

    else:
        # Manual Mode
        gc_thickness = plot_params.get("gc_thick", 1.05)
        skew_thickness = plot_params.get("skew_thick", 0.505)

    # Re-verify manual Skew doesn't hit ID ring
    skew_baseline_r = R * skew_inner_radius
    id_ring_baseline_r = R * id_ring_radius
    gap = 0.01 * R

    if original_skew_arr is not None and len(original_skew_arr) > 0:
        skew_clipped = np.clip(original_skew_arr, -1.0, 1.0)
        max_skew_outward = np.max(skew_clipped)
        if max_skew_outward > 1e-9:
            max_allowed_skew_outer = id_ring_baseline_r - gap
            current_max_skew_outer = skew_baseline_r + R * skew_thickness * max_skew_outward
            if current_max_skew_outer > max_allowed_skew_outer:
                # Clamp manual thickness if it collides with ID ring
                skew_thickness = (max_allowed_skew_outer - skew_baseline_r) / (R * max_skew_outward)

    # Filter BLAST data
    homology_filtered_df = original_blast_df[
        (original_blast_df['pident'] >= homology_thresh_min) &
        (original_blast_df['pident'] <= homology_thresh_max)]

    if show_blast_coding and show_blast_noncoding:
        filtered_blast_df = homology_filtered_df
    elif show_blast_coding:
        filtered_blast_df = homology_filtered_df[homology_filtered_df['is_coding']]
    elif show_blast_noncoding:
        filtered_blast_df = homology_filtered_df[~homology_filtered_df['is_coding']]
    else:
        filtered_blast_df = pd.DataFrame(columns=homology_filtered_df.columns)

    if limit_region and not filtered_blast_df.empty:
        # Filter for visibility in region: (Start < RegionEnd) AND (End > RegionStart)
        mask = (filtered_blast_df['qstart'] < region_end) & (filtered_blast_df['qend'] > region_start)
        filtered_blast_df = filtered_blast_df[mask]

    # Color scaling setup (Now uses filtered data for min/max)
    color_vmin = filtered_blast_df['pident'].min() if (
            autofit_color_range and not filtered_blast_df.empty) else homology_thresh_min
    color_vmax = filtered_blast_df['pident'].max() if (
            autofit_color_range and not filtered_blast_df.empty) else homology_thresh_max

    # Colormap setup
    LOWERCASE_CMAPS = {'plasma', 'viridis', 'inferno', 'magma', 'cividis', 'turbo', 'coolwarm', 'gist_earth', 'terrain',
                       'hsv', 'jet'}
    base_name = colormap_name.replace(" (Reversed)", "")
    final_name = base_name.lower() if base_name.lower() in LOWERCASE_CMAPS else base_name
    cmap = plt.get_cmap(final_name)
    cmap = cmap.reversed() if colormap_name.endswith(" (Reversed)") else cmap

    nc_colormap_name = plot_params.get("nc_colormap", "Viridis")
    base_name_nc = nc_colormap_name.replace(" (Reversed)", "")
    final_name_nc = base_name_nc.lower() if base_name_nc.lower() in LOWERCASE_CMAPS else base_name_nc
    nc_cmap = plt.get_cmap(final_name_nc)
    nc_cmap = nc_cmap.reversed() if nc_colormap_name.endswith(" (Reversed)") else nc_cmap

    gc_colormap_name = plot_params.get("gc_colormap", "Grey")
    if plot_params.get("gc_ring_color_only", False) and gc_colormap_name == "Grey": gc_colormap_name = "Coolwarm"
    gc_cmap, gc_norm = None, None
    if gc_colormap_name != "Grey" and original_gc_arr is not None and len(original_gc_arr) > 0:
        base_name_gc = gc_colormap_name.replace(" (Reversed)", "")
        final_name_gc = base_name_gc.lower() if base_name_gc.lower() in LOWERCASE_CMAPS else base_name_gc
        gc_cmap = plt.get_cmap(final_name_gc)
        gc_cmap = gc_cmap.reversed() if gc_colormap_name.endswith(" (Reversed)") else gc_cmap
        gc_norm = mcolors.Normalize(vmin=original_gc_arr.min(), vmax=original_gc_arr.max())

    # Draw Gene Features (Outer Ring)
    allow_code_fallback = data.get("allow_code_fallback", False)
    features_to_plot_colored, features_to_label = [], []
    feature_plot_map = []

    for f in original_gene_feats:
        is_in_region = (not limit_region) or (f.end > region_start and f.start < region_end)
        if is_in_region:
            is_roi = (f.color == '#e53935')
            identity = getattr(f, 'best_identity', 0)
            is_homologous = homology_thresh_min <= identity <= homology_thresh_max

            if is_roi: feature_color = '#e53935'
            elif is_homologous: feature_color = '#43a047'
            else: feature_color = '#cccccc'

            gf = GraphicFeature(start=f.start, end=f.end, strand=f.strand, color=feature_color, label=None)
            gf._feature_ref = f
            features_to_plot_colored.append(gf)
            feature_plot_map.append(f)

            is_baseline = getattr(f, 'is_baseline', True)
            if is_baseline:
                current_label = _label_for_feature(f, allow_code_fallback)
                if is_roi and not current_label:
                    current_label = _nice_label(_get_attr_value(f.attributes, "gene") or
                                                _get_attr_value(f.attributes, "product") or
                                                f"feature:{f.start}-{f.end}")
                if current_label and (
                        (is_roi and show_roi_labels) or (is_homologous and not is_roi and show_homology_labels)):
                    final_label_text = f"{current_label} ({identity:.1f}%)" if is_homologous else current_label
                    features_to_label.append({'feature': f, 'label': final_label_text, 'color': feature_color})

    # Custom Labels
    if 'custom_labels' in data and data['custom_labels']:
        for label_entry in data['custom_labels']:
            if not label_entry.get('visible', True): continue
            original_name = label_entry['name']
            user_color = label_entry['color']
            positions = label_entry['positions']
            label_type = label_entry.get('type')

            for start_pos, end_pos in positions:
                if label_type == 'range' and start_pos != end_pos:
                    final_label_text = f"{original_name} ({start_pos}-{end_pos})"
                elif label_type == 'coordinate' or (label_type == 'range' and start_pos == end_pos):
                    final_label_text = f"{original_name} ({start_pos})"
                elif label_type == 'sequence':
                    if start_pos == end_pos:
                        final_label_text = f"{original_name} ({start_pos})"
                    else:
                        final_label_text = f"{original_name} ({start_pos}-{end_pos})"
                else:
                    final_label_text = original_name

                if limit_region:
                    if end_pos < region_start or start_pos > region_end: continue

                class CustomFeature:
                    def __init__(self, start, end, entry):
                        self.start = start
                        self.end = end
                        self.best_identity = 0
                        self.is_custom_label = True
                        self.custom_color = entry.get('color', 'black')
                        self.font_family = entry.get('font_family', 'sans-serif')
                        self.is_bold = entry.get('is_bold', False)
                        self.is_italic = entry.get('is_italic', False)
                        self.is_thicker = entry.get('is_thicker', False)

                mock_feature = CustomFeature(start_pos, end_pos, label_entry)
                features_to_label.append({'feature': mock_feature, 'label': final_label_text, 'color': user_color})

    # Plot Genes
    features_for_graphic = []
    plot_idx = 0
    for f in original_gene_feats:
        is_in_region = (not limit_region) or (f.end > region_start and f.start < region_end)
        if not is_in_region: continue
        if plot_idx < len(features_to_plot_colored):
            gf = features_to_plot_colored[plot_idx]
            if hasattr(f, 'is_baseline') and f.is_baseline:
                features_for_graphic.append(gf)
            else:
                gf_no_label = GraphicFeature(start=gf.start, end=gf.end, strand=gf.strand, color=gf.color, label=None)
                features_for_graphic.append(gf_no_label)
            plot_idx += 1

    graphic.features = features_for_graphic
    _before = len(ax.patches)
    graphic.plot(ax=ax)
    _dfv_patches = ax.patches[_before:]

    # Store patch info for hovering
    filtered_dfv_patches = []
    for patch in _dfv_patches:
        if isinstance(patch, mpatches.Circle) and np.isclose(patch.radius, R): continue
        filtered_dfv_patches.append(patch)

    patch_theta_list = []
    for patch in filtered_dfv_patches:
        if hasattr(patch, 'theta1') and hasattr(patch, 'theta2'):
            theta_mid = (patch.theta1 + patch.theta2) / 2.0
            patch_theta_list.append((patch, theta_mid))

    for idx, feature_data in enumerate(features_to_plot_colored):
        feature_obj = feature_plot_map[idx] if idx < len(feature_plot_map) else None
        center_bp = (feature_data.start + feature_data.end) / 2.0
        feature_theta = 90.0 - 360.0 * ((center_bp - float(plot_top_position)) / plot_seqlen)

        # Match patch to feature
        matching_patches = []
        for patch, patch_theta in patch_theta_list:
            diff = abs(feature_theta - patch_theta)
            if diff > 180: diff = 360 - diff
            if diff < 1.0: matching_patches.append((patch, diff))

        if matching_patches:
            matching_patches.sort(key=lambda x: x[1])
            for patch, _ in matching_patches:
                gene_feature_patches_local.append((patch, feature_data.start, feature_data.end, feature_obj))
        else:
            best_patch = None
            min_diff = float('inf')
            for patch, patch_theta in patch_theta_list:
                diff = abs(feature_theta - patch_theta)
                if diff > 180: diff = 360 - diff
                if diff < min_diff:
                    min_diff, best_patch = diff, patch
            if best_patch:
                gene_feature_patches_local.append((best_patch, feature_data.start, feature_data.end, feature_obj))

    # Remove gene borders if requested
    if remove_gene_borders:
        for p in _dfv_patches:
            fc = None
            if hasattr(p, "get_facecolor"):
                try: fc = p.get_facecolor()
                except Exception: fc = None
            if not fc or (isinstance(fc, tuple) and len(fc) >= 4 and fc[3] == 0): continue

            # Don't remove baseline circle border
            if hasattr(p, "theta1") and hasattr(p, "theta2"):
                try:
                    span = abs(float(p.theta2) - float(p.theta1))
                    if span >= 359.0: continue
                except Exception:
                    pass
            if hasattr(p, "set_edgecolor"):
                try: p.set_edgecolor("none")
                except Exception: pass
            if hasattr(p, "set_linewidth"):
                try: p.set_linewidth(0)
                except Exception: pass

    # Draw Limited Region Baseline
    if limit_region:
        # Remove full circle baseline
        baseline_circle = None
        target_center = (0, -R)
        for patch in ax.patches:
            if isinstance(patch, mpatches.Circle) and patch.center == target_center and np.isclose(patch.radius, R):
                baseline_circle = patch
                break
        if baseline_circle:
            baseline_circle.remove()
            if region_end is None: region_end = plot_seqlen
            t1_deg = 90.0 - 360.0 * ((region_end - float(plot_top_position)) / plot_seqlen)
            t2_deg = 90.0 - 360.0 * ((region_start - float(plot_top_position)) / plot_seqlen)
            baseline_arc = mpatches.Arc(xy=(0, -R), width=R * 2, height=R * 2, angle=0.0, theta1=t1_deg, theta2=t2_deg,
                                        fill=False, edgecolor='black', linestyle='-', lw=1, zorder=0.1)
            ax.add_patch(baseline_arc)

    # Draw Inner Rings
    identity_ring_artists_local = []

    if plot_params.get("id_ring_precise_mode", False):
        identity_ring_artists_local = draw_precise_identity_ring(
            ax, graphic, data, plot_params, nc_cmap, id_ring_radius, calculated_id_thickness,
            limit_region=limit_region, region_start=region_start, region_end=region_end
        )
    else:
        if scores_to_plot:
            artists_before = set(ax.patches) | set(ax.collections)
            # Extract coding regions from gene features
            coding_regions_list = [(f.start, f.end) for f in original_gene_feats]

            draw_identity_ring(
                ax, graphic, scores_to_plot, plot_seqlen, plot_top_position, cmap=nc_cmap,
                color_only_mode=plot_params.get("id_ring_color_only", False),
                auto_scale_color=plot_params.get("nc_auto_fit_color", True),
                inner_radius=id_ring_radius, thickness=calculated_id_thickness,
                baseline_alpha=id_alpha, step_size=nc_step_size,
                limit_region=limit_region, region_start=region_start, region_end=region_end,
                coding_regions=coding_regions_list, include_coding=include_coding, show_noncoding=show_noncoding
            )
            artists_after = set(ax.patches) | set(ax.collections)
            identity_ring_artists_local = list(artists_after - artists_before)

    ax._identity_ring_artists = identity_ring_artists_local

    # Draw BLAST Hits (Vectorized)
    draw_blast_hits_as_ring(
        ax, filtered_blast_df, plot_seqlen, graphic, homology_min=homology_thresh_min,
        homology_max=homology_thresh_max, color_min=color_vmin, color_max=color_vmax, cmap=cmap,
        radius=blast_ring_radius, thickness=blast_ring_thickness,
        show_blast_separators=show_blast_separators, limit_region=limit_region,
        region_start=region_start, region_end=region_end
    )

    # Draw GC Content and Skew (Vectorized calculation)
    draw_gc_and_skew_plot(
        ax, graphic, original_gc_arr, original_skew_arr, plot_seqlen, inner_radius=gc_inner,
        skew_inner_radius=skew_inner_radius, gc_thickness=gc_thickness, skew_thickness=skew_thickness,
        top_position=plot_top_position, gc_baseline_alpha=gc_alpha, skew_baseline_alpha=skew_alpha,
        gc_cmap=gc_cmap, gc_norm=gc_norm,
        color_only_mode=plot_params.get("gc_ring_color_only", False), limit_region=limit_region,
        region_start=region_start, region_end=region_end
    )

    # Draw Custom Markers & Calculate Max Gene Radius
    if 'custom_labels' in data and data['custom_labels']:
        _draw_custom_label_markers(ax, data['custom_labels'], plot_seqlen, graphic, R,
                                   limit_region=limit_region, region_start=region_start, region_end=region_end)

    # Calculate Precise Max Radius
    max_gene_radius_found = R  # Default to baseline
    if 'filtered_dfv_patches' in locals():
        for patch in filtered_dfv_patches:
            if hasattr(patch, 'r'):
                if patch.r > max_gene_radius_found:
                    max_gene_radius_found = patch.r

    max_gene_radius_frac = max_gene_radius_found / R if R > 0 else 1.0

    # Zoom & Pan Logic
    final_zoom = zoom_factor
    x_center_pan, y_center_pan = 0, -R
    auto_zoom_scalar = 1.0
    calculated_this_run_x_pan = 0.0
    calculated_this_run_y_pan = -R

    if limit_region:
        if plot_seqlen != full_seqlen:
            auto_zoom_scalar = 1.1
        else:
            region_span_bp = region_end - region_start
            arc_fraction = region_span_bp / full_seqlen
            if arc_fraction < 0.8:
                # 1. Angular Zoom (Base)
                angular_zoom = min(1.0 / np.sqrt(arc_fraction), 8.0)

                # 2. Radial Zoom Constraint (Prevents cutting off tracks/labels)
                # Determine lowest visible ring radius
                lowest_r_frac = 1.0
                if plot_params.get("gc_inner") is not None:
                    lowest_r_frac = min(lowest_r_frac, plot_params.get("gc_inner"))
                if plot_params.get("skew_inner_radius") is not None:
                    lowest_r_frac = min(lowest_r_frac, plot_params.get("skew_inner_radius"))

                min_r_data = R * lowest_r_frac
                # Calculate required height from innermost ring to outermost gene
                radial_height = max(max_gene_radius_found - min_r_data, 0.1 * R)

                # Calculate max zoom that fits this height in the viewport
                # We add 25% buffer for labels/margins
                required_view_height = radial_height * 1.25
                max_radial_zoom = (3.2 * R) / required_view_height

                # Apply constraint (factor in user zoom)
                auto_zoom_scalar = min(angular_zoom, max_radial_zoom / max(0.1, zoom_factor))

        final_zoom = auto_zoom_scalar * zoom_factor

        if recalculate_auto_view:
            region_midpoint_bp = (region_start + region_end) / 2.0
            mid_angle_rad = np.deg2rad(90.0 - 360.0 * ((region_midpoint_bp - float(plot_top_position)) / plot_seqlen))

            lowest_r = 1.0
            if plot_params.get("gc_inner") is not None:
                lowest_r = min(lowest_r, plot_params.get("gc_inner"))
            if plot_params.get("skew_inner_radius") is not None:
                lowest_r = min(lowest_r, plot_params.get("skew_inner_radius"))

            # Calculate a "Target Radius" halfway between the genes (1.0) and the bottom ring
            target_radius_factor = (1.0 + lowest_r) / 2.0

            # Calculate coordinates based on this mid-point radius
            target_R = R * target_radius_factor

            calculated_this_run_x_pan = target_R * np.cos(mid_angle_rad)
            calculated_this_run_y_pan = target_R * np.sin(mid_angle_rad) - R

            # Apply values to current view
            x_center_pan = calculated_this_run_x_pan
            y_center_pan = calculated_this_run_y_pan
        else:
            x_center_pan, y_center_pan = cached_x_center_pan, cached_y_center_pan
            calculated_this_run_x_pan, calculated_this_run_y_pan = cached_x_center_pan, cached_y_center_pan

    plot_limit = (R * 1.6) / final_zoom
    ax.set_xlim(x_center_pan - plot_limit, x_center_pan + plot_limit)
    ax.set_ylim(y_center_pan - plot_limit, y_center_pan + plot_limit)

    cached_xlim = plot_params.get('cached_xlim')
    cached_ylim = plot_params.get('cached_ylim')
    if not recalculate_auto_view and limit_region and cached_xlim is not None and cached_ylim is not None:
        ax.set_xlim(cached_xlim)
        ax.set_ylim(cached_ylim)

    # Label Placement
    label_args = plot_params.copy()
    label_args.update({
        'features_to_label': features_to_label,
        'seqlen': plot_seqlen,
        'graphic': graphic,
        'homology_min': homology_thresh_min,
        'homology_max': homology_thresh_max,
        'color_min': color_vmin,
        'color_max': color_vmax,
        'cmap': cmap,
        'auto_zoom_scalar': auto_zoom_scalar,
        'cached_auto_zoom_scalar': plot_params.get('cached_auto_zoom_scalar', None),
        'recalculate_auto_view': recalculate_auto_view,
        'plot_center_x': x_center_pan,
        'plot_center_y': y_center_pan,
        'plot_radius_limit': plot_limit,
        'limit_region': limit_region,
        'region_start': region_start,
        'region_end': region_end,
        'max_gene_radius_frac': max_gene_radius_frac
    })

    if use_smart_layout:
        final_fontsize, text_objects = place_labels_smartly(ax, **label_args)
    else:
        final_fontsize, text_objects = place_labels_vertically(ax, **label_args)

    limits_before_pass2 = (ax.get_xlim(), ax.get_ylim())

    # Refine View for Limited Regions with Labels
    if limit_region and text_objects:
        try:
            fig.canvas.draw()
            renderer = fig.canvas.get_renderer()
            start_angle_deg = 90.0 - 360.0 * ((region_start - float(plot_top_position)) / plot_seqlen)
            end_angle_deg = 90.0 - 360.0 * ((region_end - float(plot_top_position)) / plot_seqlen)
            angle_samples = np.linspace(np.deg2rad(end_angle_deg), np.deg2rad(start_angle_deg), 100)
            x_samples = R * np.cos(angle_samples)
            arc_left_x, arc_right_x = np.min(x_samples), np.max(x_samples)
            arc_midpoint_x = (arc_left_x + arc_right_x) / 2.0

            left_labels = [t for t in text_objects if t is not None and t.get_position()[0] < arc_midpoint_x]
            right_labels = [t for t in text_objects if t is not None and t.get_position()[0] > arc_midpoint_x]
            has_dual_sided_labels = len(left_labels) > 0 and len(right_labels) > 0

            plot_artists = ax.patches + ax.lines + ax.collections
            if plot_artists:
                plot_bboxes = [a.get_window_extent(renderer=renderer) for a in plot_artists]
                plot_valid_bboxes = [b for b in plot_bboxes if b.width > 0 and b.height > 0]

                if plot_valid_bboxes:
                    inv = ax.transData.inverted()
                    plot_bbox = mtransforms.Bbox.union(plot_valid_bboxes).transformed(inv)
                    plot_x0, plot_y0 = plot_bbox.p0
                    plot_x1, plot_y1 = plot_bbox.p1

                    text_bboxes = [t.get_window_extent(renderer=renderer) for t in text_objects if t is not None]
                    if text_bboxes:
                        label_bbox = mtransforms.Bbox.union(text_bboxes).transformed(inv)
                        label_x0, label_y0 = label_bbox.p0
                        label_x1, label_y1 = label_bbox.p1

                        if has_dual_sided_labels:
                            left_extent = min(plot_x0, label_x0, arc_left_x)
                            right_extent = max(plot_x1, label_x1, arc_right_x)
                            left_distance = arc_midpoint_x - left_extent
                            right_distance = right_extent - arc_midpoint_x
                            half_width = max(left_distance, right_distance)

                            view_x0 = arc_midpoint_x - half_width
                            view_x1 = arc_midpoint_x + half_width
                            content_y0 = min(label_y0, plot_y0)
                            content_y1 = max(label_y1, plot_y1)
                            pad_x = (view_x1 - view_x0) * 0.08
                            pad_y = (content_y1 - content_y0) * 0.05
                            final_x0 = view_x0 - pad_x
                            final_x1 = view_x1 + pad_x
                            final_y0 = content_y0 - pad_y
                            final_y1 = content_y1 + pad_y
                        else:
                            content_x0 = min(plot_x0, label_x0)
                            content_x1 = max(plot_x1, label_x1)
                            content_y0 = min(plot_y0, label_y0)
                            content_y1 = max(plot_y1, label_y1)
                            pad_x = (content_x1 - content_x0) * 0.08
                            pad_y = (content_y1 - content_y0) * 0.05
                            final_x0 = content_x0 - pad_x
                            final_x1 = content_x1 + pad_x
                            final_y0 = content_y0 - pad_y
                            final_y1 = content_y1 + pad_y

                        ax.set_xlim(final_x0, final_x1)
                        ax.set_ylim(final_y0, final_y1)

                        if recalculate_auto_view:
                            calculated_this_run_x_pan = (final_x0 + final_x1) / 2.0
                            calculated_this_run_y_pan = (final_y0 + final_y1) / 2.0
        except Exception:
            pass

    wga_label = "Non-Coding ID"
    wga_id_for_title = data.get("noncoding_avg_id", 0.0)
    identity_algo = data.get("identity_algorithm_used", "Mauve + SibeliaZ Fallback")
    if "Legacy" in identity_algo:
        wga_label = "Legacy WGA ID"
        wga_id_for_title = data.get("global_wga_id", 0.0)

    ani, dDDH = _calculate_genome_stats(filtered_blast_df)

    anib_val = data.get("anib_score", 0.0)
    anib_cov = data.get("anib_coverage", 0.0)

    title_str = (f"Comparison of {record.id} ({full_seqlen:,} bp) with {ref_record.id} ({ref_seqlen:,} bp)\n"
                 f"dDDH: {dDDH:.2f}%  |  ANIb: {anib_val:.2f}% (Cov: {anib_cov:.1f}%)")

    if wga_id_for_title > 0.0: title_str += f"  |  {wga_label}: {wga_id_for_title:.2f}%"
    ax.set_title(title_str, weight="bold", size=12, y=1.02)

    all_ids_for_colorbar = []
    if plot_params.get("id_ring_precise_mode", False):
        if identity_ring_artists_local:
            for artist in identity_ring_artists_local:
                if isinstance(artist, matplotlib.collections.PatchCollection):
                    # For collections, look at the lcb_info_list
                    if hasattr(artist, 'lcb_info_list'):
                        for info in artist.lcb_info_list:
                            if info.get('identity', 0) > 0:
                                all_ids_for_colorbar.append(info['identity'])
                elif hasattr(artist, 'lcb_info'):
                    val = artist.lcb_info.get('identity', 0)
                    if val > 0: all_ids_for_colorbar.append(val)
    else:
        if scores_to_plot:
            all_ids_for_colorbar = [s for _, s in scores_to_plot]

    legend, colorbars = _create_legend_and_colorbars(
        fig, ax, data, plot_params, filtered_blast_df, scores_to_plot,
        bed_genes=[f for f in original_gene_feats if f.color == '#e53935'],
        homolog_genes=[f for f in original_gene_feats if f.color == '#43a047'],
        gc_cmap=gc_cmap, gc_norm=gc_norm, nc_cmap=nc_cmap,
        color_vmin=color_vmin, color_vmax=color_vmax, id_label=id_label,
        all_ids=all_ids_for_colorbar,
        limit_region=limit_region, region_start=region_start, region_end=region_end
    )

    # Draw Start/End Line (if enabled)
    if show_start_end_line:
        if not limit_region or (region_start <= 0 < region_end):
            angle_rad = np.deg2rad(90.0 + 360.0 * (plot_top_position / plot_seqlen))
            x_start, y_start = 0, -R
            end_radius = R * 1.15
            x_end = end_radius * np.cos(angle_rad)
            y_end = (end_radius * np.sin(angle_rad)) - R
            ax.plot([x_start, x_end], [y_start, y_end], color='black', alpha=0.5, linewidth=0.75, zorder=2)

    # Draw Coordinate Numbers (if enabled)
    if plot_params.get("show_coords", False):
        # Calculate number of gene levels based on how far out they extend
        max_gene_radius = R
        for patch in filtered_dfv_patches:
            if hasattr(patch, 'r') and patch.r > max_gene_radius:
                max_gene_radius = patch.r

        gene_levels = int((max_gene_radius - R) / (0.1 * R)) + 1
        gene_levels = max(1, gene_levels)
        text_radius = R * (1.08 + 0.08 * (gene_levels - 1))

        # Logic Selection
        format_coord_final = None
        coord_positions = []

        if limit_region:
            region_fraction = (region_end - region_start) / full_seqlen
            arc_degrees = 360.0 * region_fraction

            if arc_degrees <= 2: num_coords_to_show = 2
            elif arc_degrees < 10: num_coords_to_show = 3
            elif arc_degrees < 20: num_coords_to_show = 5
            elif plot_seqlen != full_seqlen: num_coords_to_show = 20
            else: num_coords_to_show = max(10, int(50 * region_fraction))

            coord_positions = np.linspace(region_start, region_end, num_coords_to_show)

            best_precision = 1
            for precision in [1000, 100, 50, 10]:
                is_fit = True
                for val in coord_positions:
                    if int(round(val)) % precision != 0:
                        is_fit = False
                        break
                if is_fit: best_precision = precision; break

            def format_limited(bp):
                val = int(round(bp))
                if best_precision == 1000: return f"{val // 1000}k"
                return f"{val}"

            format_coord_final = format_limited
        else:
            num_coords = 50
            coord_positions = np.linspace(0, full_seqlen, num_coords, endpoint=False)

            chosen_strategy = '10'
            for strategy in ['1k', '500', '100', '50']:
                labels = []
                for x in coord_positions:
                    if strategy == '1k': labels.append(f"{int(round(x / 1000))}k")
                    elif strategy == '500': labels.append(f"{int(round(x / 500) * 500)}")
                    elif strategy == '100': labels.append(f"{int(round(x / 100) * 100)}")
                    elif strategy == '50': labels.append(f"{int(round(x / 50) * 50)}")
                if len(labels) == len(set(labels)):
                    chosen_strategy = strategy
                    break

            def format_full(bp):
                val = int(round(bp))
                if chosen_strategy == '1k': return f"{val // 1000}k"
                elif chosen_strategy == '500': return f"{int(round(val / 500) * 500)}"
                elif chosen_strategy == '100': return f"{int(round(val / 100) * 100)}"
                elif chosen_strategy == '50': return f"{int(round(val / 50) * 50)}"
                return f"{int(round(val / 10) * 10)}"

            format_coord_final = format_full

        for bp_pos in coord_positions:
            if limit_region and not (region_start <= bp_pos <= region_end): continue

            angle_deg = 90.0 - 360.0 * ((bp_pos - float(plot_top_position)) / plot_seqlen)
            angle_rad = np.deg2rad(angle_deg)
            x = text_radius * np.cos(angle_rad)
            y = text_radius * np.sin(angle_rad) - R
            label = format_coord_final(bp_pos)
            ha, va = 'center', 'center'
            if 0 < angle_deg < 180: va = 'bottom'
            if 180 < angle_deg < 360: va = 'top'
            if 90 < angle_deg < 270: ha = 'right'
            if angle_deg < 90 or angle_deg > 270: ha = 'left'

            ax.text(x, y, label, ha=ha, va=va, fontsize=8, color='black', clip_on=False, zorder=3)

    new_thicknesses = {"id_thickness": calculated_id_thickness, "gc_thickness": gc_thickness,
                       "skew_thickness": skew_thickness, "label_fontsize": final_fontsize,
                       "auto_zoom_scalar": auto_zoom_scalar}
    calculated_pan_tuple = (calculated_this_run_x_pan, calculated_this_run_y_pan)
    hover_data = {"graphic_ref": graphic, "last_calculated_values": new_thicknesses,
                  "id_label": id_label, "gene_feature_patches": gene_feature_patches_local,
                  "identity_ring_artists": identity_ring_artists_local,
                  "scores_plotted_for_hover": scores_to_plot}

    return fig, legend, [], colorbars, new_thicknesses, calculated_pan_tuple, limits_before_pass2, hover_data

# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# Plotting & Drawing Functions
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────

def _estimate_dfv_radius(ax, graphic):
    R = getattr(graphic, "radius", None)
    if isinstance(R, (int, float)) and R > 0: return float(R)
    y0, y1 = ax.get_ylim()
    return abs(y0) if y0 < 0 < y1 else 1.0


def compute_top_position_for_roi(roi_features, seqlen, base_top=0.0, desired_angle_deg=0.0, manual_deg=0.0,
                                 limit_region=False, region_start=0, region_end=None):
    """
    Computes the top position for plot rotation, with special handling for limited regions.
    """
    # Determine the base position to center on
    if limit_region and region_end is not None:
        # 1. Use the region midpoint as the feature to center
        target_position_bp = (region_start + region_end) / 2.0
    elif roi_features:
        # 2. Use the average position of ROI features
        centers = [(((f.start + f.end) / 2.0) % seqlen) for f in roi_features]
        # Calculate circular mean of positions
        thetas_bp = 2 * np.pi * (np.asarray(centers) / seqlen)
        mean_theta_bp = np.arctan2(np.mean(np.sin(thetas_bp)), np.mean(np.cos(thetas_bp)))
        target_position_bp = (mean_theta_bp * seqlen / (2 * np.pi)) % seqlen
    else:
        # 3. No features, just use the base
        target_position_bp = base_top

    # Calculate the base rotation
    base_top_calculated = target_position_bp - (seqlen * (90.0 - desired_angle_deg) / 360.0)
    # Add manual rotation and return
    final_top_position = (base_top_calculated + (manual_deg / 360.0) * seqlen) % seqlen
    return final_top_position


def draw_identity_ring(ax, graphic, score_data, seqlen, top_position, cmap, auto_scale_color=True,
                       color_only_mode=False, inner_radius=0.2, thickness=0.2, baseline_alpha=1.0,
                       step_size=50, limit_region=False, region_start=0, region_end=None,
                       coding_regions=None, include_coding=True, show_noncoding=True):
    """Draws identity data with multiple viewing modes and corrected radius/thickness logic.
       Baseline now represents 0% identity when color_only_mode is False.
    """
    if limit_region:
        if region_end is None: region_end = seqlen
        score_data = [(p, s) for p, s in score_data if region_start <= p < region_end]
    if not score_data: return

    R = _estimate_dfv_radius(ax, graphic)
    all_scores = np.array([score for _, score in score_data])
    if len(all_scores) == 0: return
    r_inner = R * inner_radius
    data_min, data_max = all_scores.min(), all_scores.max()

    # Color Normalization
    if auto_scale_color and data_max > data_min:
        norm = mcolors.Normalize(vmin=data_min, vmax=data_max)
    else:
        norm = mcolors.Normalize(vmin=0, vmax=100)

    # Draw Baseline FIRST (in both Color-Only and Bar modes)
    if baseline_alpha > 0.0:  # Only draw if not completely transparent
        r_baseline = r_inner
        baseline_color = 'gray'
        if limit_region:
            if region_end is None: region_end = seqlen
            t1_deg = 90.0 - 360.0 * ((region_end - float(top_position)) / seqlen)
            t2_deg = 90.0 - 360.0 * ((region_start - float(top_position)) / seqlen)

            baseline_arc = mpatches.Arc(
                xy=(0, -R), width=r_baseline * 2, height=r_baseline * 2,
                angle=0.0, theta1=t1_deg, theta2=t2_deg,
                fill=False, edgecolor=baseline_color, linestyle='--', lw=1, zorder=0.45, alpha=baseline_alpha
            )
            ax.add_patch(baseline_arc)
        else:
            baseline_circle = mpatches.Circle((0, -R), r_baseline, fill=False, edgecolor=baseline_color,
                                              linestyle='--', lw=1, zorder=0.45, alpha=baseline_alpha)
            ax.add_patch(baseline_circle)

    if color_only_mode:
        # Color-Only Mode
        r_start = r_inner
        r_end = r_inner + (R * thickness)
        arcs_to_draw = []
        if score_data and len(score_data) > 1:
            current_arc = [score_data[0]]
            for i in range(1, len(score_data)):
                if (score_data[i][0] - score_data[i - 1][0]) > (step_size * 1.5):
                    arcs_to_draw.append(current_arc)
                    current_arc = []
                current_arc.append(score_data[i])
            arcs_to_draw.append(current_arc)
        elif score_data:
            arcs_to_draw.append(score_data)

        drawn_scores = []
        for arc_data in arcs_to_draw:
            for pos, score in arc_data:
                drawn_scores.append(score)

        total_polygons = 0
        for arc_data in arcs_to_draw:
            if len(arc_data) < 1:
                continue
            positions = np.array([p for p, s in arc_data])
            scores = np.array([s for p, s in arc_data])
            angles = np.deg2rad(90.0 - 360.0 * ((positions - float(top_position)) / seqlen))

            # Draw polygons for ALL points (including the last one)
            for i in range(len(positions)):
                # Calculate end angle: next point's angle, or extrapolate using step_size for last point
                if i + 1 < len(angles):
                    angle_end = angles[i + 1]
                else:
                    # For the last point, calculate where it should extend to based on step_size
                    next_pos = positions[i] + step_size
                    if limit_region and region_end is not None:
                        next_pos = min(next_pos, region_end)
                    angle_end = np.deg2rad(90.0 - 360.0 * ((next_pos - float(top_position)) / seqlen))

                    # Clip angle based on coding/non-coding boundaries if restrictions apply
                    if coding_regions is not None and (not include_coding or not show_noncoding):
                        current_pos = positions[i]
                        clipped_end = next_pos

                        if not include_coding and show_noncoding:
                            # Showing only non-coding: clip at coding region start
                            for c_start, c_end in coding_regions:
                                if current_pos < c_start < clipped_end:
                                    clipped_end = min(clipped_end, c_start)

                        elif include_coding and not show_noncoding:
                            # Showing only coding: find which coding region we're in and clip at its end
                            in_coding_region = None
                            for cs, ce in coding_regions:
                                if cs <= current_pos < ce:
                                    in_coding_region = (cs, ce)
                                    break

                            if in_coding_region:
                                cs, ce = in_coding_region
                                # Clip at the end of current coding region
                                clipped_end = min(clipped_end, ce)

                                next_coding_start = None
                                for ncs, nce in coding_regions:
                                    if ncs >= ce and (next_coding_start is None or ncs < next_coding_start):
                                        next_coding_start = ncs

                                if next_coding_start is not None and next_coding_start <= next_pos:
                                    clipped_end = next_coding_start

                        # Recalculate angle_end with clipped position
                        if clipped_end != next_pos:
                            angle_end = np.deg2rad(90.0 - 360.0 * ((clipped_end - float(top_position)) / seqlen))

                color = cmap(norm(scores[i]))

                verts = [
                    (r_end * np.cos(angles[i]), r_end * np.sin(angles[i]) - R),
                    (r_end * np.cos(angle_end), r_end * np.sin(angle_end) - R),
                    (r_start * np.cos(angle_end), r_start * np.sin(angle_end) - R),
                    (r_start * np.cos(angles[i]), r_start * np.sin(angles[i]) - R),
                ]
                polygon = mpatches.Polygon(verts, closed=True, facecolor=color, edgecolor=color, zorder=0.4)
                ax.add_patch(polygon)

                total_polygons += 1
    else:
        # Bar Mode (Baseline = 0%)
        r_baseline = r_inner

        positions, scores = np.array([p for p, s in score_data]), np.array([s for p, s in score_data])

        angles = np.deg2rad(90.0 - 360.0 * ((positions - float(top_position)) / seqlen))
        normalized_scores_0_100 = norm(scores)
        r_scores = r_baseline + (R * thickness * normalized_scores_0_100)
        x_starts, y_starts = r_baseline * np.cos(angles), r_baseline * np.sin(angles) - R
        x_ends, y_ends = r_scores * np.cos(angles), r_scores * np.sin(angles) - R
        lines = [[(x_starts[i], y_starts[i]), (x_ends[i], y_ends[i])] for i in range(len(positions))]
        line_colors = cmap(norm(scores))
        if lines:
            line_collection = LineCollection(lines, colors=line_colors, linewidths=1, zorder=0.4)
            ax.add_collection(line_collection)


def draw_gc_and_skew_plot(ax, graphic, gc_arr, skew_arr, seqlen, inner_radius=0.40, skew_inner_radius=0.60,
                          gc_thickness=1.20, skew_thickness=0.80, top_position=None,
                          gc_baseline_alpha=1.0, skew_baseline_alpha=1.0, gc_cmap=None, gc_norm=None,
                          color_only_mode=False, limit_region=False, region_start=0, region_end=None):
    if gc_arr is None or len(gc_arr) == 0: return
    if top_position is None: top_position = getattr(graphic, "top_position", 0)
    R = _estimate_dfv_radius(ax, graphic)

    if limit_region:
        # Dynamic Local Sampling: Grab points ONLY from the visible region
        r_start = int(region_start)
        r_end = int(region_end) if region_end is not None else seqlen

        # Handle wrap-around (e.g. start=4.9Mb, end=100bp)
        if r_start > r_end:
            idx_part1 = np.arange(r_start, len(gc_arr))
            idx_part2 = np.arange(0, r_end)
            full_region_idx = np.concatenate((idx_part1, idx_part2))
        else:
            # Clamp to bounds just in case
            s = max(0, r_start)
            e = min(len(gc_arr), r_end)
            full_region_idx = np.arange(s, e)

        # Downsample to max 2000 points within THIS region
        if len(full_region_idx) > 2000:
            sub_indices = np.linspace(0, len(full_region_idx) - 1, 2000, dtype=int)
            idx = full_region_idx[sub_indices]
        else:
            idx = full_region_idx

        if len(idx) < 2: return  # Region too small to plot

        # Extract data directly (No closure point needed for partial arc)
        pos = idx
        gc = gc_arr[idx]
        skew = skew_arr[idx] if skew_arr is not None else np.zeros(len(idx))

    else:
        # Original Global Sampling (Low Res for Full View)
        n = min(2000, len(gc_arr))
        idx = np.linspace(0, len(gc_arr) - 1, n, dtype=int)
        pos_orig = np.linspace(0, seqlen, n, endpoint=False)
        gc_orig = gc_arr[idx]
        skew_orig = skew_arr[idx] if skew_arr is not None and len(skew_arr) > 0 else np.zeros(n)

        # Append closure point for full circle
        pos = np.append(pos_orig, seqlen)
        gc = np.append(gc_orig, gc_orig[0])
        skew = np.append(skew_orig, skew_orig[0])

    angle_deg = 90.0 - 360.0 * ((pos - float(top_position)) / seqlen)
    theta = np.deg2rad(angle_deg)

    # GC Content Plot
    r_inner_gc = R * inner_radius

    if gc_cmap is None or gc_norm is None:
        # Draw a single diverging polygon (Grey style)
        # Use median from the visible data for better local contrast
        gc_median = np.median(gc)

        if limit_region:
            # For limited region, we don't need complex break logic if we pre-filtered indices
            # But we keep it robust for safety
            r_outer_gc = r_inner_gc + R * gc_thickness * (gc - gc_median)

            # Simple single polygon for the region
            verts = np.column_stack([
                np.concatenate(
                    [r_outer_gc * np.cos(theta), np.full_like(r_outer_gc, r_inner_gc) * np.cos(theta[::-1])]),
                np.concatenate(
                    [r_outer_gc * np.sin(theta) - R, np.full_like(r_outer_gc, r_inner_gc) * np.sin(theta[::-1]) - R])
            ])
            ax.add_patch(mpatches.Polygon(verts, closed=True, facecolor="0.7", alpha=0.5, edgecolor="none", zorder=0.5))
        else:
            # Full circle plot
            r_outer_gc = r_inner_gc + R * gc_thickness * (gc - gc_median)
            verts = np.column_stack([
                np.concatenate(
                    [r_outer_gc * np.cos(theta), np.full_like(r_outer_gc, r_inner_gc) * np.cos(theta[::-1])]),
                np.concatenate(
                    [r_outer_gc * np.sin(theta) - R, np.full_like(r_outer_gc, r_inner_gc) * np.sin(theta[::-1]) - R])
            ])
            ax.add_patch(mpatches.Polygon(verts, closed=True, facecolor="0.7", alpha=0.5, edgecolor="none", zorder=0.5))
    else:
        # Draw colored segments
        if color_only_mode:
            # Uniform Thickness (Color Only)
            r_start = r_inner_gc
            r_end = r_start + (R * gc_thickness)
            for i in range(len(theta) - 1):
                verts = [
                    (r_end * np.cos(theta[i]), r_end * np.sin(theta[i]) - R),
                    (r_end * np.cos(theta[i + 1]), r_end * np.sin(theta[i + 1]) - R),
                    (r_start * np.cos(theta[i + 1]), r_start * np.sin(theta[i + 1]) - R),
                    (r_start * np.cos(theta[i]), r_start * np.sin(theta[i]) - R),
                ]
                color = gc_cmap(gc_norm(gc[i]))
                polygon = mpatches.Polygon(verts, closed=True, facecolor=color, edgecolor=color, alpha=0.7, zorder=0.5)
                ax.add_patch(polygon)
        else:
            # Diverging Bars (Color + Radius)
            gc_median = np.median(gc)
            r_outer_gc = r_inner_gc + R * gc_thickness * (gc - gc_median)
            for i in range(len(theta) - 1):
                if np.isnan(r_outer_gc[i]) or np.isnan(r_outer_gc[i + 1]): continue

                verts = [
                    (r_outer_gc[i] * np.cos(theta[i]), r_outer_gc[i] * np.sin(theta[i]) - R),
                    (r_outer_gc[i + 1] * np.cos(theta[i + 1]), r_outer_gc[i + 1] * np.sin(theta[i + 1]) - R),
                    (r_inner_gc * np.cos(theta[i + 1]), r_inner_gc * np.sin(theta[i + 1]) - R),
                    (r_inner_gc * np.cos(theta[i]), r_inner_gc * np.sin(theta[i]) - R),
                ]
                color = gc_cmap(gc_norm(gc[i]))
                polygon = mpatches.Polygon(verts, closed=True, facecolor=color, edgecolor='none', alpha=0.5,
                                           zorder=0.5)
                ax.add_patch(polygon)

    # Draw GC baseline as Arc if region is limited
    if limit_region:
        if region_end is None: region_end = seqlen
        t1_deg = 90.0 - 360.0 * ((region_end - float(top_position)) / seqlen)
        t2_deg = 90.0 - 360.0 * ((region_start - float(top_position)) / seqlen)

        gc_arc = mpatches.Arc(
            xy=(0, -R), width=r_inner_gc * 2, height=r_inner_gc * 2,
            angle=0.0, theta1=t1_deg, theta2=t2_deg,
            fill=False, edgecolor='gray', linestyle='--', lw=0.8, zorder=0.6, alpha=gc_baseline_alpha
        )
        ax.add_patch(gc_arc)
    else:
        ax.add_patch(mpatches.Circle((0, -R), r_inner_gc, fill=False, edgecolor='gray',
                                     linestyle='--', lw=0.8, zorder=0.6, alpha=gc_baseline_alpha))

    # GC Skew Plot
    r_skew_baseline = R * skew_inner_radius
    r_skew = r_skew_baseline + R * skew_thickness * np.clip(skew, -1.0, 1.0)
    x_coords, y_coords = r_skew * np.cos(theta), r_skew * np.sin(theta) - R

    # The np.nan values will break zero_crossings. We must use the masking method.
    zero_crossings = np.where(np.diff(np.sign(np.nan_to_num(skew))))[0]

    # Use the new length of the pos array
    split_indices = np.sort(np.concatenate([[0], zero_crossings + 1, [len(pos) - 1]]))

    for i in range(len(split_indices) - 1):
        start, end = split_indices[i], split_indices[i + 1]
        if start >= end: continue
        end_slice = end + 1

        segment_skew_for_color = skew[start:end_slice]
        if np.all(np.isnan(segment_skew_for_color)): continue
        color = '#33a02c' if np.nanmean(segment_skew_for_color) >= 0 else '#fb9a99'

        segment_x = x_coords[start:end_slice]
        segment_y = y_coords[start:end_slice]
        ax.plot(segment_x, segment_y, lw=1.2, color=color, alpha=0.9, zorder=1)

    # Draw Skew baseline
    if limit_region:
        if region_end is None: region_end = seqlen
        t1_deg = 90.0 - 360.0 * ((region_end - float(top_position)) / seqlen)
        t2_deg = 90.0 - 360.0 * ((region_start - float(top_position)) / seqlen)
        skew_arc = mpatches.Arc(
            xy=(0, -R), width=r_skew_baseline * 2, height=r_skew_baseline * 2, angle=0.0, theta1=t1_deg, theta2=t2_deg,
            fill=False, edgecolor='gray', linestyle='--', lw=0.8, zorder=0.9, alpha=skew_baseline_alpha)
        ax.add_patch(skew_arc)
    else:
        ax.add_patch(mpatches.Circle((0, -R), r_skew_baseline, fill=False, edgecolor='gray',
                                     linestyle='--', lw=0.8, zorder=0.9, alpha=skew_baseline_alpha))


def _batch_assign_homology(features, blast_df, homology_threshold):
    """
    Instantly matches BLAST hits to ALL genes using Numpy broadcasting.
    """
    if not features or blast_df is None or blast_df.empty:
        return

    # 1. Extract Arrays
    # Gene coordinates (N genes)
    gene_starts = np.array([f.start for f in features])
    gene_ends = np.array([f.end for f in features])

    # BLAST coordinates and data (M hits)
    hit_starts = np.minimum(blast_df['qstart'].values, blast_df['qend'].values)
    hit_ends = np.maximum(blast_df['qstart'].values, blast_df['qend'].values)
    hit_pidents = blast_df['pident'].values

    # 2. Vectorized Overlap Check (N x M boolean matrix)
    # Broadcasting: (N, 1) compared against (1, M)
    # Overlap logic: (Gene_Start < Hit_End) AND (Gene_End > Hit_Start)
    overlaps = (gene_starts[:, None] < hit_ends) & (gene_ends[:, None] > hit_starts)

    # 3. Find Best Hit for Every Gene
    # Create matrix of identities where overlaps exist (0 elsewhere)
    identity_matrix = np.zeros_like(overlaps, dtype=float)
    np.putmask(identity_matrix, overlaps, hit_pidents)

    # Max identity per gene (axis 1 = across all hits)
    best_identities = np.max(identity_matrix, axis=1)
    best_hit_indices = np.argmax(identity_matrix, axis=1)

    # 4. Bulk Update Features
    # Identify which genes actually have a match > 0
    has_hit_mask = best_identities > 0

    # We only loop through genes that actually have hits to update properties
    # (This loop is fast because it's just property assignment, no searching)
    for i, feature in enumerate(features):
        if has_hit_mask[i]:
            best_id = best_identities[i]
            hit_idx = best_hit_indices[i]

            # Update Feature Properties
            feature.best_identity = best_id

            # Get specific hit details from the dataframe using the index
            # We use .iloc for direct row access
            hit_row = blast_df.iloc[hit_idx]
            feature.best_sstart = hit_row['sstart']
            feature.best_send = hit_row['send']
            feature.best_sseqid = hit_row['sseqid']
            feature.best_qseqid = hit_row['qseqid']

            # Apply Green Color if Threshold Met
            if best_id >= homology_threshold:
                feature.color = "#43a047"  # Green

                # Update Label to include percentage if not already present
                base_label = getattr(feature, 'label', None)
                if base_label:
                    # Update label logic handled by plot routine,
                    # but we mark it as homologous here
                    pass


def draw_blast_hits_as_ring(ax, blast_df, seqlen, graphic, homology_min, homology_max, color_min, color_max, cmap,
                            radius=1.0, thickness=0.15, show_blast_separators=False,
                            limit_region=False, region_start=0, region_end=None):
    """
    Optimized drawing of BLAST hits using Numpy vectorization and LineCollections.
    """
    R = _estimate_dfv_radius(ax, graphic)
    top_position = getattr(graphic, "top_position", 0)

    # 1. Filter Data using Vectorized Pandas Operations
    # We create a view, not a copy, to stay fast
    mask = (blast_df['pident'] >= homology_min) & (blast_df['pident'] <= homology_max)
    hits = blast_df[mask]

    if hits.empty: return

    # 2. Extract Columns to Numpy Arrays (Instant access)
    qstarts = hits['qstart'].values
    qends = hits['qend'].values
    pidents = hits['pident'].values

    # Calculate start/end positions (min/max of qstart/qend)
    start_pos = np.minimum(qstarts, qends)
    end_pos = np.maximum(qstarts, qends)

    # 3. Handle Region Limiting (Vectorized)
    if limit_region:
        if region_end is None: region_end = seqlen

        # Create a mask for hits that overlap the region
        # Logic: Hit End > Region Start  AND  Hit Start < Region End
        region_mask = (end_pos > region_start) & (start_pos < region_end)

        # Apply mask to all arrays
        if not np.any(region_mask): return

        hits = hits[region_mask]  # Keep df for metadata
        start_pos = start_pos[region_mask]
        end_pos = end_pos[region_mask]
        pidents = pidents[region_mask]

        # Clip coordinates to the region bounds
        start_pos = np.clip(start_pos, region_start, region_end)
        end_pos = np.clip(end_pos, region_start, region_end)

    # 4. Calculate Angles (Theta) Vectorized
    # Standard formula: 90 - 360 * (pos - top) / seqlen
    # We cast to float to ensure precision
    norm_start = (start_pos - float(top_position)) / seqlen
    norm_end = (end_pos - float(top_position)) / seqlen

    theta1 = 90.0 - 360.0 * norm_start
    theta2 = 90.0 - 360.0 * norm_end

    # 5. Calculate Colors Vectorized
    vmin, vmax = color_min, color_max
    if vmax > vmin:
        norm_vals = (pidents - vmin) / (vmax - vmin)
    else:
        norm_vals = np.ones_like(pidents)

    # Get all colors at once (cmap accepts numpy arrays)
    colors = cmap(norm_vals)

    # 6. Create Wedges Loop (Optimized)
    # We zip numpy arrays which is much faster than iterrows
    # We use hits.itertuples(index=False) for fast metadata access without Series overhead
    new_patches = []

    for row, t1, t2, c in zip(hits.itertuples(index=False), theta1, theta2, colors):
        wedge = mpatches.Wedge(
            center=(0, -R), r=R * radius, theta1=t2, theta2=t1,
            width=R * thickness, facecolor=c, edgecolor=None, zorder=0
        )
        # Attach row metadata for tooltips (compatible with existing hover logic)
        wedge.blast_info = row
        new_patches.append(wedge)

    # Batch add is slightly cleaner, though add_patch is standard for wedges
    for p in new_patches:
        ax.add_patch(p)

    # 7. Optimized Separators (LineCollection)
    if show_blast_separators and len(start_pos) > 1:
        # Merge intervals to find visual junctions
        intervals = sorted(zip(start_pos, end_pos))
        merged_intervals = []
        if intervals:
            curr_start, curr_end = intervals[0]
            tolerance = 5
            for next_start, next_end in intervals[1:]:
                if next_start <= (curr_end + tolerance):
                    curr_end = max(curr_end, next_end)
                else:
                    merged_intervals.append((curr_start, curr_end))
                    curr_start, curr_end = next_start, next_end
            merged_intervals.append((curr_start, curr_end))

        # Identify internal junctions
        external = set(p for iv in merged_intervals for p in iv)
        all_pts = set(p for iv in intervals for p in iv)
        junctions = list(all_pts - external)

        if junctions:
            # Vectorized Line Calculation
            j_arr = np.array(junctions)

            # Calculate angles for all junctions
            j_angles_deg = 90.0 - 360.0 * ((j_arr - float(top_position)) / seqlen)
            j_angles_rad = np.deg2rad(j_angles_deg)

            r_inner = R * (radius - thickness)
            r_outer = R * radius
            cos_a = np.cos(j_angles_rad)
            sin_a = np.sin(j_angles_rad)

            # Start points (Inner)
            x0 = r_inner * cos_a
            y0 = r_inner * sin_a - R

            # End points (Outer)
            x1 = r_outer * cos_a
            y1 = r_outer * sin_a - R

            # Stack into (NumLines, 2, 2) for LineCollection
            # Structure: [ [(x0, y0), (x1, y1)], ... ]
            points = np.stack([np.stack([x0, y0], axis=-1),
                               np.stack([x1, y1], axis=-1)], axis=1)

            # Create and add collection (One object instead of hundreds)
            lc = mcoll.LineCollection(points, colors='white', linewidths=0.6, zorder=0.1)
            ax.add_collection(lc)


def draw_precise_identity_ring(ax, graphic, data, plot_params, nc_cmap, calculated_radius_frac,
                               calculated_thickness_frac, limit_region=False, region_start=0, region_end=None):
    """
    Optimized rendering of raw alignment blocks using a Sweep-Line algorithm
    and Vectorized PatchCollections. Correctly fills gaps at start/end.
    """
    # 1. Get Parameters
    include_coding = plot_params.get("id_ring_include_coding", False)
    show_noncoding = plot_params.get("id_ring_show_noncoding", True)
    color_only_mode = plot_params.get("id_ring_color_only", False)

    if not include_coding and not show_noncoding:
        return []

    # 2. Select Data Source
    selected_algorithm = plot_params.get("identity_algorithm", "Mauve + SibeliaZ Fallback")

    # Use intergenic data if coding is excluded
    suffix = "intergenic_lcb_data" if (not include_coding and show_noncoding) else "lcb_data"
    mauve_lcbs = data.get("mauve_data", {}).get(suffix, [])
    sibeliaz_lcbs = data.get("sibeliaz_data", {}).get(suffix, [])

    # Filter by algorithm
    raw_segments = []
    if "Mauve" in selected_algorithm or "Combined" in selected_algorithm:
        for lcb in mauve_lcbs: lcb['_algo'] = 'Mauve'; raw_segments.append(lcb)
    if "SibeliaZ" in selected_algorithm or "Combined" in selected_algorithm:
        for lcb in sibeliaz_lcbs: lcb['_algo'] = 'SibeliaZ'; raw_segments.append(lcb)

    R = _estimate_dfv_radius(ax, graphic)
    top_position = getattr(graphic, "top_position", 0)
    seqlen = getattr(graphic, "sequence_length", data["seqlen"])
    r_inner = R * calculated_radius_frac
    thickness = R * calculated_thickness_frac

    # Determine Plot Bounds (1-based inclusive)
    if limit_region:
        g_start = region_start + 1
        g_end = region_end if region_end is not None else seqlen
    else:
        g_start = 1
        g_end = seqlen

    # 3. Pre-process Segments (Split Wraps & Filter)
    processed_segments = []

    for seg in raw_segments:
        start = int(seg['query_start'])
        end = int(seg['query_end'])
        identity = seg.get('identity_incl_gaps', 0.0)

        # Handle 0% identity blocks if they exist in source
        if identity < 0: continue

        # Handle Wrap-around (Split into two linear segments)
        segments_to_add = []
        if start > end:
            segments_to_add.append((start, seqlen, identity, seg))  # Part 1
            segments_to_add.append((1, end, identity, seg))  # Part 2
        else:
            segments_to_add.append((start, end, identity, seg))

        for s, e, ident, original_obj in segments_to_add:
            # Strict filtering against viewing window
            if e < g_start or s > g_end:
                continue

            # Clip coordinates to view window
            s_clipped = max(s, g_start)
            e_clipped = min(e, g_end)

            processed_segments.append({
                'start': s_clipped, 'end': e_clipped, 'identity': ident, 'obj': original_obj
            })

    final_segments = []

    # 4. Sweep-Line Algorithm
    if not processed_segments:
        # Case: No data at all -> Entire region is a gap
        final_segments.append({
            'start': g_start, 'end': g_end, 'identity': 0.0, 'obj': None
        })
    else:
        # Events: (position, type, index). type: 1=Start, -1=End
        events = []
        for i, seg in enumerate(processed_segments):
            events.append((seg['start'], 1, i))
            events.append((seg['end'] + 1, -1, i))  # Exclusive end for sweep

        events.sort()  # Sort by position, then type
        active_indices = set()
        prev_pos = g_start

        for pos, type, idx in events:
            # Ensure we don't go backwards or process before start
            if pos < prev_pos:
                # Just update active set if event happens before our current window
                if type == 1: active_indices.add(idx)
                else: active_indices.discard(idx)
                continue

            if pos > prev_pos:
                interval_end = min(pos - 1, g_end)

                if interval_end >= prev_pos:
                    if active_indices:
                        best_idx = max(active_indices, key=lambda i: processed_segments[i]['identity'])
                        best_seg = processed_segments[best_idx]

                        final_segments.append({
                            'start': prev_pos,
                            'end': interval_end,
                            'identity': best_seg['identity'],
                            'obj': best_seg['obj']
                        })
                    else:
                        final_segments.append({
                            'start': prev_pos,
                            'end': interval_end,
                            'identity': 0.0,
                            'obj': None
                        })

            if type == 1:
                active_indices.add(idx)
            else:
                active_indices.discard(idx)

            prev_pos = pos
            if prev_pos > g_end: break

        if prev_pos <= g_end:
            final_segments.append({
                'start': prev_pos, 'end': g_end,
                'identity': 0.0, 'obj': None
            })

    # 5. Vectorized Drawing
    if not final_segments: return []

    starts = np.array([s['start'] for s in final_segments])
    ends = np.array([s['end'] for s in final_segments])
    identities = np.array([s['identity'] for s in final_segments])

    # Calculate Angles
    norm_starts = (starts - float(top_position)) / seqlen
    norm_ends = (ends - float(top_position)) / seqlen

    theta1 = 90.0 - 360.0 * norm_starts
    theta2 = 90.0 - 360.0 * norm_ends

    # Calculate Colors
    valid_ids = identities[identities > 0]
    if plot_params.get("nc_auto_fit_color", True) and len(valid_ids) > 0:
        vmin, vmax = valid_ids.min(), valid_ids.max()
        if vmin > 0: vmin = 0  # Anchor to 0 if data exists
    else:
        vmin, vmax = 0, 100

    if vmax <= vmin: vmax = 100

    norm = mcolors.Normalize(vmin=vmin, vmax=vmax)
    colors = nc_cmap(norm(identities))

    wedges = []

    if color_only_mode:
        width = thickness
        for t1, t2, c, seg in zip(theta1, theta2, colors, final_segments):
            w = mpatches.Wedge(center=(0, -R), r=r_inner + width, theta1=t2, theta2=t1, width=width)
            wedges.append(w)
            w.lcb_info = _make_lcb_info(seg)
    else:
        heights = np.clip((identities - vmin) / (vmax - vmin), 0.0, 1.0)
        widths = thickness * heights
        r_outer_vals = r_inner + widths

        for t1, t2, w_val, r_val, seg in zip(theta1, theta2, widths, r_outer_vals, final_segments):
            if w_val <= 1e-9:
                continue

            w = mpatches.Wedge(center=(0, -R), r=r_val, theta1=t2, theta2=t1, width=w_val)
            wedges.append(w)
            w.lcb_info = _make_lcb_info(seg)

    # 6. Batch Add using PatchCollection
    if wedges:
        collection = mcoll.PatchCollection(wedges, match_original=False)

        if not color_only_mode:
            mask = (thickness * np.clip((identities - vmin) / (vmax - vmin), 0, 1)) > 1e-9
            final_colors = colors[mask]
        else:
            final_colors = colors

        collection.set_facecolor(final_colors)
        collection.set_edgecolor(final_colors)
        collection.set_linewidth(0.1)
        collection.set_zorder(0.3)
        collection.lcb_info_list = [w.lcb_info for w in wedges]

        ax.add_collection(collection)
        return [collection]

    return []


def _make_lcb_info(seg):
    """Helper to structure metadata for tooltips"""
    obj = seg['obj']
    if obj:
        return {
            'algorithm': obj.get('_algo', 'Unknown'),
            'start': obj['query_start'],
            'end': obj['query_end'],
            'ref_start': obj.get('ref_start'),
            'ref_stop': obj.get('ref_stop'),
            'identity': seg['identity']
        }
    else:
        return {
            'algorithm': 'Gap',
            'start': seg['start'],
            'end': seg['end'],
            'identity': 0.0
        }


def _create_label_anchors(ax, features_to_label, seqlen, graphic, limit_region=False, region_start=0, region_end=None):
    """Generates a list of 'anchor' dictionaries for features that need labels."""
    R = _estimate_dfv_radius(ax, graphic)
    top_position = getattr(graphic, "top_position", 0)
    num_labels = len(features_to_label)
    # Generate colors
    dot_colors = []
    hue = 0.0
    golden_ratio_conjugate = 0.61803398875
    sv_tiers = [(0.9, 0.9), (0.75, 0.8), (0.8, 0.75)]
    for i in range(num_labels):
        hue = (hue + golden_ratio_conjugate) % 1.0
        saturation, value = sv_tiers[i % len(sv_tiers)]
        dot_colors.append(mcolors.hsv_to_rgb([hue, saturation, value]))
    anchors = []
    for i, item in enumerate(features_to_label):
        f = item['feature']
        center_midpoint = ((int(f.start) + int(f.end)) / 2.0) % seqlen
        # Calculate the angle in polar coordinates
        theta_rad = np.deg2rad(90.0 - 360.0 * ((center_midpoint - float(top_position)) / seqlen))
        # Calculate the position in screen coordinates
        xa, ya = R * np.cos(theta_rad), R * np.sin(theta_rad) - R
        # Determine side based on actual screen X position
        side = "right" if xa >= 0 else "left"

        anchor_data = {
            "label": item['label'],
            "x": xa,
            "y": ya,
            "side": side,
            "line_color": item['color'],
            "unique_dot_color": dot_colors[i],
            "feature": f,
            "theta_rad": theta_rad,
            "genomic_pos": center_midpoint
        }

        # Add custom style properties if this is a custom label
        if getattr(f, 'is_custom_label', False):
            anchor_data["font_family"] = getattr(f, 'font_family', 'sans-serif')
            anchor_data["is_bold"] = getattr(f, 'is_bold', False)
            anchor_data["is_italic"] = getattr(f, 'is_italic', False)
            anchor_data["is_thicker"] = getattr(f, 'is_thicker', False)

        anchors.append(anchor_data)
    return anchors


def _draw_label_connector_and_text(ax, anchor, R, **kwargs):
    if 'label_x' not in anchor: return None
    homology_min, homology_max = kwargs['homology_min'], kwargs['homology_max']
    color_min, color_max = kwargs['color_min'], kwargs['color_max']
    cmap = kwargs['cmap']
    fontsize = kwargs.get('label_fontsize', 9)
    show_label_boxes = kwargs.get('show_label_boxes', True)
    label_line_radial_len = kwargs.get('label_line_radial_len', 1.15)
    label_line_horizontal_len = kwargs.get('label_line_horizontal_len', 0.1)
    connect_dots_with_curve = kwargs.get('connect_dots_with_curve', False)
    curve_tension = kwargs.get('curve_tension', 0.5)
    show_connector_dots = kwargs.get('show_connector_dots', False)
    color_lines_by_homology = kwargs.get('color_lines_by_homology', False)
    color_dots_by_homology = kwargs.get('color_dots_by_homology', False)
    limit_region = kwargs.get('limit_region', False)

    p1 = (anchor["x"], anchor["y"])
    p4 = (anchor["label_x"], anchor["label_y"])

    # Calculate P3 (Start of the horizontal line segment near text)
    if anchor['side'] == 'right':
        p3_x = p4[0] - (R * label_line_horizontal_len)
    else:
        p3_x = p4[0] + (R * label_line_horizontal_len)
    p3 = (p3_x, p4[1])

    # Geometry Calculation for P2 (Elbow)
    # Use Radial Projection with Clamping for ALL modes (Full and Limited).
    # Radial projection ensures we respect the calculated 'label_line_radial_len'
    # to clear gene layers. Clamping prevents the S-shape overshoot.

    raw_p2_x = R * label_line_radial_len * np.cos(anchor["theta_rad"])
    raw_p2_y = R * label_line_radial_len * np.sin(anchor["theta_rad"]) - R

    # Clamp Horizontal Position to prevent "S-shape" overshoot/loops.
    # Ensure the radial point P2 never goes horizontally past P3 (the text anchor).
    # This effectively "bends" the radial line vertical if the text is pulled in tight.
    if anchor['side'] == 'right':
        safe_x = min(raw_p2_x, p3[0])
        safe_x = max(safe_x, p1[0])  # Don't go inwards of P1
    else:  # left
        safe_x = max(raw_p2_x, p3[0])
        safe_x = min(safe_x, p1[0])  # Don't go inwards of P1

    p2 = (safe_x, raw_p2_y)

    # Check if this is a custom label
    is_custom_label = getattr(anchor['feature'], 'is_custom_label', False)

    is_thicker = anchor.get('is_thicker', False)
    font_family = anchor.get('font_family', 'sans-serif')
    font_weight = 'bold' if anchor.get('is_bold', False) else 'normal'
    font_style = 'italic' if anchor.get('is_italic', False) else 'normal'

    line_weight = 1.5 if is_thicker else 0.8
    bbox_lw = 1.2 if is_thicker else 0.5

    if is_custom_label:
        # For custom labels, use the user's chosen color (ignore homology coloring)
        custom_color = getattr(anchor['feature'], 'custom_color', anchor['line_color'])
        final_dot_color = custom_color
        final_line_color = custom_color
    else:
        # For GFF labels, apply homology coloring if enabled
        identity = getattr(anchor['feature'], 'best_identity', 0)
        is_homologous = homology_min <= identity <= homology_max
        homology_color = None
        if is_homologous and (color_dots_by_homology or color_lines_by_homology):
            norm_val = (identity - color_min) / (color_max - color_min) if color_max > color_min else 1.0
            homology_color = cmap(norm_val)

        final_dot_color = anchor['unique_dot_color']
        if color_dots_by_homology and homology_color:
            final_dot_color = homology_color
        final_line_color = anchor['unique_dot_color']
        if color_lines_by_homology and homology_color:
            final_line_color = homology_color

        # GFF labels always use default font
        font_family = 'sans-serif'
        font_weight = 'normal'
        font_style = 'normal'

    if connect_dots_with_curve:
        # Dynamic Tension Adjustment
        # If points are very close (due to zoom/limit region), high tension causes loops.
        dist = np.sqrt((p1[0] - p4[0]) ** 2 + (p1[1] - p4[1]) ** 2)

        dynamic_tension = curve_tension
        if limit_region or dist < (R * 0.5):
            dynamic_tension = curve_tension * 0.3

        # Define the control points for the Bézier curve
        control1 = (p2[0] + (p2[0] - p1[0]) * dynamic_tension, p2[1] + (p2[1] - p1[1]) * dynamic_tension)
        control2 = (p3[0] + (p3[0] - p4[0]) * dynamic_tension, p3[1] + (p3[1] - p4[1]) * dynamic_tension)

        # Combine all parts of the line into a single path
        path_data = [
            (mpath.Path.MOVETO, p1),
            (mpath.Path.LINETO, p2),
            (mpath.Path.CURVE4, control1),
            (mpath.Path.CURVE4, control2),
            (mpath.Path.CURVE4, p3),
            (mpath.Path.LINETO, p4)
        ]
        codes, verts = zip(*path_data)
        path = mpath.Path(verts, codes)

        patch = PathPatch(path, facecolor='none', edgecolor=final_line_color, lw=line_weight, alpha=0.7, zorder=0.7,
                          clip_on=False)
        ax.add_patch(patch)
    else:
        ax.plot([p1[0], p2[0], p3[0], p4[0]], [p1[1], p2[1], p3[1], p4[1]], color=final_line_color, lw=line_weight,
                alpha=0.7, zorder=0.7, clip_on=False)

    if show_connector_dots:
        ax.plot(p2[0], p2[1], 'o', markersize=6, color=final_dot_color, markeredgecolor='white', markeredgewidth=0.5,
                zorder=0.8, clip_on=False)
        ax.plot(p3[0], p3[1], 'o', markersize=6, color=final_dot_color, markeredgecolor='white', markeredgewidth=0.5,
                zorder=0.8, clip_on=False)

    ha = 'left' if anchor['side'] == 'right' else 'right'

    text_kwargs = {
        'fontsize': fontsize,
        'va': "center",
        'ha': ha,
        'bbox': None,
        'fontfamily': font_family,
        'weight': font_weight,
        'style': font_style
    }

    if show_label_boxes:
        box_edge_color = anchor["line_color"]
        if color_lines_by_homology and is_homologous: box_edge_color = final_line_color
        text_kwargs['bbox'] = dict(boxstyle="round,pad=0.25", fc="white", alpha=0.8, lw=bbox_lw, ec=box_edge_color)

    # Create and return the text object
    text_obj = ax.text(p4[0], p4[1], anchor["label"], clip_on=False, **text_kwargs)
    return text_obj


def place_labels_smartly(ax, features_to_label, seqlen, graphic, **kwargs):
    """Intelligently places feature labels around the circular plot to minimize overlap.

    This function attempts to arrange labels in two columns (left and right) and
    iteratively adjusts the font size and spacing to prevent labels from
    colliding with each other, especially when the plot is zoomed or in a
    limited region."""
    R = _estimate_dfv_radius(ax, graphic)
    limit_region = kwargs.get('limit_region', False)
    top_position = getattr(graphic, "top_position", 0)

    # Pass limit_region to anchor creation
    anchors = _create_label_anchors(ax, features_to_label, seqlen, graphic,
                                    limit_region=limit_region,
                                    region_start=kwargs.get('region_start', 0),
                                    region_end=kwargs.get('region_end', seqlen))
    if not anchors:
        return kwargs.get('label_fontsize', 9), []
    # Get parameters
    manual_fontsize = kwargs.get('label_fontsize', 9)
    autofit_label_fontsize = kwargs.get('autofit_label_fontsize', False)
    label_spread = kwargs.get('label_spread', 1.0)

    # Apply multiplier for limited regions
    label_distance_factor = kwargs.get("label_distance_factor", 0.25)
    if limit_region: label_distance_factor *= 3.0  # Apply 3x multiplier for limited regions
    cached_zoom = kwargs.get('cached_auto_zoom_scalar', None)
    recalculate = kwargs.get('recalculate_auto_view', True)

    # Get zoom scalar
    if cached_zoom is not None and not recalculate:
        combined_zoom_scalar = cached_zoom
    else:
        xlim_min, xlim_max = ax.get_xlim()
        ylim_min, ylim_max = ax.get_ylim()
        visible_width = xlim_max - xlim_min
        visible_height = ylim_max - ylim_min
        full_plot_diameter = R * 2 * 1.6
        effective_view_dimension = max(visible_width, visible_height, 1e-6)
        combined_zoom_scalar = full_plot_diameter / effective_view_dimension
    new_kwargs = kwargs.copy()

    # Apply zoom boost to spread
    zoom_spread_boost = 1.0 + 0.25 * max(0, combined_zoom_scalar - 1.0)
    effective_label_spread = label_spread * zoom_spread_boost

    # Count labels per side and detect dual-sided configuration
    left_anchors_count = sum(1 for a in anchors if a['side'] == 'left')
    right_anchors_count = sum(1 for a in anchors if a['side'] == 'right')
    n_max_global = max(left_anchors_count, right_anchors_count)  # Use n_max_global for font calc

    # Calculate font size with spacing awareness
    # Force canvas draw if we recalculated the view to ensure ax_bbox is accurate
    if recalculate or limit_region:
        try: ax.figure.canvas.draw()
        except Exception: pass  # Ignore errors during draw

    renderer = ax.figure.canvas.get_renderer()
    ax_bbox = ax.get_window_extent()
    base_available_pixels = ax_bbox.height * 0.95

    if autofit_label_fontsize and n_max_global > 0:
        available_pixels = base_available_pixels * effective_label_spread

        # Calculate font size that ensures labels won't overlap
        padding_factor = 0.35
        max_text_height_px = available_pixels / (n_max_global * (1.0 + padding_factor))
        calculated_fontsize = max_text_height_px * (72 / ax.figure.dpi)

        # Apply zoom scaling
        if combined_zoom_scalar > 1.0:
            if combined_zoom_scalar < 2.0: calculated_fontsize /= np.sqrt(combined_zoom_scalar)
            else: calculated_fontsize /= combined_zoom_scalar

        # When auto-fitting, use a high cap (16pt), NOT the manual slider value.
        final_fontsize = np.clip(calculated_fontsize, 3.0, 16.0)

    else:  # Not auto-fitting, just use the manual slider value
        final_fontsize = manual_fontsize
        # But, scale it down if zoomed in
        if combined_zoom_scalar > 1.0:
            if combined_zoom_scalar < 2.0: final_fontsize = max(manual_fontsize / np.sqrt(combined_zoom_scalar), 6.0)
            else: final_fontsize = max(manual_fontsize / combined_zoom_scalar, 4.0)

    max_iterations, iteration = 4, 0
    overlap_found = True

    while overlap_found and iteration < max_iterations:
        iteration += 1
        new_kwargs['label_fontsize'] = final_fontsize

        # Calculate spacing parameters
        dummy_text = ax.text(0, 0, "Xg", fontsize=final_fontsize, clip_on=False)
        bbox = dummy_text.get_window_extent(renderer=renderer)
        dummy_text.remove()
        inv = ax.transData.inverted()
        label_pixel_height = bbox.height

        min_padding_pixels = label_pixel_height * 0.60
        min_total_step_pixels = label_pixel_height + min_padding_pixels
        min_label_step_data = abs(inv.transform((0, min_total_step_pixels))[1] - inv.transform((0, 0))[1])
        # This is the pure label height in data coordinates, for overlap checking
        label_height_data = abs(inv.transform((0, label_pixel_height))[1] - inv.transform((0, 0))[1])

        # Get view bounds
        ylim_min, ylim_max = ax.get_ylim()
        xlim_min, xlim_max = ax.get_xlim()
        view_height = abs(ylim_max - ylim_min)
        view_center_y = (ylim_min + ylim_max) / 2.0

        # Scale line lengths
        safe_zoom_scalar = max(combined_zoom_scalar, 1e-6)
        max_gene_frac = kwargs.get('max_gene_radius_frac', 1.0)
        base_multiplier = max_gene_frac + 0.02
        user_total_multiplier = kwargs.get('label_line_radial_len', 1.2)

        default_old_base = 1.11
        ext_multiplier = max(0.01, user_total_multiplier - default_old_base)
        scaled_ext_multiplier = ext_multiplier / safe_zoom_scalar
        new_kwargs['label_line_radial_len'] = base_multiplier + scaled_ext_multiplier
        new_kwargs['label_line_horizontal_len'] = kwargs.get('label_line_horizontal_len', 0.1) / safe_zoom_scalar
        new_kwargs['label_distance_factor'] = label_distance_factor / safe_zoom_scalar
        new_kwargs['label_spread'] = effective_label_spread

        # Calculate horizontal offset from PLOT EDGES
        horizontal_offset = R * new_kwargs['label_distance_factor']

        if limit_region:
            all_plot_artists = ax.patches + ax.lines + ax.collections

            # Try to get visual extent
            plot_bbox = None
            if all_plot_artists:
                try:
                    # Filter out artists with 0 width/height (unrendered or empty)
                    plot_bboxes = [a.get_window_extent(renderer=renderer) for a in all_plot_artists]
                    valid_bboxes = [b for b in plot_bboxes if b.width > 0 and b.height > 0]
                    if valid_bboxes:
                        plot_bbox = mtransforms.Bbox.union(valid_bboxes).transformed(inv)
                except Exception:
                    pass  # Fallback to geometric calc if renderer fails

            if plot_bbox:
                plot_left_edge = plot_bbox.p0[0]
                plot_right_edge = plot_bbox.p1[0]
            else:
                # Geometric Fallback
                plot_left_edge = max(xlim_min, -R)
                plot_right_edge = min(xlim_max, R)
        else:
            plot_left_edge = xlim_min
            plot_right_edge = xlim_max

        # Process each side
        max_overlap_this_run = 0.0
        for side in ["left", "right"]:
            anchor_list_side = [a for a in anchors if a['side'] == side]
            if not anchor_list_side:
                continue

            # Detect plot orientation (Per Side)
            min_angle, max_angle, is_near_top, is_near_bottom = _calculate_plot_angular_extent(
                anchor_list_side, top_position, seqlen
            )

            # Calculate label bounds based on actual feature positions (Per Side)
            if limit_region:
                anchor_y_values = [a['y'] for a in anchor_list_side]
                if not anchor_y_values: continue

                min_feature_y = min(anchor_y_values)
                max_feature_y = max(anchor_y_values)
                feature_y_center = (min_feature_y + max_feature_y) / 2.0
                feature_y_span = max_feature_y - min_feature_y
                padding_multiplier = 1.3 + (effective_label_spread - 1.0) * 0.5
                available_height_data = max(feature_y_span * padding_multiplier,
                                            min_label_step_data * (len(anchor_list_side) + 2))
                half_spread_height = available_height_data / 2.0
                view_height = abs(ylim_max - ylim_min)
                feature_span_ratio = feature_y_span / view_height if view_height > 0 else 0

                top_30_threshold = max_feature_y - feature_y_span * 0.3
                bottom_30_threshold = min_feature_y + feature_y_span * 0.3
                top_concentrated = sum(1 for y in anchor_y_values if y >= top_30_threshold)
                bottom_concentrated = sum(1 for y in anchor_y_values if y <= bottom_30_threshold)
                total_features = len(anchor_list_side)
                top_ratio = top_concentrated / total_features if total_features > 0 else 0
                bottom_ratio = bottom_concentrated / total_features if total_features > 0 else 0
                needs_top_bias = (feature_span_ratio > 0.4 and top_ratio > 0.3) or is_near_top
                needs_bottom_bias = (feature_span_ratio > 0.4 and bottom_ratio > 0.3) or is_near_bottom

                if needs_top_bias and needs_bottom_bias:
                    push_amount = R * 0.15 / max(1.0, combined_zoom_scalar * 0.5)
                    outer_top_limit = max_feature_y + half_spread_height * 0.8 + push_amount
                    outer_bottom_limit = min_feature_y - half_spread_height * 0.8 - push_amount
                elif needs_top_bias:
                    push_amount = R * 0.25 / max(1.0, combined_zoom_scalar * 0.5)
                    outer_top_limit = max_feature_y + half_spread_height * 1.2 + push_amount
                    outer_bottom_limit = feature_y_center - half_spread_height
                elif needs_bottom_bias:
                    push_amount = R * 0.25 / max(1.0, combined_zoom_scalar * 0.5)
                    outer_top_limit = feature_y_center + half_spread_height
                    outer_bottom_limit = min_feature_y - half_spread_height * 1.2 - push_amount
                else:
                    outer_top_limit = feature_y_center + half_spread_height
                    outer_bottom_limit = feature_y_center - half_spread_height
            else:
                available_height_data = view_height * 0.95 * effective_label_spread
                half_spread_height = available_height_data / 2.0
                outer_top_limit = view_center_y + half_spread_height
                outer_bottom_limit = view_center_y - half_spread_height

            edge_buffer = abs(inv.transform((0, label_pixel_height * 0.5))[1] - inv.transform((0, 0))[1])
            final_top_bound = outer_top_limit - edge_buffer
            final_bottom_bound = outer_bottom_limit + edge_buffer

            if final_top_bound < final_bottom_bound: final_top_bound = final_bottom_bound = view_center_y

            sorted_anchors_by_feature = sorted(anchor_list_side, key=lambda a: a['y'], reverse=True)
            n = len(sorted_anchors_by_feature)

            if side == "right": text_x_pos = plot_right_edge + horizontal_offset
            else: text_x_pos = plot_left_edge - horizontal_offset

            if n > 1: label_y_positions = np.linspace(final_top_bound, final_bottom_bound, n)
            else: label_y_positions = [(final_top_bound + final_bottom_bound) / 2.0]

            for i, anchor in enumerate(sorted_anchors_by_feature):
                anchor['label_y'] = label_y_positions[i]
                anchor['label_x'] = text_x_pos

            sorted_anchors = sorted(anchor_list_side, key=lambda a: a['label_y'])
            for i in range(len(sorted_anchors) - 1):
                y_lower_center = sorted_anchors[i]['label_y']
                y_upper_center = sorted_anchors[i + 1]['label_y']
                half_label_height_data = label_height_data / 2.0

                top_edge_of_lower_label = y_lower_center + half_label_height_data
                bottom_edge_of_upper_label = y_upper_center - half_label_height_data
                gap_between_labels = bottom_edge_of_upper_label - top_edge_of_lower_label
                min_padding_data = min_label_step_data - label_height_data
                if gap_between_labels < min_padding_data * 0.95:
                    shortfall = min_padding_data - gap_between_labels
                    overlap_percent = max(shortfall, 0) / min_label_step_data
                    max_overlap_this_run = max(max_overlap_this_run, overlap_percent)
                    break

        if max_overlap_this_run > 0.05:  # 5% overlap threshold
            overlap_found = True  # Triggers another iteration
            reduction_factor = 1.0 - (max_overlap_this_run * 0.8)
            final_fontsize = max(3.0, final_fontsize * reduction_factor)
        else:
            overlap_found = False
            if iteration > 1:
                print(f"INFO: Label placement finalized at {final_fontsize:.1f}pt after {iteration} iterations.")

    all_text_objects = []
    for anchor in anchors:
        if 'label_y' in anchor:
            text_obj = _draw_label_connector_and_text(ax, anchor, R, **new_kwargs)
            if text_obj: all_text_objects.append(text_obj)
    return final_fontsize, all_text_objects


def place_labels_vertically(ax, features_to_label, seqlen, graphic, **kwargs):
    """
    Places labels in simple, evenly-spaced vertical columns on either side of the plot.
    """
    R = _estimate_dfv_radius(ax, graphic)
    anchors = _create_label_anchors(ax, features_to_label, seqlen, graphic)
    ylim_min, ylim_max = ax.get_ylim()
    xlim_min, xlim_max = ax.get_xlim()

    # Get base params
    manual_fontsize = kwargs.get('label_fontsize', 9)
    label_spread = kwargs.get('label_spread', 1.0)
    label_distance_factor = kwargs.get("label_distance_factor", 0.25)

    # Use cached auto_zoom_scalar if available, otherwise calculate from current limits
    cached_zoom = kwargs.get('cached_auto_zoom_scalar', None)
    recalculate = kwargs.get('recalculate_auto_view', True)

    if cached_zoom is not None and not recalculate: combined_zoom_scalar = cached_zoom
    else:
        visible_width = xlim_max - xlim_min
        visible_height = ylim_max - ylim_min
        full_plot_diameter = R * 2 * 1.6
        combined_zoom_scalar = full_plot_diameter / max(visible_width, visible_height)

    new_kwargs = kwargs.copy()
    label_spread *= (1.0 + 0.15 * max(0, combined_zoom_scalar - 1.0))
    new_kwargs['label_spread'] = label_spread
    safe_zoom_scalar = max(combined_zoom_scalar, 1e-6)

    # Dynamic base multiplier based on actual gene height
    max_gene_frac = kwargs.get('max_gene_radius_frac', 1.0)
    base_multiplier = max_gene_frac + 0.02

    user_total_multiplier = kwargs.get('label_line_radial_len', 1.15)
    default_old_base = 1.12
    ext_multiplier = max(0.01, user_total_multiplier - default_old_base)
    scaled_ext_multiplier = ext_multiplier / safe_zoom_scalar
    new_kwargs['label_line_radial_len'] = base_multiplier + scaled_ext_multiplier
    new_kwargs['label_line_horizontal_len'] = kwargs.get('label_line_horizontal_len', 0.1) / safe_zoom_scalar
    new_kwargs['label_distance_factor'] = label_distance_factor / safe_zoom_scalar
    final_fontsize = max(manual_fontsize / np.sqrt(combined_zoom_scalar),
                         6.0) if combined_zoom_scalar > 1.0 else manual_fontsize
    new_kwargs['label_fontsize'] = final_fontsize

    left_anchors = sorted([a for a in anchors if a['side'] == 'left'], key=lambda a: a['y'], reverse=True)
    right_anchors = sorted([a for a in anchors if a['side'] == 'right'], key=lambda a: a['y'], reverse=True)

    for side, anchor_list in [("left", left_anchors), ("right", right_anchors)]:
        if not anchor_list: continue
        label_y_positions = np.linspace(ylim_max * 0.95 * label_spread, ylim_min * 0.95 * label_spread,
                                        len(anchor_list))
        horizontal_offset = R * new_kwargs['label_distance_factor']
        text_x_pos = xlim_max + horizontal_offset if side == 'right' else xlim_min - horizontal_offset

        for anchor, y_pos in zip(anchor_list, label_y_positions):
            anchor['label_x'] = text_x_pos
            anchor['label_y'] = y_pos

    all_text_objects = []
    for anchor in anchors:
        text_obj = _draw_label_connector_and_text(ax, anchor, R, **new_kwargs)
        if text_obj:
            all_text_objects.append(text_obj)

    return final_fontsize, all_text_objects


def _calculate_plot_angular_extent(anchors, top_position, seqlen):
    """
    Calculates the angular extent of features and determines if they're near 12 or 6 o'clock.
    Returns: (min_angle, max_angle, is_near_top, is_near_bottom)
    """
    if not anchors: return 0, 0, False, False

    angles = []
    for anchor in anchors:
        genomic_pos = anchor.get('genomic_pos', 0)
        angle_deg = (90.0 - 360.0 * ((genomic_pos - float(top_position)) / seqlen)) % 360
        angles.append(angle_deg)

    min_angle, max_angle = min(angles), max(angles)
    is_near_top = any(70 <= angle <= 110 for angle in angles)
    is_near_bottom = any(250 <= angle <= 290 for angle in angles)

    return min_angle, max_angle, is_near_top, is_near_bottom


def _draw_custom_label_markers(ax, custom_labels_list, seqlen, graphic, R, limit_region=False, region_start=0,
                               region_end=None):
    """
    Manually draws wedges for custom labels instead of using GraphicFeature.
    This allows precise control over visibility and centering.

    NOTE: Coordinates are expected to be 1-indexed (inclusive endpoints).
    """
    top_position = getattr(graphic, "top_position", 0)

    for label_entry in custom_labels_list:
        try:
            positions = label_entry.get("positions", [])
            if not positions: continue

            start_pos, end_pos = positions[0]
            color = label_entry.get("color", "black")
            label_name = label_entry.get("name", "unnamed")

            print(f"\n[DEBUG CUSTOM LABEL] Processing: {label_name}")
            print(f"  Start pos: {start_pos}, End pos: {end_pos}")
            print(f"  Sequence length: {seqlen}")
            print(f"  Top position: {top_position}")

            if start_pos <= 0 or end_pos <= 0: continue
            if limit_region and (end_pos < region_start or start_pos > region_end): continue

            is_point = (end_pos == start_pos + 1)

            if is_point:
                mid_point = start_pos + 0.5
                mid_angle_deg = 90.0 - 360.0 * ((mid_point - float(top_position)) / seqlen)
                visual_arc_width_deg = 2.0
                theta1 = mid_angle_deg + (visual_arc_width_deg / 2.0)
                theta2 = mid_angle_deg - (visual_arc_width_deg / 2.0)
                wedge_thickness = 0.05 * R
                print(f"  Point label - theta1: {theta1:.2f}°, theta2: {theta2:.2f}°")
            else:
                # Calculate angles for start and end positions
                theta_start = 90.0 - 360.0 * ((start_pos - float(top_position)) / seqlen)
                theta_end = 90.0 - 360.0 * ((end_pos - float(top_position)) / seqlen)
                print(f"  Range label - initial theta_start: {theta_start:.2f}°, theta_end: {theta_end:.2f}°")

                # Calculate the angular span
                angular_span = abs(theta_start - theta_end)
                print(f"  Angular span: {angular_span:.2f}°")

                # If the span is > 180 degrees, we're wrapping around - fix it
                if angular_span > 180:
                    print(f"  WARNING: Angular span > 180°, correcting...")
                    # Swap to use the shorter arc
                    if theta_start > theta_end: theta_end += 360
                    else: theta_start += 360
                    print(f"  Corrected theta_start: {theta_start:.2f}°, theta_end: {theta_end:.2f}°")

                theta1 = min(theta_start, theta_end)
                theta2 = max(theta_start, theta_end)
                print(f"  Final theta1: {theta1:.2f}°, theta2: {theta2:.2f}°")
                print(f"  Final span: {abs(theta2 - theta1):.2f}°")
                wedge_thickness = 0.1 * R

            wedge_radius = 1.02 * R

            # Matplotlib Wedge goes counterclockwise from theta1 to theta2
            # We need theta1 < theta2 to draw the short arc
            # Our coordinate system has angles decreasing as we go around (90° at top, decreasing clockwise)
            # So we need to use the MINIMUM as theta1 and MAXIMUM as theta2
            actual_theta1 = min(theta1, theta2)
            actual_theta2 = max(theta1, theta2)

            print(f"  Using theta1={actual_theta1:.2f}°, theta2={actual_theta2:.2f}° for wedge")

            wedge = mpatches.Wedge(
                center=(0, -R),
                r=wedge_radius,
                theta1=actual_theta1,
                theta2=actual_theta2,
                width=wedge_thickness,
                facecolor=color,
                edgecolor=color,
                alpha=0.4,
                zorder=0.2
            )
            ax.add_patch(wedge)

        except Exception as e:
            print(f"Warning: Could not draw custom marker for '{label_entry.get('name')}': {e}")


# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# Report & Export Functions
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────

def _format_blast_df_for_excel(blast_df):
    """
    Formats the enriched BLAST DataFrame to match the LCB sheet structure.
    Optimized using NumPy vectorization instead of row-wise apply.
    """
    if blast_df.empty: return pd.DataFrame()

    df = blast_df.copy()

    if 'sstart' in df.columns and 'send' in df.columns:
        sstart = df['sstart'].values
        send = df['send'].values
        df['ref_start'] = np.minimum(sstart, send)
        df['ref_stop'] = np.maximum(sstart, send)

    df = df.rename(columns={
        "qstart": "Query Start",
        "qend": "Query End",
        "qseq": "Aligned Query Sequence",
        "sseq": "Aligned Reference Sequence",
        "query_strand": "Query Strand",
        "ref_strand": "Ref Strand",
        "identity_excl_gaps": "Identity (%) Excl Gaps",
        "identity_incl_gaps": "Identity (%) Incl Gaps",
        "difference_string": "Difference String",
        "query_mismatches_marked": "Query with Mismatches (#)",
        "query_differences_only": "Query Differences Only",
        "ref_differences_only": "Reference Differences Only",
    })

    if 'Aligned Query Sequence' in df.columns:
        df['Original Query Sequence'] = df['Aligned Query Sequence'].astype(str).str.replace('-', '', regex=False)
    if 'Aligned Reference Sequence' in df.columns:
        df['Original Reference Sequence'] = df['Aligned Reference Sequence'].astype(str).str.replace('-', '',
                                                                                                     regex=False)

    final_cols = [
        "Query Start", "Query End", "Query Strand",
        "Ref Start", "Ref Stop", "Ref Strand",
        "Identity (%) Excl Gaps", "Identity (%) Incl Gaps",
        "Original Query Sequence", "Original Reference Sequence",
        "Aligned Query Sequence", "Aligned Reference Sequence",
        "Difference String", "Query with Mismatches (#)",
        "Query Differences Only", "Reference Differences Only",
        "sseqid", "evalue", "bitscore"
    ]

    existing_cols = [col for col in final_cols if col in df.columns]
    return df[existing_cols].sort_values(by="Query Start").reset_index(drop=True)


def _adjust_excel_column_widths(filepath):
    """
    Adjust Excel column widths after file creation and format number columns.
    - Homology Report (Genes): Auto-fit all columns to fit all data
    - All other sheets: Auto-fit non-sequence columns to all data,
                        sequence columns to headers only
    - Identity columns: Format to show up to 12 decimal places
    """

    # Sequence column keywords to identify columns that should fit headers only
    sequence_keywords = [
        "Query Sequence", "Reference Sequence", "Aligned", "Difference String",
        "Mismatches", "Differences Only"
    ]

    # Identity column keywords for number formatting
    identity_keywords = ["Identity"]

    def is_sequence_column(col_name):
        """Check if column contains sequence data"""
        return any(keyword in col_name for keyword in sequence_keywords)

    def is_identity_column(col_name):
        """Check if column contains identity values that should be formatted"""
        return any(keyword in col_name for keyword in identity_keywords)

    try:
        wb = load_workbook(filepath)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Determine if this is the Homology Report sheet
            is_homology_sheet = "Homology Report" in sheet_name

            # Get all columns
            for col_idx, col in enumerate(ws.columns, 1):
                col_letter = col[0].column_letter

                # Get header (first row)
                header = ws.cell(row=1, column=col_idx).value
                if header is None:
                    continue

                header_str = str(header)

                # Format identity columns with up to 12 decimal places, no trailing zeros or decimal point
                if is_identity_column(header_str):
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        val = cell.value
                        # Check if value is a whole number or has decimals
                        if val is not None and isinstance(val, (int, float)):
                            if val == int(val):
                                # Whole number - no decimals
                                cell.number_format = '0'
                            else:
                                # Has decimals - show up to 12 decimal places without trailing zeros
                                cell.number_format = '0.############'

                # Determine if we should fit headers only or all data
                fit_headers_only = False
                if not is_homology_sheet and is_sequence_column(header_str):
                    fit_headers_only = True

                # Calculate width
                if fit_headers_only:
                    width = len(header_str)
                else:
                    # Get all values in column (excluding header)
                    col_values = [ws.cell(row=row_idx, column=col_idx).value
                                 for row_idx in range(2, ws.max_row + 1)]
                    max_length = len(header_str)
                    for val in col_values:
                        val_len = len(str(val)) if val is not None else 0
                        max_length = max(max_length, val_len)
                    width = min(max_length, 50)

                # Set minimum width
                width = max(width, 8)

                # Apply width
                ws.column_dimensions[col_letter].width = width

        # Save the workbook
        wb.save(filepath)
        print(f"[EXCEL] Column widths and number formatting adjusted for: {filepath}")

    except Exception as e:
        print(f"[WARNING] Could not adjust Excel column widths: {e}")
        traceback.print_exc()


def _save_figure_with_aspect_ratio(fig, ax, filepath, max_width, max_height):
    renderer = fig.canvas.get_renderer()
    tight_bbox_px = ax.get_tightbbox(renderer)
    px_to_inches = fig.dpi_scale_trans.inverted()
    tight_bbox_inches = tight_bbox_px.transformed(px_to_inches)
    content_w_inches, content_h_inches = tight_bbox_inches.width, tight_bbox_inches.height

    if content_w_inches <= 0 or content_h_inches <= 0: raise ValueError("Plot content has zero or negative size.")

    content_aspect_ratio = content_w_inches / content_h_inches
    target_w = max_width
    target_h = int(target_w / content_aspect_ratio)

    if target_h > max_height:
        target_h = max_height
        target_w = int(target_h * content_aspect_ratio)

    dpi = target_w / content_w_inches
    pad_in_inches = 8 / dpi
    fig.savefig(filepath, dpi=dpi, bbox_inches='tight', pad_inches=pad_in_inches)
    return target_w, target_h


# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# Threading & Concurrency
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────

def _run_background_processing(data_params, plot_params, out_png, log_queue=None):
    def status(msg, color="black"):
        if log_queue: log_queue.put(("status", (msg, color)))

    try:
        data = _process_data(**data_params, log_queue=log_queue)
        payload = {"data": data, "plot_params": plot_params, "out_png": out_png}
        if log_queue: log_queue.put(("plot_and_save", payload))
    except Exception as e:
        if "cancelled by user" in str(e):
            status("Processing cancelled by user.", "blue")
            if log_queue: log_queue.put(("done", None))
            return

        status(f"Error during data processing: {e}", "red")
        traceback.print_exc()
        if log_queue: log_queue.put(("done", None))


class StreamLogger:
    """
    A simple logger class that captures a stream (like stdout or stderr)
    to an in-memory buffer (StringIO) while also passing the output to
    the original stream so it still appears in the console.
    """

    def __init__(self, original_stream):
        self.original_stream = original_stream
        self.log_buffer = io.StringIO()

    def write(self, text):
        try: self.original_stream.write(text)
        except Exception: pass
        self.log_buffer.write(text)

    def flush(self):
        try: self.original_stream.flush()
        except Exception: pass

    def get_value(self):
        return self.log_buffer.getvalue()

    def fileno(self):
        try: return self.original_stream.fileno()
        except Exception: return -1


# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# GUI: Core Components
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────

class CollapsibleFrame(ttk.Frame):
    """A collapsible frame widget for tkinter."""

    def __init__(self, parent, text="", expanded=False, style_prefix=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.columnconfigure(0, weight=1)
        self.text = text
        self._expanded = tk.BooleanVar(value=expanded)
        self.header = ttk.Frame(self)
        self.header.grid(row=0, column=0, sticky='ew')
        header_style = f"{style_prefix}.Header.TLabel" if style_prefix else "Header.TLabel"
        self.toggle_button = ttk.Label(self.header, text=f"▶ {self.text}", style=header_style, cursor="hand2")
        self.toggle_button.pack(side="left", fill="x", expand=True, pady=2)
        self.toggle_button.bind("<Button-1>", self._toggle)
        self.content_frame = ttk.Frame(self, padding=(10, 5))
        if expanded: self._expand()
        else: self._collapse()

    @property
    def interior(self):
        return self.content_frame

    def _toggle(self, event):
        if self._expanded.get(): self._collapse()
        else: self._expand()

    def _expand(self):
        self.toggle_button.configure(text=f"▼ {self.text}")
        self.content_frame.grid(row=1, column=0, sticky='nsew')
        self._expanded.set(True)

    def _collapse(self):
        self.toggle_button.configure(text=f"▶ {self.text}")
        self.content_frame.grid_forget()
        self._expanded.set(False)


def _create_tooltip(widget, text):
    """Create a simple hover tooltip for tkinter widgets."""
    if not text:
        return

    tip_window = [None]

    def on_enter(event=None):
        if tip_window[0]:
            return
        x, y = widget.winfo_rootx() + 10, widget.winfo_rooty() + 25
        tip_window[0] = tw = tk.Toplevel(widget)
        tw.wm_overrideredirect(True)

        if sys.platform == "darwin":
            # Make sure it appears on top and is opaque enough to render text
            tw.wm_attributes("-topmost", True)
            tw.wm_attributes("-alpha", 1.0)
            # macOS sometimes needs this for the window to display text
            tw.wm_attributes("-transparent", False)
        elif sys.platform.startswith("linux"):
            tw.attributes("-topmost", True)
        else:
            tw.wm_attributes("-topmost", True)

        label = tk.Label(tw, text=text, background="#ffffe0", foreground="#000000", relief="solid", borderwidth=1,
                         font=("TkDefaultFont", 10, "normal"), wraplength=300, justify="left", padx=4, pady=2)
        label.pack()
        tw.update_idletasks()
        tw.geometry(f"+{x}+{y}")

        # Ensure it becomes visible
        if sys.platform == "darwin":
            tw.lift()
            tw.update_idletasks()
            tw.update()

    def on_leave(event=None):
        if tip_window[0]:
            tip_window[0].destroy()
            tip_window[0] = None

    widget.bind("<Enter>", on_enter)
    widget.bind("<Leave>", on_leave)


class Tooltip:
    """Simple tooltip manager for Matplotlib events."""

    def __init__(self, canvas):
        self.canvas = canvas
        self.widget = canvas.get_tk_widget()
        self.tip_window = None
        self.id = None

    def show_tip(self, text):
        """Display text in tooltip window (cross-platform)."""
        if self.tip_window or not text: return

        x, y = self.widget.winfo_pointerx() + 15, self.widget.winfo_pointery() + 15
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)

        # Platform-specific attributes Before adding content
        if sys.platform == "darwin":  # macOS
            # Make sure it appears on top and is opaque enough to render text
            tw.wm_attributes("-topmost", True)
            tw.wm_attributes("-alpha", 1.0)
            # macOS sometimes needs this for the window to display text
            tw.wm_attributes("-transparent", False)
        elif sys.platform.startswith("linux"):
            tw.attributes("-topmost", True)
        else:
            tw.wm_attributes("-topmost", True)

        # Add the label
        label = tk.Label(
            tw,
            text=text,
            justify="left",
            background="#ffffe0",
            foreground="#000000",
            relief="solid",
            borderwidth=1,
            font=("TkDefaultFont", 10, "normal"),
            padx=4, pady=2,
        )
        label.pack()

        tw.update_idletasks()
        tw.geometry(f"+{x}+{y}")

        # Ensure it becomes visible
        if sys.platform == "darwin":
            tw.lift()
            tw.update_idletasks()
            tw.update()

    def hide_tip(self):
        tw = self.tip_window
        self.tip_window = None
        if tw: tw.destroy()


# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────
# GUI: Core Components
# ──────────────────────────────────────────────────────────────────────────────────────────────────────────────────

class PlottingWindow(tk.Toplevel):
    def __init__(self, master, data):
        super().__init__(master)

        # Create a unique style prefix for this window
        self.style_prefix = f"PlotSidebar.{id(self)}"
        style = ttk.Style(self)
        style.configure(f"{self.style_prefix}.Header.TLabel", font=("Helvetica", 10, "bold"))

        self.MAX_FONT_SIZE = 10  # The default/max font size
        self.MIN_FONT_SIZE = 8  # The minimum allowable font size
        self.current_sidebar_font_size = self.MAX_FONT_SIZE
        self.sidebar_resize_timer = None
        self.sidebar_is_expanded = True
        self.last_sidebar_width = 400
        self.title("CGCT Interactive Plot Viewer")
        self.state('zoomed')
        self.master_app = master
        self.master_app.current_plot_window = self
        self.data = data
        self.coding_intervals_for_hover = sorted([(f.start, f.end) for f in self.data.get('gene_feats', [])])
        self.last_rotation = 0.0
        self.resolution_options = RESOLUTION_OPTIONS
        self.data_params = self.master_app._get_data_params()
        self.plot_tooltips = {}
        if hasattr(master, 'ui_tooltips'): self.plot_tooltips.update(master.ui_tooltips)

        self.plot_params = {
            "autorotate_roi": tk.BooleanVar(value=True),
            "manual_rot_deg": tk.DoubleVar(value=0.0),
            "gc_inner": tk.DoubleVar(value=0.3),
            "gc_thick": tk.DoubleVar(value=1.05),
            "skew_thick": tk.DoubleVar(value=0.505),
            "skew_inner_radius": tk.DoubleVar(value=0.5),
            "show_roi_labels": tk.BooleanVar(value=True),
            "show_homology_labels": tk.BooleanVar(value=True),
            "homology_thresh": tk.DoubleVar(value=70.0),
            "homology_thresh_max": tk.DoubleVar(value=100.0),
            "autofit_color_range": tk.BooleanVar(value=True),
            "zoom": tk.DoubleVar(value=1.0),
            "show_label_boxes": tk.BooleanVar(value=True),
            "label_spread": tk.DoubleVar(value=1.0),
            "label_fontsize": tk.DoubleVar(value=9.0),
            "color_lines_by_homology": tk.BooleanVar(value=False),
            "color_dots_by_homology": tk.BooleanVar(value=False),
            "label_line_radial_len": tk.DoubleVar(value=1.15),
            "label_line_horizontal_len": tk.DoubleVar(value=0.1),
            "label_distance_factor": tk.DoubleVar(value=0.25),
            "connect_dots_with_curve": tk.BooleanVar(value=True),
            "curve_tension": tk.DoubleVar(value=1.88),
            "show_connector_dots": tk.BooleanVar(value=False),
            "use_smart_layout": tk.BooleanVar(value=True),
            "show_blast_separators": tk.BooleanVar(value=False),
            "homology_colormap": tk.StringVar(value="Plasma"),
            "save_resolution": tk.StringVar(value="4K UHD (3840x2160)"),
            "custom_save_width": tk.StringVar(value="3840"),
            "custom_save_height": tk.StringVar(value="2160"),
            "nc_colormap": tk.StringVar(value="Viridis"),
            "id_ring_include_coding": tk.BooleanVar(value=False),
            "id_ring_radius": tk.DoubleVar(value=0.7),
            "id_ring_thickness": tk.DoubleVar(value=0.20),
            "id_ring_precise_mode": tk.BooleanVar(value=data.get("id_ring_precise_mode", False)),
            "blast_ring_thickness": tk.DoubleVar(value=0.15),
            "blast_ring_radius": tk.DoubleVar(value=1.0),
            "id_ring_color_only": tk.BooleanVar(value=True),
            "id_ring_show_noncoding": tk.BooleanVar(value=True),
            "autofit_label_fontsize": tk.BooleanVar(value=True),
            "show_start_end_line": tk.BooleanVar(value=False),
            "nc_auto_fit_color": tk.BooleanVar(value=True),
            "auto_thickness_rings": tk.BooleanVar(value=True),
            "blast_task": tk.StringVar(value=self.data_params.get("blast_task", "blastn")),
            "blast_word_size": tk.StringVar(value=self.data_params.get("blast_word_size", "11")),
            "blast_reward": tk.StringVar(value=self.data_params.get("blast_reward", "2")),
            "blast_penalty": tk.StringVar(value=self.data_params.get("blast_penalty", "-3")),
            "blast_evalue": tk.StringVar(value=self.data_params.get("blast_evalue", "10")),
            "blast_num_threads": tk.StringVar(value=self.data_params.get("blast_num_threads", _get_cpu_count())),
            "show_blast_coding": tk.BooleanVar(value=True),
            "show_blast_noncoding": tk.BooleanVar(value=False),
            "nc_window_size": tk.IntVar(value=self.data_params.get("nc_window_size", 500)),
            "nc_step_size": tk.IntVar(value=self.data_params.get("nc_step_size", 50)),
            "id_baseline_transparency_pct": tk.DoubleVar(value=100.0),  # 100% when color-only enabled by default
            "gc_baseline_transparency_pct": tk.DoubleVar(value=50.0),
            "skew_baseline_transparency_pct": tk.DoubleVar(value=50.0),
            "gc_colormap": tk.StringVar(value="Grey"),
            "gc_ring_color_only": tk.BooleanVar(value=False),
            "gc_window_size": tk.IntVar(value=self.data_params.get("window", 500)),
            "limit_to_region": tk.BooleanVar(value=False),
            "region_start": tk.StringVar(value="1"),
            "region_end": tk.StringVar(value=str(data.get("seqlen", ""))),
            "identity_algorithm": tk.StringVar(
                value=self.data_params.get("identity_algorithm", "Mauve + SibeliaZ Fallback")),
            "remove_gene_borders": tk.BooleanVar(value=False),
            "run_legacy_report_on_export": tk.BooleanVar(value=False),
            "show_coords": tk.BooleanVar(value=True),
            "subsequence_start": tk.StringVar(value="1"),
            "subsequence_end": tk.StringVar(value=str(data.get("seqlen", ""))),
            "label_font_family": tk.StringVar(value="sans-serif"),
            "label_font_bold": tk.BooleanVar(value=False),
            "label_font_italic": tk.BooleanVar(value=False),
            "label_thicker_lines": tk.BooleanVar(value=False),
        }

        # Apply initial plot parameters from main window
        initial_params = data.get("initial_plot_params", {})
        self._apply_initial_params(initial_params)

        # Widget Layout and Scrollbar Setup
        main_frame = tk.Frame(self)
        main_frame.pack(fill="both", expand=True)
        m_paned_window = ttk.PanedWindow(main_frame, orient="horizontal")
        m_paned_window.pack(fill="both", expand=True)
        self.m_paned_window = m_paned_window
        self.controls_container = tk.Frame(m_paned_window, padx=10, pady=10)
        m_paned_window.add(self.controls_container, weight=1)
        self.controls_container.grid_rowconfigure(0, weight=1)
        self.controls_container.grid_columnconfigure(0, weight=1)
        self.controls_container.grid_columnconfigure(1, weight=0)
        self.controls_canvas = tk.Canvas(self.controls_container, highlightthickness=0)
        self.controls_scrollbar = ttk.Scrollbar(self.controls_container, orient="vertical",
                                                command=self.controls_canvas.yview)
        self.scrollable_frame = ttk.Frame(self.controls_canvas)
        self.scrollable_frame_window = self.controls_canvas.create_window((0, 0), window=self.scrollable_frame,
                                                                          anchor="nw")
        self.controls_canvas.configure(yscrollcommand=self.controls_scrollbar.set)
        self.scrollbar_is_visible = False  # To track the scrollbar's state
        self.scrollable_frame.bind("<Configure>", self._on_frame_configure)
        self.controls_canvas.bind("<Configure>", self._on_canvas_configure)
        self.controls_canvas.grid(row=0, column=0, sticky='nsew')
        plot_frame = tk.Frame(m_paned_window)
        m_paned_window.add(plot_frame, weight=3)
        self.fig = plt.Figure(figsize=(10, 8), dpi=100)
        self.ax = self.fig.add_subplot(111)
        plot_frame.grid_rowconfigure(0, weight=1)
        plot_frame.grid_columnconfigure(0, weight=1)
        self.canvas = FigureCanvasTkAgg(self.fig, master=plot_frame)
        self.canvas.get_tk_widget().grid(row=0, column=0, sticky='nsew')
        toolbar = SafeNavigationToolbar2Tk(self.canvas, plot_frame, pack_toolbar=False)
        toolbar.grid(row=1, column=0, sticky='ew')

        self.toggle_sidebar_button = tk.Button(
            plot_frame,
            text="«",
            command=self._toggle_sidebar,
            width=2,
            relief="raised",
            borderwidth=1
        )
        self.toggle_sidebar_button.place(x=5, y=5)
        self.fig.canvas.mpl_connect('motion_notify_event', self.on_plot_hover)
        self.tooltip = Tooltip(self.canvas)
        self.gene_feature_patches = []
        self.identity_ring_artists = []
        self.scores_plotted_for_hover = []
        self.graphic_ref = None
        self.last_calculated_values = {}
        self.log_queue = queue.Queue()
        self.initial_load_complete = False
        self.user_set_id_thickness = None
        self.sidebar_is_expanded = True
        self.last_sidebar_width = 400
        self.COLLAPSED_WIDTH = 30
        self.MAX_FONT_SIZE = 11
        self.MIN_FONT_SIZE = 7
        self.current_sidebar_font_size = self.MAX_FONT_SIZE
        self.sidebar_resize_timer = None
        self.initial_gc_thick = None
        self.resize_timer = None
        self.id_radius_slider = None
        self.id_thickness_slider = None
        self.gc_thickness_slider = None
        self.skew_thickness_slider = None
        self.gc_inner_radius_slider = None
        self.skew_inner_radius_slider = None
        self._add_controls(self.scrollable_frame)
        if 'custom_labels' not in self.data: self.data['custom_labels'] = []
        self.last_region_limited = False
        self.last_region_start = ""
        self.last_region_end = ""
        self.cached_x_center_pan = 0.0
        self.cached_y_center_pan = -1.0
        self.cached_xlim = None
        self.cached_ylim = None
        self.cached_auto_zoom_scalar = 1.0
        self.export_button = None
        self.protocol("WM_DELETE_WINDOW", self._on_closing)
        self.fig.canvas.mpl_connect('resize_event', self._on_resize)

        def bind_recursive(widget):
            widget.bind('<MouseWheel>', self._on_mousewheel)
            widget.bind('<Button-4>', self._on_mousewheel)
            widget.bind('<Button-5>', self._on_mousewheel)
            for child in widget.winfo_children():
                bind_recursive(child)

        bind_recursive(self.scrollable_frame)

        # Trigger the initial plot with applied parameters
        self.after(100, self.update_plot)

    def on_plot_hover(self, event: MouseEvent):
        """Handle mouse hover events over the plot."""
        # Hide previous tooltip if it exists
        if hasattr(self, 'tooltip') and self.tooltip and self.tooltip.tip_window:
            self.tooltip.hide_tip()

        # Check if mouse is inside the axes and we have plot data
        if event.inaxes != self.ax or not hasattr(self, 'graphic_ref') or self.graphic_ref is None: return

        tooltip_text = None

        if hasattr(self, 'gene_feature_patches'):
            candidate_patches = []

            for patch_data in self.gene_feature_patches:
                patch = patch_data[0]
                start = patch_data[1]
                end = patch_data[2]
                feature_obj = patch_data[3] if len(patch_data) > 3 else None

                # Make sure patch is still valid
                if patch.axes is None: continue

                contains, _ = patch.contains(event)
                if contains: candidate_patches.append((patch, start, end, feature_obj))

            # If we have candidates, select the best match
            if candidate_patches:
                best_match = None

                if len(candidate_patches) == 1:
                    # Only one candidate - use it
                    best_match = candidate_patches[0]
                else:
                    # Multiple overlapping patches - need to disambiguate
                    # Strategy: Match based on which patch's angular extent matches its genomic coordinates
                    seqlen = self.data["seqlen"]
                    plot_top_position = getattr(self.graphic_ref, "top_position", 0)
                    best_score = -1

                    for patch, start, end, feature_obj in candidate_patches:
                        # Calculate expected angular extent for this feature's genomic coordinates
                        expected_theta1 = 90.0 - 360.0 * ((start - float(plot_top_position)) / seqlen)
                        expected_theta2 = 90.0 - 360.0 * ((end - float(plot_top_position)) / seqlen)
                        expected_angular_size = abs(expected_theta2 - expected_theta1)

                        # Get actual angular extent of the patch
                        if hasattr(patch, 'theta1') and hasattr(patch, 'theta2'):
                            actual_theta1 = patch.theta1
                            actual_theta2 = patch.theta2
                            actual_angular_size = abs(actual_theta2 - actual_theta1)
                            size_diff = abs(actual_angular_size - expected_angular_size)

                            # Also consider if the angles themselves match
                            angle1_diff = min(abs(actual_theta1 - expected_theta1),
                                              abs(actual_theta1 - expected_theta2))
                            angle2_diff = min(abs(actual_theta2 - expected_theta1),
                                              abs(actual_theta2 - expected_theta2))

                            # Combined score (lower is better, so negate for "best")
                            score = -(size_diff + angle1_diff + angle2_diff)

                            if score > best_score:
                                best_score = score
                                best_match = (patch, start, end, feature_obj)
                        else:
                            coord_range = end - start
                            score = -coord_range
                            if score > best_score:
                                best_score = score
                                best_match = (patch, start, end, feature_obj)

                # Generate tooltip for the best match
                if best_match:
                    patch, start, end, feature_obj = best_match

                    # Use the stored feature object directly for accurate data
                    if feature_obj:
                        feature_identity = getattr(feature_obj, 'best_identity', None)
                        tooltip_text = _generate_contextual_tooltip(feature_obj, start, end, feature_identity)
                    else:
                        # Fallback: try to find by coordinates
                        for f in self.data.get('gene_feats', []):
                            if f.start == start and f.end == end:
                                feature_identity = getattr(f, 'best_identity', None)
                                tooltip_text = _generate_contextual_tooltip(f, start, end, feature_identity)
                                break

            # Fallback: If no match found in stored patches, directly check all axis patches
            if tooltip_text is None:
                for patch in self.ax.patches:
                    # Skip patches that don't have our custom attributes
                    if not hasattr(patch, 'axes') or patch.axes is None: continue
                    if hasattr(patch, 'blast_info') or hasattr(patch, 'lcb_info'): continue

                    contains, _ = patch.contains(event)
                    if contains:
                        if hasattr(patch, 'theta1') and hasattr(patch, 'theta2'):
                            # Find closest feature by angle
                            theta_mid = (patch.theta1 + patch.theta2) / 2.0
                            plot_top_position = getattr(self.graphic_ref, "top_position", 0)
                            seqlen = self.data["seqlen"]
                            plot_angle_deg = (90.0 - theta_mid) % 360.0
                            bp_pos_float = ((plot_angle_deg / 360.0) * seqlen + plot_top_position) % seqlen
                            bp_pos = int(round(bp_pos_float))

                            # Find closest feature
                            closest_feature = None
                            min_dist = float('inf')
                            for f in self.data.get('gene_feats', []):
                                feat_center = (f.start + f.end) / 2.0
                                dist = min(abs(bp_pos - feat_center), seqlen - abs(bp_pos - feat_center))
                                if dist < min_dist:
                                    min_dist = dist
                                    closest_feature = f

                            if closest_feature and min_dist < 500:
                                feature_identity = getattr(closest_feature, 'best_identity', None)
                                tooltip_text = _generate_contextual_tooltip(closest_feature, closest_feature.start,
                                                                            closest_feature.end, feature_identity)
                                break

        if tooltip_text is None:
            for patch in self.ax.patches:
                # Check if this patch is a BLAST hit
                is_inside, _ = patch.contains(event)
                if is_inside and hasattr(patch, 'blast_info'):
                    info = patch.blast_info

                    overlapping_gene_label = "N/A"
                    hit_start, hit_end = min(info.qstart, info.qend), max(info.qstart, info.qend)
                    allow_fallback = self.data.get("allow_code_fallback", False)
                    for feature in self.data['gene_feats']:
                        if feature.start < hit_end and feature.end > hit_start:
                            label = _label_for_feature(feature, allow_code_fallback=allow_fallback)
                            if label:
                                overlapping_gene_label = label
                                break

                    # Build the tooltip text
                    tooltip_text = (
                        f"BLAST Hit\n"
                        f"Identity: {info.pident:.2f}%\n"
                        f"Ref: {info.sseqid}\n"
                        f"Qry Coords: {info.qstart} - {info.qend}\n"
                        f"Ref Coords: {info.sstart} - {info.send}\n"
                        f"Overlaps: {overlapping_gene_label}"
                    )
                    break  # Found the topmost BLAST hit, stop iterating

        # Check Identity Ring (Precise or Sliding Window)
        if tooltip_text is None:

            if self.plot_params["id_ring_precise_mode"].get():
                for artist in reversed(self.identity_ring_artists):
                    if artist.axes is None: continue

                    is_inside, info_dict = artist.contains(event)
                    if is_inside:
                        if isinstance(artist, matplotlib.collections.PatchCollection):
                            if 'ind' in info_dict and len(info_dict['ind']) > 0:
                                # Get the top-most (last) index
                                idx = info_dict['ind'][-1]
                                if hasattr(artist, 'lcb_info_list'): info = artist.lcb_info_list[idx]
                                else: continue
                        else:
                            # Fallback for individual patches (old method)
                            if hasattr(artist, 'lcb_info'): info = artist.lcb_info
                            else: continue

                        # Build tooltip
                        tooltip_parts = [
                            f"Algorithm: {info['algorithm']}",
                            f"Qry Coords: {info['start']} - {info['end']}"
                        ]
                        if info.get('ref_start') is not None:
                            tooltip_parts.append(f"Ref Coords: {info['ref_start']} - {info['ref_stop']}")
                        tooltip_parts.append(f"Identity: {info['identity']:.1f}%")
                        tooltip_text = "\n".join(tooltip_parts)
                        break
            else:
                if hasattr(self, 'scores_plotted_for_hover') and self.scores_plotted_for_hover:
                    R = _estimate_dfv_radius(self.ax, self.graphic_ref)
                    center_x, center_y = 0, -R
                    dx = event.xdata - center_x
                    dy = event.ydata - center_y
                    radius = np.sqrt(dx ** 2 + dy ** 2)
                    angle_rad = np.arctan2(dy, dx)

                    # Get ring bounds from plot params AND calculated values
                    id_params = self.plot_params
                    is_color_only = id_params["id_ring_color_only"].get()
                    inner_r_frac = id_params["id_ring_radius"].get()
                    # Use the thickness actually used in the last plot draw
                    thickness_frac = self.last_calculated_values.get("id_thickness",
                                                                     id_params["id_ring_thickness"].get())

                    ring_inner_radius = R * inner_r_frac
                    ring_outer_radius = ring_inner_radius + R * thickness_frac
                    if ring_inner_radius <= radius <= ring_outer_radius:
                        angle_deg_event = np.rad2deg(angle_rad)
                        plot_angle_deg = (90.0 - angle_deg_event) % 360.0
                        top_pos = self.graphic_ref.top_position
                        seqlen = self.data["seqlen"]
                        bp_pos_float = ((plot_angle_deg / 360.0) * seqlen + top_pos) % seqlen
                        bp_pos = int(round(bp_pos_float))
                        scores_available = self.scores_plotted_for_hover
                        if scores_available:
                            min_dist = float('inf')
                            closest_score_val = None
                            closest_pos_val = None
                            for pos, score in scores_available:
                                if score == 0: continue

                                dist = min(abs(bp_pos - pos), seqlen - abs(bp_pos - pos))
                                if dist < min_dist:
                                    min_dist = dist
                                    closest_score_val = score
                                    closest_pos_val = pos
                            step_size_bp = self.plot_params["nc_step_size"].get()
                            max_allowable_dist = step_size_bp * 1.5

                            if closest_score_val is not None and min_dist <= max_allowable_dist:
                                window_size = self.plot_params.get("nc_window_size", {}).get() if hasattr(
                                    self.plot_params.get("nc_window_size", 0), 'get') else self.plot_params.get(
                                    "nc_window_size", 50)
                                half_window = window_size / 2
                                angular_dist_from_center = min(abs(bp_pos - closest_pos_val),
                                                               seqlen - abs(bp_pos - closest_pos_val))
                                is_within_window = angular_dist_from_center <= half_window

                                if is_within_window:
                                    is_within_bar_height = True
                                    if not is_color_only:
                                        normalized_score = np.clip(closest_score_val / 100.0, 0.0, 1.0)
                                        bar_outer_radius = ring_inner_radius + (R * thickness_frac * normalized_score)
                                        tolerance = R * thickness_frac * 0.05
                                        is_within_bar_height = (
                                                ring_inner_radius - tolerance <= radius <= bar_outer_radius + tolerance)

                                    if is_within_bar_height:
                                        is_coding_pos = _is_coding(closest_pos_val + 1, self.coding_intervals_for_hover)
                                        if is_coding_pos: id_label = "Coding Region"
                                        else: id_label = "Non-Coding Region"
                                        display_pos = closest_pos_val + 1  # Convert 0-based to 1-based for display
                                        tooltip_text = f"{id_label}:\nPosition: ~{display_pos}\nIdentity: {closest_score_val:.1f}%"

        if tooltip_text:
            if not hasattr(self, 'tooltip') or self.tooltip is None:
                self.tooltip = Tooltip(self.canvas)
            self.tooltip.show_tip(tooltip_text)

    def _toggle_sidebar(self):
        """Collapses or expands the sidebar using pane forget/insert."""
        if self.sidebar_is_expanded:
            current_width = self.controls_container.winfo_width()
            if current_width > 50: self.last_sidebar_width = current_width
            self.m_paned_window.forget(self.controls_container)
            self.toggle_sidebar_button.configure(text="»")
            self.sidebar_is_expanded = False
        else:
            self.m_paned_window.insert(0, self.controls_container)
            self.m_paned_window.sashpos(0, self.last_sidebar_width)
            self.toggle_sidebar_button.configure(text="«")
            self.sidebar_is_expanded = True
            self.after(50, self._update_scrollbar_visibility)

    def _on_closing(self):
        """Handle the window closing event to notify the main app."""
        try:
            if hasattr(self, 'fig') and self.fig is not None:
                import matplotlib.pyplot as plt
                plt.close(self.fig)
                self.fig = None

            if hasattr(self, 'canvas') and self.canvas is not None:
                self.canvas.get_tk_widget().destroy()
                self.canvas = None

            # Clear tkinter variables to prevent garbage collection on background threads
            if hasattr(self, 'plot_params') and self.plot_params:
                for var in self.plot_params.values():
                    try:
                        if hasattr(var, 'set'): pass
                    except: pass
                self.plot_params.clear()

            if hasattr(self, 'resize_timer') and self.resize_timer is not None:
                self.after_cancel(self.resize_timer)
                self.resize_timer = None

            if hasattr(self, 'sidebar_resize_timer') and self.sidebar_resize_timer is not None:
                self.after_cancel(self.sidebar_resize_timer)
                self.sidebar_resize_timer = None

        except Exception as e: print(f"[WARNING] Error during plot window cleanup: {e}")

        finally:
            self.master_app._enable_buttons()
            self.master_app.current_plot_window = None
            self.destroy()

    def _on_resize(self, event=None):
        """
        Handles the resize event. On the first run (startup maximization), it draws the plot
        immediately. On subsequent runs (manual resizing), it debounces the event.
        """
        if not self.initial_load_complete:
            # Auto-select best available algorithm on initial load
            self._auto_select_algorithm()
            self.update_plot()
            self.after(100, self._adjust_sidebar_font)
            return

        if self.resize_timer: self.after_cancel(self.resize_timer)
        self.resize_timer = self.after(300, self.update_plot)

    def _on_sidebar_resize_debounced(self, event=None):
        """Debounces the sidebar resize event to avoid excessive calculations."""
        if self.sidebar_resize_timer: self.after_cancel(self.sidebar_resize_timer)
        self.sidebar_resize_timer = self.after(150, self._adjust_sidebar_font)

    def _on_manual_fontsize_change(self, event=None):
        """Called when user manually adjusts font size slider."""
        self.plot_params["autofit_label_fontsize"].set(False)
        self.update_plot()

    def _on_toggle_precise_mode(self):
        """
        Handles logic when the 'Show Raw Alignment Blocks' checkbox is toggled.
        If enabling raw blocks, it switches away from legacy algorithms.
        Also recalculates auto-fit color ranges based on the new data.
        """
        if self.plot_params["id_ring_precise_mode"].get():
            algo = self.plot_params["identity_algorithm"].get()
            legacy_algos = ["Legacy (Global k-mer %)", "Legacy (1:1 Base %)"]
            if algo in legacy_algos:
                self.plot_params["identity_algorithm"].set("Mauve + SibeliaZ Fallback")
                self._switch_identity_algorithm()
                return
        self.update_plot()

    def _on_smart_layout_toggle(self):
        """Called when user toggles smart layout checkbox."""
        self.update_plot()

    def _on_autofit_fontsize_toggle(self):
        """Called when user toggles auto-fit font size checkbox."""
        if self.plot_params["autofit_label_fontsize"].get(): self.update_plot()

    def _adjust_sidebar_font(self):
        """
        Calculates the optimal font size by measuring text against available width,
        accounting for UI element buffers (checkbox icons, entry boxes) and indentation.
        """
        try:
            available_width = self.scrollable_frame.winfo_width()
            if available_width <= 50: return  # Guard against resize-to-zero or initialization

            content_items = []
            # Buffers for different UI elements (pixels)
            BUFFER_CHECKBOX = 45
            BUFFER_LABEL_ENTRY = 110
            BUFFER_HEADER = 40
            INDENT_PER_LEVEL = 12

            def traverse_widgets(widget, depth=0):
                w_class = widget.winfo_class()
                indent_px = depth * INDENT_PER_LEVEL

                try:
                    if w_class == 'TCheckbutton':
                        text = str(widget.cget('text'))
                        content_items.append((text, BUFFER_CHECKBOX + indent_px))

                    elif w_class in ('TLabel', 'Label'):
                        text = str(widget.cget('text'))
                        try:
                            f_info = widget.cget("font")
                            is_bold = "bold" in str(f_info).lower()
                        except: is_bold = False

                        if is_bold: content_items.append((text, BUFFER_HEADER + indent_px))
                        else: content_items.append((text, BUFFER_LABEL_ENTRY + indent_px))

                    elif w_class == 'TLabelframe':
                        text = str(widget.cget('text'))
                        content_items.append((text, BUFFER_HEADER + indent_px))

                except Exception: pass
                next_depth = depth + 1 if w_class in ('TLabelframe', 'Frame', 'Labelframe') else depth
                for child in widget.winfo_children(): traverse_widgets(child, next_depth)

            traverse_widgets(self.scrollable_frame)

            if not content_items: return
            try: default_font_family = tkfont.nametofont("TkDefaultFont").actual()['family']
            except tk.TclError: default_font_family = "Helvetica"

            target_width = available_width - 15
            low, high, optimal_size = self.MIN_FONT_SIZE, self.MAX_FONT_SIZE, self.MIN_FONT_SIZE

            while low <= high:
                mid = (low + high) // 2
                test_bold_font = tkfont.Font(family=default_font_family, size=mid, weight='bold')

                fits_all = True
                for text, buffer_px in content_items:
                    if not text: continue
                    text_width = test_bold_font.measure(text)

                    if (text_width + buffer_px) > target_width:
                        fits_all = False
                        break

                if fits_all: optimal_size, low = mid, mid + 1
                else: high = mid - 1

            if optimal_size != self.current_sidebar_font_size:
                self._update_sidebar_font_styles(optimal_size)
                self.current_sidebar_font_size = optimal_size
        except Exception: return

    def _update_sidebar_font_styles(self, size):
        """
        Applies the new font size to all relevant widgets within this window using
        the option database and explicit configuration for existing widgets.
        """
        try:
            default_font_family = tkfont.nametofont("TkDefaultFont").actual()['family']
            text_font_family = tkfont.nametofont("TkTextFont").actual()['family']
        except tk.TclError: default_font_family, text_font_family = "Helvetica", "Helvetica"

        # Update the option database for future widgets (and some current ones)
        self.option_add("*TLabel.font", (default_font_family, size))
        self.option_add("*TButton.font", (default_font_family, size))
        self.option_add("*TCheckbutton.font", (default_font_family, size))  # Explicitly add Checkbutton
        self.option_add("*TLabelframe.Label.font", (default_font_family, size, "bold"))
        self.option_add("*TScale.font", (default_font_family, size))
        self.option_add("*TEntry.font", (text_font_family, size))
        self.option_add("*TCombobox.font", (text_font_family, size))
        self.option_add("*TCombobox*Listbox.font", (text_font_family, size))

        style = ttk.Style(self)
        style.configure(f"{self.style_prefix}.Header.TLabel", font=(default_font_family, size, "bold"))
        self._recursive_widget_update(self.scrollable_frame, default_font_family, size)

    def _recursive_widget_update(self, parent, font_family, size):
        """Recursively forces font updates on widgets."""
        normal_font = (font_family, size)
        for child in parent.winfo_children():
            try:
                w_class = child.winfo_class()
                if w_class == 'Label': child.configure(font=normal_font)
                elif w_class == 'TCheckbutton':
                    style_name = child.cget('style')
                    if style_name: ttk.Style().configure(style_name, font=normal_font)
                    else: child.state(['!disabled'])
            except tk.TclError: pass
            if child.winfo_children(): self._recursive_widget_update(child, font_family, size)

    def _update_tk_widget_fonts(self, parent, size):
        """Recursively finds and updates old tk.Label widgets."""
        try: default_font_family = tkfont.nametofont("TkDefaultFont").actual()['family']
        except tk.TclError: default_font_family = "Helvetica"
        new_font = (default_font_family, size)

        for child in parent.winfo_children():
            widget_class = str(child.winfo_class())
            if widget_class == 'Label':
                try: child.configure(font=new_font)
                except tk.TclError: pass
            if child.winfo_children(): self._update_tk_widget_fonts(child, size)

    def _on_mousewheel(self, event):
        """Handle mouse wheel scrolling for the control panel canvas."""
        if hasattr(self, 'controls_canvas') and self.controls_canvas:
            self.controls_canvas.yview_scroll(-1, "units")
        elif event.num == 5 or (hasattr(event, 'delta') and event.delta < 0):
            self.controls_canvas.yview_scroll(1, "units")

    def _on_blast_task_change_interactive(self, event=None):
        """Updates BLAST parameter entry boxes in the interactive window and enables/disables task-specific options."""
        task = str(self.plot_params["blast_task"].get())
        if task in BLAST_DEFAULTS:
            defaults = BLAST_DEFAULTS[task]
            self.plot_params["blast_word_size"].set(defaults["word_size"])
            self.plot_params["blast_reward"].set(defaults["reward"])
            self.plot_params["blast_penalty"].set(defaults["penalty"])
            self.plot_params["blast_evalue"].set(defaults["evalue"])
            self.plot_params["blast_num_threads"].set(defaults["num_threads"])
        self._update_blast_param_states_sidebar()

    def _update_blast_param_states_sidebar(self):
        """Enable/disable BLAST parameters in sidebar based on task selection."""
        pass

    def _rerun_blast_and_update(self):
        """Re-runs BLAST only with new parameters, preserving identity ring data for speed."""
        for child in self.scrollable_frame.winfo_children():
            try: child.configure(state='disabled')
            except tk.TclError: pass

        self.master_app.status_label.config(text="Re-running BLAST (preserving identity rings)...", fg="black")

        def threaded_task():
            try:
                thread_log_queue = queue.Queue()

                # Extract paths from data_params
                fasta_path = Path(self.data_params["fasta"])
                ref_fasta_path = Path(self.data_params["ref_fasta"])
                annot_path = Path(self.data_params["annot"])

                # Get current BLAST parameters
                blast_task = self.plot_params["blast_task"].get()
                blast_word_size = self.plot_params["blast_word_size"].get()
                blast_reward = self.plot_params["blast_reward"].get()
                blast_penalty = self.plot_params["blast_penalty"].get()
                blast_evalue = self.plot_params["blast_evalue"].get()
                blast_num_threads = self.plot_params["blast_num_threads"].get()

                # Re-run BLAST only, preserving all identity data
                updated_data = _rerun_blast_only(
                    data=self.data.copy(),
                    fasta_path=fasta_path,
                    ref_fasta_path=ref_fasta_path,
                    annot_path=annot_path,
                    fasta_id=self.data["record"].id,
                    seqlen=self.data["seqlen"],
                    blast_task=blast_task,
                    blast_word_size=blast_word_size,
                    blast_reward=blast_reward,
                    blast_penalty=blast_penalty,
                    blast_evalue=blast_evalue,
                    blast_num_threads=blast_num_threads,
                    homology_threshold=self.data_params["homology_threshold"],
                    allow_code_fallback=self.data_params["allow_code_fallback"],
                    manual_find=self.data_params["manual_find"],
                    manual_replace=self.data_params["manual_replace"],
                    highlight_bed_path=self.data_params.get("highlight_bed_path"),
                    log_queue=thread_log_queue
                )

                self.master_app.log_queue.put(("reprocess_done", updated_data))
            except Exception as e:
                self.master_app.log_queue.put(("reprocess_fail", str(e)))
                traceback.print_exc()

        threading.Thread(target=threaded_task, daemon=True).start()

    def _on_resolution_change(self, event=None):
        if self.plot_params["save_resolution"].get() == "Custom Resolution":
            self.custom_res_frame.pack(fill="x", pady=5)
        else: self.custom_res_frame.pack_forget()


    def _auto_select_algorithm(self):
        """Automatically select the best available identity algorithm on initial load."""
        has_mauve = bool(self.data.get("mauve_data", {}).get("coding_scores"))
        has_sibeliaz = bool(self.data.get("sibeliaz_data", {}).get("coding_scores"))
        has_combined = bool(self.data.get("combined_data", {}).get("coding_scores"))

        current_selection = self.plot_params["identity_algorithm"].get()
        print(f"[DEBUG] Auto-selecting algorithm. Current: {current_selection}")
        print(f"[DEBUG] Available: Mauve={has_mauve}, SibeliaZ={has_sibeliaz}, Combined={has_combined}")

        # If current selection has data, keep it
        if current_selection == "Mauve + SibeliaZ Fallback" and has_combined:
            return
        elif current_selection == "Mauve" and has_mauve:
            return
        elif current_selection == "SibeliaZ" and has_sibeliaz:
            return
        elif current_selection.startswith("Legacy"):
            # Legacy algorithms don't need auto-selection
            return

        # Otherwise, pick the best available in priority order
        if has_combined:
            self.plot_params["identity_algorithm"].set("Mauve + SibeliaZ Fallback")
            print("[INFO] Auto-selected: Mauve + SibeliaZ Fallback")
        elif has_mauve:
            self.plot_params["identity_algorithm"].set("Mauve")
            print("[INFO] Auto-selected: Mauve")
        elif has_sibeliaz:
            self.plot_params["identity_algorithm"].set("SibeliaZ")
            print("[INFO] Auto-selected: SibeliaZ")
        else:
            print("[WARNING] No WGA identity data available")

    def _switch_identity_algorithm(self):
        """Switch between pre-calculated identity algorithms without re-running alignments."""
        selected_algo = self.plot_params["identity_algorithm"].get()
        legacy_algos = ["Legacy (Global k-mer %)", "Legacy (1:1 Base %)"]

        # 1. Handle Legacy vs Raw Block Mode compatibility
        if selected_algo in legacy_algos:
            self.plot_params["id_ring_precise_mode"].set(False)

        # 2. Map Selection to Data Keys
        if selected_algo == "Legacy (Global k-mer %)": data_key = "legacy_data"
        elif selected_algo == "Legacy (1:1 Base %)": data_key = "legacy_base_data"
        elif selected_algo == "Mauve": data_key = "mauve_data"
        elif selected_algo == "SibeliaZ": data_key = "sibeliaz_data"
        else: data_key = "combined_data"  # "Mauve + SibeliaZ Fallback"

        # 3. Check if Data Exists and Auto-Fallback
        has_mauve = bool(self.data.get("mauve_data", {}).get("coding_scores"))
        has_sibeliaz = bool(self.data.get("sibeliaz_data", {}).get("coding_scores"))
        has_combined = bool(self.data.get("combined_data", {}).get("coding_scores"))

        # Check if selected algorithm has data, if not, auto-fallback
        if data_key == "combined_data" and not has_combined:
            if has_mauve:
                print("[INFO] Combined data not available, switching to Mauve")
                data_key = "mauve_data"
                self.plot_params["identity_algorithm"].set("Mauve")
                selected_algo = "Mauve"
            elif has_sibeliaz:
                print("[INFO] Combined data not available, switching to SibeliaZ")
                data_key = "sibeliaz_data"
                self.plot_params["identity_algorithm"].set("SibeliaZ")
                selected_algo = "SibeliaZ"
        elif data_key == "mauve_data" and not has_mauve:
            if has_sibeliaz:
                print("[INFO] Mauve data not available, switching to SibeliaZ")
                data_key = "sibeliaz_data"
                self.plot_params["identity_algorithm"].set("SibeliaZ")
                selected_algo = "SibeliaZ"
            elif has_combined:
                print("[INFO] Mauve data not available, switching to Combined")
                data_key = "combined_data"
                self.plot_params["identity_algorithm"].set("Mauve + SibeliaZ Fallback")
                selected_algo = "Mauve + SibeliaZ Fallback"
        elif data_key == "sibeliaz_data" and not has_sibeliaz:
            if has_mauve:
                print("[INFO] SibeliaZ data not available, switching to Mauve")
                data_key = "mauve_data"
                self.plot_params["identity_algorithm"].set("Mauve")
                selected_algo = "Mauve"
            elif has_combined:
                print("[INFO] SibeliaZ data not available, switching to Combined")
                data_key = "combined_data"
                self.plot_params["identity_algorithm"].set("Mauve + SibeliaZ Fallback")
                selected_algo = "Mauve + SibeliaZ Fallback"

        # 4. Final Validation
        if data_key not in self.data or not (
                self.data[data_key].get("coding_scores") or self.data[data_key].get("non_coding_scores")):
            # Only show error if we truly have NO data for this selection
            messagebox.showwarning(
                "Data Not Available",
                f"The selected algorithm ({selected_algo}) did not produce valid data.\n"
                f"This may happen if the alignment failed or was not run."
            )
            # Revert to whatever was previously working
            current_used = self.data.get("identity_algorithm_used", "Mauve + SibeliaZ Fallback")
            self.plot_params["identity_algorithm"].set(current_used)
            return

        # 5. Apply the Switch
        cached = self.data[data_key]
        self.data["coding_scores"] = cached["coding_scores"]
        self.data["non_coding_scores"] = cached["non_coding_scores"]
        self.data["global_wga_id"] = cached["global_wga_id"]
        self.data["noncoding_avg_id"] = cached["noncoding_avg_id"]
        self.data["wga_id_intergenic_incl_gaps"] = cached.get("wga_id_intergenic_incl_gaps", 0.0)
        self.data["mauve_lcb_data"] = self.data.get("mauve_data", {}).get("lcb_data", [])
        self.data["sibeliaz_lcb_data"] = self.data.get("sibeliaz_data", {}).get("lcb_data", [])
        self.data["mauve_intergenic_lcb_data"] = self.data.get("mauve_data", {}).get("intergenic_lcb_data", [])
        self.data["sibeliaz_intergenic_lcb_data"] = self.data.get("sibeliaz_data", {}).get("intergenic_lcb_data", [])

        self.data["identity_algorithm_used"] = selected_algo
        self.update_plot()
        self.master_app.status_label.config(text=f"Switched to {selected_algo} (cached data).", fg="blue")

    def _recalculate_identity_and_update(self):
        """
        Fast recalculation using cached arrays and the optimized sampler.
        """
        # 1. Get parameters
        try:
            new_window = self.plot_params["nc_window_size"].get()
            new_step = self.plot_params["nc_step_size"].get()
        except (ValueError, TypeError):
            return

        algo = self.plot_params["identity_algorithm"].get()

        # 2. Get Cached Data
        if algo == "Mauve":
            raw_array = self.data.get("mauve_g_array")
        elif algo == "SibeliaZ":
            raw_array = self.data.get("sibeliaz_g_array")
        else:  # Combined
            raw_array = self.data.get("combined_g_array")

        if raw_array is None:
            print("[INFO] Cache missing, running full reprocess...")
            self._reprocess_data_and_update()
            return

        # 3. Call the UNIFIED optimized function
        self.master_app.status_label.config(text="Instantly recalculating...", fg="blue")

        coding, non_coding, avg = _sample_identity_profile(
            raw_array,
            self.data["seqlen"],
            new_window,
            new_step,
            self.data["gene_feats"],
            self.log_queue
        )

        # 4. Update Data Structure
        if algo == "Mauve":
            self.data["mauve_data"]["coding_scores"] = coding
            self.data["mauve_data"]["non_coding_scores"] = non_coding
            self.data["mauve_data"]["noncoding_avg_id"] = avg
        elif algo == "SibeliaZ":
            self.data["sibeliaz_data"]["coding_scores"] = coding
            self.data["sibeliaz_data"]["non_coding_scores"] = non_coding
            self.data["sibeliaz_data"]["noncoding_avg_id"] = avg
        else:  # Combined
            self.data["combined_data"]["coding_scores"] = coding
            self.data["combined_data"]["non_coding_scores"] = non_coding
            self.data["combined_data"]["noncoding_avg_id"] = avg

        self.data["coding_scores"] = coding
        self.data["non_coding_scores"] = non_coding
        self.data["noncoding_avg_id"] = avg

        self.update_plot()
        self.master_app.status_label.config(text="Identity ring updated.", fg="green")

    def _recalculate_gc_and_update(self):
        for child in self.scrollable_frame.winfo_children():
            try: child.configure(state='disabled')
            except tk.TclError: pass
        self.master_app.status_label.config(text="Recalculating GC/Skew rings...", fg="black")
        new_window_size = self.plot_params["gc_window_size"].get()

        def threaded_task():
            try:
                new_gc_arr, new_skew_arr = calculate_gc_skew_data(
                    sequence=str(self.data["record"].seq).upper(),
                    window_size=new_window_size
                )
                self.master_app.log_queue.put(("gc_recalc_done", (new_gc_arr, new_skew_arr)))
            except Exception as e:
                self.master_app.log_queue.put(("reprocess_fail", str(e)))
                traceback.print_exc()
        threading.Thread(target=threaded_task, daemon=True).start()

    def _on_gc_color_only_toggle(self):
        """
        Manages the GC Ring visualization modes (Color-Only vs. Diverging).
        It enforces the fixed Color-Only geometry or the calculated Diverging thickness,
        and crucially, ensures the GC Skew ring runs its auto-adjustment in response
        to *any* change in GC ring geometry.
        """
        is_color_only = self.plot_params["gc_ring_color_only"].get()
        current_thickness = self.plot_params["gc_thick"].get()

        if is_color_only and self.plot_params["gc_colormap"].get() == "Grey":
            self.plot_params["gc_colormap"].set(str("Coolwarm (Reversed)"))
        if is_color_only:
            self.user_set_gc_thickness = current_thickness if current_thickness > 0.5 else 1.05
            self.plot_params["auto_thickness_rings"].set(True)
            self._set_slider_value_with_range_adjustment(self.gc_thickness_slider,
                                                         self.plot_params["gc_thick"], 0.15)
            self._set_slider_value_with_range_adjustment(self.gc_inner_radius_slider,
                                                         self.plot_params["gc_inner"], 0.15)
            self.gc_thickness_slider.configure(from_=0.05, to=0.5)
        else:
            self.plot_params["auto_thickness_rings"].set(True)
            restore_thick = self.user_set_gc_thickness if (hasattr(self,
                                                                  'user_set_gc_thickness') and
                                                           self.user_set_gc_thickness is not None) else 1.05

            self.gc_thickness_slider.configure(from_=0.1, to=2.0)
            self._set_slider_value_with_range_adjustment(self.gc_inner_radius_slider,
                                                         self.plot_params["gc_inner"], 0.3)
            try:
                gc_arr = self.data["gc_arr"]
                R_data_units = _estimate_dfv_radius(self.ax,
                                                    self.graphic_ref) if (hasattr(self, 'graphic_ref')
                                                                          and self.graphic_ref is not None) else 1.0
                gc_median = np.median(gc_arr) if gc_arr is not None and len(gc_arr) > 0 else 0.5
                largest_gc_dev = np.max(np.abs(gc_arr - gc_median)) if gc_arr is not None and len(gc_arr) > 0 else 0.1
                gc_inner_frac = self.plot_params["gc_inner"].get()
                skew_inner_frac = self.plot_params["skew_inner_radius"].get()
                available_space_frac = skew_inner_frac - gc_inner_frac

                if largest_gc_dev > 1e-9 and R_data_units > 0:
                    required_thickness_factor = (0.5 * available_space_frac) / largest_gc_dev
                    restore_thick = np.clip(required_thickness_factor * (1 / R_data_units), 0.1, 2.0)
            except Exception as e:
                print(f"WARNING: GC thickness calculation failed: {e}")
                pass
            self._set_slider_value_with_range_adjustment(self.gc_thickness_slider,
                                                         self.plot_params["gc_thick"], restore_thick)
        self.update_plot()

    def _on_id_color_only_toggle(self):
        """
        Adjusts Identity Ring appearance and handles thickness storage/restoration.
        Auto-calculates thickness when switching color mode OFF.
        Restores user-set thickness when switching color mode ON.
        """
        is_color_only = self.plot_params["id_ring_color_only"].get()
        current_thickness = self.plot_params["id_ring_thickness"].get()

        if is_color_only:
            self.user_set_id_thickness = current_thickness
            self.plot_params["id_baseline_transparency_pct"].set(100)
            restore_thickness = self.user_set_id_thickness if self.user_set_id_thickness is not None else 0.20
            self._set_slider_value_with_range_adjustment(self.id_radius_slider, self.plot_params["id_ring_radius"], 0.7)
            self._set_slider_value_with_range_adjustment(self.id_thickness_slider,
                                                         self.plot_params["id_ring_thickness"], restore_thickness)
        else:
            self.plot_params["id_baseline_transparency_pct"].set(50)
            self.plot_params["auto_thickness_rings"].set(True)
            self.update_plot()
            if hasattr(self, 'last_calculated_values') and 'id_thickness' in self.last_calculated_values:
                calculated_thickness = self.last_calculated_values['id_thickness']
                self._set_slider_value_with_range_adjustment(
                    self.id_thickness_slider,
                    self.plot_params["id_ring_thickness"],
                    calculated_thickness)
            self.plot_params["auto_thickness_rings"].set(False)
            self.update_plot()
            return
        self.update_plot()

    def _on_manual_ring_adjustment(self):
        """
        Generic handler called when user manually adjusts any ring parameter
        (thickness, radius, line width, etc).
        Disables auto-thickness so manual values override auto-calculation.
        """
        self.plot_params["auto_thickness_rings"].set(False)
        self.update_plot()

    def _reprocess_data_and_update(self):
        for child in self.scrollable_frame.winfo_children():
            try: child.configure(state='disabled')
            except tk.TclError: pass

        self.master_app.status_label.config(text="Re-processing data for interactive plot...", fg="black")
        self.data_params["identity_algorithm"] = self.plot_params["identity_algorithm"].get()
        self.data_params["nc_window_size"] = self.plot_params["nc_window_size"].get()
        self.data_params["nc_step_size"] = self.plot_params["nc_step_size"].get()

        def threaded_task():
            try:
                thread_log_queue = queue.Queue()
                new_data = _process_data(**self.data_params, log_queue=thread_log_queue)
                self.master_app.log_queue.put(("reprocess_done", new_data))
            except Exception as e:
                self.master_app.log_queue.put(("reprocess_fail", str(e)))
                traceback.print_exc()
        threading.Thread(target=threaded_task, daemon=True).start()

    def _on_toggle_region_limit(self):
        """
        When the 'Limit to Region' checkbox is enabled, this function populates the
        start/end boxes with the full sequence range if they are currently empty.
        """
        if self.plot_params["limit_to_region"].get():
            is_start_empty = not self.plot_params["region_start"].get()
            is_end_empty = not self.plot_params["region_end"].get()
            if is_start_empty and is_end_empty:
                full_seqlen = self.data.get("seqlen", 0)
                self.plot_params["region_start"].set("1")
                self.plot_params["region_end"].set(str(full_seqlen))
        self.update_plot()

    def _run_legacy_recalculation_kmer(self):
        """Disables controls and runs the legacy K-MER calculation in a thread."""
        for child in self.scrollable_frame.winfo_children():
            try: child.configure(state='disabled')
            except tk.TclError: pass
        self.master_app.status_label.config(text="Recalculating Legacy (Global k-mer %) identity...", fg="black")

        try:
            query_seq = str(self.data["record"].seq).upper()
            KMER_SIZE_FOR_PLOT = 10
            ref_kmers_global_set = _get_kmers(self.data["ref_seq_concatenated"], KMER_SIZE_FOR_PLOT)
            gene_features = self.data["gene_feats"]
            window_size = self.plot_params["nc_window_size"].get()
            step_size = self.plot_params["nc_step_size"].get()
        except Exception as e:
            messagebox.showerror("Error", f"Could not prepare data for legacy k-mer calculation: {e}")
            self._enable_controls()
            return

        def threaded_task():
            try:
                coding_scores, non_coding_scores, global_avg_id = calculate_global_kmer_ring_data(
                    query_seq, ref_kmers_global_set, gene_features,
                    window_size, step_size, self.master_app.log_queue
                )
                payload = (coding_scores, non_coding_scores, global_avg_id)
                self.master_app.log_queue.put(("legacy_kmer_recalc_done", payload))
            except Exception as e:
                self.master_app.log_queue.put(("status", (f"Legacy k-mer recalc failed: {e}", "red")))
                self.master_app.log_queue.put(("legacy_recalc_fail", None))
                traceback.print_exc()
        threading.Thread(target=threaded_task, daemon=True).start()

    def _run_legacy_recalculation_base(self):
        """Disables controls and runs the legacy 1:1 BASE COMPARISON calculation in a thread."""
        for child in self.scrollable_frame.winfo_children():
            try: child.configure(state='disabled')
            except tk.TclError: pass
        self.master_app.status_label.config(text="Recalculating Legacy (1:1 Base %) identity...", fg="black")

        try:
            query_seq = str(self.data["record"].seq).upper()
            ref_seq_concatenated = self.data["ref_seq_concatenated"]
            gene_features = self.data["gene_feats"]
            window_size = self.plot_params["nc_window_size"].get()
            step_size = self.plot_params["nc_step_size"].get()
        except Exception as e:
            messagebox.showerror("Error", f"Could not prepare data for base comparison: {e}")
            self._enable_controls()
            return

        def threaded_task():
            try:
                coding_scores, non_coding_scores, global_avg_id = calculate_legacy_base_comparison_ring_data(
                    query_seq, ref_seq_concatenated, gene_features,
                    window_size, step_size, self.master_app.log_queue
                )
                payload = (coding_scores, non_coding_scores, global_avg_id)
                self.master_app.log_queue.put(("legacy_base_recalc_done", payload))
            except Exception as e:
                self.master_app.log_queue.put(("status", (f"Legacy base recalc failed: {e}", "red")))
                self.master_app.log_queue.put(("legacy_recalc_fail", None))
                traceback.print_exc()
        threading.Thread(target=threaded_task, daemon=True).start()

    def _enable_controls(self):
        """Helper to re-enable all controls in the scrollable frame."""
        for child in self.scrollable_frame.winfo_children():
            try: child.configure(state='normal')
            except tk.TclError: pass

    def _zoom_to_roi_region(self):
        """
        Finds all highlighted ROI features (red), calculates their total span,
        and updates the region text boxes to zoom the plot to that area.

        This will *only* find red features loaded from the BED file.
        """
        all_features = self.data.get('gene_feats', [])
        roi_features = [f for f in all_features if f.color == "#e53935"]
        if not roi_features:
            messagebox.showinfo("No ROI Found", "No highlighted ROI features (from BED file) were found in the data.")
            return

        min_start = min(f.start for f in roi_features)
        max_end = max(f.end for f in roi_features)
        full_seqlen = self.data.get("seqlen", max_end)
        final_start = max(1, min_start - 1000)
        final_end = min(full_seqlen, max_end + 1000)

        self.plot_params["limit_to_region"].set(True)
        self.plot_params["region_start"].set(str(final_start))
        self.plot_params["region_end"].set(str(final_end))
        self.update_plot()

    def _set_slider_value_with_range_adjustment(self, slider_widget, var, value):
        """
        Set a slider value and dynamically adjust its range if the value is outside bounds.
        - If value < minimum: set minimum to value/2, keep maximum at default
        - If value > maximum: set maximum to value*1.5, keep minimum at default
        """
        try:
            val = float(value)
        except (ValueError, TypeError):
            return

        if slider_widget is None:
            var.set(value)
            return

        if not hasattr(slider_widget, 'default_from'):
            var.set(value)
            return

        default_from = slider_widget.default_from
        default_to = slider_widget.default_to
        current_from = slider_widget.cget("from")
        current_to = slider_widget.cget("to")

        if val < default_from:
            new_min = val / 2.0
            slider_widget.configure(from_=new_min)
            if current_to != default_to: slider_widget.configure(to=default_to)
        elif val > default_to:
            new_max = val * 1.5
            slider_widget.configure(to=new_max)
            if current_from != default_from: slider_widget.configure(from_=default_from)
        else:
            if current_from != default_from or current_to != default_to:
                slider_widget.configure(from_=default_from, to=default_to)
        var.set(value)

    def _apply_initial_params(self, initial_params):
        """
        Apply initial parameters from main window to plot_params.
        Respects auto-options:
        - If autofit_label_fontsize is True, ignore label_fontsize
        - If use_smart_layout is True, ignore label_spread, label_distance_factor, label_line_radial_len,
        label_line_horizontal_len, curve_tension
        - If auto_thickness_rings is True, ignore gc_thick and skew_thick
        """
        if not initial_params:
            return

        # Boolean flags for auto-options
        autofit_fontsize = initial_params.get("autofit_label_fontsize", True)
        use_smart_layout = initial_params.get("use_smart_layout", True)
        auto_thickness = initial_params.get("auto_thickness_rings", True)

        # Parameters that should be ignored if certain auto-options are enabled
        ignore_if_autofit = {"label_fontsize"}
        ignore_if_smart_layout = {"label_spread", "label_distance_factor", "label_line_radial_len", "label_line_horizontal_len", "curve_tension"}
        ignore_if_auto_thickness = {"gc_thick", "skew_thick"}

        for key, value in initial_params.items():
            # Skip parameters based on auto-option settings
            if autofit_fontsize and key in ignore_if_autofit: continue
            if use_smart_layout and key in ignore_if_smart_layout: continue
            if auto_thickness and key in ignore_if_auto_thickness: continue

            # Apply parameter if it exists in plot_params
            if key in self.plot_params:
                self.plot_params[key].set(value)

        # Adjust GC ring settings if color-only mode is enabled
        if initial_params.get("gc_ring_color_only", False):
            self.plot_params["gc_thick"].set(0.15)
            self.plot_params["gc_inner"].set(0.15)

        # Adjust ID ring settings if color-only mode is enabled
        if initial_params.get("id_ring_color_only", False):
            self.plot_params["id_ring_radius"].set(0.7)

    def _add_slider_with_entry(self, parent, label, var, from_, to, command=None, tooltip_text=None):
        """
        Adds a new layout with a label, text entry, and slider.
        The text entry and slider are linked to the same 'var'.
        """
        frame = ttk.Frame(parent)
        frame.pack(fill="x", pady=2, padx=5)
        frame.columnconfigure(1, weight=1)
        top_frame = ttk.Frame(frame)
        top_frame.grid(row=0, column=0, columnspan=2, sticky="ew")
        top_frame.columnconfigure(1, weight=1)
        lbl = tk.Label(top_frame, text=label, width=22, anchor='w')
        lbl.pack(side="left")
        if tooltip_text: _create_tooltip(lbl, tooltip_text)

        entry = ttk.Entry(top_frame, textvariable=var)
        entry.pack(side="left", padx=5, fill='x', expand=True)
        if tooltip_text: _create_tooltip(entry, tooltip_text)
        effective_command = command or (lambda e: self.update_plot())

        default_from, default_to = from_, to

        def adjust_slider_range_for_value(value):
            """Dynamically adjust slider range if value is outside bounds."""
            try: val = float(value)
            except (ValueError, tk.TclError): return

            current_from, current_to = scale.cget("from"), scale.cget("to")

            if val < current_from:
                new_min = val / 2.0
                scale.configure(from_=new_min)
                scale.configure(to=default_to)
            elif val > current_to:
                new_max = val * 1.5
                scale.configure(to=new_max)
                scale.configure(from_=default_from)
            elif current_from != default_from or current_to != default_to:
                if default_from <= val <= default_to:
                    all_within_default = (default_from <= val <= default_to)
                    if all_within_default: scale.configure(from_=default_from, to=default_to)

        def on_enter_pressed(event):
            try: float(var.get())
            except (ValueError, tk.TclError): var.set(default_from)
            adjust_slider_range_for_value(var.get())
            effective_command(event)

        entry.bind("<Return>", on_enter_pressed)

        scale = ttk.Scale(frame, variable=var, from_=from_, to=to, orient="horizontal",
                          command=effective_command)
        scale.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(2, 5))
        if tooltip_text: _create_tooltip(scale, tooltip_text)
        scale.adjust_range = adjust_slider_range_for_value
        scale.default_from = default_from
        scale.default_to = default_to

        return scale

    def _add_controls(self, parent):
        def add_dropdown(p, label, var, options, tooltip_key=None):
            frame = tk.Frame(p)
            frame.pack(fill="x", pady=2, padx=5)
            lbl = tk.Label(frame, text=label, width=22, anchor='w')
            lbl.pack(side="left")
            if tooltip_key and tooltip_key in self.plot_tooltips: _create_tooltip(lbl, self.plot_tooltips[tooltip_key])
            combo = ttk.Combobox(frame, textvariable=var, values=options, state='readonly')
            combo.pack(side="left", expand=True, fill="x", padx=(5, 0))
            if tooltip_key and tooltip_key in self.plot_tooltips:
                _create_tooltip(combo, self.plot_tooltips[tooltip_key])
            combo.bind("<<ComboboxSelected>>", lambda e: self.update_plot())

        colormap_options = ["Plasma", "Viridis", "Inferno", "Magma", "Cividis", "Turbo", "Coolwarm", "RdYlBu",
                            "Gist_earth", "Terrain", "Hsv", "Jet"]
        full_cmaps = colormap_options + [f"{c} (Reversed)" for c in colormap_options]

        view_frame = ttk.LabelFrame(parent, text="General View & Orientation", padding=(10, 5))
        view_frame.pack(fill="x", pady=(0, 10))
        self._add_slider_with_entry(view_frame, "Zoom", self.plot_params["zoom"], 0.25, 6.0,
                                    tooltip_text=self.plot_tooltips.get("zoom"))
        self._add_slider_with_entry(view_frame,
                                    "Manual Rotation", self.plot_params["manual_rot_deg"], 0, 360,
                                    tooltip_text=self.plot_tooltips.get("manual_rot_deg"))
        def add_check_with_tooltip(p, label, key, command, **kwargs):
            cb = ttk.Checkbutton(p, text=label, variable=self.plot_params[key], command=command)
            cb.pack(anchor="w", padx=5, **kwargs)
            if key in self.plot_tooltips:
                _create_tooltip(cb, self.plot_tooltips[key])

        add_check_with_tooltip(view_frame, "Autorotate to ROI", "autorotate_roi", self.update_plot)
        add_check_with_tooltip(view_frame, "Show Coordinate Numbers", "show_coords", self.update_plot)
        blast_section_interactive = CollapsibleFrame(parent, text="BLAST Parameters", expanded=True,
                                                     style_prefix=self.style_prefix)
        blast_section_interactive.pack(fill="x", pady=5)
        b_interior = blast_section_interactive.interior
        task_frame = tk.Frame(b_interior)
        task_frame.pack(fill="x", pady=2, padx=5)
        lbl_task = tk.Label(task_frame, text="Task", width=15, anchor='w')
        lbl_task.pack(side="left")
        if "blast_task" in self.plot_tooltips: _create_tooltip(lbl_task, self.plot_tooltips["blast_task"])
        task_combo_interactive = ttk.Combobox(task_frame, textvariable=self.plot_params["blast_task"],
                                              values=list(BLAST_DEFAULTS.keys()), state='readonly')
        task_combo_interactive.pack(side="left", expand=True, fill="x", padx=(5, 0))
        if "blast_task" in self.plot_tooltips: _create_tooltip(task_combo_interactive, self.plot_tooltips["blast_task"])
        task_combo_interactive.bind("<<ComboboxSelected>>", self._on_blast_task_change_interactive)

        blast_entry_frame1 = tk.Frame(b_interior)
        blast_entry_frame1.pack(fill="x", pady=2, padx=5)
        lbl_ws = tk.Label(blast_entry_frame1, text="Word Size", width=15, anchor='w')
        lbl_ws.pack(side="left")
        if "blast_word_size" in self.plot_tooltips: _create_tooltip(lbl_ws, self.plot_tooltips["blast_word_size"])
        entry_ws = ttk.Entry(blast_entry_frame1, textvariable=self.plot_params["blast_word_size"], width=10)
        entry_ws.pack(side="left", padx=(5, 10))
        if "blast_word_size" in self.plot_tooltips: _create_tooltip(entry_ws, self.plot_tooltips["blast_word_size"])
        entry_ws.bind("<FocusOut>",
                      _create_bounded_entry_validator(self.plot_params["blast_word_size"],
                                                      7, 28, "BLAST Word Size"))
        lbl_rew = tk.Label(blast_entry_frame1, text="Reward", width=15, anchor='w')
        lbl_rew.pack(side="left")
        if "blast_reward" in self.plot_tooltips: _create_tooltip(lbl_rew, self.plot_tooltips["blast_reward"])
        entry_rew = ttk.Entry(blast_entry_frame1, textvariable=self.plot_params["blast_reward"], width=10)
        entry_rew.pack(side="left", padx=(5, 0))
        if "blast_reward" in self.plot_tooltips: _create_tooltip(entry_rew, self.plot_tooltips["blast_reward"])
        entry_rew.bind("<FocusOut>", _create_positive_validator(self.plot_params["blast_reward"], "BLAST Reward"))

        blast_entry_frame2 = tk.Frame(b_interior)
        blast_entry_frame2.pack(fill="x", pady=2, padx=5)
        lbl_pen = tk.Label(blast_entry_frame2, text="Penalty", width=15, anchor='w')
        lbl_pen.pack(side="left")
        if "blast_penalty" in self.plot_tooltips: _create_tooltip(lbl_pen, self.plot_tooltips["blast_penalty"])
        entry_pen = ttk.Entry(blast_entry_frame2, textvariable=self.plot_params["blast_penalty"], width=10)
        entry_pen.pack(side="left", padx=(5, 10))
        if "blast_penalty" in self.plot_tooltips: _create_tooltip(entry_pen, self.plot_tooltips["blast_penalty"])
        entry_pen.bind("<FocusOut>",
                       _create_bounded_entry_validator(self.plot_params["blast_penalty"],
                                                       -10, 0, "BLAST Penalty"))

        lbl_eval = tk.Label(blast_entry_frame2, text="E-value", width=15, anchor='w')
        lbl_eval.pack(side="left")
        if "blast_evalue" in self.plot_tooltips: _create_tooltip(lbl_eval, self.plot_tooltips["blast_evalue"])
        entry_eval = ttk.Entry(blast_entry_frame2, textvariable=self.plot_params["blast_evalue"], width=10)
        entry_eval.pack(side="left", padx=(5, 0))
        if "blast_evalue" in self.plot_tooltips: _create_tooltip(entry_eval, self.plot_tooltips["blast_evalue"])
        entry_eval.bind("<FocusOut>",
                        _create_positive_validator(self.plot_params["blast_evalue"],
                                                   "BLAST E-value", is_float=True))
        blast_entry_frame_threads = tk.Frame(b_interior)
        blast_entry_frame_threads.pack(fill="x", pady=2, padx=5)
        lbl_threads = tk.Label(blast_entry_frame_threads, text="Num Threads", width=15, anchor='w')
        lbl_threads.pack(side="left")
        if "blast_num_threads" in self.plot_tooltips:
            _create_tooltip(lbl_threads, self.plot_tooltips["blast_num_threads"])
        entry_threads = ttk.Entry(blast_entry_frame_threads, textvariable=self.plot_params["blast_num_threads"],
                                  width=10)
        entry_threads.pack(side="left", padx=(5, 0))
        if "blast_num_threads" in self.plot_tooltips:
            _create_tooltip(entry_threads, self.plot_tooltips["blast_num_threads"])
        entry_threads.bind("<FocusOut>", _create_bounded_entry_validator(self.plot_params["blast_num_threads"], 1,
                                                                         max(os.cpu_count() or 1, 512),
                                                                         "BLAST Num Threads"))
        self._on_blast_task_change_interactive()
        self._update_blast_param_states_sidebar()
        ttk.Button(b_interior, text="Re-run BLAST & Update Plot", command=self._rerun_blast_and_update).pack(pady=10,
                                                                                                             fill="x",
                                                                                                             padx=5)
        ttk.Separator(b_interior).pack(fill="x", pady=5)
        region_check = ttk.Checkbutton(b_interior, text="Limit Plot to Specific Region",
                                       variable=self.plot_params["limit_to_region"],
                                       command=self._on_toggle_region_limit)
        region_check.pack(anchor="w", padx=5, pady=(5, 2))
        if "limit_to_region" in self.plot_tooltips: _create_tooltip(region_check, self.plot_tooltips["limit_to_region"])
        region_frame_interactive = tk.Frame(b_interior)
        region_frame_interactive.pack(fill="x", pady=2, padx=5)
        lbl_reg_start = tk.Label(region_frame_interactive, text="Region Start:", width=15, anchor='w')
        lbl_reg_start.pack(side="left")
        if "region_start" in self.plot_tooltips:
            _create_tooltip(lbl_reg_start, self.plot_tooltips["region_start"])
        region_start_entry = ttk.Entry(region_frame_interactive, textvariable=self.plot_params["region_start"],
                                       width=10)
        region_start_entry.pack(side="left", padx=(5, 10))
        if "region_start" in self.plot_tooltips:
            _create_tooltip(region_start_entry, self.plot_tooltips["region_start"])
        region_start_entry.bind("<Return>", lambda e: self.update_plot())
        region_start_entry.bind("<FocusOut>",
                                _create_positive_validator(self.plot_params["region_start"], "Region Start"))
        lbl_reg_end = tk.Label(region_frame_interactive, text="Region End:", width=15, anchor='w')
        lbl_reg_end.pack(side="left")
        if "region_end" in self.plot_tooltips: _create_tooltip(lbl_reg_end, self.plot_tooltips["region_end"])
        region_end_entry = ttk.Entry(region_frame_interactive, textvariable=self.plot_params["region_end"], width=10)
        region_end_entry.pack(side="left", padx=(5, 0))
        if "region_end" in self.plot_tooltips:
            _create_tooltip(region_end_entry, self.plot_tooltips["region_end"])
        region_end_entry.bind("<Return>", lambda e: self.update_plot())
        region_end_entry.bind("<FocusOut>", _create_positive_validator(self.plot_params["region_end"],
                                                                       "Region End"))
        ttk.Button(b_interior, text="Zoom to ROI", command=self._zoom_to_roi_region).pack(fill="x", pady=(5, 0), padx=5)
        blast_section = CollapsibleFrame(parent, text="Homology & BLAST Ring", expanded=True,
                                         style_prefix=self.style_prefix)
        blast_section.pack(fill="x", pady=5)
        b_interior2 = blast_section.interior
        ttk.Label(b_interior2, text="Filtering & Display", font=("Helvetica", 9, "bold")).pack(anchor="w")
        add_check_with_tooltip(b_interior2, "Show Homology Labels", "show_homology_labels", self.update_plot)
        add_check_with_tooltip(b_interior2, "Show ROI Labels", "show_roi_labels", self.update_plot)
        add_check_with_tooltip(b_interior2, "Show BLAST Hits in Coding Regions",
                               "show_blast_coding", self.update_plot)
        add_check_with_tooltip(b_interior2, "Show BLAST Hits in Non-Coding Regions", "show_blast_noncoding",
                               self.update_plot)
        add_check_with_tooltip(b_interior2, "Show Start/End Line", "show_start_end_line", self.update_plot)
        add_check_with_tooltip(b_interior2, "Remove Gene Outlines", "remove_gene_borders", self.update_plot)

        min_frame = tk.Frame(b_interior2)
        min_frame.pack(fill="x", pady=2, padx=5)
        tk.Label(min_frame, text="Min Homology %", width=22, anchor='w').pack(side="left")
        min_entry = ttk.Entry(min_frame, textvariable=self.plot_params["homology_thresh"], width=8)
        min_entry.pack(side="left", padx=(5, 5))
        min_entry.bind("<Return>", lambda e: self.update_plot())
        min_entry.bind("<FocusOut>",
                       _create_bounded_entry_validator(self.plot_params["homology_thresh"],
                                                       0, 100, "Min Homology %", is_float=True))
        max_frame = tk.Frame(b_interior2)
        max_frame.pack(fill="x", pady=2, padx=5)
        tk.Label(max_frame, text="Max Homology %", width=22, anchor='w').pack(side="left")
        max_entry = ttk.Entry(max_frame, textvariable=self.plot_params["homology_thresh_max"], width=8)
        max_entry.pack(side="left", padx=(5, 5))
        max_entry.bind("<Return>", lambda e: self.update_plot())
        max_entry.bind("<FocusOut>", _create_bounded_entry_validator(self.plot_params["homology_thresh_max"],
                                                                     0, 100,
                                                                     "Max Homology %", is_float=True))
        ttk.Separator(b_interior2).pack(fill="x", pady=5)
        ttk.Label(b_interior2, text="Color & Appearance", font=("Helvetica", 9, "bold")).pack(anchor="w")
        add_dropdown(b_interior2, "Homology Colormap", self.plot_params["homology_colormap"], full_cmaps,
                     tooltip_key="homology_colormap")

        add_check_with_tooltip(b_interior2, "Auto-fit Color Range to visible hits", "autofit_color_range",
                               self.update_plot)
        ttk.Separator(b_interior2).pack(fill="x", pady=5)
        ttk.Label(b_interior2, text="BLAST Ring Geometry", font=("Helvetica", 9, "bold")).pack(anchor="w")
        self._add_slider_with_entry(b_interior2, "BLAST Ring Radius",
                                    self.plot_params["blast_ring_radius"], 0.8, 1.5,
                                    tooltip_text=self.plot_tooltips.get("blast_ring_radius"))
        self._add_slider_with_entry(b_interior2, "BLAST Ring Thickness",
                                    self.plot_params["blast_ring_thickness"], 0.05,
                                    0.5, command=lambda e: self._on_manual_ring_adjustment(),
                                    tooltip_text=self.plot_tooltips.get("blast_ring_thickness"))
        add_check_with_tooltip(b_interior2, "Show BLAST Hit Separators", "show_blast_separators",
                               self.update_plot)
        label_section = CollapsibleFrame(parent, text="Label & Connector Formatting", expanded=True,
                                         style_prefix=self.style_prefix)
        label_section.pack(fill="x", pady=5)
        l_interior = label_section.interior
        ttk.Label(l_interior, text="Layout & Spacing", font=("Helvetica", 9, "bold")).pack(anchor="w")
        add_check_with_tooltip(l_interior, "Use Smart Label Layout", "use_smart_layout",
                               self._on_smart_layout_toggle)
        add_check_with_tooltip(l_interior, "Auto-Fit Label Font Size", "autofit_label_fontsize",
                               self._on_autofit_fontsize_toggle)
        self._add_slider_with_entry(l_interior, "Label Spread",
                                    self.plot_params["label_spread"], 0.2, 2.5,
                                    tooltip_text=self.plot_tooltips.get("label_spread"))
        self._add_slider_with_entry(l_interior, "Label Distance from Plot",
                                    self.plot_params["label_distance_factor"],
                                    0.000, 1.5, tooltip_text=self.plot_tooltips.get("label_distance_factor"))
        self.fontsize_slider = self._add_slider_with_entry(
            l_interior, "Label Font Size", self.plot_params["label_fontsize"], 5, 16,
            command=self._on_manual_fontsize_change, tooltip_text=self.plot_tooltips.get("label_fontsize")
        )
        ttk.Separator(l_interior).pack(fill="x", pady=5)
        ttk.Label(l_interior, text="Appearance", font=("Helvetica", 9, "bold")).pack(anchor="w")
        add_check_with_tooltip(l_interior, "Show Label Boxes", "show_label_boxes", self.update_plot)
        add_check_with_tooltip(l_interior, "Color Connector Lines by Homology", "color_lines_by_homology",
                               self.update_plot)
        add_check_with_tooltip(l_interior, "Show Connector Dots", "show_connector_dots", self.update_plot)
        add_check_with_tooltip(l_interior, "Color Connector Dots by Homology", "color_dots_by_homology",
                               self.update_plot)
        ttk.Separator(l_interior).pack(fill="x", pady=5)
        ttk.Label(l_interior, text="Connector Line Geometry", font=("Helvetica", 9, "bold")).pack(anchor="w")
        add_check_with_tooltip(l_interior, "Use Curved Connectors", "connect_dots_with_curve",
                               self.update_plot)
        self._add_slider_with_entry(l_interior, "Curve Tension",
                                    self.plot_params["curve_tension"], 0.0, 2.5,
                                    tooltip_text=self.plot_tooltips.get("curve_tension"))
        self._add_slider_with_entry(l_interior, "Label Radial Line Length",
                                    self.plot_params["label_line_radial_len"],
                                    1.01, 1.5, tooltip_text=self.plot_tooltips.get("label_line_radial_len"))
        self._add_slider_with_entry(l_interior, "Label Horizontal Line",
                                    self.plot_params["label_line_horizontal_len"],
                                    0.0, 0.5,
                                    tooltip_text=self.plot_tooltips.get("label_line_horizontal_len"))
        self.inner_rings_section = CollapsibleFrame(parent, text="Inner Rings", expanded=True,
                                                    style_prefix=self.style_prefix)
        self.inner_rings_section.pack(fill="x", pady=5)
        self.inner_rings_section.toggle_button.unbind("<Button-1>")
        self.inner_rings_section.toggle_button.bind("<Button-1>", self.inner_rings_section._toggle)
        self.identity_section = CollapsibleFrame(self.inner_rings_section.interior, text="Identity Ring",
                                                 expanded=True, style_prefix=self.style_prefix)
        self.identity_section.pack(fill="x", pady=(0, 5), padx=5)
        i_interior = self.identity_section.interior
        algo_options = ["Mauve + SibeliaZ Fallback", "Mauve", "SibeliaZ", "Legacy (Global k-mer %)",
                        "Legacy (1:1 Base %)"]
        id_algo_frame = tk.Frame(i_interior)
        id_algo_frame.pack(fill="x", pady=2, padx=5)
        lbl_algo = tk.Label(id_algo_frame, text="Algorithm", width=22, anchor='w')
        lbl_algo.pack(side="left")
        if "identity_algorithm" in self.plot_tooltips:
            _create_tooltip(lbl_algo, self.plot_tooltips["identity_algorithm"])
        id_algo_combo = ttk.Combobox(id_algo_frame, textvariable=self.plot_params["identity_algorithm"],
                                     values=algo_options, state='readonly')
        id_algo_combo.pack(side="left", expand=True, fill="x", padx=(5, 0))
        if "identity_algorithm" in self.plot_tooltips:
            _create_tooltip(id_algo_combo, self.plot_tooltips["identity_algorithm"])
        id_algo_combo.bind("<<ComboboxSelected>>", lambda e: self._switch_identity_algorithm())
        add_check_with_tooltip(i_interior, 'Show Raw Alignment Blocks', "id_ring_precise_mode",
                               self._on_toggle_precise_mode)
        add_check_with_tooltip(i_interior, "Include Coding Regions in Analysis", "id_ring_include_coding",
                               self.update_plot)
        add_check_with_tooltip(i_interior, "Show Non-Coding Regions", "id_ring_show_noncoding",
                               self.update_plot)
        add_check_with_tooltip(i_interior, 'Use "Color-Only" Mode', "id_ring_color_only",
                               self._on_id_color_only_toggle)
        add_check_with_tooltip(i_interior, "Auto-fit Color Range", "nc_auto_fit_color", self.update_plot)
        add_dropdown(i_interior, "Identity Ring Colormap", self.plot_params["nc_colormap"], full_cmaps,
                     tooltip_key="nc_colormap")
        def on_id_change(e):
            if not self.plot_params["id_ring_color_only"].get(): self.plot_params["auto_thickness_rings"].set(False)
            self.update_plot()
        self.id_radius_slider = self._add_slider_with_entry(i_interior, "Identity Ring Radius",
                                                            self.plot_params["id_ring_radius"], 0.1, 1.5,
                                                            command=on_id_change,
                                                            tooltip_text=self.plot_tooltips.get("id_ring_radius"))
        self.id_thickness_slider = self._add_slider_with_entry(i_interior, "Identity Ring Thickness",
                                                               self.plot_params["id_ring_thickness"], 0.05,
                                                               1.0, command=on_id_change,
                                                               tooltip_text=self.plot_tooltips.get("id_ring_thickness"))
        self._add_slider_with_entry(i_interior, "ID Baseline Transparency (%)",
                                    self.plot_params["id_baseline_transparency_pct"], 0, 100,
                                    tooltip_text=self.plot_tooltips.get("id_baseline_transparency_pct"))
        id_ring_calc_frame = tk.Frame(i_interior)
        id_ring_calc_frame.pack(fill="x", pady=(5, 0), padx=5)
        lbl_win_sz = tk.Label(id_ring_calc_frame, text="Window Size (bp):", anchor='w')
        lbl_win_sz.pack(side="left")
        if "nc_window_size" in self.plot_tooltips:
            _create_tooltip(lbl_win_sz, self.plot_tooltips["nc_window_size"])

        entry_win_sz = ttk.Entry(id_ring_calc_frame, textvariable=self.plot_params['nc_window_size'], width=8)
        entry_win_sz.pack(side="left", padx=5)
        if "nc_window_size" in self.plot_tooltips:
            _create_tooltip(entry_win_sz, self.plot_tooltips["nc_window_size"])
        entry_win_sz.bind("<FocusOut>", _create_bounded_entry_validator(self.plot_params['nc_window_size'],
                                                                        10, 100000,
                                                                        "NC Window Size"))
        lbl_step_sz = tk.Label(id_ring_calc_frame, text="Step Size (bp):", anchor='w')
        lbl_step_sz.pack(side="left", padx=5)
        if "nc_step_size" in self.plot_tooltips:
            _create_tooltip(lbl_step_sz, self.plot_tooltips["nc_step_size"])
        entry_step_sz = ttk.Entry(id_ring_calc_frame, textvariable=self.plot_params['nc_step_size'], width=8)
        entry_step_sz.pack(side="left", padx=5)
        if "nc_step_size" in self.plot_tooltips:
            _create_tooltip(entry_step_sz, self.plot_tooltips["nc_step_size"])
        entry_step_sz.bind("<FocusOut>",
                           _create_bounded_entry_validator(self.plot_params['nc_step_size'], 1, 100000, "NC Step Size"))
        ttk.Button(i_interior, text="Recalculate Identity Ring",
                   command=self._recalculate_identity_and_update).pack(pady=10, fill="x", padx=5)
        self.skew_section = CollapsibleFrame(self.inner_rings_section.interior, text="GC Skew Ring", expanded=True,
                                             style_prefix=self.style_prefix)
        self.skew_section.pack(fill="x", pady=(0, 5), padx=5)
        s_interior = self.skew_section.interior

        def on_gc_thickness_change(event=None):
            """When user manually changes GC thickness, disable auto-thickness for GC."""
            self.plot_params["auto_thickness_rings"].set(False)
            self.update_plot()

        def on_skew_thickness_change(event=None):
            """When user manually changes Skew thickness, disable auto-thickness for Skew."""
            self.plot_params["auto_thickness_rings"].set(False)
            self.update_plot()

        self.skew_thickness_slider = self._add_slider_with_entry(s_interior, "GC Skew Thickness",
                                                                 self.plot_params["skew_thick"], 0.001, 1.0,
                                                                 command=on_skew_thickness_change,
                                                                 tooltip_text=self.plot_tooltips.get("skew_thick"))
        self.skew_inner_radius_slider = self._add_slider_with_entry(s_interior, "GC Skew Inner Radius",
                                                                    self.plot_params["skew_inner_radius"], 0.34,
                                                                    1.35,
                                                                    command=lambda e: self._on_manual_ring_adjustment(),
                                                                    tooltip_text=self.plot_tooltips.get(
                                                                        "skew_inner_radius"))
        self._add_slider_with_entry(s_interior, "Baseline Transparency (%)",
                                    self.plot_params["skew_baseline_transparency_pct"], 0, 100,
                                    tooltip_text=self.plot_tooltips.get("skew_baseline_transparency_pct"))

        gc_section = CollapsibleFrame(self.inner_rings_section.interior, text="GC Content Ring", expanded=True,
                                      style_prefix=self.style_prefix)
        gc_section.pack(fill="x", pady=(0, 5), padx=5)
        gc_interior = gc_section.interior
        gc_colormap_options = ["Grey"] + colormap_options + [f"{c} (Reversed)" for c in colormap_options]
        add_check_with_tooltip(gc_interior, 'Use "Color-Only" Mode', "gc_ring_color_only",
                               self._on_gc_color_only_toggle)
        add_dropdown(gc_interior, "GC Content Colormap", self.plot_params["gc_colormap"], gc_colormap_options,
                     tooltip_key="gc_colormap")
        self.gc_thickness_slider = self._add_slider_with_entry(
            gc_interior, "GC Content Thickness", self.plot_params["gc_thick"], 0.1, 2.0,
            command=on_gc_thickness_change,
            tooltip_text=self.plot_tooltips.get("gc_thick"))
        self.gc_inner_radius_slider = self._add_slider_with_entry(gc_interior, "GC Content Inner Radius",
                                                                  self.plot_params["gc_inner"], 0.1, 0.9,
                                                                  command=lambda e: self._on_manual_ring_adjustment(),
                                                                  tooltip_text=self.plot_tooltips.get("gc_inner"))
        self._add_slider_with_entry(gc_interior, "Baseline Transparency (%)",
                                    self.plot_params["gc_baseline_transparency_pct"], 0, 100,
                                    tooltip_text=self.plot_tooltips.get("gc_baseline_transparency_pct"))
        ttk.Separator(self.inner_rings_section.interior).pack(fill="x", pady=5, padx=5)
        gc_recalc_frame = tk.Frame(self.inner_rings_section.interior)
        gc_recalc_frame.pack(fill="x", pady=(5, 5), padx=5)
        tk.Label(gc_recalc_frame, text="GC Window (bp):", anchor='w').pack(side="left")
        gc_entry = ttk.Entry(gc_recalc_frame, textvariable=self.plot_params['gc_window_size'], width=8)
        gc_entry.pack(side="left", padx=5)
        gc_entry.bind("<FocusOut>",
                      _create_bounded_entry_validator(self.plot_params['gc_window_size'],
                                                      10, 100000, "GC Window Size"))
        ttk.Button(gc_recalc_frame, text="Recalculate GC/Skew", command=self._recalculate_gc_and_update).pack(
            side="left", expand=True, fill="x", padx=5)

        export_frame = ttk.LabelFrame(parent, text="Export Options", padding=(10, 5))
        export_frame.pack(pady=20, fill="x")
        res_frame = tk.Frame(export_frame)
        res_frame.pack(fill="x", pady=2)
        tk.Label(res_frame, text="Resolution", width=15, anchor='w').pack(side="left")
        res_combo = ttk.Combobox(res_frame, textvariable=self.plot_params["save_resolution"],
                                 values=list(self.resolution_options.keys()), state='readonly')
        res_combo.pack(side="left", expand=True, fill="x")
        res_combo.bind("<<ComboboxSelected>>", self._on_resolution_change)
        self.custom_res_frame = tk.Frame(export_frame)
        width_frame = tk.Frame(self.custom_res_frame)
        width_frame.pack(fill="x", pady=2)
        tk.Label(width_frame, text="Width (px)", width=15, anchor='w').pack(side="left")
        width_entry = ttk.Entry(width_frame, textvariable=self.plot_params["custom_save_width"], width=10)
        width_entry.pack(side="left")
        width_entry.bind("<FocusOut>",
                         _create_bounded_entry_validator(self.plot_params["custom_save_width"], 100, 20000,
                                                         "Custom Width"))
        height_frame = tk.Frame(self.custom_res_frame)
        height_frame.pack(fill="x", pady=2)
        tk.Label(height_frame, text="Height (px)", width=15, anchor='w').pack(side="left")
        height_entry = ttk.Entry(height_frame, textvariable=self.plot_params["custom_save_height"], width=10)
        height_entry.pack(side="left")
        height_entry.bind("<FocusOut>",
                          _create_bounded_entry_validator(self.plot_params["custom_save_height"], 100, 20000,
                                                          "Custom Height"))
        self._on_resolution_change()
        ttk.Checkbutton(export_frame, text="Include Legacy k-mer Report (Slow)",
                        variable=self.plot_params["run_legacy_report_on_export"]).pack(pady=(5, 5), anchor="w", padx=5)
        button_frame = tk.Frame(parent)
        button_frame.pack(pady=10, fill="x")
        self.export_button = ttk.Button(button_frame, text="Export Homology Report",
                                        command=self.export_homology_report)
        self.export_button.pack(pady=(0, 5), fill="x")
        save_plot_button = ttk.Button(button_frame, text="Save Plot", command=self.save_plot)
        save_plot_button.pack(fill="x")

        labels_frame = ttk.LabelFrame(parent, text="Custom Labels", padding=(10, 5))
        labels_frame.pack(fill="x", pady=(15, 10))
        labels_frame.columnconfigure(1, weight=1)
        labels_frame.columnconfigure(3, weight=1)

        # --- Row 0: Name ---
        tk.Label(labels_frame, text="Name:", width=10, anchor='w').grid(row=0, column=0, sticky="w", padx=2, pady=2)
        self.label_name_var = tk.StringVar()
        ttk.Entry(labels_frame, textvariable=self.label_name_var).grid(row=0, column=1, columnspan=3, sticky="ew",
                                                                       padx=2)

        # --- Row 1: Input ---
        tk.Label(labels_frame, text="Input:", width=10, anchor='w').grid(row=1, column=0, sticky="w", padx=2, pady=2)
        self.label_input_var = tk.StringVar()
        ttk.Entry(labels_frame, textvariable=self.label_input_var).grid(row=1, column=1, columnspan=3, sticky="ew",
                                                                        padx=2)

        # --- Row 2: Color and Font ---
        tk.Label(labels_frame, text="Color:", width=10, anchor='w').grid(row=2, column=0, sticky="w", padx=2, pady=2)
        self.label_color_var = tk.StringVar(value="black")
        color_options = ["black", "red", "blue", "green", "orange", "purple", "brown", "pink", "gray", "cyan"]
        ttk.Combobox(labels_frame, textvariable=self.label_color_var, values=color_options,
                     state="readonly", width=15).grid(row=2, column=1, sticky="ew", padx=2)

        tk.Label(labels_frame, text="Font:", width=6, anchor='w').grid(row=2, column=2, sticky="w", padx=(10, 2),
                                                                       pady=2)
        font_families = sorted(list(tkfont.families()))
        common_fonts = ["sans-serif", "Arial", "Times New Roman", "Courier New", "Verdana"]
        display_fonts = common_fonts + [f for f in font_families if f not in common_fonts]

        ttk.Combobox(labels_frame, textvariable=self.plot_params["label_font_family"], values=display_fonts,
                     state="readonly", width=15).grid(row=2, column=3, sticky="ew", padx=2)

        # --- Row 3: Checkboxes ---
        style_frame = ttk.Frame(labels_frame)
        style_frame.grid(row=3, column=0, columnspan=4, sticky="ew", pady=3)

        ttk.Checkbutton(style_frame, text="Bold", variable=self.plot_params["label_font_bold"]).pack(side="left",
                                                                                                     padx=5)
        ttk.Checkbutton(style_frame, text="Italic", variable=self.plot_params["label_font_italic"]).pack(side="left",
                                                                                                         padx=10)
        ttk.Checkbutton(style_frame, text="Thicker Lines", variable=self.plot_params["label_thicker_lines"]).pack(
            side="left", padx=10)

        # --- Row 4: Add Button ---
        ttk.Button(labels_frame, text="Add Label", command=self._add_custom_label).grid(row=4, column=0, columnspan=4,
                                                                                        sticky="ew", pady=5, padx=2)

        # --- Row 5: Listbox ---
        list_frame = ttk.Frame(labels_frame)
        list_frame.grid(row=5, column=0, columnspan=4, sticky="ew", pady=5, padx=2)
        list_frame.columnconfigure(0, weight=1)
        list_scroll = ttk.Scrollbar(list_frame, orient="vertical")
        list_scroll.grid(row=0, column=1, sticky="ns")
        self.custom_labels_listbox = tk.Listbox(list_frame, yscrollcommand=list_scroll.set, height=5)
        self.custom_labels_listbox.grid(row=0, column=0, sticky="nsew")
        list_scroll.config(command=self.custom_labels_listbox.yview)

        # --- Row 6: Delete Button ---
        ttk.Button(labels_frame, text="Delete Selected", command=self._delete_custom_label).grid(row=6, column=0,
                                                                                                 columnspan=4,
                                                                                                 sticky="ew", pady=2,
                                                                                                 padx=2)

        subsequence_frame = ttk.LabelFrame(parent, text="Export FASTA Subsequence", padding=(10, 5))
        subsequence_frame.pack(fill="x", pady=(15, 5))
        sub_entry_frame = tk.Frame(subsequence_frame)
        sub_entry_frame.pack(fill="x", pady=2, padx=5)
        tk.Label(sub_entry_frame, text="Start:", width=10, anchor='w').pack(side="left")
        sub_start_entry = ttk.Entry(sub_entry_frame, textvariable=self.plot_params["subsequence_start"], width=10)
        sub_start_entry.pack(side="left", padx=(5, 10))
        sub_start_entry.bind("<FocusOut>",
                             _create_positive_validator(self.plot_params["subsequence_start"],
                                                        "Subsequence Start"))
        tk.Label(sub_entry_frame, text="End:", width=10, anchor='w').pack(side="left")
        sub_end_entry = ttk.Entry(sub_entry_frame, textvariable=self.plot_params["subsequence_end"], width=10)
        sub_end_entry.pack(side="left", padx=(5, 0))
        sub_end_entry.bind("<FocusOut>",
                           _create_positive_validator(self.plot_params["subsequence_end"],
                                                      "Subsequence End"))
        sub_button = ttk.Button(subsequence_frame, text="Export Subsequence...", command=self._export_subsequence)
        sub_button.pack(fill="x", pady=(10, 5), padx=5)

    def _update_scrollbar_visibility(self):
        """Checks if the scrollbar is needed and updates its state, avoiding redundant calls."""
        should_be_visible = self.scrollable_frame.winfo_reqheight() > self.controls_canvas.winfo_height()
        if should_be_visible and not self.scrollbar_is_visible:
            self.controls_scrollbar.grid(row=0, column=1, sticky='ns')
            self.scrollbar_is_visible = True
        elif not should_be_visible and self.scrollbar_is_visible:
            self.controls_scrollbar.grid_remove()
            self.scrollbar_is_visible = False

    def _on_frame_configure(self, event):
        """Event handler for when the scrollable frame's content changes size."""
        self.controls_canvas.configure(scrollregion=self.controls_canvas.bbox("all"))
        self._update_scrollbar_visibility()

    def _on_canvas_configure(self, event):
        """
        Event handler for when the canvas itself changes size.
        Updates the scrollable frame's width, checks scrollbar,
        and triggers the font resize check.
        """
        self.controls_canvas.itemconfig(self.scrollable_frame_window, width=event.width)
        self._update_scrollbar_visibility()
        self._on_sidebar_resize_debounced(event)

    def update_plot(self):
        try:
            if hasattr(self, 'colorbars'):
                for cbar, cax in self.colorbars:
                    try:
                        if cax and cax in self.fig.axes: cax.remove()
                    except (KeyError, ValueError) as e: pass

            if hasattr(self, 'legend') and self.legend:
                try: self.legend.remove()
                except (KeyError, ValueError) as e: pass

            params = {key: var.get() for key, var in self.plot_params.items()}
            params['is_initial_load'] = not self.initial_load_complete
            if params.get("limit_to_region", False):
                try:
                    region_start = int(params.get("region_start"))
                    region_end = int(params.get("region_end"))
                    new_region_start = region_start
                    new_region_end = region_end
                    all_features = self.data.get('gene_feats', [])

                    for f in all_features:
                        if f.start < region_end < f.end: new_region_end = max(new_region_end, f.end)
                        if f.start < region_start < f.end: new_region_start = min(new_region_start, f.start)

                    full_seqlen = self.data.get("seqlen", new_region_end)

                    if new_region_start != region_start:
                        final_start = max(1, new_region_start)  # ✅ 1-based minimum
                        self.plot_params["region_start"].set(str(final_start))
                        params["region_start"] = str(final_start)  # Update params dict too
                        print(f"[DEBUG EXPANSION] Region start expanded: {region_start} -> {final_start}")

                    if new_region_end != region_end:
                        final_end = min(full_seqlen, new_region_end)  # ✅ 1-based
                        self.plot_params["region_end"].set(str(final_end))
                        params["region_end"] = str(final_end)  # Update params dict too
                        print(f"[DEBUG EXPANSION] Region end expanded: {region_end} -> {final_end}")

                except (ValueError, TypeError): pass

            # Use the updated region values after auto-expansion (not the original params)
            current_limited = params.get("limit_to_region", False)
            current_start = self.plot_params.get("region_start").get() if current_limited else ""
            current_end = self.plot_params.get("region_end").get() if current_limited else ""

            val = params.get("manual_rot_deg", 0.0)
            current_rotation = float(val) if val is not None else 0.0
            last_rotation_float = float(self.last_rotation) if self.last_rotation is not None else 0.0
            current_gc_inner = params.get("gc_inner", 0.3)
            current_skew_inner = params.get("skew_inner_radius", 0.5)
            if not hasattr(self, 'last_gc_inner'): self.last_gc_inner = current_gc_inner
            if not hasattr(self, 'last_skew_inner'): self.last_skew_inner = current_skew_inner

            inner_radius_changed = (current_gc_inner != self.last_gc_inner or
                                    current_skew_inner != self.last_skew_inner)

            # Check if autofit or smart layout is enabled (requires recalculation for proper label sizing)
            autofit_enabled = params.get("autofit_label_fontsize", False)
            smart_layout_enabled = params.get("use_smart_layout", False)

            region_changed = (
                    current_limited != self.last_region_limited or
                    current_start != self.last_region_start or
                    current_end != self.last_region_end or
                    not self.initial_load_complete or
                    (current_limited and abs(current_rotation - last_rotation_float) > 0.1) or
                    inner_radius_changed
            )

            if (autofit_enabled or smart_layout_enabled) and region_changed: recalculate_labels = True
            else: recalculate_labels = region_changed

            params['cached_x_center_pan'] = self.cached_x_center_pan
            params['cached_y_center_pan'] = self.cached_y_center_pan
            params['cached_xlim'] = self.cached_xlim if hasattr(self, 'cached_xlim') else None
            params['cached_ylim'] = self.cached_ylim if hasattr(self, 'cached_ylim') else None
            params['cached_auto_zoom_scalar'] = self.cached_auto_zoom_scalar if hasattr(self,
                                                                                        'cached_auto_zoom_scalar') else 1.0

            (fig, self.legend, legend_elements, self.colorbars, calculated_values,
             calculated_pan, limits_before_pass2, hover_data) = _draw_plot(
                self.ax, self.data, params, recalculate_auto_view=recalculate_labels
            )

            self.graphic_ref = hover_data["graphic_ref"]
            self.last_calculated_values = hover_data["last_calculated_values"]
            self.last_calculated_values["id_label"] = hover_data["id_label"]
            self.gene_feature_patches = hover_data["gene_feature_patches"]
            self.identity_ring_artists = hover_data["identity_ring_artists"]
            self.scores_plotted_for_hover = hover_data["scores_plotted_for_hover"]

            if region_changed:
                self.cached_x_center_pan, self.cached_y_center_pan = calculated_pan
                self.last_region_limited = current_limited
                self.last_region_start = current_start
                self.last_region_end = current_end
                self.last_rotation = current_rotation
                self.last_gc_inner = current_gc_inner
                self.last_skew_inner = current_skew_inner
                self.cached_xlim, self.cached_ylim = limits_before_pass2
                self.cached_auto_zoom_scalar = calculated_values.get('auto_zoom_scalar', 1.0)

            if not self.initial_load_complete:
                self._set_slider_value_with_range_adjustment(
                    self.id_thickness_slider if hasattr(self, 'id_thickness_slider') else None,
                    self.plot_params["id_ring_thickness"],
                    calculated_values["id_thickness"])
                self._set_slider_value_with_range_adjustment(
                    self.gc_thickness_slider,
                    self.plot_params["gc_thick"],
                    calculated_values["gc_thickness"])
                self._set_slider_value_with_range_adjustment(
                    self.skew_thickness_slider,
                    self.plot_params["skew_thick"],
                    calculated_values["skew_thickness"])
                self.initial_gc_thick = calculated_values["gc_thickness"]
                self.initial_load_complete = True

            if self.plot_params["autofit_label_fontsize"].get():
                calculated_fontsize = calculated_values.get("label_fontsize", self.plot_params["label_fontsize"].get())
                self.plot_params["label_fontsize"].set(calculated_fontsize)

            if self.initial_load_complete and self.plot_params.get("auto_thickness_rings",
                                                                   tk.BooleanVar(value=True)).get():
                self._set_slider_value_with_range_adjustment(
                    self.gc_thickness_slider,
                    self.plot_params["gc_thick"],
                    calculated_values.get("gc_thickness", self.plot_params["gc_thick"].get()))
                self._set_slider_value_with_range_adjustment(
                    self.skew_thickness_slider,
                    self.plot_params["skew_thick"],
                    calculated_values.get("skew_thickness", self.plot_params["skew_thick"].get()))

            if self.colorbars:
                for cbar, cax in self.colorbars: cax.set_zorder(-1)
            self.canvas.draw()
        except Exception as e:
            print(f"Error updating plot: {e}")
            traceback.print_exc()

    def save_plot(self):
        filepath = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[
                ("PNG", "*.png"),
                ("TIFF", "*.tiff"),
                ("SVG", "*.svg"),
                ("JPEG", "*.jpg"),
                ("PDF", "*.pdf"),
                ("EPS", "*.eps"),
                ("All files", "*.*")
            ],
            initialfile="interactive_plot.png"
        )
        if not filepath: return

        resolution_key = str(self.plot_params["save_resolution"].get())
        if resolution_key == "Custom Resolution":
            try:
                max_width = int(self.plot_params["custom_save_width"].get())
                max_height = int(self.plot_params["custom_save_height"].get())
                if max_width <= 0 or max_height <= 0: raise ValueError("Dimensions must be positive.")
            except (ValueError, TypeError):
                messagebox.showerror("Invalid Input", "Please enter valid integer values for custom width and height.")
                return
        else: max_width, max_height = self.resolution_options[resolution_key]

        try:
            target_w, target_h = _save_figure_with_aspect_ratio(
                self.fig, self.ax, filepath, max_width, max_height)
            messagebox.showinfo("Success", f"Plot (Target: {target_w}x{target_h}px) saved to:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save plot:\n{e}")
            traceback.print_exc()

    def export_homology_report(self):
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel file", "*.xlsx")], initialfile="homology_report.xlsx")
        if not filepath: return
        run_legacy = self.plot_params["run_legacy_report_on_export"].get()
        if not run_legacy:
            self._perform_excel_save(filepath, include_legacy=False)
            return

        if self.data.get("legacy_data", {}).get("lcb_data"):
            self.master_app.status_label.config(text="Exporting with cached legacy data...", fg="blue")
            self._perform_excel_save(filepath, include_legacy=True)
            return

        self.master_app.status_label.config(text="Running legacy k-mer report (can be slow)...", fg="orange")
        if self.export_button: self.export_button.config(state="disabled")
        self.master_app.log_queue.put(("run_legacy_for_export", filepath))

    def _export_subsequence(self):
        """
        Export a subsequence (1-indexed range) to FASTA format.

        COORDINATE CONVENTION:
        - User input: 1-indexed inclusive (e.g., 100-200 means positions 100, 101, ..., 200)
        - Validation: Checks that start >= 1 and end <= seqlen
        - Storage: Keeps as 1-indexed for display in filename and header
        - Extraction: Converts to 0-indexed for Python string slicing
        """
        try:
            start_val = int(self.plot_params["subsequence_start"].get())  # 1-indexed
            end_val = int(self.plot_params["subsequence_end"].get())  # 1-indexed
        except (ValueError, tk.TclError):
            messagebox.showerror("Invalid Input", "Start and End values must be integers.")
            return

        if not self.data:
            messagebox.showerror("Error", "Sequence data not found.")
            return

        query_record = self.data.get("record")
        if not query_record:
            messagebox.showerror("Error", "Query record not found in data.")
            return

        full_seq = str(query_record.seq)
        seq_len = len(full_seq)

        is_out_of_bounds = False
        if start_val < 1 or end_val > seq_len: is_out_of_bounds = True
        if start_val > end_val:
            messagebox.showerror("Invalid Input",
                                 f"Start value ({start_val}) must be less than or equal to End value ({end_val}).")
            return

        extract_start = max(0, start_val - 1)
        extract_end = min(seq_len, end_val)
        original_fasta_path = Path(self.master_app.vars["fasta"].get())
        clean_stem = "".join(c if c.isalnum() or c == '-' else '_' for c in original_fasta_path.stem)
        default_filename = f"{clean_stem}-{start_val}-{end_val}.fasta"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".fasta",
            filetypes=[("FASTA file", "*.fasta"), ("All files", "*.*")],
            initialfile=default_filename
        )

        if not filepath: return

        try:
            subsequence = full_seq[extract_start:extract_end]
            new_header = f">{query_record.id} [subsequence {start_val}-{end_val}]"
            with open(filepath, 'w') as f:
                f.write(f"{new_header}\n")
                line_width = 70
                for i in range(0, len(subsequence), line_width):
                    f.write(f"{subsequence[i:i + line_width]}\n")
            if is_out_of_bounds:
                messagebox.showwarning("Values Clamped",
                                       f"Warning: Your values were outside the valid range (1-{seq_len}).\n"
                                       f"The subsequence from {start_val} to {end_val} was exported to:\n{filepath}")
            else:
                messagebox.showinfo("Success",
                                    f"Subsequence ({start_val}-{end_val}) saved to:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to write FASTA file: {e}")
            traceback.print_exc()

    def _validate_and_parse_label_input(self, text, query_seq, seqlen):
        """
        Validates and parses label input.
        Raises ValueError with a user-friendly message on failure.

        Returns 1-indexed (start, end) tuples in positions list.
        """
        if not text: raise ValueError("Input cannot be empty")

        cleaned = text.replace(" ", "").replace("#", "")
        normalized_text = "".join(cleaned.split())
        allowed_chars = set("atcgnATCGN" + "0123456789" + "-")
        if not all(char in allowed_chars for char in normalized_text):
            raise ValueError("Invalid characters in input. Only A, T, C, G, N, numbers, and hyphens are allowed.")
        if any(c in "atcgnATCGN" for c in normalized_text):
            if not all(c in "atcgnATCGN" for c in normalized_text):
                raise ValueError(
                    "Invalid sequence. Sequences can only contain A, T, C, G, or N, "
                    "and cannot be mixed with numbers or hyphens.")

            sequence_upper = normalized_text.upper()
            positions = []
            seq_upper = query_seq.upper()
            start = 0
            while True:
                idx = seq_upper.find(sequence_upper, start)
                if idx == -1: break
                positions.append((idx + 1, idx + len(sequence_upper)))
                start = idx + 1
            if not positions: raise ValueError("Sequence not found in query genome")
            return 'sequence', positions, sequence_upper

        if '-' in normalized_text:
            parts = normalized_text.split('-')
            if len(parts) == 2:
                try:
                    start, end = int(parts[0]), int(parts[1])
                    if start > end: start, end = end, start
                    if start < 1 or end > seqlen:
                        raise ValueError(f"Coordinates are out of bounds. Must be between 1 and {seqlen}.")
                    return 'range', [(start, end)], None
                except ValueError as e:
                    if "out of bounds" in str(e): raise e
                    raise ValueError("Invalid range format. Use: start-end (e.g., 1000-2000)")
            else: raise ValueError("Invalid range format. Use: start-end (e.g., 1000-2000)")
        try:
            coord = int(normalized_text)
            if coord < 1 or coord > seqlen:
                raise ValueError(f"Coordinate is out of bounds. Must be between 1 and {seqlen}.")
            return 'coordinate', [(coord, coord)], None
        except ValueError as e:
            if "out of bounds" in str(e): raise e
            raise ValueError("Invalid input. Use: coordinate (123), range (100-200), or sequence (ATCG)")

    def _add_custom_label(self):
        """Add a custom label from user input after validation."""
        name = self.label_name_var.get().strip()
        input_text = self.label_input_var.get().strip()
        color = self.label_color_var.get()

        if not name:
            messagebox.showwarning("Missing Name", "Please enter a label name.")
            return
        if not input_text:
            messagebox.showwarning("Missing Input", "Please enter coordinates, range, or sequence.")
            return
        try:
            query_seq = str(self.data['record'].seq)
            seqlen = self.data['seqlen']
            label_type, positions, sequence = self._validate_and_parse_label_input(input_text, query_seq, seqlen)
            if len(positions) >= 5:
                match_type = 'sequence' if label_type == 'sequence' else 'input'
                msg = (f"Found {len(positions)} matches for this {match_type}.\n\n"
                       f"Do you want to add labels at all {len(positions)} positions?")

                response = messagebox.askyesno("Multiple Matches", msg)
                if not response:
                    return

            font_family = self.plot_params["label_font_family"].get()
            is_bold = self.plot_params["label_font_bold"].get()
            is_italic = self.plot_params["label_font_italic"].get()
            is_thicker = self.plot_params["label_thicker_lines"].get()

            label_entry = {
                'name': name,
                'type': label_type,
                'positions': positions,
                'sequence': sequence,
                'color': color,
                'visible': True,
                'font_family': font_family,
                'is_bold': is_bold,
                'is_italic': is_italic,
                'is_thicker': is_thicker
            }

            self.data['custom_labels'].append(label_entry)

            if label_type == 'sequence':
                display_text = (f"{name} (seq: {sequence[:10]}{'...' if len(sequence) > 10 else ''}, "
                                f"{len(positions)} match{'es' if len(positions) > 1 else ''})")
            elif label_type == 'range': display_text = f"{name} ({positions[0][0]}-{positions[0][1]})"
            else: display_text = f"{name} ({positions[0][0]})"
            self.custom_labels_listbox.insert(tk.END, display_text)
            self.label_name_var.set("")
            self.label_input_var.set("")
            self.update_plot()
        except ValueError as e: messagebox.showwarning("Invalid Input", str(e))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add label: {e}")
            traceback.print_exc()

    def _delete_custom_label(self):
        """Delete selected label from list."""
        selection = self.custom_labels_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a label to delete.")
            return
        idx = selection[0]
        del self.data['custom_labels'][idx]
        self.custom_labels_listbox.delete(idx)
        self.update_plot()

    def _perform_excel_save(self, filepath, include_legacy=True):
        try:
            homology_threshold = self.plot_params["homology_thresh"].get()
            homology_hits = []
            for f in self.data["gene_feats"]:
                if f.label and getattr(f, 'best_identity', 0) >= homology_threshold:
                    gene_name = _nice_label(f.label, 100)
                    # Extract gene ID from feature attributes
                    gene_id = "N/A"
                    attrs = getattr(f, "attributes", {}) or {}
                    gene_id_value = _get_attr_value(attrs, "ID")
                    if gene_id_value:
                        gene_id = gene_id_value
                    query_start = int(f.start)
                    query_stop = int(f.end)
                    ref_start = min(getattr(f, 'best_sstart', 0), getattr(f, 'best_send', 0))
                    ref_stop = max(getattr(f, 'best_sstart', 0), getattr(f, 'best_send', 0))
                    identity = f"{f.best_identity:.2f}"
                    query_accession = getattr(f, 'best_qseqid', 'N/A')
                    reference_accession = getattr(f, 'best_sseqid', 'N/A')
                    homology_hits.append((gene_name, gene_id, query_start, query_stop, ref_start, ref_stop, identity, query_accession, reference_accession))
            homology_df = pd.DataFrame(sorted(homology_hits, key=lambda x: x[2]),
                                       columns=["Gene Name", "ID", "Query Start", "Query Stop", "Reference Start", "Reference Stop", "Identity (%)",
                                                "Query Accession", "Reference Accession"])
            enriched_blast_df = self.data.get("blast_df", pd.DataFrame())
            blast_all_df = _format_blast_df_for_excel(enriched_blast_df)
            blast_intergenic_df = _format_blast_df_for_excel(
                enriched_blast_df[~enriched_blast_df['is_coding']]
            )
            def process_coding_lcb_df(lcb_list):
                """Process coding LCB data with gene_name as first column."""
                df = pd.DataFrame(lcb_list)
                if not df.empty:
                    df = df.rename(
                        columns={
                            "gene_name": "Gene Name",
                            "gene_id": "Gene ID",
                            "query_start": "Query Start",
                            "query_end": "Query End",
                            "query_strand": "Query Strand",
                            "ref_start": "Ref Start",
                            "ref_stop": "Ref Stop",
                            "ref_strand": "Ref Strand",
                            "identity_excl_gaps": "Identity (%) Excl Gaps",
                            "identity_incl_gaps": "Identity (%) Incl Gaps",
                            "original_query_sequence": "Original Query Sequence",
                            "original_reference_sequence": "Original Reference Sequence",
                            "aligned_query_sequence": "Aligned Query Sequence",
                            "aligned_reference_sequence": "Aligned Reference Sequence",
                        }
                    )
                    if "Query Strand" in df.columns:
                        df["Query Strand"] = df["Query Strand"].replace(
                            {1: "+", -1: "-", "1": "+", "-1": "-"}
                        )
                    if "Ref Strand" in df.columns:
                        df["Ref Strand"] = df["Ref Strand"].replace(
                            {1: "+", -1: "-", "1": "+", "-1": "-"}
                        )
                    cols = [
                        "Gene Name",
                        "Gene ID",
                        "Query Start",
                        "Query End",
                        "Query Strand",
                        "Ref Start",
                        "Ref Stop",
                        "Ref Strand",
                        "Original Query Sequence",
                        "Original Reference Sequence",
                        "Identity (%) Excl Gaps",
                        "Identity (%) Incl Gaps",
                        "Aligned Query Sequence",
                        "Aligned Reference Sequence",
                    ]
                    if "difference_string" in df.columns:
                        df = df.rename(
                            columns={
                                "difference_string": "Difference String",
                                "query_mismatches_marked": "Query with Mismatches (#)",
                                "query_differences_only": "Query Differences Only",
                                "ref_differences_only": "Reference Differences Only",
                            }
                        )
                        cols.extend(
                            [
                                "Difference String",
                                "Query with Mismatches (#)",
                                "Query Differences Only",
                                "Reference Differences Only",
                            ]
                        )

                    final_cols = [c for c in cols if c in df.columns]
                    df = df[final_cols].copy()
                    df = df.sort_values(by="Query Start").reset_index(drop=True)
                return df

            def process_lcb_df(lcb_list):
                df = pd.DataFrame(lcb_list)
                if not df.empty:
                    df = df.rename(
                        columns={
                            "query_start": "Query Start",
                            "query_end": "Query End",
                            "query_strand": "Query Strand",
                            "ref_start": "Ref Start",
                            "ref_stop": "Ref Stop",
                            "ref_strand": "Ref Strand",
                            "identity_excl_gaps": "Identity (%) Excl Gaps",
                            "identity_incl_gaps": "Identity (%) Incl Gaps",
                            "original_query_sequence": "Original Query Sequence",
                            "original_reference_sequence": "Original Reference Sequence",
                            "aligned_query_sequence": "Aligned Query Sequence",
                            "aligned_reference_sequence": "Aligned Reference Sequence",
                        }
                    )

                    if "Query Strand" in df.columns:
                        df["Query Strand"] = df["Query Strand"].replace(
                            {1: "+", -1: "-", "1": "+", "-1": "-"}
                        )
                    if "Ref Strand" in df.columns:
                        df["Ref Strand"] = df["Ref Strand"].replace(
                            {1: "+", -1: "-", "1": "+", "-1": "-"}
                        )
                    cols = [
                        "Query Start",
                        "Query End",
                        "Query Strand",
                        "Ref Start",
                        "Ref Stop",
                        "Ref Strand",
                        "Original Query Sequence",
                        "Original Reference Sequence",
                        "Identity (%) Excl Gaps",
                        "Identity (%) Incl Gaps",
                        "Aligned Query Sequence",
                        "Aligned Reference Sequence",
                    ]
                    if "difference_string" in df.columns:
                        df = df.rename(
                            columns={
                                "difference_string": "Difference String",
                                "query_mismatches_marked": "Query with Mismatches (#)",
                                "query_differences_only": "Query Differences Only",
                                "ref_differences_only": "Reference Differences Only",
                            }
                        )
                        cols.extend(
                            [
                                "Difference String",
                                "Query with Mismatches (#)",
                                "Query Differences Only",
                                "Reference Differences Only",
                            ]
                        )

                    final_cols = [c for c in cols if c in df.columns]
                    df = df[final_cols].copy()
                    df = df.sort_values(by="Query Start").reset_index(drop=True)
                return df

            mauve_lcb_df = process_lcb_df(self.data.get("mauve_data", {}).get("lcb_data", []))

            mauve_coding_raw = self.data.get("mauve_data", {}).get("coding_lcb_data", [])
            mauve_coding_lcb_df = process_coding_lcb_df(mauve_coding_raw)
            mauve_intergenic_raw = self.data.get("mauve_data", {}).get("intergenic_lcb_data", [])
            mauve_intergenic_lcb_df = process_lcb_df(mauve_intergenic_raw)
            sibeliaz_lcb_df = process_lcb_df(self.data.get("sibeliaz_data", {}).get("lcb_data", []))
            sibeliaz_coding_raw = self.data.get("sibeliaz_data", {}).get("coding_lcb_data", [])
            sibeliaz_coding_lcb_df = process_coding_lcb_df(sibeliaz_coding_raw)

            sibeliaz_intergenic_lcb_df = process_lcb_df(
                self.data.get("sibeliaz_data", {}).get("intergenic_lcb_data", [])
            )
            combined_intergenic_data = []

            mauve_intergenic = self.data.get("mauve_data", {}).get("intergenic_lcb_data", [])
            for lcb in mauve_intergenic:
                lcb_copy = lcb.copy()
                lcb_copy['algorithm'] = 'Mauve'
                combined_intergenic_data.append(lcb_copy)

            sibeliaz_intergenic = self.data.get("sibeliaz_data", {}).get("intergenic_lcb_data", [])
            for lcb in sibeliaz_intergenic:
                lcb_copy = lcb.copy()
                lcb_copy['algorithm'] = 'SibeliaZ'
                combined_intergenic_data.append(lcb_copy)

            # Standardize BLAST column names to match Mauve/SibeliaZ before combining
            blast_intergenic_records = blast_intergenic_df.to_dict('records')
            for record in blast_intergenic_records:
                record['algorithm'] = 'BLAST'
                # Convert capitalized column names back to lowercase with underscores to match Mauve/SibeliaZ format
                if 'Original Query Sequence' in record:
                    record['original_query_sequence'] = record.pop('Original Query Sequence')
                if 'Original Reference Sequence' in record:
                    record['original_reference_sequence'] = record.pop('Original Reference Sequence')
                if 'Difference String' in record:
                    record['difference_string'] = record.pop('Difference String')
                if 'Query with Mismatches (#)' in record:
                    record['query_mismatches_marked'] = record.pop('Query with Mismatches (#)')
                if 'Query Differences Only' in record:
                    record['query_differences_only'] = record.pop('Query Differences Only')
                if 'Reference Differences Only' in record:
                    record['ref_differences_only'] = record.pop('Reference Differences Only')
                if 'Aligned Query Sequence' in record:
                    record['aligned_query_sequence'] = record.pop('Aligned Query Sequence')
                if 'Aligned Reference Sequence' in record:
                    record['aligned_reference_sequence'] = record.pop('Aligned Reference Sequence')
                combined_intergenic_data.append(record)

            legacy_lcb_df, legacy_intergenic_lcb_df = pd.DataFrame(), pd.DataFrame()

            if include_legacy:
                _compare_legacy_to_wga(
                    self.data.get("legacy_data", {}).get("intergenic_lcb_data", []),
                    self.data.get("mauve_data", {}).get("intergenic_lcb_data", []),
                    self.data.get("sibeliaz_data", {}).get("intergenic_lcb_data", [])
                )

                legacy_lcb_df = process_lcb_df(self.data.get("legacy_data", {}).get("lcb_data", []))
                legacy_intergenic_lcb_df = process_lcb_df(
                    self.data.get("legacy_data", {}).get("intergenic_lcb_data", []))
                if not legacy_intergenic_lcb_df.empty:
                    legacy_intergenic_lcb_df['sequence_length'] = legacy_intergenic_lcb_df[
                        'aligned_query_sequence'].str.len()
                    legacy_intergenic_lcb_df = legacy_intergenic_lcb_df.sort_values(
                        by=["sequence_length", "query_start"],
                        ascending=[False, True]
                    ).reset_index(drop=True)

                legacy_intergenic = self.data.get("legacy_data", {}).get("intergenic_lcb_data", [])
                for lcb in legacy_intergenic:
                    seq_len = len(lcb.get('aligned_query_sequence', ''))
                    if seq_len < 15: continue
                    lcb_copy = lcb.copy()
                    lcb_copy['algorithm'] = 'Legacy k-mer'
                    combined_intergenic_data.append(lcb_copy)

            combined_intergenic_df = pd.DataFrame(combined_intergenic_data)
            if not combined_intergenic_df.empty:
                # Single rename to standardize column names from all sources
                combined_intergenic_df = combined_intergenic_df.rename(columns={
                    "algorithm": "Algorithm",
                    "query_start": "Query Start",
                    "query_end": "Query End",
                    "query_strand": "Query Strand",
                    "ref_start": "Ref Start",
                    "ref_stop": "Ref Stop",
                    "ref_strand": "Ref Strand",
                    "identity_excl_gaps": "Identity (%) Excl Gaps",
                    "identity_incl_gaps": "Identity (%) Incl Gaps",
                    "original_query_sequence": "Original Query Sequence",
                    "original_reference_sequence": "Original Reference Sequence",
                    "aligned_query_sequence": "Aligned Query Sequence",
                    "aligned_reference_sequence": "Aligned Reference Sequence",
                })

                if 'Query Strand' in combined_intergenic_df.columns:
                    combined_intergenic_df['Query Strand'] = combined_intergenic_df['Query Strand'].replace(
                        {1: '+', -1: '-', '1': '+', '-1': '-'})
                if 'Ref Strand' in combined_intergenic_df.columns:
                    combined_intergenic_df['Ref Strand'] = combined_intergenic_df['Ref Strand'].replace(
                        {1: '+', -1: '-', '1': '+', '-1': '-'})

                cols = ["Algorithm", "Query Start", "Query End", "Query Strand",
                        "Ref Start", "Ref Stop", "Ref Strand",
                        "Identity (%) Excl Gaps", "Identity (%) Incl Gaps",
                        "Original Query Sequence", "Original Reference Sequence",
                        "Aligned Query Sequence", "Aligned Reference Sequence",
                        "Difference String", "Query with Mismatches (#)",
                        "Query Differences Only", "Reference Differences Only"]

                final_cols = [c for c in cols if c in combined_intergenic_df.columns]
                combined_intergenic_df = combined_intergenic_df[final_cols].copy()
                # Remove duplicate columns if any
                combined_intergenic_df = combined_intergenic_df.loc[:, ~combined_intergenic_df.columns.duplicated()]
                if not combined_intergenic_df.empty:
                    combined_intergenic_df = combined_intergenic_df.sort_values(by="Query Start").reset_index(drop=True)

            seqlen = self.data["seqlen"]
            gc_window_size = self.data.get("gc_window_size", self.plot_params["gc_window_size"].get())

            if self.data.get("gc_arr") is not None and len(self.data["gc_arr"]) > 0:
                gc_vals = self.data["gc_arr"][::gc_window_size] * 100.0
                skew_vals = self.data["skew_arr"][::gc_window_size]

                n_windows = len(gc_vals)
                starts = np.arange(n_windows) * gc_window_size + 1
                ends = np.minimum(starts + gc_window_size - 1, seqlen)

                gc_full_df = pd.DataFrame({
                    "Window Start": starts,
                    "Window End": ends,
                    "GC Content (%)": gc_vals,
                    "GC Skew": skew_vals
                })
            else:
                gc_full_df = pd.DataFrame(columns=["Window Start", "Window End", "GC Content (%)", "GC Skew"])

            gc_skew_df = gc_full_df[["Window Start", "Window End", "GC Skew"]]
            gc_content_df = gc_full_df[["Window Start", "Window End", "GC Content (%)"]]
            MIN_IDENTITY_THRESHOLD = 60.0
            identity_col = "Identity (%) Excl Gaps"

            def filter_by_identity(df, threshold=MIN_IDENTITY_THRESHOLD, identity_column=identity_col):
                """Filter dataframe to keep only rows with identity >= threshold."""
                if identity_column in df.columns:
                    rows_before = len(df)
                    df_filtered = df[df[identity_column] >= threshold].reset_index(drop=True)
                    rows_after = len(df_filtered)
                    if rows_before > rows_after:
                        print(f"[IDENTITY FILTER] Filtered from {rows_before} to "
                              f"{rows_after} rows (removed "
                              f"{rows_before - rows_after} rows with <{threshold}% identity)")
                    return df_filtered
                return df

            print(f"\n[EXCEL EXPORT] Applying minimum "
                  f"{MIN_IDENTITY_THRESHOLD}% identity filter to all alignment sheets...")
            homology_df = filter_by_identity(homology_df)
            blast_all_df = filter_by_identity(blast_all_df)
            blast_intergenic_df = filter_by_identity(blast_intergenic_df)
            mauve_lcb_df = filter_by_identity(mauve_lcb_df)
            mauve_coding_lcb_df = filter_by_identity(mauve_coding_lcb_df)
            mauve_intergenic_lcb_df = filter_by_identity(mauve_intergenic_lcb_df)
            sibeliaz_lcb_df = filter_by_identity(sibeliaz_lcb_df)
            sibeliaz_coding_lcb_df = filter_by_identity(sibeliaz_coding_lcb_df)
            sibeliaz_intergenic_lcb_df = filter_by_identity(sibeliaz_intergenic_lcb_df)
            legacy_lcb_df = filter_by_identity(legacy_lcb_df)
            legacy_intergenic_lcb_df = filter_by_identity(legacy_intergenic_lcb_df)
            combined_intergenic_df = filter_by_identity(combined_intergenic_df)

            with pd.ExcelWriter(filepath) as writer:
                homology_df.to_excel(writer, sheet_name="Homology Report (Genes)", index=False)
                blast_all_df.to_excel(writer, sheet_name="Alignment Blocks (BLAST)", index=False)
                blast_intergenic_df.to_excel(writer, sheet_name="Intergenic Blocks (BLAST)", index=False)
                mauve_lcb_df.to_excel(writer, sheet_name="Alignment Blocks (Mauve)", index=False)
                mauve_coding_lcb_df.to_excel(writer, sheet_name="Coding Blocks (Mauve)", index=False)
                mauve_intergenic_lcb_df.to_excel(writer, sheet_name="Intergenic Blocks (Mauve)", index=False)
                sibeliaz_lcb_df.to_excel(writer, sheet_name="Alignment Blocks (SibeliaZ)", index=False)
                sibeliaz_coding_lcb_df.to_excel(writer, sheet_name="Coding Blocks (SibeliaZ)", index=False)
                sibeliaz_intergenic_lcb_df.to_excel(writer, sheet_name="Intergenic Blocks (SibeliaZ)", index=False)

                if include_legacy:
                    legacy_lcb_df.to_excel(writer, sheet_name="Alignment Blocks (Legacy)", index=False)
                    legacy_intergenic_lcb_df.to_excel(writer, sheet_name="Intergenic Blocks (Legacy)", index=False)
                combined_intergenic_df.to_excel(writer, sheet_name="Combined Intergenic Blocks", index=False)
                gc_skew_df.to_excel(writer, sheet_name="GC Skew", index=False)
                gc_content_df.to_excel(writer, sheet_name="GC Content", index=False)

            # Adjust column widths
            _adjust_excel_column_widths(filepath)

            try:
                log_filepath = Path(filepath).with_suffix('.log.txt')

                stdout_log = self.master_app.log_capture.get_value()
                stderr_log = self.master_app.error_capture.get_value()

                with open(log_filepath, 'w', encoding='utf-8') as f:
                    f.write("--- CGCT Log File ---\n")
                    f.write(f"Timestamp: {datetime.datetime.now().isoformat()}\n")
                    f.write(f"Report File: {Path(filepath).name}\n")
                    f.write("\n" + "=" * 80 + "\n")
                    f.write(" STDOUT LOG (Debug prints, status messages)\n")
                    f.write("=" * 80 + "\n\n")
                    f.write(stdout_log)
                    f.write("\n\n" + "=" * 80 + "\n")
                    f.write(" STDERR LOG (Errors, tracebacks)\n")
                    f.write("=" * 80 + "\n\n")
                    f.write(stderr_log)

            except Exception as e_log:
                print(f"WARNING: Could not save log file to {log_filepath.name}")
                print(f"  Error: {e_log}")

            num_sheets = 10 + (2 if include_legacy else 0)
            messagebox.showinfo("Success", f"{num_sheets}-sheet report saved to:\n{filepath}")
            self.master_app.status_label.config(text="Report saved.", fg="green")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export report:\n{e}")
            self.master_app.status_label.config(text="Report export failed.", fg="red")
            traceback.print_exc()
        finally:
            if self.export_button:
                self.export_button.config(state="normal")


# ──────────────────────────────────────────────────────────────────────────────
# GUI: Main Application
# ──────────────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.is_loading_settings = False
        self.original_stdout = sys.stdout
        self.original_stderr = sys.stderr
        self.log_capture = StreamLogger(sys.stdout)
        self.error_capture = StreamLogger(sys.stderr)
        sys.stdout = self.log_capture
        sys.stderr = self.error_capture
        self.title("Circular Genome Comparison Tool")
        self.geometry("1600x850")
        self.resizable(True, True)
        self.settings_file = Path.home() / ".plasmid_plotter_settings.json"
        main = tk.Frame(self, padx=10, pady=10)
        main.pack(fill="both", expand=True)
        main.grid_columnconfigure(1, weight=1)
        self.resolution_options = RESOLUTION_OPTIONS

        self.vars = {
            "fasta": tk.StringVar(), "annot": tk.StringVar(), "ref": tk.StringVar(),
            "out": tk.StringVar(value=str(Path.cwd() / "cgct_plot.png")),
            "bed": tk.StringVar(),
            "xmfa_file": tk.StringVar(),
            "open_image": tk.BooleanVar(value=True),
            "force_mauve": tk.BooleanVar(value=False),
            "gc_window_size": tk.IntVar(value=500),
            "allow_code_fallback": tk.BooleanVar(value=False),
            "homology_thresh": tk.DoubleVar(value=70.0),
            "homology_thresh_max": tk.DoubleVar(value=100.0),
            "show_roi_labels": tk.BooleanVar(value=True),
            "show_homology_labels": tk.BooleanVar(value=True),
            "autofit_color_range": tk.BooleanVar(value=True),
            "autorotate_roi": tk.BooleanVar(value=True),
            "manual_rot_deg": tk.DoubleVar(value=0.0),
            "homology_colormap": tk.StringVar(value="Plasma"),
            "show_blast_separators": tk.BooleanVar(value=False),
            "use_smart_layout": tk.BooleanVar(value=True),
            "show_label_boxes": tk.BooleanVar(value=True),
            "label_spread": tk.DoubleVar(value=1.0),
            "label_fontsize": tk.DoubleVar(value=9.0),
            "label_line_radial_len": tk.DoubleVar(value=1.15),
            "label_line_horizontal_len": tk.DoubleVar(value=0.1),
            "label_distance_factor": tk.DoubleVar(value=0.25),
            "connect_dots_with_curve": tk.BooleanVar(value=True),
            "curve_tension": tk.DoubleVar(value=1.88),
            "show_connector_dots": tk.BooleanVar(value=False),
            "color_lines_by_homology": tk.BooleanVar(value=False),
            "color_dots_by_homology": tk.BooleanVar(value=False),
            "gc_inner": tk.DoubleVar(value=0.3),
            "gc_thick": tk.DoubleVar(value=1.05),
            "skew_thick": tk.DoubleVar(value=0.505),
            "skew_inner_radius": tk.DoubleVar(value=0.5),
            "gc_colormap": tk.StringVar(value="Grey"),
            "gc_ring_color_only": tk.BooleanVar(value=False),
            "manual_find": tk.StringVar(value="manA"),
            "manual_replace": tk.StringVar(),
            "save_resolution": tk.StringVar(value="4K UHD (3840x2160)"),
            "custom_save_width": tk.StringVar(value="3840"),
            "custom_save_height": tk.StringVar(value="2160"),
            "nc_window_size": tk.IntVar(value=500),
            "nc_step_size": tk.IntVar(value=50),
            "id_ring_include_coding": tk.BooleanVar(value=False),
            "id_ring_radius": tk.DoubleVar(value=0.7),
            "id_ring_thickness": tk.DoubleVar(value=0.20),
            "blast_ring_thickness": tk.DoubleVar(value=0.15),
            "blast_ring_radius": tk.DoubleVar(value=1.0),
            "id_ring_color_only": tk.BooleanVar(value=True),
            "id_ring_precise_mode": tk.BooleanVar(value=False),
            "id_ring_show_noncoding": tk.BooleanVar(value=True),
            "autofit_label_fontsize": tk.BooleanVar(value=True),
            "show_start_end_line": tk.BooleanVar(value=False),
            "nc_auto_fit_color": tk.BooleanVar(value=True),
            "blast_task": tk.StringVar(value="blastn"),
            "blast_word_size": tk.StringVar(value="11"),
            "blast_reward": tk.StringVar(value="2"),
            "blast_penalty": tk.StringVar(value="-3"),
            "blast_evalue": tk.StringVar(value="10"),
            "blast_num_threads": tk.StringVar(value=_get_cpu_count()),
            "auto_thickness_rings": tk.BooleanVar(value=True),
            "limit_to_region": tk.BooleanVar(value=False),
            "region_start": tk.StringVar(value="1"),
            "region_end": tk.StringVar(value=""),
            "identity_algorithm": tk.StringVar(value="Mauve + SibeliaZ Fallback"),
            "nc_colormap": tk.StringVar(value="Viridis"),
            "id_baseline_transparency_pct": tk.DoubleVar(value=100.0),
            "gc_baseline_transparency_pct": tk.DoubleVar(value=50.0),
            "skew_baseline_transparency_pct": tk.DoubleVar(value=50.0),
            "show_blast_coding": tk.BooleanVar(value=True),
            "show_blast_noncoding": tk.BooleanVar(value=False),
            "remove_gene_borders": tk.BooleanVar(value=False),
            "sibeliaz_maf": tk.StringVar(),
            "force_sibeliaz": tk.BooleanVar(value=False),
            "run_legacy_report_on_save": tk.BooleanVar(value=False),
            "show_coords": tk.BooleanVar(value=True),
            "subsequence_start": tk.StringVar(value="1"),
            "subsequence_end": tk.StringVar(value=""),
            "mauve_disable_backbone": tk.BooleanVar(value=False),
            "mauve_mums": tk.BooleanVar(value=False),
            "mauve_collinear": tk.BooleanVar(value=False),
            "mauve_skip_refinement": tk.BooleanVar(value=False),
            "mauve_skip_gapped_alignment": tk.BooleanVar(value=False),
            "mauve_seed_family": tk.BooleanVar(value=False),
            "mauve_no_weight_scaling": tk.BooleanVar(value=False),
            "mauve_mem_clean": tk.BooleanVar(value=False),
            "mauve_seed_weight": tk.StringVar(value=""),
            "mauve_max_gapped_aligner_length": tk.StringVar(value=""),
            "mauve_max_breakpoint_distance_scale": tk.StringVar(value=""),
            "mauve_conservation_distance_scale": tk.StringVar(value=""),
            "mauve_bp_dist_estimate_min_score": tk.StringVar(value=""),
            "mauve_gap_open": tk.StringVar(value=""),
            "mauve_gap_extend": tk.StringVar(value=""),
            "mauve_weight": tk.StringVar(value=""),
            "mauve_min_scaled_penalty": tk.StringVar(value=""),
            "mauve_hmm_p_go_homologous": tk.StringVar(value=""),
            "mauve_hmm_p_go_unrelated": tk.StringVar(value=""),
        }

        self.defaults = {key: var.get() for key, var in self.vars.items()}
        self.file_keys = {"fasta", "annot", "ref", "out", "bed", "xmfa_file", "sibeliaz_maf"}

        self.mauve_tooltips = {
            "mauve_disable_backbone": "Disable automatic backbone detection",
            "mauve_mums": "Find maximal unique matches only, skip LCB detection",
            "mauve_collinear": "Assume sequences are collinear (no rearrangements)",
            "mauve_skip_refinement": "Skip LCB refinement step",
            "mauve_skip_gapped_alignment": "Skip gapped alignment of local collinear blocks",
            "mauve_seed_family": "Use seed family instead of MUMs",
            "mauve_no_weight_scaling": "Disable weight scaling based on breakpoint proximity",
            "mauve_mem_clean": "Release memory more aggressively during processing",
            "mauve_seed_weight": "Length of k-mer for MUM search. Leave empty for progressiveMauve default",
            "mauve_max_gapped_aligner_length": "Maximum region length (bp) for gapped aligner. Leave empty for progressiveMauve default",
            "mauve_max_breakpoint_distance_scale": "Scaling factor for LCB weighting (disabled if No Weight Scaling). Range: [0,1]. Leave empty for progressiveMauve default",
            "mauve_conservation_distance_scale": "Scaling for conservation distance weighting. Range: [0,1]. Leave empty for progressiveMauve default",
            "mauve_bp_dist_estimate_min_score": "Minimum LCB score for breakpoint distance estimation. Leave empty for progressiveMauve default",
            "mauve_gap_open": "Gap opening penalty for internal aligner. Leave empty for progressiveMauve default",
            "mauve_gap_extend": "Gap extension penalty for internal aligner. Leave empty for progressiveMauve default",
            "mauve_weight": "Minimum pairwise LCB score threshold. Leave empty for progressiveMauve default",
            "mauve_min_scaled_penalty": "Minimum breakpoint penalty after scaling. Leave empty for progressiveMauve default",
            "mauve_hmm_p_go_homologous": "HMM transition probability (unrelated→homologous). Leave empty for progressiveMauve default",
            "mauve_hmm_p_go_unrelated": "HMM transition probability (homologous→unrelated). Leave empty for progressiveMauve default",
        }

        self.ui_tooltips = {
            "fasta": "Query plasmid sequence in FASTA format",
            "annot": "Gene annotation file in GFF3 or GenBank format",
            "ref": "Reference sequence in FASTA format",
            "bed": "Optional BED file to highlight regions on the plot",
            "xmfa_file": "Pre-computed Mauve alignment file to skip alignment step",
            "sibeliaz_maf": "Pre-computed SibeliaZ MAF file to skip alignment",
            "manual_find": "Gene name to find for label replacement",
            "manual_replace": "New label text to display instead of gene name",
            "blast_task": "BLAST algorithm variant: blastn (balanced, default), megablast (fast), dc-megablast, blastn-short",
            "blast_word_size": "Length of exact matching seed words. Larger=faster but less sensitive",
            "blast_reward": "Score for matching nucleotides",
            "blast_penalty": "Score penalty for mismatches (use negative value)",
            "blast_evalue": "E-value threshold for reporting hits. Lower=more stringent",
            "blast_num_threads": "Number of CPU threads to use for BLAST searches",
            "limit_to_region": "Limit plot to a specific region instead of full sequence",
            "region_start": "Start position for region limit (1-indexed)",
            "region_end": "End position for region limit (1-indexed)",
            "identity_algorithm": "Algorithm for sequence identity computation: Mauve (fast), SibeliaZ (sensitive), or legacy k-mer methods",
            "gc_window_size": "Window size (bp) for GC% content calculation. Smaller=more detail but noisier",
            "allow_code_fallback": "Fall back to simpler code if primary code fails",
            "homology_thresh": "Minimum homology score (0-100%) threshold for display",
            "show_roi_labels": "Display labels for regions of interest",
            "show_homology_labels": "Display homology score labels on plot",
            "autofit_color_range": "Automatically adjust color scale to data range",
            "autorotate_roi": "Automatically rotate plot to keep regions readable",
            "manual_rot_deg": "Manual rotation angle in degrees (0-360)",
            "homology_colormap": "Color scheme for homology visualization",
            "show_blast_separators": "Show vertical lines separating BLAST hits",
            "use_smart_layout": "Intelligently position labels to avoid overlap",
            "show_label_boxes": "Show background boxes behind labels for readability",
            "label_spread": "Label spread factor for positioning (1.0=normal)",
            "label_fontsize": "Font size for labels in points",
            "label_line_radial_len": "Radial length of label connector lines",
            "label_line_horizontal_len": "Horizontal offset for label positioning",
            "label_distance_factor": "Distance factor for label placement",
            "connect_dots_with_curve": "Connect homology points with smooth curves",
            "curve_tension": "Tension parameter for curve smoothness (1-3, higher=smoother)",
            "show_connector_dots": "Show dots at homology connection points",
            "color_lines_by_homology": "Color connector lines based on homology score",
            "color_dots_by_homology": "Color dots based on homology score",
            "gc_inner": "Inner radius for GC% ring (0-1)",
            "gc_thick": "Thickness of GC% ring",
            "skew_thick": "Thickness of GC skew ring",
            "skew_inner_radius": "Inner radius for GC skew ring (0-1)",
            "skew_baseline_transparency_pct": "Baseline transparency for GC skew ring (0-100%)",
            "gc_colormap": "Color scheme for GC% visualization",
            "gc_ring_color_only": "Show only GC% color without values",
            "gc_baseline_transparency_pct": "Baseline transparency for GC ring (0-100%)",
            "nc_window_size": "Window size (bp) for nucleotide composition analysis",
            "nc_step_size": "Step size (bp) for sliding window analysis",
            "nc_colormap": "Color scheme for nucleotide composition visualization",
            "id_ring_include_coding": "Include coding regions in identity ring",
            "id_ring_radius": "Radial position of identity ring (0-1, where 1=outer)",
            "id_ring_thickness": "Thickness of identity ring",
            "id_ring_color_only": "Show only identity color without percentage text",
            "id_ring_precise_mode": "Show raw alignment blocks instead of sliding window",
            "id_ring_show_noncoding": "Show identity for non-coding regions",
            "id_baseline_transparency_pct": "Baseline transparency for identity ring (0-100%)",
            "blast_ring_thickness": "Thickness of BLAST homology ring",
            "blast_ring_radius": "Radial position of BLAST ring",
            "show_blast_coding": "Show BLAST hits in coding regions",
            "show_blast_noncoding": "Show BLAST hits in non-coding regions",
            "remove_gene_borders": "Remove visual borders between genes",
            "autofit_label_fontsize": "Automatically adjust font size to fit labels",
            "show_start_end_line": "Show line indicating sequence start/end",
            "auto_thickness_rings": "Automatically adjust ring thickness based on data",
            "save_resolution": "Output resolution preset or select custom",
            "custom_save_width": "Custom output width in pixels",
            "custom_save_height": "Custom output height in pixels",
            "open_image": "Automatically open generated plot after saving",
            "force_mauve": "Force re-computation of Mauve alignment even if cached",
            "force_sibeliaz": "Force re-computation of SibeliaZ alignment even if cached",
            "run_legacy_report_on_save": "Generate legacy k-mer analysis report when saving",
            "show_coords": "Show sequence coordinates on outer ring",
            "subsequence_start": "Start position for FASTA export (1-indexed, inclusive)",
            "subsequence_end": "End position for FASTA export (1-indexed, inclusive)",
        }

        self.mauve_entry_widgets = {}
        self.vars["fasta"].trace_add("write", self._check_for_cached_xmfa)
        self.vars["fasta"].trace_add("write", self._check_for_cached_maf)
        self.vars["fasta"].trace_add("write", self._on_fasta_load_update_coords)
        self.vars["ref"].trace_add("write", self._check_for_cached_xmfa)
        self.vars["ref"].trace_add("write", self._check_for_cached_maf)

        main.grid_columnconfigure(0, weight=0, minsize=90)
        main.grid_columnconfigure(1, weight=1)
        main.grid_columnconfigure(2, weight=0, minsize=60)
        main.grid_columnconfigure(3, weight=0, minsize=90)
        main.grid_columnconfigure(4, weight=1)
        main.grid_columnconfigure(5, weight=0, minsize=60)

        lbl_query = tk.Label(main, text="Query Plasmid FASTA")
        lbl_query.grid(row=0, column=0, sticky="e", pady=3, padx=(5, 2))
        if "fasta" in self.ui_tooltips: _create_tooltip(lbl_query, self.ui_tooltips["fasta"])
        entry_query = tk.Entry(main, textvariable=self.vars["fasta"])
        entry_query.grid(row=0, column=1, padx=5, sticky="ew")
        if "fasta" in self.ui_tooltips: _create_tooltip(entry_query, self.ui_tooltips["fasta"])
        self.vars["fasta"].trace_add("write", lambda *args: self._update_region_and_subsequence_defaults())
        tk.Button(main, text="Browse...", command=lambda: self._browse_file(self.vars["fasta"])).grid(row=0, column=2,
                                                                                                      padx=2)
        lbl_ref = tk.Label(main, text="Reference FASTA")
        lbl_ref.grid(row=0, column=3, sticky="e", pady=3, padx=(5, 2))
        if "ref" in self.ui_tooltips:
            _create_tooltip(lbl_ref, self.ui_tooltips["ref"])
        entry_ref = tk.Entry(main, textvariable=self.vars["ref"])
        entry_ref.grid(row=0, column=4, padx=5, sticky="ew")
        if "ref" in self.ui_tooltips:
            _create_tooltip(entry_ref, self.ui_tooltips["ref"])
        tk.Button(main, text="Browse...", command=lambda: self._browse_file(self.vars["ref"])).grid(row=0, column=5,
                                                                                                    padx=2)

        lbl_annot = tk.Label(main, text="Annotation (GFF/GBK)")
        lbl_annot.grid(row=1, column=0, sticky="e", pady=3, padx=(5, 2))
        if "annot" in self.ui_tooltips: _create_tooltip(lbl_annot, self.ui_tooltips["annot"])
        entry_annot = tk.Entry(main, textvariable=self.vars["annot"])
        entry_annot.grid(row=1, column=1, padx=5, sticky="ew")
        if "annot" in self.ui_tooltips: _create_tooltip(entry_annot, self.ui_tooltips["annot"])
        tk.Button(main, text="Browse...", command=lambda: self._browse_file(self.vars["annot"])).grid(row=1, column=2,
                                                                                                      padx=2)
        lbl_bed = tk.Label(main, text="Highlight Regions (.bed)")
        lbl_bed.grid(row=1, column=3, sticky="e", pady=3, padx=(5, 2))
        if "bed" in self.ui_tooltips: _create_tooltip(lbl_bed, self.ui_tooltips["bed"])
        entry_bed = tk.Entry(main, textvariable=self.vars["bed"])
        entry_bed.grid(row=1, column=4, padx=5, sticky="ew")
        if "bed" in self.ui_tooltips: _create_tooltip(entry_bed, self.ui_tooltips["bed"])
        tk.Button(main, text="Browse...", command=lambda: self._browse_file(self.vars["bed"])).grid(row=1, column=5,
                                                                                                    padx=2)
        lbl_mauve = tk.Label(main, text="Pre-computed Mauve (.xmfa)")
        lbl_mauve.grid(row=2, column=0, sticky="e", pady=3, padx=(5, 2))
        if "xmfa_file" in self.ui_tooltips: _create_tooltip(lbl_mauve, self.ui_tooltips["xmfa_file"])
        entry_mauve = tk.Entry(main, textvariable=self.vars["xmfa_file"])
        entry_mauve.grid(row=2, column=1, padx=5, sticky="ew")
        if "xmfa_file" in self.ui_tooltips: _create_tooltip(entry_mauve, self.ui_tooltips["xmfa_file"])
        tk.Button(main, text="Browse...", command=lambda: self._browse_file(self.vars["xmfa_file"])).grid(row=2,
                                                                                                          column=2,
                                                                                                          padx=2)
        lbl_sibeliaz = tk.Label(main, text="Pre-computed SibeliaZ (.maf)")
        lbl_sibeliaz.grid(row=2, column=3, sticky="e", pady=3, padx=(5, 2))
        if "sibeliaz_maf" in self.ui_tooltips:
            _create_tooltip(lbl_sibeliaz, self.ui_tooltips["sibeliaz_maf"])
        entry_sibeliaz = tk.Entry(main, textvariable=self.vars["sibeliaz_maf"])
        entry_sibeliaz.grid(row=2, column=4, padx=5, sticky="ew")
        if "sibeliaz_maf" in self.ui_tooltips:
            _create_tooltip(entry_sibeliaz, self.ui_tooltips["sibeliaz_maf"])
        tk.Button(main, text="Browse...", command=lambda: self._browse_file(self.vars["sibeliaz_maf"])).grid(row=2,
                                                                                                             column=5,
                                                                                                             padx=2)
        row_counter = 3
        override_frame = tk.LabelFrame(main, text="Manual Label Override", padx=5, pady=5)
        override_frame.grid(row=row_counter, column=0, columnspan=6, sticky="ew", pady=10)
        row_counter += 1
        lbl1 = tk.Label(override_frame, text="If Gene Name is:")
        lbl1.pack(side="left", padx=5)
        if "manual_find" in self.ui_tooltips: _create_tooltip(lbl1, self.ui_tooltips["manual_find"])
        entry1 = tk.Entry(override_frame, textvariable=self.vars["manual_find"], width=15)
        entry1.pack(side="left", padx=5)
        if "manual_find" in self.ui_tooltips: _create_tooltip(entry1, self.ui_tooltips["manual_find"])
        lbl2 = tk.Label(override_frame, text="Replace Label with:")
        lbl2.pack(side="left", padx=5)
        if "manual_replace" in self.ui_tooltips: _create_tooltip(lbl2, self.ui_tooltips["manual_replace"])
        entry2 = tk.Entry(override_frame, textvariable=self.vars["manual_replace"], width=30)
        entry2.pack(side="left", padx=5, expand=True, fill="x")
        if "manual_replace" in self.ui_tooltips: _create_tooltip(entry2, self.ui_tooltips["manual_replace"])
        blast_params_frame = ttk.LabelFrame(main, text="BLAST Parameters")
        blast_params_frame.grid(row=row_counter, column=0, columnspan=6, sticky="ew", pady=10, padx=5)

        for i in range(20): blast_params_frame.grid_columnconfigure(i, weight=(1 if i % 2 == 1 else 0))
        blast_row, blast_col = 0, 0

        lbl_task = tk.Label(blast_params_frame, text="Task:")
        lbl_task.grid(row=blast_row, column=blast_col, sticky="w", padx=(5, 2), pady=3)
        if "blast_task" in self.ui_tooltips:
            _create_tooltip(lbl_task, self.ui_tooltips["blast_task"])
        task_combo = ttk.Combobox(blast_params_frame, textvariable=self.vars["blast_task"],
                                  values=list(BLAST_DEFAULTS.keys()), state='readonly', width=12)
        task_combo.grid(row=blast_row, column=blast_col + 1, sticky="ew", padx=(0, 10))
        if "blast_task" in self.ui_tooltips:
            _create_tooltip(task_combo, self.ui_tooltips["blast_task"])
        task_combo.bind("<<ComboboxSelected>>", self._on_blast_task_change)
        blast_col += 2
        blast_entry_row1 = [("Word Size", "blast_word_size"), ("Reward", "blast_reward"),
                            ("Penalty", "blast_penalty"), ("E-value", "blast_evalue"),
                            ("Num Threads", "blast_num_threads")]
        for label, key in blast_entry_row1:
            lbl = tk.Label(blast_params_frame, text=f"{label}:")
            lbl.grid(row=blast_row, column=blast_col, sticky="w", padx=(5, 2), pady=3)
            if key in self.ui_tooltips: _create_tooltip(lbl, self.ui_tooltips[key])
            entry = ttk.Entry(blast_params_frame, textvariable=self.vars[key], width=8)
            entry.grid(row=blast_row, column=blast_col + 1, sticky="ew", padx=(0, 10))
            if key in self.ui_tooltips: _create_tooltip(entry, self.ui_tooltips[key])

            if key == "blast_word_size":
                entry.bind("<FocusOut>", _create_bounded_entry_validator(self.vars[key],
                                                                         7, 28, "BLAST Word Size"))
            elif key == "blast_reward":
                entry.bind("<FocusOut>", _create_positive_validator(self.vars[key], "BLAST Reward"))
            elif key == "blast_penalty":
                entry.bind("<FocusOut>", _create_bounded_entry_validator(self.vars[key],
                                                                         -10, 0, "BLAST Penalty"))
            elif key == "blast_evalue":
                entry.bind("<FocusOut>", _create_positive_validator(self.vars[key],
                                                                    "BLAST E-value", is_float=True))
            elif key == "blast_num_threads":
                entry.bind("<FocusOut>",
                           _create_bounded_entry_validator(self.vars[key], 1, max(os.cpu_count() or 1, 512),
                                                           "BLAST Num Threads"))
            blast_col += 2
        blast_row = 2
        region_frame = tk.Frame(blast_params_frame)
        region_frame.grid(row=blast_row, column=0, columnspan=20, sticky="ew", pady=(5, 3), padx=5)
        cb_limit = ttk.Checkbutton(region_frame, text="Limit Plot to Region:", variable=self.vars["limit_to_region"])
        cb_limit.pack(side="left", padx=(0, 10))
        if "limit_to_region" in self.ui_tooltips: _create_tooltip(cb_limit, self.ui_tooltips["limit_to_region"])
        lbl_start = tk.Label(region_frame, text="Start:")
        lbl_start.pack(side="left", padx=(0, 2))
        if "region_start" in self.ui_tooltips: _create_tooltip(lbl_start, self.ui_tooltips["region_start"])
        entry_start = ttk.Entry(region_frame, textvariable=self.vars["region_start"], width=10)
        entry_start.pack(side="left", padx=(0, 10))
        if "region_start" in self.ui_tooltips: _create_tooltip(entry_start, self.ui_tooltips["region_start"])
        entry_start.bind("<FocusOut>", _create_positive_validator(self.vars["region_start"], "Region Start"))
        entry_start.bind("<KeyRelease>", self._on_region_value_change)
        lbl_end = tk.Label(region_frame, text="End:")
        lbl_end.pack(side="left", padx=(0, 2))
        if "region_end" in self.ui_tooltips: _create_tooltip(lbl_end, self.ui_tooltips["region_end"])
        entry_end = ttk.Entry(region_frame, textvariable=self.vars["region_end"], width=10)
        entry_end.pack(side="left", padx=(0, 10))
        if "region_end" in self.ui_tooltips: _create_tooltip(entry_end, self.ui_tooltips["region_end"])
        entry_end.bind("<FocusOut>", _create_positive_validator(self.vars["region_end"], "Region End"))
        entry_end.bind("<KeyRelease>", self._on_region_value_change)
        lbl_algo = tk.Label(region_frame, text="Identity Algorithm:")
        lbl_algo.pack(side="left", padx=(15, 2))
        if "identity_algorithm" in self.ui_tooltips: _create_tooltip(lbl_algo, self.ui_tooltips["identity_algorithm"])
        algo_options = ["Mauve + SibeliaZ Fallback", "Mauve", "SibeliaZ", "Legacy (Global k-mer %)",
                        "Legacy (1:1 Base %)"]
        algo_combo = ttk.Combobox(region_frame, textvariable=self.vars["identity_algorithm"], values=algo_options,
                                  state='readonly', width=25)
        algo_combo.pack(side="left", padx=(0, 10))
        if "identity_algorithm" in self.ui_tooltips:
            _create_tooltip(algo_combo, self.ui_tooltips["identity_algorithm"])

        self._update_blast_param_states()
        row_counter += 1
        params_frame = ttk.LabelFrame(main, text="Plotting Parameters")
        params_frame.grid(row=row_counter, column=0, columnspan=6, sticky="ew", pady=10, padx=5)

        for i in range(8): params_frame.grid_columnconfigure(i, weight=1, uniform="param_col")

        colormap_options = ["Plasma", "Viridis", "Inferno", "Magma", "Cividis", "Turbo", "Coolwarm", "RdYlBu",
                            "Gist_earth", "Terrain", "Hsv", "Jet"]
        full_cmaps = colormap_options + [f"{c} (Reversed)" for c in colormap_options]
        gc_colormap_options = ["Grey"] + colormap_options + [f"{c} (Reversed)" for c in colormap_options if c != "Grey"]

        check_options = [
            ("Autorotate to ROI", "autorotate_roi"), ("Auto-Calc. Thickness", "auto_thickness_rings"),
            ("Show Start/End Line", "show_start_end_line"), ("Remove Gene Outlines", "remove_gene_borders"),
            ("Open Image on Save", "open_image"), ("Show Homology Labels", "show_homology_labels"),
            ("Show ROI Labels", "show_roi_labels"), ("Use Smart Label Layout", "use_smart_layout"),
            ("Auto-Fit Label Font Size", "autofit_label_fontsize"), ("Show Label Boxes", "show_label_boxes"),
            ("Use Curved Connectors", "connect_dots_with_curve"),
            ("Color Lines by Homology", "color_lines_by_homology"),
            ("Show Connector Dots", "show_connector_dots"), ("Color Dots by Homology", "color_dots_by_homology"),
            ("Allow Fallback Labels", "allow_code_fallback"), ("Show BLAST Coding", "show_blast_coding"),
            ("Show BLAST Non-Coding", "show_blast_noncoding"), ("Show BLAST Separators", "show_blast_separators"),
            ("Auto-fit Color Range", "autofit_color_range"), ("Include Coding (Identity)", "id_ring_include_coding"),
            ("Show Non-Coding (Identity)", "id_ring_show_noncoding"), ("Auto-fit ID Color", "nc_auto_fit_color"),
            ("ID Ring: Color Only", "id_ring_color_only"), ("ID Ring: Show Raw Blocks", "id_ring_precise_mode"),
            ("GC Ring: Color Only", "gc_ring_color_only"), ("Show Coordinate Numbers", "show_coords"),
            ("Force New SibeliaZ Alignment", "force_sibeliaz"),
        ]

        param_row, param_col = 0, 0
        for label, key in check_options:
            cb = ttk.Checkbutton(params_frame, text=label, variable=self.vars[key])
            cb.grid(row=param_row, column=param_col, sticky="w", padx=3, pady=2)
            if key in self.ui_tooltips: _create_tooltip(cb, self.ui_tooltips[key])
            param_col += 1
            if param_col >= 8:
                param_col = 0
                param_row += 1

        if param_col != 0: param_row += 1
        ttk.Separator(params_frame, orient='horizontal').grid(
            row=param_row, column=0, columnspan=8, sticky='ew', pady=5, padx=5)
        param_row += 1
        colormap_container = tk.Frame(params_frame)
        colormap_container.grid(row=param_row, column=0, columnspan=8, sticky="ew", padx=3, pady=3)
        for i in range(3): colormap_container.grid_columnconfigure(i, weight=1, uniform="colormap")

        homo_frame = tk.Frame(colormap_container)
        homo_frame.grid(row=0, column=0, sticky="ew", padx=3)
        homo_frame.grid_columnconfigure(1, weight=1)

        lbl_homo_cm = tk.Label(homo_frame, text="Homology", font=("TkDefaultFont", 8))
        lbl_homo_cm.grid(row=0, column=0, sticky="e", padx=(0, 3))
        if "homology_colormap" in self.ui_tooltips: _create_tooltip(lbl_homo_cm, self.ui_tooltips["homology_colormap"])
        combo_homo_cm = ttk.Combobox(homo_frame, textvariable=self.vars["homology_colormap"], values=full_cmaps,
                                     state='readonly')
        combo_homo_cm.grid(row=0, column=1, sticky="ew", padx=0)
        if "homology_colormap" in self.ui_tooltips:
            _create_tooltip(combo_homo_cm, self.ui_tooltips["homology_colormap"])
        id_frame = tk.Frame(colormap_container)
        id_frame.grid(row=0, column=1, sticky="ew", padx=3)
        id_frame.grid_columnconfigure(1, weight=1)
        lbl_id_cm = tk.Label(id_frame, text="Identity", font=("TkDefaultFont", 8))
        lbl_id_cm.grid(row=0, column=0, sticky="e", padx=(0, 3))
        if "nc_colormap" in self.ui_tooltips: _create_tooltip(lbl_id_cm, self.ui_tooltips["nc_colormap"])
        combo_id_cm = ttk.Combobox(id_frame, textvariable=self.vars["nc_colormap"], values=full_cmaps,
                                   state='readonly')
        combo_id_cm.grid(row=0, column=1, sticky="ew", padx=0)
        if "nc_colormap" in self.ui_tooltips: _create_tooltip(combo_id_cm, self.ui_tooltips["nc_colormap"])
        gc_frame = tk.Frame(colormap_container)
        gc_frame.grid(row=0, column=2, sticky="ew", padx=3)
        gc_frame.grid_columnconfigure(1, weight=1)
        lbl_gc_cm = tk.Label(gc_frame, text="GC Content", font=("TkDefaultFont", 8))
        lbl_gc_cm.grid(row=0, column=0, sticky="e", padx=(0, 3))
        if "gc_colormap" in self.ui_tooltips: _create_tooltip(lbl_gc_cm, self.ui_tooltips["gc_colormap"])
        combo_gc_cm = ttk.Combobox(gc_frame, textvariable=self.vars["gc_colormap"], values=gc_colormap_options,
                                   state='readonly')
        combo_gc_cm.grid(row=0, column=1, sticky="ew", padx=0)
        if "gc_colormap" in self.ui_tooltips: _create_tooltip(combo_gc_cm, self.ui_tooltips["gc_colormap"])
        param_row += 1
        ttk.Separator(params_frame, orient='horizontal').grid(
            row=param_row, column=0, columnspan=8, sticky='ew', pady=5, padx=5)
        param_row += 1

        text_box_options = [
            ("Min Homology %", "homology_thresh"), ("Max Homology %", "homology_thresh_max"),
            ("BLAST Ring Radius", "blast_ring_radius"), ("BLAST Ring Thickness", "blast_ring_thickness"),
            ("Curve Tension", "curve_tension"), ("Label Font Size", "label_fontsize"),
            ("Label Spread", "label_spread"), ("Label Distance", "label_distance_factor"),
            ("Label Radial Line", "label_line_radial_len"), ("Label Horizontal Line", "label_line_horizontal_len"),
            ("GC Window (bp)", "gc_window_size"), ("GC Inner Radius", "gc_inner"),
            ("GC Thickness", "gc_thick"), ("GC Baseline Trans. (%)", "gc_baseline_transparency_pct"),
            ("Skew Inner Radius", "skew_inner_radius"), ("Skew Thickness", "skew_thick"),
            ("Skew Baseline Trans. (%)", "skew_baseline_transparency_pct"), ("NC Window Size", "nc_window_size"),
            ("NC Step Size", "nc_step_size"), ("ID Baseline Trans. (%)", "id_baseline_transparency_pct"),
            ("Identity Ring Radius", "id_ring_radius"), ("Identity Ring Thickness", "id_ring_thickness")
        ]

        param_col = 0
        for i, (label, key) in enumerate(text_box_options):
            item_frame = tk.Frame(params_frame)
            item_frame.grid(row=param_row, column=param_col, sticky="ew", padx=2, pady=2)
            item_frame.grid_columnconfigure(0, weight=0, minsize=150)
            item_frame.grid_columnconfigure(1, weight=1, uniform="param_entry")
            lbl = tk.Label(item_frame, text=label, font=("TkDefaultFont", 8))
            lbl.grid(row=0, column=0, sticky="e", padx=(0, 3))
            if key in self.ui_tooltips: _create_tooltip(lbl, self.ui_tooltips[key])
            entry = ttk.Entry(item_frame, textvariable=self.vars[key], justify="right")
            entry.grid(row=0, column=1, sticky="ew", padx=0)
            if key in self.ui_tooltips: _create_tooltip(entry, self.ui_tooltips[key])

            param_col += 1
            if param_col >= 8:
                param_col = 0
                param_row += 1

        row_counter = params_frame.grid_info()['row']
        if param_col != 0: row_counter += 1
        mauve_options_frame = ttk.LabelFrame(main, text="progressiveMauve Options")
        mauve_options_frame.grid(row=row_counter, column=0, columnspan=6, sticky="ew", pady=10, padx=5)

        mauve_check_options = [
            ("Disable Backbone Detection", "mauve_disable_backbone"),
            ("Find MUMs Only", "mauve_mums"),
            ("Assume Collinear", "mauve_collinear"),
            ("Skip Refinement", "mauve_skip_refinement"),
            ("Skip Gapped Alignment", "mauve_skip_gapped_alignment"),
            ("Use Seed Family", "mauve_seed_family"),
            ("No Weight Scaling", "mauve_no_weight_scaling"),
            ("Mem Clean", "mauve_mem_clean"),
            ("Force New Alignment", "force_mauve"),
        ]

        mauve_cb_container = tk.Frame(mauve_options_frame)
        mauve_cb_container.grid(row=0, column=0, columnspan=6, sticky="ew", padx=0, pady=2)

        for i in range(9): mauve_cb_container.grid_columnconfigure(i, weight=1, uniform="mauve_checkbox")

        mauve_col = 0
        for label, key in mauve_check_options:
            cb = ttk.Checkbutton(mauve_cb_container, text=label, variable=self.vars[key])
            cb.grid(row=0, column=mauve_col, sticky="w", padx=2, pady=2)
            if key in self.mauve_tooltips: _create_tooltip(cb, self.mauve_tooltips[key])
            if key == "mauve_no_weight_scaling":
                self.vars[key].trace_add("write", lambda *args: self._update_mauve_field_states())
            if key != "force_mauve":
                self.vars[key].trace_add("write", lambda *args: self._on_mauve_param_change())
            mauve_col += 1

        ttk.Separator(mauve_options_frame, orient='horizontal').grid(
            row=1, column=0, columnspan=6, sticky='ew', pady=5, padx=5)

        for i in range(6): mauve_options_frame.grid_columnconfigure(i, weight=1, uniform="mauve_text_col")

        mauve_text_options = [
            ("Seed Weight", "mauve_seed_weight"),
            ("Max Gapped Aligner Length", "mauve_max_gapped_aligner_length"),
            ("Max Breakpoint Distance Scale [0,1]", "mauve_max_breakpoint_distance_scale"),
            ("Conservation Distance Scale [0,1]", "mauve_conservation_distance_scale"),
            ("BP Distance Est Min Score", "mauve_bp_dist_estimate_min_score"),
            ("Gap Open Penalty", "mauve_gap_open"),
            ("Gap Extend Penalty", "mauve_gap_extend"),
            ("Min Pairwise LCB Score", "mauve_weight"),
            ("Min Scaled Penalty", "mauve_min_scaled_penalty"),
            ("HMM P(go homologous) [0.0001]", "mauve_hmm_p_go_homologous"),
            ("HMM P(go unrelated) [0.000001]", "mauve_hmm_p_go_unrelated"),
        ]

        mauve_row, mauve_col = 2, 0
        for label, key in mauve_text_options:
            item_frame = tk.Frame(mauve_options_frame)
            item_frame.grid(row=mauve_row, column=mauve_col, sticky="ew", padx=2, pady=2)
            item_frame.grid_columnconfigure(0, weight=0, minsize=180)
            item_frame.grid_columnconfigure(1, weight=1, uniform="mauve_entry_box")
            label_widget = tk.Label(item_frame, text=label)
            label_widget.grid(row=0, column=0, sticky="e", padx=(0, 3))
            if key in self.mauve_tooltips: _create_tooltip(label_widget, self.mauve_tooltips[key])
            entry_widget = ttk.Entry(item_frame, textvariable=self.vars[key], justify="right")
            entry_widget.grid(row=0, column=1, sticky="ew", padx=0)
            self.mauve_entry_widgets[key] = entry_widget
            if key in self.mauve_tooltips:  _create_tooltip(entry_widget, self.mauve_tooltips[key])
            self.vars[key].trace_add("write", lambda *args: self._on_mauve_param_change())

            if key == "mauve_seed_weight":
                entry_widget.bind("<FocusOut>", _create_positive_validator(self.vars[key], "Mauve Seed Weight"))
            elif key == "mauve_max_gapped_aligner_length":
                entry_widget.bind("<FocusOut>",
                                  _create_positive_validator(self.vars[key], "Mauve Max Gapped Aligner Length"))
            elif key in ["mauve_max_breakpoint_distance_scale", "mauve_conservation_distance_scale"]:
                entry_widget.bind("<FocusOut>",
                                  _create_bounded_entry_validator(self.vars[key], 0, 1, label, is_float=True))
            elif key == "mauve_bp_dist_estimate_min_score":
                entry_widget.bind("<FocusOut>",
                                  _create_positive_validator(self.vars[key], "Mauve BP Distance Est Min Score"))
            elif key in ["mauve_gap_open", "mauve_gap_extend"]:
                entry_widget.bind("<FocusOut>", _create_nonnegative_validator(self.vars[key], label))
            elif key == "mauve_weight":
                entry_widget.bind("<FocusOut>",
                                  _create_nonnegative_validator(self.vars[key], "Mauve Min Pairwise LCB Score"))
            elif key == "mauve_min_scaled_penalty":
                entry_widget.bind("<FocusOut>",
                                  _create_nonnegative_validator(self.vars[key], "Mauve Min Scaled Penalty"))
            elif key in ["mauve_hmm_p_go_homologous", "mauve_hmm_p_go_unrelated"]:
                entry_widget.bind("<FocusOut>",
                                  _create_bounded_entry_validator(self.vars[key], 0, 1, label, is_float=True))

            mauve_col += 1
            if mauve_col >= 6:
                mauve_col = 0
                mauve_row += 1

        self._update_mauve_field_states()

        row_counter = mauve_options_frame.grid_info()['row'] + 1
        row_counter += 1
        save_opts_frame = ttk.LabelFrame(main, text="Save & Run")
        save_opts_frame.grid(row=row_counter, column=0, columnspan=6, sticky="ew", pady=10, padx=5)
        save_opts_frame.grid_columnconfigure(1, weight=1)
        save_opts_frame.grid_columnconfigure(3, weight=4)
        lbl_res = tk.Label(save_opts_frame, text="Resolution")
        lbl_res.grid(row=0, column=0, sticky="w", padx=(5, 0), pady=3)
        if "save_resolution" in self.ui_tooltips: _create_tooltip(lbl_res, self.ui_tooltips["save_resolution"])
        res_combo = ttk.Combobox(save_opts_frame, textvariable=self.vars["save_resolution"],
                                 values=list(self.resolution_options.keys()), state='readonly', width=30)
        res_combo.grid(row=0, column=1, sticky="ew", padx=5, pady=3)
        if "save_resolution" in self.ui_tooltips: _create_tooltip(res_combo, self.ui_tooltips["save_resolution"])
        res_combo.bind("<<ComboboxSelected>>", self._on_main_resolution_change)
        lbl_out = tk.Label(save_opts_frame, text="Output File")
        lbl_out.grid(row=0, column=2, sticky="w", padx=(10, 0), pady=3)
        if "out" in self.ui_tooltips: _create_tooltip(lbl_out, "Output file path for the generated plot")
        entry_out = tk.Entry(save_opts_frame, textvariable=self.vars["out"])
        entry_out.grid(row=0, column=3, sticky="ew", padx=5, pady=3)
        if "out" in self.ui_tooltips: _create_tooltip(entry_out, "Output file path for the generated plot")
        tk.Button(save_opts_frame, text="Browse...",
                  command=lambda: self._browse_save_as(self.vars["out"])).grid(row=0, column=4, sticky="e", padx=5,
                                                                               pady=3)
        self.custom_res_frame_main = tk.Frame(save_opts_frame)
        self.custom_res_frame_main.grid(row=1, column=1, columnspan=2, sticky="ew", pady=3, padx=5)
        lbl_width = tk.Label(self.custom_res_frame_main, text="Width (px)")
        lbl_width.pack(side="left", padx=(5, 0))
        if "custom_save_width" in self.ui_tooltips: _create_tooltip(lbl_width, self.ui_tooltips["custom_save_width"])
        entry_width = ttk.Entry(self.custom_res_frame_main, textvariable=self.vars["custom_save_width"], width=8)
        entry_width.pack(side="left", padx=5)
        if "custom_save_width" in self.ui_tooltips: _create_tooltip(entry_width, self.ui_tooltips["custom_save_width"])
        entry_width.bind("<FocusOut>",
                         _create_bounded_entry_validator(self.vars["custom_save_width"],
                                                         100, 20000, "Custom Width"))
        lbl_height = tk.Label(self.custom_res_frame_main, text="Height (px)")
        lbl_height.pack(side="left")
        if "custom_save_height" in self.ui_tooltips: _create_tooltip(lbl_height, self.ui_tooltips["custom_save_height"])
        entry_height = ttk.Entry(self.custom_res_frame_main, textvariable=self.vars["custom_save_height"], width=8)
        entry_height.pack(side="left", padx=5)
        if "custom_save_height" in self.ui_tooltips: _create_tooltip(entry_height, self.ui_tooltips["custom_save_height"])
        entry_height.bind("<FocusOut>",
                          _create_bounded_entry_validator(self.vars["custom_save_height"],
                                                          100, 20000, "Custom Height"))
        cb_frame = tk.Frame(save_opts_frame)
        cb_frame.grid(row=2, column=0, columnspan=5, sticky="ew", pady=3, padx=5)
        cb_frame.grid_columnconfigure(1, weight=1)
        cb_legacy = ttk.Checkbutton(cb_frame, text="Include Legacy k-mer Report (Slow, for Excel)",
                                    variable=self.vars["run_legacy_report_on_save"])
        cb_legacy.grid(row=0, column=0, sticky="w", padx=5)
        if "run_legacy_report_on_save" in self.ui_tooltips:
            _create_tooltip(cb_legacy, self.ui_tooltips["run_legacy_report_on_save"])
        subsequence_frame = tk.Frame(cb_frame)
        subsequence_frame.grid(row=0, column=1, sticky="ew", padx=(20, 5))
        lbl_sub_start = tk.Label(subsequence_frame, text="Export Subsequence: Start:")
        lbl_sub_start.pack(side="left")
        if "subsequence_start" in self.ui_tooltips:
            _create_tooltip(lbl_sub_start, self.ui_tooltips["subsequence_start"])
        sub_start_entry = ttk.Entry(subsequence_frame, textvariable=self.vars["subsequence_start"], width=10)
        sub_start_entry.pack(side="left", padx=5)
        if "subsequence_start" in self.ui_tooltips:
            _create_tooltip(sub_start_entry, self.ui_tooltips["subsequence_start"])
        sub_start_entry.bind("<FocusOut>",
                             _create_positive_validator(self.vars["subsequence_start"], "Subsequence Start"))
        lbl_sub_end = tk.Label(subsequence_frame, text="End:")
        lbl_sub_end.pack(side="left")
        if "subsequence_end" in self.ui_tooltips: _create_tooltip(lbl_sub_end, self.ui_tooltips["subsequence_end"])
        sub_end_entry = ttk.Entry(subsequence_frame, textvariable=self.vars["subsequence_end"], width=10)
        sub_end_entry.pack(side="left", padx=5)
        if "subsequence_end" in self.ui_tooltips: _create_tooltip(sub_end_entry, self.ui_tooltips["subsequence_end"])
        sub_end_entry.bind("<FocusOut>", _create_positive_validator(self.vars["subsequence_end"],
                                                                    "Subsequence End"))
        sub_button = ttk.Button(subsequence_frame, text="Export...", command=self._export_subsequence)
        sub_button.pack(side="left", padx=5)

        self.button_frame = tk.Frame(save_opts_frame)
        self.button_frame.grid(row=3, column=0, columnspan=5, sticky="ew", pady=10)
        self.load_button = tk.Button(self.button_frame, text="Use Previous Values", command=self._load_settings)
        self.load_button.pack(side="left", padx=(5, 10))
        self.reset_defaults_button = tk.Button(self.button_frame, text="Use Default Values",
                                               command=self._reset_to_defaults)
        self.reset_defaults_button.pack(side="left", padx=(0, 10))
        self.save_button = tk.Button(self.button_frame, text="Generate & Save Plot and Homology Report",
                                     command=self._run_save)
        self.save_button.pack(side="left", padx=5, expand=True, fill="x")
        self.interactive_button = tk.Button(self.button_frame, text="Show Interactive Plot",
                                            command=self._run_interactive, font=("Helvetica", 10, "bold"))
        self.interactive_button.pack(side="right", padx=5, expand=True, fill="x")

        row_counter += 1

        self.status_label = tk.Label(main, text="Select files and click an action.", justify="center")
        self.status_label.grid(row=row_counter, column=0, columnspan=6, pady=5, sticky="ew")
        self.log_queue = queue.Queue()
        self.current_plot_window = None
        self.protocol("WM_DELETE_WINDOW", self.on_main_app_closing)
        self.after(100, self.process_queue)
        self._on_main_resolution_change()

    def on_main_app_closing(self):
        """
        Called when the main application window is closed.
        Restores original stdout/stderr before destroying the window.
        """
        sys.stdout = self.original_stdout
        sys.stderr = self.original_stderr
        self.destroy()

    def _on_fasta_load_update_coords(self, *args):
        """Updates region_end and subsequence_end when a FASTA file is loaded."""
        fasta_path_str = self.vars["fasta"].get()
        if not fasta_path_str or not Path(fasta_path_str).is_file():
            self.vars["region_end"].set("")
            self.vars["subsequence_end"].set("")
            return

        try:
            record = SeqIO.read(fasta_path_str, "fasta")
            seqlen = len(record.seq)
            if not self.vars["region_end"].get():
                self.vars["region_end"].set(str(seqlen))
            if not self.vars["subsequence_end"].get():
                self.vars["subsequence_end"].set(str(seqlen))
        except Exception as e:
            print(f"DEBUG: Could not read FASTA for seqlen: {e}")
            self.vars["region_end"].set("")
            self.vars["subsequence_end"].set("")

    def _export_subsequence(self):
        """
        Export a subsequence (1-indexed range) to FASTA format.

        COORDINATE CONVENTION:
        - User input: 1-indexed inclusive (e.g., 100-200 means positions 100, 101, ..., 200)
        - Validation: Checks that start >= 1 and end <= seqlen
        - Storage: Keeps as 1-indexed for display in filename and header
        - Extraction: Converts to 0-indexed for Python string slicing
        """
        try:
            start_val = int(self.vars["subsequence_start"].get())  # 1-indexed
            end_val = int(self.vars["subsequence_end"].get())  # 1-indexed
        except (ValueError, tk.TclError):
            messagebox.showerror("Invalid Input", "Start and End values must be integers.")
            return

        fasta_path_str = self.vars["fasta"].get()
        if not fasta_path_str or not Path(fasta_path_str).is_file():
            messagebox.showerror("Error", "Query FASTA file not found.")
            return

        try:
            query_record = SeqIO.read(fasta_path_str, "fasta")
            full_seq = str(query_record.seq)
            seq_len = len(full_seq)
        except Exception as e:
            messagebox.showerror("Error", f"Could not read Query FASTA file: {e}")
            return

        is_out_of_bounds = False
        if start_val < 1 or end_val > seq_len: is_out_of_bounds = True

        if start_val > end_val:
            messagebox.showerror("Invalid Input",
                                 f"Start value ({start_val}) must be less than or equal to End value ({end_val}).")
            return

        extract_start = max(0, start_val - 1)
        extract_end = min(seq_len, end_val)
        original_fasta_path = Path(self.vars["fasta"].get())
        clean_stem = "".join(c if c.isalnum() or c == '-' else '_' for c in original_fasta_path.stem)
        default_filename = f"{clean_stem}-{start_val}-{end_val}.fasta"
        filepath = filedialog.asksaveasfilename(
            defaultextension=".fasta",
            filetypes=[("FASTA file", "*.fasta"), ("All files", "*.*")],
            initialfile=default_filename
        )

        if not filepath: return

        try:
            subsequence = full_seq[extract_start:extract_end]
            new_header = f">{query_record.id} [subsequence {start_val}-{end_val}]"
            with open(filepath, 'w') as f:
                f.write(f"{new_header}\n")
                line_width = 70
                for i in range(0, len(subsequence), line_width):
                    f.write(f"{subsequence[i:i + line_width]}\n")

            if is_out_of_bounds:
                messagebox.showwarning("Values Clamped",
                                       f"Warning: Your values were outside the valid range (1-{seq_len}).\n"
                                       f"The subsequence from {start_val} to {end_val} was exported to:\n{filepath}"
                                       )
            else:
                messagebox.showinfo("Success",
                                    f"Subsequence ({start_val}-{end_val}) saved to:\n{filepath}"
                                    )
        except Exception as e:
            messagebox.showerror("Error", f"Failed to write FASTA file: {e}")
            traceback.print_exc()

    def _on_region_value_change(self, event=None):
        """Automatically check the 'Limit Plot to Region' checkbox when user modifies region start or end values."""
        region_start = self.vars["region_start"].get().strip()
        region_end = self.vars["region_end"].get().strip()
        if region_start and region_start != "1": self.vars["limit_to_region"].set(True)
        elif region_end: self.vars["limit_to_region"].set(True)

    def _on_blast_task_change(self, event=None):
        """Updates BLAST parameter entry boxes to the defaults for the selected task and enables/disables task-specific options."""
        task = self.vars["blast_task"].get()
        if task in BLAST_DEFAULTS:
            defaults = BLAST_DEFAULTS[task]
            self.vars["blast_word_size"].set(defaults["word_size"])
            self.vars["blast_reward"].set(defaults["reward"])
            self.vars["blast_penalty"].set(defaults["penalty"])
            self.vars["blast_evalue"].set(defaults["evalue"])
            self.vars["blast_num_threads"].set(defaults["num_threads"])
        self._update_blast_param_states()

    def _update_blast_param_states(self):
        """Enable/disable BLAST parameters based on task selection."""
        pass

    def _enable_buttons(self):
        """Re-enables the main action buttons."""
        self.load_button.config(state="normal")
        self.save_button.config(state="normal")
        self.interactive_button.config(state="normal")

    def _on_main_resolution_change(self, event=None):
        if self.vars["save_resolution"].get() == "Custom Resolution":
            self.custom_res_frame_main.grid(row=1, column=1, columnspan=2, sticky="ew", pady=3, padx=5)
        else: self.custom_res_frame_main.grid_forget()

    def _on_mauve_param_change(self):
        if not self.is_loading_settings: self.vars["force_mauve"].set(True)

    def _check_for_cached_xmfa(self, *args):
        if self.is_loading_settings: return

        query_fasta_str, ref_fasta_str = self.vars["fasta"].get(), self.vars["ref"].get()
        q_valid = query_fasta_str.lower().endswith((".fasta", ".fa", ".fna", ".fsa"))
        r_valid = ref_fasta_str.lower().endswith((".fasta", ".fa", ".fna", ".fsa"))

        if not q_valid or not r_valid:
            self.vars["xmfa_file"].set("")
            return

        try:
            query_fasta_path = Path(query_fasta_str)
            ref_fasta_path = Path(ref_fasta_str)
            query_stem = query_fasta_path.stem
            ref_stem = ref_fasta_path.stem
            output_dir = query_fasta_path.parent / "mauve_alignments"
            xmfa_file_path = output_dir / f"{query_stem}_vs_{ref_stem}.xmfa"

            if xmfa_file_path.is_file(): self.vars["xmfa_file"].set(str(xmfa_file_path))
            else: self.vars["xmfa_file"].set("")

        except Exception as e:
            print(f"DEBUG: Error during cache check: {e}")
            self.vars["xmfa_file"].set("")

    def _check_for_cached_maf(self, *args):
        if self.is_loading_settings: return

        query_fasta_str, ref_fasta_str = self.vars["fasta"].get(), self.vars["ref"].get()
        q_valid = query_fasta_str.lower().endswith((".fasta", ".fa", ".fna", ".fsa"))
        r_valid = ref_fasta_str.lower().endswith((".fasta", ".fa", ".fna", ".fsa"))

        if not q_valid or not r_valid:
            self.vars["sibeliaz_maf"].set("")
            return

        try:
            query_fasta_path = Path(query_fasta_str)
            ref_fasta_path = Path(ref_fasta_str)
            query_stem = query_fasta_path.stem
            ref_stem = ref_fasta_path.stem
            output_dir = query_fasta_path.parent / "sibeliaz_alignments"
            maf_file_path = output_dir / f"{query_stem}_vs_{ref_stem}.maf"

            if maf_file_path.is_file(): self.vars["sibeliaz_maf"].set(str(maf_file_path))
            else: self.vars["sibeliaz_maf"].set("")

        except Exception as e:
            print(f"DEBUG: Error during SibeliaZ cache check: {e}")
            self.vars["sibeliaz_maf"].set("")

    def _load_settings(self):
        if not self.settings_file.exists():
            messagebox.showinfo("No Settings Found", "No previous settings file was found.")
            return

        self.is_loading_settings = True
        try:
            with open(self.settings_file, "r") as f:
                settings = json.load(f)
            saved_force_mauve = settings.pop("force_mauve", None)
            for key, value in settings.items():
                if key in self.vars: self.vars[key].set(value)
            self.is_loading_settings = False
            if not self.vars["xmfa_file"].get(): self._check_for_cached_xmfa()
            if not self.vars["sibeliaz_maf"].get(): self._check_for_cached_maf()
            self.is_loading_settings = True
            if saved_force_mauve is not None: self.vars["force_mauve"].set(saved_force_mauve)
            self.status_label.config(text="Loaded previous values.", fg="blue")
        except Exception as e: messagebox.showerror("Error", f"Could not load settings file: {e}")
        finally: self.is_loading_settings = False

    def _reset_to_defaults(self):
        mauve_keys = [
            "mauve_disable_backbone", "mauve_mums", "mauve_collinear",
            "mauve_skip_refinement", "mauve_skip_gapped_alignment",
            "mauve_seed_family", "mauve_no_weight_scaling", "mauve_mem_clean",
            "mauve_seed_weight", "mauve_max_gapped_aligner_length",
            "mauve_max_breakpoint_distance_scale", "mauve_conservation_distance_scale",
            "mauve_bp_dist_estimate_min_score", "mauve_gap_open", "mauve_gap_extend",
            "mauve_weight", "mauve_min_scaled_penalty",
            "mauve_hmm_p_go_homologous", "mauve_hmm_p_go_unrelated"
        ]

        mauve_has_changed = False
        for k in mauve_keys:
            current_val = self.vars[k].get()
            if str(current_val) != str(self.defaults[k]):
                mauve_has_changed = True
                break

        self.is_loading_settings = True
        try:
            for key, default_value in self.defaults.items():
                if key not in self.file_keys:
                    try: self.vars[key].set(default_value)
                    except Exception as e: print(f"Error resetting {key}: {e}")

            if mauve_has_changed: self.vars["force_mauve"].set(True)
            else: self.vars["force_mauve"].set(False)
            self._update_region_and_subsequence_defaults()
            self.status_label.config(text="Reset to default values (file directories preserved).", fg="blue")
        finally: self.is_loading_settings = False

    def _save_settings(self):
        settings = {key: var.get() for key, var in self.vars.items()}
        try:
            with open(self.settings_file, "w") as f: json.dump(settings, f, indent=4)
        except Exception as e: print(f"Error saving settings: {e}")

    def _update_region_and_subsequence_defaults(self):
        """Update region_start/end and subsequence_start/end based on query FASTA length."""
        fasta_path = self.vars["fasta"].get()
        if not fasta_path or not Path(fasta_path).is_file(): return

        try:
            seqs = list(SeqIO.parse(fasta_path, "fasta"))
            if not seqs: return
            seqlen = len(seqs[0].seq)

            self.vars["region_start"].set("1")
            self.vars["region_end"].set(str(seqlen))
            self.vars["subsequence_start"].set("1")
            self.vars["subsequence_end"].set(str(seqlen))
        except Exception as e: print(f"Error updating region/subsequence defaults: {e}")

    def _perform_plot_and_save(self, payload):
        data = payload["data"]
        plot_params = payload["plot_params"]
        out_png = payload["out_png"]
        self.status_label.config(text="Generating high-resolution plot and report...", fg="black")

        try:
            unique_plot_path = get_unique_filepath(Path(out_png))
            report_xlsx_path = unique_plot_path.with_suffix('.xlsx')
            resolution_key = str(self.vars["save_resolution"].get())

            if resolution_key == "Custom Resolution":
                try:
                    max_width = int(self.vars["custom_save_width"].get())
                    max_height = int(self.vars["custom_save_height"].get())
                    assert max_width > 0 and max_height > 0
                except:
                    self.status_label.config(text="Error: Custom width/height invalid.", fg="red")
                    self.log_queue.put("done", None)
                    return
            else: max_width, max_height = self.resolution_options[resolution_key]
            fig, ax = plt.subplots(figsize=(10, 8), dpi=100)
            plot_params['show_coords'] = self.vars['show_coords'].get()
            _draw_plot(ax, data, plot_params)
            _save_figure_with_aspect_ratio(fig, ax, unique_plot_path, max_width, max_height)
            plt.close(fig)
            self.status_label.config(text=f"Success! Plot saved to {unique_plot_path.name}", fg="green")
            if self.vars["open_image"].get():
                self.status_label.config(text=f"Opening {unique_plot_path.name}...", fg="blue")
                try:
                    if sys.platform == "win32": os.startfile(unique_plot_path)
                    elif sys.platform == "darwin": subprocess.call(["open", unique_plot_path])
                    else: subprocess.call(["xdg-open", unique_plot_path])
                except Exception as e: self.status_label.config(text=f"Error opening image: {e}", fg="orange")

            homology_hits = []
            for f in data["gene_feats"]:
                if f.label and f.color == '#43a047':
                    identity = getattr(f, 'best_identity', 0)
                    if plot_params['homology_thresh'] <= identity <= 100.0:
                        gene_name = _nice_label(f.label, 100)
                        # Extract gene ID from feature attributes
                        gene_id = "N/A"
                        attrs = getattr(f, "attributes", {}) or {}
                        gene_id_value = _get_attr_value(attrs, "ID")
                        if gene_id_value:
                            gene_id = gene_id_value
                        query_start = int(f.start)
                        query_stop = int(f.end)
                        ref_start = min(getattr(f, 'best_sstart', 0), getattr(f, 'best_send', 0))
                        ref_stop = max(getattr(f, 'best_sstart', 0), getattr(f, 'best_send', 0))
                        query_accession = getattr(f, 'best_qseqid', 'N/A')
                        reference_accession = getattr(f, 'best_sseqid', 'N/A')
                        homology_hits.append((gene_name, gene_id, query_start, query_stop, ref_start, ref_stop, f"{identity:.2f}", query_accession, reference_accession))
            homology_df = pd.DataFrame(sorted(homology_hits, key=lambda x: x[2]),
                                       columns=["Gene Name", "ID", "Query Start", "Query Stop", "Reference Start", "Reference Stop", "Identity (%)",
                                                "Query Accession", "Reference Accession"])

            enriched_blast_df = data.get("blast_df", pd.DataFrame())
            blast_all_df = _format_blast_df_for_excel(enriched_blast_df)
            blast_intergenic_df = _format_blast_df_for_excel(
                enriched_blast_df[~enriched_blast_df['is_coding']]
            )

            def process_coding_lcb_df(lcb_list):
                """Process coding LCB data with gene_name as first column."""
                df = pd.DataFrame(lcb_list)
                if not df.empty:
                    df = df.rename(
                        columns={
                            "gene_name": "Gene Name",
                            "gene_id": "Gene ID",
                            "query_start": "Query Start",
                            "query_end": "Query End",
                            "query_strand": "Query Strand",
                            "ref_start": "Ref Start",
                            "ref_stop": "Ref Stop",
                            "ref_strand": "Ref Strand",
                            "identity_excl_gaps": "Identity (%) Excl Gaps",
                            "identity_incl_gaps": "Identity (%) Incl Gaps",
                            "original_query_sequence": "Original Query Sequence",
                            "original_reference_sequence": "Original Reference Sequence",
                            "aligned_query_sequence": "Aligned Query Sequence",
                            "aligned_reference_sequence": "Aligned Reference Sequence",
                        }
                    )
                    if "Query Strand" in df.columns:
                        df["Query Strand"] = df["Query Strand"].replace(
                            {1: "+", -1: "-", "1": "+", "-1": "-"}
                        )
                    if "Ref Strand" in df.columns:
                        df["Ref Strand"] = df["Ref Strand"].replace(
                            {1: "+", -1: "-", "1": "+", "-1": "-"}
                        )
                    cols = [
                        "Gene Name",
                        "Gene ID",
                        "Query Start",
                        "Query End",
                        "Query Strand",
                        "Ref Start",
                        "Ref Stop",
                        "Ref Strand",
                        "Original Query Sequence",
                        "Original Reference Sequence",
                        "Identity (%) Excl Gaps",
                        "Identity (%) Incl Gaps",
                        "Aligned Query Sequence",
                        "Aligned Reference Sequence",
                    ]
                    if "difference_string" in df.columns:
                        df = df.rename(columns={
                            "difference_string": "Difference String",
                            "query_mismatches_marked": "Query with Mismatches (#)",
                            "query_differences_only": "Query Differences Only",
                            "ref_differences_only": "Reference Differences Only"})
                        cols.extend(["Difference String", "Query with Mismatches (#)", "Query Differences Only",
                                     "Reference Differences Only"])

                    final_cols = [c for c in cols if c in df.columns]
                    df = df[final_cols].copy()
                    df = df.sort_values(by="Query Start").reset_index(drop=True)
                return df

            def process_lcb_df(lcb_list):
                df = pd.DataFrame(lcb_list)
                if not df.empty:
                    df = df.rename(columns={
                        "query_start": "Query Start",
                        "query_end": "Query End",
                        "query_strand": "Query Strand",
                        "ref_start": "Ref Start",
                        "ref_stop": "Ref Stop",
                        "ref_strand": "Ref Strand",
                        "identity_excl_gaps": "Identity (%) Excl Gaps",
                        "identity_incl_gaps": "Identity (%) Incl Gaps",
                        "original_query_sequence": "Original Query Sequence",
                        "original_reference_sequence": "Original Reference Sequence",
                        "aligned_query_sequence": "Aligned Query Sequence",
                        "aligned_reference_sequence": "Aligned Reference Sequence",
                    })

                    if 'Query Strand' in df.columns:
                        df['Query Strand'] = df['Query Strand'].replace({1: '+', -1: '-', '1': '+', '-1': '-'})
                    if 'Ref Strand' in df.columns:
                        df['Ref Strand'] = df['Ref Strand'].replace({1: '+', -1: '-', '1': '+', '-1': '-'})

                    cols = [
                        "Query Start",
                        "Query End",
                        "Query Strand",
                        "Ref Start",
                        "Ref Stop",
                        "Ref Strand",
                        "Original Query Sequence",
                        "Original Reference Sequence",
                        "Identity (%) Excl Gaps",
                        "Identity (%) Incl Gaps",
                        "Aligned Query Sequence",
                        "Aligned Reference Sequence",
                    ]

                    if "difference_string" in df.columns:
                        df = df.rename(columns={
                            "difference_string": "Difference String",
                            "query_mismatches_marked": "Query with Mismatches (#)",
                            "query_differences_only": "Query Differences Only",
                            "ref_differences_only": "Reference Differences Only"})
                        cols.extend(["Difference String", "Query with Mismatches (#)", "Query Differences Only",
                                     "Reference Differences Only"])

                    final_cols = [c for c in cols if c in df.columns]
                    df = df[final_cols].copy()
                    df = df.sort_values(by="Query Start").reset_index(drop=True)
                return df

            mauve_lcb_df = process_lcb_df(data.get("mauve_data", {}).get("lcb_data", []))
            mauve_coding_lcb_df = process_coding_lcb_df(data.get("mauve_data", {}).get("coding_lcb_data", []))
            mauve_intergenic_lcb_df = process_lcb_df(data.get("mauve_data", {}).get("intergenic_lcb_data", []))
            sibeliaz_lcb_df = process_lcb_df(data.get("sibeliaz_data", {}).get("lcb_data", []))
            sibeliaz_coding_lcb_df = process_coding_lcb_df(data.get("sibeliaz_data", {}).get("coding_lcb_data", []))
            sibeliaz_intergenic_lcb_df = process_lcb_df(data.get("sibeliaz_data", {}).get("intergenic_lcb_data", []))
            combined_intergenic_data = []
            mauve_intergenic = data.get("mauve_data", {}).get("intergenic_lcb_data", [])
            for lcb in mauve_intergenic:
                lcb_copy = lcb.copy()
                lcb_copy['algorithm'] = 'Mauve'
                combined_intergenic_data.append(lcb_copy)
            sibeliaz_intergenic = data.get("sibeliaz_data", {}).get("intergenic_lcb_data", [])
            for lcb in sibeliaz_intergenic:
                lcb_copy = lcb.copy()
                lcb_copy['algorithm'] = 'SibeliaZ'
                combined_intergenic_data.append(lcb_copy)
            # Standardize BLAST column names to match Mauve/SibeliaZ before combining
            blast_intergenic_records = blast_intergenic_df.to_dict('records')
            for record in blast_intergenic_records:
                record['algorithm'] = 'BLAST'
                # Convert capitalized column names back to lowercase with underscores to match Mauve/SibeliaZ format
                if 'Original Query Sequence' in record:
                    record['original_query_sequence'] = record.pop('Original Query Sequence')
                if 'Original Reference Sequence' in record:
                    record['original_reference_sequence'] = record.pop('Original Reference Sequence')
                if 'Difference String' in record:
                    record['difference_string'] = record.pop('Difference String')
                if 'Query with Mismatches (#)' in record:
                    record['query_mismatches_marked'] = record.pop('Query with Mismatches (#)')
                if 'Query Differences Only' in record:
                    record['query_differences_only'] = record.pop('Query Differences Only')
                if 'Reference Differences Only' in record:
                    record['ref_differences_only'] = record.pop('Reference Differences Only')
                if 'Aligned Query Sequence' in record:
                    record['aligned_query_sequence'] = record.pop('Aligned Query Sequence')
                if 'Aligned Reference Sequence' in record:
                    record['aligned_reference_sequence'] = record.pop('Aligned Reference Sequence')
                combined_intergenic_data.append(record)
            include_legacy = self.vars["run_legacy_report_on_save"].get()
            legacy_lcb_df = pd.DataFrame()
            legacy_intergenic_lcb_df = pd.DataFrame()
            if include_legacy:
                _compare_legacy_to_wga(
                    data.get("legacy_data", {}).get("intergenic_lcb_data", []),
                    data.get("mauve_data", {}).get("intergenic_lcb_data", []),
                    data.get("sibeliaz_data", {}).get("intergenic_lcb_data", [])
                )
                legacy_lcb_df = process_lcb_df(data.get("legacy_data", {}).get("lcb_data", []))
                legacy_intergenic_lcb_df = process_lcb_df(
                    data.get("legacy_data", {}).get("intergenic_lcb_data", []))
                if not legacy_intergenic_lcb_df.empty:
                    legacy_intergenic_lcb_df['sequence_length'] = legacy_intergenic_lcb_df[
                        'aligned_query_sequence'].str.len()
                    legacy_intergenic_lcb_df = legacy_intergenic_lcb_df.sort_values(
                        by=["sequence_length", "query_start"], ascending=[False, True]).reset_index(drop=True)
                legacy_intergenic = data.get("legacy_data", {}).get("intergenic_lcb_data", [])
                for lcb in legacy_intergenic:
                    seq_len = len(lcb.get('aligned_query_sequence', ''))
                    if seq_len < 15: continue
                    lcb_copy = lcb.copy()
                    lcb_copy['algorithm'] = 'Legacy k-mer'
                    combined_intergenic_data.append(lcb_copy)
            combined_intergenic_df = pd.DataFrame(combined_intergenic_data)
            if not combined_intergenic_df.empty:
                # Single rename to standardize column names from all sources
                combined_intergenic_df = combined_intergenic_df.rename(columns={
                    "algorithm": "Algorithm",
                    "query_start": "Query Start",
                    "query_end": "Query End",
                    "query_strand": "Query Strand",
                    "ref_start": "Ref Start",
                    "ref_stop": "Ref Stop",
                    "ref_strand": "Ref Strand",
                    "identity_excl_gaps": "Identity (%) Excl Gaps",
                    "identity_incl_gaps": "Identity (%) Incl Gaps",
                    "original_query_sequence": "Original Query Sequence",
                    "original_reference_sequence": "Original Reference Sequence",
                    "aligned_query_sequence": "Aligned Query Sequence",
                    "aligned_reference_sequence": "Aligned Reference Sequence",
                })
                if 'Query Strand' in combined_intergenic_df.columns:
                    combined_intergenic_df['Query Strand'] = combined_intergenic_df['Query Strand'].replace(
                        {1: '+', -1: '-', '1': '+', '-1': '-'})
                if 'Ref Strand' in combined_intergenic_df.columns:
                    combined_intergenic_df['Ref Strand'] = combined_intergenic_df['Ref Strand'].replace(
                        {1: '+', -1: '-', '1': '+', '-1': '-'})
                cols = ["Algorithm", "Query Start", "Query End", "Query Strand", "Ref Start", "Ref Stop", "Ref Strand",
                        "Identity (%) Excl Gaps", "Identity (%) Incl Gaps",
                        "Original Query Sequence", "Original Reference Sequence",
                        "Aligned Query Sequence",
                        "Aligned Reference Sequence",
                        "Difference String", "Query with Mismatches (#)", "Query Differences Only",
                        "Reference Differences Only"]
                final_cols = [c for c in cols if c in combined_intergenic_df.columns]
                combined_intergenic_df = combined_intergenic_df[final_cols].copy()
                # Remove duplicate columns if any
                combined_intergenic_df = combined_intergenic_df.loc[:, ~combined_intergenic_df.columns.duplicated()]
                if not combined_intergenic_df.empty:
                    combined_intergenic_df = combined_intergenic_df.sort_values(by="Query Start").reset_index(drop=True)

                seqlen = data["seqlen"]
                gc_window_size = data.get("gc_window_size", 500)

                if data.get("gc_arr") is not None:
                    gc_vals = data["gc_arr"][::gc_window_size] * 100.0  # Convert to %
                    skew_vals = data["skew_arr"][::gc_window_size]

                    n_windows = len(gc_vals)
                    starts = np.arange(n_windows) * gc_window_size + 1
                    ends = np.minimum(starts + gc_window_size - 1, seqlen)

                    gc_full_df = pd.DataFrame({
                        "Window Start": starts,
                        "Window End": ends,
                        "GC Content (%)": gc_vals,
                        "GC Skew": skew_vals
                    })
                else:
                    # Fallback if no data exists
                    gc_full_df = pd.DataFrame(columns=["Window Start", "Window End", "GC Content (%)", "GC Skew"])
            gc_skew_df = gc_full_df[["Window Start", "Window End", "GC Skew"]]
            gc_content_df = gc_full_df[["Window Start", "Window End", "GC Content (%)"]]
            MIN_IDENTITY_THRESHOLD = 60.0
            identity_col = "Identity (%) Excl Gaps"

            def filter_by_identity(df, threshold=MIN_IDENTITY_THRESHOLD, identity_column=identity_col):
                """Filter dataframe to keep only rows with identity >= threshold."""
                if identity_column in df.columns:
                    rows_before = len(df)
                    df_filtered = df[df[identity_column] >= threshold].reset_index(drop=True)
                    rows_after = len(df_filtered)
                    if rows_before > rows_after:
                        print(
                            f"[IDENTITY FILTER] Filtered from {rows_before} to "
                            f"{rows_after} rows (removed {rows_before - rows_after} rows with <{threshold}% identity)")
                    return df_filtered
                return df

            print(
                f"\n[EXCEL EXPORT] Applying minimum {MIN_IDENTITY_THRESHOLD}% identity filter to all alignment sheets...")
            homology_df = filter_by_identity(homology_df)
            blast_all_df = filter_by_identity(blast_all_df)
            blast_intergenic_df = filter_by_identity(blast_intergenic_df)
            mauve_lcb_df = filter_by_identity(mauve_lcb_df)
            mauve_coding_lcb_df = filter_by_identity(mauve_coding_lcb_df)
            mauve_intergenic_lcb_df = filter_by_identity(mauve_intergenic_lcb_df)
            sibeliaz_lcb_df = filter_by_identity(sibeliaz_lcb_df)
            sibeliaz_coding_lcb_df = filter_by_identity(sibeliaz_coding_lcb_df)
            sibeliaz_intergenic_lcb_df = filter_by_identity(sibeliaz_intergenic_lcb_df)
            legacy_lcb_df = filter_by_identity(legacy_lcb_df)
            legacy_intergenic_lcb_df = filter_by_identity(legacy_intergenic_lcb_df)
            combined_intergenic_df = filter_by_identity(combined_intergenic_df)

            with pd.ExcelWriter(report_xlsx_path) as writer:
                homology_df.to_excel(writer, sheet_name="Homology Report (Genes)", index=False)
                blast_all_df.to_excel(writer, sheet_name="Alignment Blocks (BLAST)", index=False)
                blast_intergenic_df.to_excel(writer, sheet_name="Intergenic Blocks (BLAST)", index=False)
                mauve_lcb_df.to_excel(writer, sheet_name="Alignment Blocks (Mauve)", index=False)
                mauve_coding_lcb_df.to_excel(writer, sheet_name="Coding Blocks (Mauve)", index=False)
                mauve_intergenic_lcb_df.to_excel(writer, sheet_name="Intergenic Blocks (Mauve)", index=False)
                sibeliaz_lcb_df.to_excel(writer, sheet_name="Alignment Blocks (SibeliaZ)", index=False)
                sibeliaz_coding_lcb_df.to_excel(writer, sheet_name="Coding Blocks (SibeliaZ)", index=False)
                sibeliaz_intergenic_lcb_df.to_excel(writer, sheet_name="Intergenic Blocks (SibeliaZ)", index=False)
                if include_legacy:
                    legacy_lcb_df.to_excel(writer, sheet_name="Alignment Blocks (Legacy)", index=False)
                    legacy_intergenic_lcb_df.to_excel(writer, sheet_name="Intergenic Blocks (Legacy)", index=False)
                combined_intergenic_df.to_excel(writer, sheet_name="Combined Intergenic Blocks", index=False)
                gc_skew_df.to_excel(writer, sheet_name="GC Skew", index=False)
                gc_content_df.to_excel(writer, sheet_name="GC Content", index=False)

            # Adjust column widths
            _adjust_excel_column_widths(report_xlsx_path)
        except Exception as e:
            self.status_label.config(text=f"Error during plot/report generation: {e}", fg="red")
            traceback.print_exc()
        finally: self.log_queue.put(("done", None))

    def process_queue(self):
        try:
            while True:
                msg_type, data = self.log_queue.get_nowait()

                if msg_type == "status":
                    msg, color = data
                    self.status_label.config(text=msg, fg=color)
                    self.status_label.update_idletasks()

                elif msg_type == "ask_yes_no":
                    # Unpack the request
                    message, event, result_container = data

                    # Show the modal dialog
                    response = messagebox.askyesno("Confirm", message)

                    # Store result and wake up the background thread
                    result_container["result"] = response
                    event.set()

                elif msg_type == "plot_and_save":
                    self._perform_plot_and_save(data)

                elif msg_type == "launch_interactive":
                    self.status_label.config(text="Launching interactive plot window...", fg="blue")
                    try:
                        self.current_plot_window = PlottingWindow(self, data)
                        self.status_label.config(text="Interactive plot window opened.", fg="green")
                    except Exception as e:
                        self.status_label.config(text=f"Error launching interactive window: {e}", fg="red")
                        traceback.print_exc()
                        self._enable_buttons()

                elif msg_type == "reprocess_done":
                    if self.current_plot_window and self.current_plot_window.winfo_exists():
                        new_data = data
                        if hasattr(self.current_plot_window, 'data'):
                            old_labels = self.current_plot_window.data.get('custom_labels', [])
                            new_data['custom_labels'] = old_labels
                        else:
                            new_data['custom_labels'] = []
                        self.current_plot_window.data = new_data
                        self.current_plot_window.last_region_limited = None
                        self.current_plot_window.last_region_start = None
                        self.current_plot_window.last_region_end = None
                        self.current_plot_window.last_rotation = None
                        self.current_plot_window.initial_load_complete = False
                        self.current_plot_window._enable_controls()
                        self.current_plot_window.update_plot()
                        self.status_label.config(text="Interactive plot updated with new BLAST results.", fg="blue")
                    else:
                        self.status_label.config(text="Reprocessing done, but plot window was closed.", fg="grey")

                elif msg_type == "reprocess_fail":
                    if self.current_plot_window and self.current_plot_window.winfo_exists():
                        self.current_plot_window._enable_controls()
                    self.status_label.config(text=f"Reprocessing failed: {data}", fg="red")

                elif msg_type == "gc_recalc_done":
                    if self.current_plot_window and self.current_plot_window.winfo_exists():
                        new_gc_arr, new_skew_arr = data
                        self.current_plot_window.data["gc_arr"] = new_gc_arr
                        self.current_plot_window.data["skew_arr"] = new_skew_arr
                        self.current_plot_window.data_params["window"] = self.current_plot_window.plot_params[
                            "gc_window_size"].get()
                        if "auto_thickness_rings" in self.current_plot_window.plot_params:
                            self.current_plot_window.plot_params["auto_thickness_rings"].set(True)
                        self.current_plot_window._enable_controls()
                        self.current_plot_window.update_plot()
                        self.status_label.config(text="GC/Skew rings updated.", fg="blue")
                    else:
                        self.status_label.config(text="GC recalc done, but plot window was closed.", fg="grey")

                elif msg_type == "run_legacy_for_export":
                    filepath = data

                    def legacy_export_thread():
                        try:
                            if not self.current_plot_window or not self.current_plot_window.winfo_exists():
                                self.log_queue.put(("status", ("Plot window closed, aborting report.", "orange")))
                                return
                            self.log_queue.put(("status", ("Running legacy k-mer report (can be slow)...", "orange")))
                            thread_data = self.current_plot_window.data
                            query_seq = str(thread_data["record"].seq).upper()
                            ref_records_dict_thread = thread_data["ref_records_dict"]
                            gene_features = thread_data["gene_feats"]
                            mauve_lcb_data = thread_data.get("mauve_data", {}).get("lcb_data", [])
                            sibeliaz_lcb_data = thread_data.get("sibeliaz_data", {}).get("lcb_data", [])
                            MIN_K_FOR_REPORT = 15
                            max_k_mauve = _get_max_exact_match([mauve_lcb_data])
                            max_k_sibeliaz = _get_max_exact_match([sibeliaz_lcb_data])
                            max_k = max(max_k_mauve, max_k_sibeliaz, MIN_K_FOR_REPORT)
                            max_k = min(max_k, 200)
                            all_blocks, intergenic_blocks = generate_hierarchical_legacy_report(
                                query_seq, ref_records_dict_thread, gene_features, max_k, MIN_K_FOR_REPORT,
                                self.log_queue
                            )
                            payload = (filepath, all_blocks, intergenic_blocks)
                            self.log_queue.put(("legacy_export_ready", payload))
                        except Exception as e:
                            self.log_queue.put(("status", (f"Background report failed: {e}", "red")))
                            traceback.print_exc()
                            self.log_queue.put(("legacy_export_ready", (filepath, None, None)))

                    threading.Thread(target=legacy_export_thread, daemon=True).start()

                elif msg_type == "legacy_export_ready":
                    filepath, all_blocks, intergenic_blocks = data
                    if not self.current_plot_window or not self.current_plot_window.winfo_exists():
                        self.status_label.config(text="Legacy report finished, but plot window was closed.", fg="grey")
                        continue
                    if all_blocks is None:
                        self.status_label.config(text="Legacy report failed. Check console.", fg="red")
                        messagebox.showerror("Error", "Legacy k-mer report generation failed.")
                        if self.current_plot_window.export_button:
                            self.current_plot_window.export_button.config(state="normal")
                        continue
                    self.status_label.config(text="Legacy report complete. Saving Excel file...", fg="blue")
                    self.current_plot_window.data["legacy_data"]["lcb_data"] = all_blocks
                    self.current_plot_window.data["legacy_data"]["intergenic_lcb_data"] = intergenic_blocks
                    self.current_plot_window._perform_excel_save(filepath, include_legacy=True)

                elif msg_type == "legacy_kmer_recalc_done":
                    if not self.current_plot_window or not self.current_plot_window.winfo_exists():
                        self.status_label.config(text="Legacy k-mer recalc done, but plot window closed.", fg="grey")
                        continue
                    coding_scores, non_coding_scores, global_avg_id = data
                    plot_data = self.current_plot_window.data
                    plot_data["legacy_data"]["coding_scores"] = coding_scores
                    plot_data["legacy_data"]["non_coding_scores"] = non_coding_scores
                    plot_data["legacy_data"]["global_wga_id"] = global_avg_id
                    plot_data["legacy_data"]["noncoding_avg_id"] = global_avg_id
                    plot_data["coding_scores"] = coding_scores
                    plot_data["non_coding_scores"] = non_coding_scores
                    plot_data["global_wga_id"] = global_avg_id
                    plot_data["noncoding_avg_id"] = global_avg_id
                    plot_data["identity_algorithm_used"] = "Legacy (Global k-mer %)"
                    self.current_plot_window._enable_controls()
                    self.current_plot_window.update_plot()
                    self.status_label.config(text="Switched to Legacy (Global k-mer %).", fg="blue")

                elif msg_type == "legacy_base_recalc_done":
                    if not self.current_plot_window or not self.current_plot_window.winfo_exists():
                        self.status_label.config(text="Legacy base recalc done, but plot window closed.", fg="grey")
                        continue
                    coding_scores, non_coding_scores, global_avg_id = data
                    plot_data = self.current_plot_window.data
                    plot_data["legacy_base_data"]["coding_scores"] = coding_scores
                    plot_data["legacy_base_data"]["non_coding_scores"] = non_coding_scores
                    plot_data["legacy_base_data"]["global_wga_id"] = global_avg_id
                    plot_data["legacy_base_data"]["noncoding_avg_id"] = global_avg_id
                    plot_data["coding_scores"] = coding_scores
                    plot_data["non_coding_scores"] = non_coding_scores
                    plot_data["global_wga_id"] = global_avg_id
                    plot_data["noncoding_avg_id"] = global_avg_id
                    plot_data["identity_algorithm_used"] = "Legacy (1:1 Base %)"
                    self.current_plot_window._enable_controls()
                    self.current_plot_window.update_plot()
                    self.status_label.config(text="Switched to Legacy (1:1 Base %).", fg="blue")

                elif msg_type == "legacy_recalc_fail":
                    if self.current_plot_window and self.current_plot_window.winfo_exists():
                        self.current_plot_window._enable_controls()
                    self.status_label.config(text="Legacy (k-mer) recalculation failed.", fg="red")

                elif msg_type == "done":
                    self._enable_buttons()

        except queue.Empty:
            pass
        finally:
            if self.winfo_exists():
                self.after(100, self.process_queue)

    def _apply_gc_color_logic(self):
        """If GC color-only is checked and the colormap is Grey, switch it to Coolwarm."""
        if self.vars["gc_ring_color_only"].get() and self.vars["gc_colormap"].get() == "Grey":
            self.vars["gc_colormap"].set("Coolwarm")

    def _update_mauve_field_states(self):
        """Enable/disable Mauve fields based on checkbox states."""
        no_weight_scaling = self.vars["mauve_no_weight_scaling"].get()
        field_to_disable = "mauve_max_breakpoint_distance_scale"

        if field_to_disable in self.mauve_entry_widgets:
            widget = self.mauve_entry_widgets[field_to_disable]
            widget.config(state="disabled" if no_weight_scaling else "normal")
            if no_weight_scaling: widget.config(foreground="gray")
            else: widget.config(foreground="black")

    def _browse_file(self, var):
        fn = filedialog.askopenfilename()
        if fn: var.set(fn)

    def _browse_save_as(self, var):
        fn = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[
                ("PNG", "*.png"),
                ("TIFF", "*.tiff"),
                ("SVG", "*.svg"),
                ("JPEG", "*.jpg"),
                ("PDF", "*.pdf"),
                ("EPS", "*.eps"),
                ("All files", "*.*")
            ]
        )
        if fn: var.set(fn)

    def _validate_inputs(self):
        """
        Validates existence and file extensions for all selected files.
        Returns True if valid, False (and shows popup) if invalid.
        """
        validation_rules = {
            "fasta": ({".fasta", ".fa", ".fna", ".fsa"}, "Query Plasmid FASTA"),
            "ref": ({".fasta", ".fa", ".fna", ".fsa"}, "Reference FASTA"),
            "annot": ({".gff", ".gff3", ".gb", ".gbk"}, "Annotation (GFF/GBK)"),
            "bed": ({".bed"}, "Highlight Regions (.bed)"),
            "xmfa_file": ({".xmfa"}, "Mauve Alignment (.xmfa)"),
            "sibeliaz_maf": ({".maf"}, "SibeliaZ Alignment (.maf)")
        }

        # 1. Check Required Files (always needed)
        required_keys = ["fasta", "annot", "ref"]
        for key in required_keys:
            path_str = self.vars[key].get().strip().replace('\\', '/')
            if not path_str:
                display_name = validation_rules[key][1]
                messagebox.showerror("Missing File", f"Please select a file for: {display_name}")
                return False

        # 2. Check All Files (Optional + Required)
        for key, (valid_exts, display_name) in validation_rules.items():

            # If "Force New Alignment" is checked, skip validating the Mauve file box
            if key == "xmfa_file" and self.vars["force_mauve"].get():
                continue

            # If "Force New SibeliaZ" is checked, skip validating the MAF file box
            if key == "sibeliaz_maf" and self.vars["force_sibeliaz"].get():
                continue

            # Fix slashes for Mac/Linux compatibility
            path_str = self.vars[key].get().strip().replace('\\', '/')

            if not path_str: continue

            path = Path(path_str)
            if not path.is_file():
                messagebox.showerror("File Not Found", f"The file for {display_name} does not exist:\n\n{path_str}")
                return False

            current_ext = path.suffix.lower()
            if current_ext not in valid_exts:
                expected_str = ", ".join(sorted(list(valid_exts)))
                messagebox.showerror(
                    "Invalid File Type",
                    f"The file selected for '{display_name}' appears to be the wrong type.\n\n"
                    f"Selected File: {path.name}\n"
                    f"Detected Type: {current_ext}\n"
                    f"Expected Types: {expected_str}\n\n"
                    f"Please check that you haven't swapped files (e.g. FASTA in the GFF field)."
                )
                return False

        return True

    def _get_data_params(self):
        return {
            "fasta": self.vars["fasta"].get(),
            "annot": self.vars["annot"].get(),
            "ref_fasta": self.vars["ref"].get(),
            "window": self.vars["gc_window_size"].get(),
            "highlight_bed_path": self.vars["bed"].get() or None,
            "allow_code_fallback": self.vars["allow_code_fallback"].get(),
            "homology_threshold": self.vars["homology_thresh"].get(),
            "manual_find": self.vars["manual_find"].get(),
            "manual_replace": self.vars["manual_replace"].get(),
            "nc_window_size": self.vars["nc_window_size"].get(),
            "nc_step_size": self.vars["nc_step_size"].get(),
            "blast_task": self.vars["blast_task"].get(),
            "blast_word_size": self.vars["blast_word_size"].get(),
            "blast_reward": self.vars["blast_reward"].get(),
            "blast_penalty": self.vars["blast_penalty"].get(),
            "blast_evalue": self.vars["blast_evalue"].get(),
            "blast_num_threads": self.vars["blast_num_threads"].get(),
            "identity_algorithm": self.vars["identity_algorithm"].get(),
            "xmfa_file": self.vars["xmfa_file"].get(),
            "force_mauve": self.vars["force_mauve"].get(),
            "sibeliaz_maf": self.vars["sibeliaz_maf"].get(),
            "force_sibeliaz": self.vars["force_sibeliaz"].get(),
            "run_legacy_report": self.vars["run_legacy_report_on_save"].get(),
            "mauve_disable_backbone": self.vars["mauve_disable_backbone"].get(),
            "mauve_mums": self.vars["mauve_mums"].get(),
            "mauve_collinear": self.vars["mauve_collinear"].get(),
            "mauve_skip_refinement": self.vars["mauve_skip_refinement"].get(),
            "mauve_skip_gapped_alignment": self.vars["mauve_skip_gapped_alignment"].get(),
            "mauve_seed_family": self.vars["mauve_seed_family"].get(),
            "mauve_no_weight_scaling": self.vars["mauve_no_weight_scaling"].get(),
            "mauve_mem_clean": self.vars["mauve_mem_clean"].get(),
            "mauve_seed_weight": self.vars["mauve_seed_weight"].get(),
            "mauve_max_gapped_aligner_length": self.vars["mauve_max_gapped_aligner_length"].get(),
            "mauve_max_breakpoint_distance_scale": self.vars["mauve_max_breakpoint_distance_scale"].get(),
            "mauve_conservation_distance_scale": self.vars["mauve_conservation_distance_scale"].get(),
            "mauve_bp_dist_estimate_min_score": self.vars["mauve_bp_dist_estimate_min_score"].get(),
            "mauve_gap_open": self.vars["mauve_gap_open"].get(),
            "mauve_gap_extend": self.vars["mauve_gap_extend"].get(),
            "mauve_weight": self.vars["mauve_weight"].get(),
            "mauve_min_scaled_penalty": self.vars["mauve_min_scaled_penalty"].get(),
            "mauve_hmm_p_go_homologous": self.vars["mauve_hmm_p_go_homologous"].get(),
            "mauve_hmm_p_go_unrelated": self.vars["mauve_hmm_p_go_unrelated"].get(),
        }

    def _get_plot_params(self):
        params = {key: var.get() for key, var in self.vars.items() if
                  key not in ["fasta", "annot", "ref", "out", "bed", "win", "manual_find", "manual_replace",
                              "xmfa_file", "force_mauve", "sibeliaz_maf"]}
        params['homology_thresh_min'] = params['homology_thresh']
        params['zoom'] = 1

        # Adjust thickness/radius values when color-only mode is enabled
        if params.get("gc_ring_color_only", False):
            params["gc_thick"] = 0.15
            params["gc_inner"] = 0.15

        if params.get("id_ring_color_only", False):
            params["id_ring_radius"] = 0.7

        return params

    def _run_save(self):
        if not self._validate_inputs(): return
        self._apply_gc_color_logic()
        self._save_settings()
        self.save_button.config(state="disabled")
        self.interactive_button.config(state="disabled")
        self.status_label.config(text="Starting data processing...")
        data_params = self._get_data_params()
        plot_params = self._get_plot_params()
        threading.Thread(target=_run_background_processing,
                         args=(data_params, plot_params, self.vars["out"].get(), self.log_queue), daemon=True).start()

    def _run_interactive(self):
        if not self._validate_inputs(): return
        self._apply_gc_color_logic()
        self._save_settings()
        self.save_button.config(state="disabled")
        self.interactive_button.config(state="disabled")
        self.status_label.config(text="Processing data for interactive plot.....")
        data_params = self._get_data_params()
        data_params['run_legacy_report'] = False
        plot_params = self._get_plot_params()

        def process_and_launch():
            try:
                try:
                    data = _process_data(**data_params, log_queue=self.log_queue)
                    # Pass the precise_mode setting to the plot window
                    data["id_ring_precise_mode"] = self.vars["id_ring_precise_mode"].get()
                    data["initial_plot_params"] = plot_params
                    self.log_queue.put(("launch_interactive", data))
                except Exception as e_process:
                    # Check for cancellation before logging as error
                    if "cancelled by user" in str(e_process):
                        self.log_queue.put(("status", ("Processing cancelled by user.", "blue")))
                        self.log_queue.put(("done", None))
                        return

                    print(f"!!! ERROR DURING _process_data: {e_process}")
                    import traceback
                    traceback.print_exc()
                    # Optionally send error back to GUI
                    self.log_queue.put(("status", (f"Error during data processing: {e_process}", "red")))
                    self.log_queue.put(("done", None))  # Make sure buttons re-enable

            except Exception as e:  # Keep outer exception handling
                self.log_queue.put(("status", (f"Error: {e}", "red")))
                self.log_queue.put(("done", None))
                traceback.print_exc()

        threading.Thread(target=process_and_launch, daemon=True).start()

# ──────────────────────────────────────────────────────────────────────────────
# Main Execution
# ──────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = App()
    try: app.mainloop()
    finally:
        if hasattr(app, 'original_stdout') and app.original_stdout: sys.stdout = app.original_stdout
        if hasattr(app, 'original_stderr') and app.original_stderr: sys.stderr = app.original_stderr
